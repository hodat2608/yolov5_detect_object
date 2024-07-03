
# import openpyxl
 
# # Define variable to load the dataframe
# dataframe = openpyxl.load_workbook("excel/2.xlsx")
 
# # Define variable to read sheet
# dataframe1 = dataframe.active
 
# # Iterate the loop to read the cell values
# # for row in range(0, dataframe1.max_row):
# #     for col in dataframe1.iter_cols(1, dataframe1.max_column):
# #         print(col[row].value)
        
# # for row in range(0, dataframe1.max_row):
# #     for col in dataframe1.iter_cols(1, 1):
# #         print(col[row].value)
        

# print(dataframe1.max_row)



# max = dataframe1.max_row
# for row, entry in enumerate(data1, start=1):
#    st.cell(row=row+max, column=1, value=entry)
# # from openpyxl.workbook import Workbook

# # headers       = ['Company','Address','Tel','Web']
# # workbook_name = 'sample.xlsx'
# # wb = Workbook()
# # page = wb.active
# # page.title = 'companies'
# # page.append(headers) # write the headers to the first line

# # # Data to write:
# # companies = [['name1','address1','tel1','web1'], ['name2','address2','tel2','web2']]

# # for info in companies:
# #     page.append(info)
# # wb.save(filename = workbook_name)


import openpyxl
import datetime

from openpyxl.styles import Alignment
from openpyxl import Workbook
from datetime import date
from openpyxl.styles import Font

today = date.today()
d1 = today.strftime("%d/%m/%Y")
now = datetime.datetime.now()
t1 = now.strftime("%H:%M:%S")
#wb = Workbook()
#ws = wb.active
#ws['A2'] = datetime.datetime.now()
# ws['B2'].font = Font(name= 'Calibri',
#                      size=40,
#                      bold=True,
#                      italic=False,
#                      strike=False,
#                      underline='none',
#                      color='4472C4'
#                      )



wb = openpyxl.load_workbook("excel/2.xlsx")
ws = wb.active
col1 = d1
col2 = t1
col3 = '3'
col4 = '4'
col5 = '5'
col6 = '6'
col7 = '7'
col8 = '8'
col9 = '9'
col10 = '10'
col11 = '11'
col12 = '24'
ws.append([col1, col2,col3, col4,col5, col6,col7, col8,col9, col10,col11, col12,])

for row in range(ws.max_row, ws.max_row+1):
    for col in range(1, ws.max_column+1):
        #print(col[row].value)
        d = ws.cell(row = row, column = col)
        #currentCell = ws.cell(col[row])
        d.alignment = Alignment(horizontal='center')
        #d.style.alignment.horizontal = 'center'
        d.font = Font(name= 'Calibri', size=12)
wb.save("excel/2.xlsx")

# currentCell = ws.cell('A1') #or currentCell = ws['A1']
# currentCell.alignment = Alignment(horizontal='center')
