from glob import glob
import os, cv2, torch, time, datetime, shutil
import numpy as np 
import pandas as pd
import PySimpleGUI as sg
from PIL import Image, ImageTk
import connect_PLC_Mitsubishi as plc
import traceback
import sqlite3
import openpyxl 
from datetime import date
from openpyxl.styles import Alignment,Font
import socket
import matplotlib.pyplot as plt
soc = socket.socket(socket.AF_INET, socket.SOCK_STREAM)


mysleep = 0.1
localtion_1 = [0.493750, 0.515833, 0.365833, 0.150000]

SCALE_X_CAM1 = 640*1.2/2048
SCALE_Y_CAM1 = 480*1.2/1536

SCALE_X_CAM2 = 640/1440
SCALE_Y_CAM2 = 480/1080


def save_values(values_1,values_2,values_3):
    with open('static/coordinate_filtering.txt', 'w') as file:
        file.write(f'width={values_1}\nheight={values_2}\nconfidence={values_3}')

def load_values():
    result = []
    with open('static/coordinate_filtering.txt') as lines:
        for line in lines:
            _, value = line.strip().split('=')
            result.append(int(value.strip()))
    return result

def check_overlap(label_1):
    x = float(label_1[0]) 
    y = float(label_1[1]) 
    w = float(label_1[2]) 
    h = float(label_1[3]) 
    t1 = x - w/2    
    t2 = y - h/2
    p1 = x + w/2     
    p2 = y + h/2
    return t1,t2,p1,p2,h,y

def check_overload(x, y, z, v, h, p, _x_px,_y_py):
    x_bc = float(_x_px[0])
    y_bc = float(_y_py[0])
    e = 0.632500 -  0.356250
    ds = 0.777083-0.255417
    i = 0.0001
    x_min = x
    x_max = z
    z_min = x_min -(e)
    z_max = x_max -(e)
    x_values = [x]  
    y_values = [y]  
    x1_values = [] 
    y1_values = [] 
    _z_value = []
    _u_value = []
    _z1_value = []
    _u1_value =[]
    k = 0.777083 
    my_result = 0
    while x < z and y < v:
        a = 0
        b = 0
        a +=3*i
        b +=i
        _z = (x - (e) + a)
        _u = (y + (ds) + b)
        _u1 = _u +(h/2)
        x = x + a
        y = y + b
        y1= y + (h/2)
        x_values.append(x)
        y_values.append(y)
        x1_values.append(x)
        y1_values.append(y1)
        _z_value.append(_z)
        _u_value.append(_u)
        _z1_value.append(_z)
        _u1_value.append(_u1)
        if ((float(p) < float(y_bc) < float(y1) and x_max > float(x_bc) > float(x)) or (float(y) < float(y_bc) < float(p) and x_min <float(x_bc) < float(x)) or 
            (float(k) < float(y_bc) < float(_u1) and z_max > float(x_bc) > float(_z)) or (float(_u) < float(y_bc) < float(k) and z_min <float(x_bc) < float(_z))):
            my_result = 1
            break      
        else:  
            my_result = 0       
    if my_result == 1:     
        return True
    else :        
        return False
   
label_1 = [0.642917, 0.235833, 0.500833, 0.288333]

def connect_plc_keyence(host, port):
    try:
        soc.connect((host, port))
        return True
    except OSError:
        print("Can't connect to PLC")
        time.sleep(3)
        print("Reconnecting....")
        return False

def run_plc_keyence(host, port):
    connected = False
    while connected == False:
        connected = connect_plc_keyence(host, port)
    print("connected") 

def read_plc_keyence(data):
    data = f'DM{data}'
    a = 'RD '
    c = '\x0D'
    d = a+ data +c
    datasend = d.encode("UTF-8")
    soc.sendall(datasend)
    data = soc.recv(1024)
    datadeco = data.decode("UTF-8")
    data1 = int(datadeco)

    return data1

#Write data
def write_plc_keyence(register, data):
    a = 'WR '
    b = ' '
    c = '\x0D'
    d = a+ f'DM{register}' + b + str(data) + c
    datasend  = d.encode("UTF-8")
    soc.sendall(datasend)
    datares = soc.recv(1024)
    #print(datares)


def removefile_1():
    os.system('rd /s /q C:\FH\camera1')
    os.system('rd /s /q C:\FH\camera2')
    print('Deleted CAM1-CAM2 Folders')

def removefile():
    directory1 = 'C:/FH/camera1/**/*'
    directory2 = 'C:/FH/camera2/**/*'
    chk1 = glob(directory1)
    for f1 in chk1:
        fname1=os.path.dirname(f1)
        shutil.rmtree(fname1)
        print('already delete folder 1')
    chk2 = glob(directory2)
    for f2 in chk2:
        fname2=os.path.dirname(f2)
        shutil.rmtree(fname2)
        print('already delete folder 2')
def removefile_cam1():
    directory1 = 'C:/FH/camera1/**/*'
    chk1 = glob(directory1)
    for f1 in chk1:
        fname1=os.path.dirname(f1)
        shutil.rmtree(fname1)
        print('already delete folder 1')
 
def removefile_cam2():
    directory2 = 'C:/FH/camera2/**/*'

    chk2 = glob(directory2)
    for f2 in chk2:
        fname2=os.path.dirname(f2)
        shutil.rmtree(fname2)
        print('already delete folder 2')

def time_to_name():
    current_time = datetime.datetime.now() 
    name_folder = str(current_time)
    name_folder = list(name_folder)
    for i in range(len(name_folder)):
        if name_folder[i] == ':':
            name_folder[i] = '-'
        if name_folder[i] == ' ':
            name_folder[i] ='_'
        if name_folder[i] == '.':
            name_folder[i] ='-'
    name_folder = ''.join(name_folder)
    return name_folder

def load_theme():
    name_themes = []
    with open('static/theme.txt') as lines:
        for line in lines:
            _, name_theme = line.strip().split(':')
            name_themes.append(name_theme)
    return name_themes

def load_choosemodel():
    with open('static/choose_model.txt') as lines:
        for line in lines:
            _, name_model = line.strip().split('=')
    return name_model

def save_theme(name_theme):
    line = 'theme:' + name_theme
    with open('static/theme.txt','w') as f:
        f.write(line)

def save_choosemodel(name_model):
    line = 'choose_model=' + name_model
    with open('static/choose_model.txt','w') as f:
        f.write(line)

def load_model(i):
    with open('static/model'+ str(i) + '.txt','r') as lines:
        for line in lines:
            _, name_model = line.strip().split('=')
    return name_model 

def save_model(i,name_model):
    line = 'model' + str(i) + '=' + name_model
    with open('static/model' + str(i) + '.txt','w') as f:
        f.write(line)

def str2bool(v):
    return v.lower() in ("yes", "true", "t", "1")

def load_all(model,i):
    values_all = []
    with open('static/all'+ str(i) + '.txt','r') as lines:
        for line in lines:
            _, name_all = line.strip().split('=')
            values_all.append(name_all)
    window['file_weights' + str(i)].update(value=values_all[0])
    window['conf_thres' + str(i)].update(value=values_all[1])
    a=1
    for item in range(len(model.names)):
        window[f'{model.names[item]}_' + str(i)].update(value=str2bool(values_all[a+1]))
        window[f'{model.names[item]}_OK_' + str(i)].update(value=str2bool(values_all[a+2]))
        window[f'{model.names[item]}_Num_' + str(i)].update(value=str(values_all[a+3]))
        window[f'{model.names[item]}_NG_' + str(i)].update(value=str2bool(values_all[a+4]))
        window[f'{model.names[item]}_Wn_' + str(i)].update(value=str(values_all[a+5]))
        window[f'{model.names[item]}_Wx_' + str(i)].update(value=str(values_all[a+6]))
        window[f'{model.names[item]}_Hn_' + str(i)].update(value=str(values_all[a+7]))
        window[f'{model.names[item]}_Hx_' + str(i)].update(value=str(values_all[a+8]))
        a += 8

def save_all(model,i):
    with open('static/all'+ str(i) + '.txt','w') as f:
        f.write('weights' + str(i) + '=' + str(values['file_weights' + str(i)]))
        f.write('\n')
        f.write('conf' + str(i) + '=' + str(values['conf_thres' + str(i)]))
        f.write('\n')

        for item in range(len(model.names)):
            f.write(str(f'{model.names[item]}_' + str(i)) + '=' + str(values[f'{model.names[item]}_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_OK_' + str(i)) + '=' + str(values[f'{model.names[item]}_OK_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_Num_' + str(i)) + '=' + str(values[f'{model.names[item]}_Num_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_NG_' + str(i)) + '=' + str(values[f'{model.names[item]}_NG_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_Wn_' + str(i)) + '=' + str(values[f'{model.names[item]}_Wn_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_Wx_' + str(i)) + '=' + str(values[f'{model.names[item]}_Wx_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_Hn_' + str(i)) + '=' + str(values[f'{model.names[item]}_Hn_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_Hx_' + str(i)) + '=' + str(values[f'{model.names[item]}_Hx_' + str(i)]))
            if item != len(model.names)-1:
                f.write('\n')

def load_all_sql(i,choose_model):
    conn = sqlite3.connect('2cam_3model.db')
    w = 0
    cursor = conn.execute("SELECT * from MYMODEL")
    for row in cursor:
        if row[0] == choose_model:
            row1_a, row1_b = row[1].strip().split('_')
            if row1_a == str(i) and row1_b == '0':
                window['file_weights' + str(i)].update(value=row[2])
                window['conf_thres' + str(i)].update(value=row[3])
                window['have_save_OK_1'].update(value=str2bool(row[4]))
                window['have_save_OK_2'].update(value=str2bool(row[5]))
                window['have_save_OK_3'].update(value=str2bool(row[6]))
                window['have_save_NG_1'].update(value=str2bool(row[7]))
                window['have_save_NG_2'].update(value=str2bool(row[8]))
                window['have_save_NG_3'].update(value=str2bool(row[9]))

                window['save_OK_1'].update(value=row[10])
                window['save_OK_2'].update(value=row[11])
                window['save_OK_3'].update(value=row[12])
        
                window['save_NG_1'].update(value=row[13])
                window['save_NG_2'].update(value=row[14])
                window['save_NG_3'].update(value=row[15])
                #print(row[2])
                model = torch.hub.load('./levu','custom', path= row[2], source='local',force_reload =False)
                if row[2][-7:] == 'edit.pt': 
                    change_label(model)

                w = 1
            if row1_a == str(i) and row[0] == choose_model and w==1:
                   
                for item in range(len(model.names)):
                    if int(row1_b) == item:
                        window[f'{model.names[item]}_' + str(i)].update(value=str2bool(row[16]))
                        window[f'{model.names[item]}_OK_' + str(i)].update(value=str2bool(row[17]))
                        window[f'{model.names[item]}_Num_' + str(i)].update(value=str(row[18]))
                        window[f'{model.names[item]}_NG_' + str(i)].update(value=str2bool(row[19]))
                        window[f'{model.names[item]}_Wn_' + str(i)].update(value=str(row[20]))
                        window[f'{model.names[item]}_Wx_' + str(i)].update(value=str(row[21]))
                        window[f'{model.names[item]}_Hn_' + str(i)].update(value=str(row[22]))
                        window[f'{model.names[item]}_Hx_' + str(i)].update(value=str(row[23]))
                        window[f'{model.names[item]}_PLC_' + str(i)].update(value=str(row[24]))
                        window[f'OK_PLC_' + str(i)].update(value=str(row[25]))
                        window[f'{model.names[item]}_Conf_' + str(i)].update(value=str(row[26]))
    
    conn.close()

def save_all_sql(model,i,choose_model):
    conn = sqlite3.connect('2cam_3model.db')
    cursor = conn.execute("SELECT * from MYMODEL")
    update = 0 
    answer = sg.popup_yes_no('Load lai thong so thiet dinh cu~ ?')
    if answer == 'Yes':
        conn.execute('UPDATE MYMODEL SET Weights=?  WHERE (ChooseModel = ? AND Camera LIKE ?)',(str(values['file_weights' + str(i)]),choose_model,str(i) + '%'))          
        update = 2            
    if answer == 'No':
        for row in cursor:
            if row[0] == choose_model:            
                row1_a, _ = row[1].strip().split('_')
                if row1_a == str(i):
                    conn.execute("DELETE FROM MYMODEL WHERE (ChooseModel = ? AND Camera LIKE ?)", (choose_model,str(i) + '%'))
                    for item in range(len(model.names)):
                        conn.execute("INSERT INTO MYMODEL (ChooseModel,Camera, Weights,Confidence,OK_Cam1,OK_Cam2,OK_Cam3,NG_Cam1,NG_Cam2,NG_Cam3,Folder_OK_Cam1,Folder_OK_Cam2,Folder_OK_Cam3,Folder_NG_Cam1,Folder_NG_Cam2,Folder_NG_Cam3,Joined,Ok,Num,NG,WidthMin, WidthMax,HeightMin,HeightMax,PLC_NG,PLC_OK,Conf) \
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (str(values['choose_model']),str(i)+ '_' +str(item) ,str(values['file_weights' + str(i)]), int(values['conf_thres' + str(i)]),str(values['have_save_OK_1']),str(values['have_save_OK_2']),str(values['have_save_OK_3']),str(values['have_save_NG_1']),str(values['have_save_NG_2']),str(values['have_save_NG_3']),str(values['save_OK_1']),str(values['save_OK_2']),str(values['save_OK_3']),str(values['save_NG_1']),str(values['save_NG_2']),str(values['save_NG_3']),str(values[f'{model.names[item]}_' + str(i)]), str(values[f'{model.names[item]}_OK_' + str(i)]), int(values[f'{model.names[item]}_Num_' + str(i)]), str(values[f'{model.names[item]}_NG_' + str(i)]), int(values[f'{model.names[item]}_Wn_' + str(i)]), int(values[f'{model.names[item]}_Wx_' + str(i)]), int(values[f'{model.names[item]}_Hn_' + str(i)]), int(values[f'{model.names[item]}_Hx_' + str(i)]), int(values[f'{model.names[item]}_PLC_' + str(i)]), int(values['OK_PLC_' + str(i)]),int(values[f'{model.names[item]}_Conf_' + str(i)])))           
                        update = 1
                    break

    # if update == 0:
    #     for item in range(len(model.names)):
    #         conn.execute("INSERT INTO MYMODEL (ChooseModel,Camera, Weights,Confidence,OK_Cam1,OK_Cam2,OK_Cam3,NG_Cam1,NG_Cam2,NG_Cam3,Folder_OK_Cam1,Folder_OK_Cam2,Folder_OK_Cam3,Folder_NG_Cam1,Folder_NG_Cam2,Folder_NG_Cam3,Joined,Ok,Num,NG,WidthMin, WidthMax,HeightMin,HeightMax,PLC_NG,PLC_OK,Conf) \
    #             VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (str(values['choose_model']),str(i)+ '_' +str(item) ,str(values['file_weights' + str(i)]), int(values['conf_thres' + str(i)]),str(values['have_save_OK_1']),str(values['have_save_OK_2']),str(values['have_save_OK_3']),str(values['have_save_NG_1']),str(values['have_save_NG_2']),str(values['have_save_NG_3']),str(values['save_OK_1']),str(values['save_OK_2']),str(values['save_OK_3']),str(values['save_NG_1']),str(values['save_NG_2']),str(values['save_NG_3']),str(values[f'{model.names[item]}_' + str(i)]), str(values[f'{model.names[item]}_OK_' + str(i)]), int(values[f'{model.names[item]}_Num_' + str(i)]), str(values[f'{model.names[item]}_NG_' + str(i)]), int(values[f'{model.names[item]}_Wn_' + str(i)]), int(values[f'{model.names[item]}_Wx_' + str(i)]), int(values[f'{model.names[item]}_Hn_' + str(i)]), int(values[f'{model.names[item]}_Hx_' + str(i)]),int(values[f'{model.names[item]}_PLC_' + str(i)]), int(values['OK_PLC_' + str(i)]),int(values[f'{model.names[item]}_Conf_' + str(i)])))
            
    for row in cursor:
        if row[0] == choose_model:
            conn.execute("UPDATE MYMODEL SET OK_Cam1 = ? , OK_Cam2 = ?,OK_Cam3 = ? , NG_Cam1 = ?,NG_Cam2 = ?, NG_Cam3 = ?, Folder_OK_Cam1 = ?, Folder_OK_Cam2 = ?,Folder_OK_Cam3 = ?, Folder_NG_Cam1 = ?, Folder_NG_Cam2 = ?,Folder_NG_Cam3 = ? WHERE ChooseModel = ? ",(str(values['have_save_OK_1']),str(values['have_save_OK_2']),str(values['have_save_OK_3']),str(values['have_save_NG_1']),str(values['have_save_NG_2']),str(values['have_save_NG_3']),str(values['save_OK_1']),str(values['save_OK_2']),str(values['save_OK_3']),str(values['save_NG_1']),str(values['save_NG_2']),str(values['save_NG_3']),choose_model))

    conn.commit()
    conn.close()
    #load model
    if update == 2:
        load_all_sql(i,choose_model)

# def save_all_sql(model,i,choose_model):
#     conn = sqlite3.connect('2cam_3model.db')
#     cursor = conn.execute("SELECT * from MYMODEL")
#     update = 0 
#     answer = sg.popup_yes_no('Muon giu thong so cai dat')
#     if answer == 'Load lai thong so thiet dinh cu~ ?':
#         conn.execute('UPDATE MYMODEL SET Weights=?  WHERE (ChooseModel = ? AND Camera LIKE ?)',(str(values['file_weights' + str(i)]),choose_model,str(i) + '%'))          
#         update = 2            
#     if answer == 'Luu lai thong so thiet dinh moi ?':
#         for row in cursor:
#             if row[0] == choose_model:            
#                 row1_a, _ = row[1].strip().split('_')
#                 if row1_a == str(i):
#                     conn.execute("DELETE FROM MYMODEL WHERE (ChooseModel = ? AND Camera LIKE ?)", (choose_model,str(i) + '%'))
#                     for item in range(len(model.names)):
#                         conn.execute("INSERT INTO MYMODEL (ChooseModel,Camera, Weights,Confidence,OK_Cam1,OK_Cam2,OK_Cam3,NG_Cam1,NG_Cam2,NG_Cam3,Folder_OK_Cam1,Folder_OK_Cam2,Folder_OK_Cam3,Folder_NG_Cam1,Folder_NG_Cam2,Folder_NG_Cam3,Joined,Ok,Num,NG,WidthMin, WidthMax,HeightMin,HeightMax,PLC_NG,PLC_OK,Conf) \
#                             VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (str(values['choose_model']),str(i)+ '_' +str(item) ,str(values['file_weights' + str(i)]), int(values['conf_thres' + str(i)]),str(values['have_save_OK_1']),str(values['have_save_OK_2']),str(values['have_save_OK_3']),str(values['have_save_NG_1']),str(values['have_save_NG_2']),str(values['have_save_NG_3']),str(values['save_OK_1']),str(values['save_OK_2']),str(values['save_OK_3']),str(values['save_NG_1']),str(values['save_NG_2']),str(values['save_NG_3']),str(values[f'{model.names[item]}_' + str(i)]), str(values[f'{model.names[item]}_OK_' + str(i)]), int(values[f'{model.names[item]}_Num_' + str(i)]), str(values[f'{model.names[item]}_NG_' + str(i)]), int(values[f'{model.names[item]}_Wn_' + str(i)]), int(values[f'{model.names[item]}_Wx_' + str(i)]), int(values[f'{model.names[item]}_Hn_' + str(i)]), int(values[f'{model.names[item]}_Hx_' + str(i)]), int(values[f'{model.names[item]}_PLC_' + str(i)]), int(values['OK_PLC_' + str(i)]),int(values[f'{model.names[item]}_Conf_' + str(i)])))           
#                         update = 1
#                     break

#     if update == 0:
#         for item in range(len(model.names)):
#             conn.execute("INSERT INTO MYMODEL (ChooseModel,Camera, Weights,Confidence,OK_Cam1,OK_Cam2,OK_Cam3,NG_Cam1,NG_Cam2,NG_Cam3,Folder_OK_Cam1,Folder_OK_Cam2,Folder_OK_Cam3,Folder_NG_Cam1,Folder_NG_Cam2,Folder_NG_Cam3,Joined,Ok,Num,NG,WidthMin, WidthMax,HeightMin,HeightMax,PLC_NG,PLC_OK,Conf) \
#                 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (str(values['choose_model']),str(i)+ '_' +str(item) ,str(values['file_weights' + str(i)]), int(values['conf_thres' + str(i)]),str(values['have_save_OK_1']),str(values['have_save_OK_2']),str(values['have_save_OK_3']),str(values['have_save_NG_1']),str(values['have_save_NG_2']),str(values['have_save_NG_3']),str(values['save_OK_1']),str(values['save_OK_2']),str(values['save_OK_3']),str(values['save_NG_1']),str(values['save_NG_2']),str(values['save_NG_3']),str(values[f'{model.names[item]}_' + str(i)]), str(values[f'{model.names[item]}_OK_' + str(i)]), int(values[f'{model.names[item]}_Num_' + str(i)]), str(values[f'{model.names[item]}_NG_' + str(i)]), int(values[f'{model.names[item]}_Wn_' + str(i)]), int(values[f'{model.names[item]}_Wx_' + str(i)]), int(values[f'{model.names[item]}_Hn_' + str(i)]), int(values[f'{model.names[item]}_Hx_' + str(i)]),int(values[f'{model.names[item]}_PLC_' + str(i)]), int(values['OK_PLC_' + str(i)]),int(values[f'{model.names[item]}_Conf_' + str(i)])))
            
#     for row in cursor:
#         if row[0] == choose_model:
#             conn.execute("UPDATE MYMODEL SET OK_Cam1 = ? , OK_Cam2 = ?,OK_Cam3 = ? , NG_Cam1 = ?,NG_Cam2 = ?, NG_Cam3 = ?, Folder_OK_Cam1 = ?, Folder_OK_Cam2 = ?,Folder_OK_Cam3 = ?, Folder_NG_Cam1 = ?, Folder_NG_Cam2 = ?,Folder_NG_Cam3 = ? WHERE ChooseModel = ? ",(str(values['have_save_OK_1']),str(values['have_save_OK_2']),str(values['have_save_OK_3']),str(values['have_save_NG_1']),str(values['have_save_NG_2']),str(values['have_save_NG_3']),str(values['save_OK_1']),str(values['save_OK_2']),str(values['save_OK_3']),str(values['save_NG_1']),str(values['save_NG_2']),str(values['save_NG_3']),choose_model))

#     conn.commit()
#     conn.close()
#     #load model
#     if update == 2:
#         load_all_sql(i,choose_model)


def excel_sang():
    today = date.today()
    mydate = today.strftime("%Y_%m_%d")
    wb = openpyxl.Workbook()

    HomNay = wb.create_sheet("Data")


    HomNay.merge_cells('A1:K1')
    HomNay.merge_cells('A2:A3')
    HomNay.merge_cells('B2:B3')
    HomNay.merge_cells('C2:K2')
    #HomNay.unmerge_cells('A2:D2')
    HomNay['A1'] = 'DỮ LIỆU MÁY NQVNHT M100 X75'
    HomNay['A1'].alignment = Alignment(horizontal='center')
    HomNay['A1'].font = Font(name= 'Calibri', size=20)
    HomNay['A2'] = 'Ngày sản xuất'
    HomNay['A2'].alignment = Alignment(horizontal='center')
    HomNay['A2'].font = Font(name= 'Calibri', size=12)
    HomNay['B2'] = 'Giờ lưu dữ liệu'
    HomNay['B2'].alignment = Alignment(horizontal='center')
    HomNay['B2'].font = Font(name= 'Calibri', size=12)
    HomNay['C2'] = 'HẠNG MỤC PHẾ PHẨM'
    HomNay['C2'].alignment = Alignment(horizontal='center')
    HomNay['C2'].font = Font(name= 'Calibri', size=12)
    HomNay['C3'] = 'Tổng số lượng sản xuất'
    HomNay['D3'] = 'Không hàn chổi'
    HomNay['E3'] = 'Hàn không đạt chổi'
    HomNay['F3'] = 'Không hàn chấu'
    HomNay['G3'] = 'Hàn không đạt chấu'
    HomNay['H3'] = 'Bụi chì'
    HomNay['I3'] = 'Dị vật'
    HomNay['J3'] = 'Phế phẩm khác'
    HomNay['K3'] = 'Tổng số lượng PP'
    # HomNay['L3'] = 'PP khác (nhiều hạng mục)'
    # HomNay['M3'] = 'Tổng số lượng PP'

    for i in range(67,77):
        HomNay[f'{str(chr(i))}3'].alignment = Alignment(horizontal='center')
        HomNay[f'{str(chr(i))}3'].font = Font(name= 'Calibri', size=12) 

    HomNay.column_dimensions['A'].width = 20
    HomNay.column_dimensions['B'].width = 20
    HomNay.column_dimensions['C'].width = 25
    HomNay.column_dimensions['D'].width = 18
    HomNay.column_dimensions['E'].width = 18
    HomNay.column_dimensions['F'].width = 18
    HomNay.column_dimensions['G'].width = 18
    HomNay.column_dimensions['H'].width = 18
    HomNay.column_dimensions['I'].width = 18
    HomNay.column_dimensions['J'].width = 18
    HomNay.column_dimensions['K'].width = 18
    # HomNay.column_dimensions['L'].width = 25
    # HomNay.column_dimensions['M'].width = 25

    wb.remove(wb['Sheet'])

    wb.save(f"excel/{mydate}_Ngay.xlsx")
    try:
        shutil.copy(f"excel/{mydate}_Ngay.xlsx", f"C:/excel/{mydate}_Ngay.xlsx")
    except:
        pass
    #write_plc_keyence(1002,0) 

def excel_dem():
    today = date.today()
    mydate = today.strftime("%Y_%m_%d")
    wb = openpyxl.Workbook()

    HomNay = wb.create_sheet("Data")


    HomNay.merge_cells('A1:K1')
    HomNay.merge_cells('A2:A3')
    HomNay.merge_cells('B2:B3')
    HomNay.merge_cells('C2:K2')
    #HomNay.unmerge_cells('A2:D2')
    HomNay['A1'] = 'DỮ LIỆU MÁY NQVNHT M100 X75'
    HomNay['A1'].alignment = Alignment(horizontal='center')
    HomNay['A1'].font = Font(name= 'Calibri', size=20)

    HomNay['A2'] = 'Ngày sản xuất'
    HomNay['A2'].alignment = Alignment(horizontal='center')
    HomNay['A2'].font = Font(name= 'Calibri', size=12)
    HomNay['B2'] = 'Giờ lưu dữ liệu'
    HomNay['B2'].alignment = Alignment(horizontal='center')
    HomNay['B2'].font = Font(name= 'Calibri', size=12)
    HomNay['C2'] = 'HẠNG MỤC PHẾ PHẨM'
    HomNay['C2'].alignment = Alignment(horizontal='center')
    HomNay['C2'].font = Font(name= 'Calibri', size=12)
    HomNay['C3'] = 'Tổng số lượng sản xuất'
    HomNay['D3'] = 'Không hàn chổi'
    HomNay['E3'] = 'Hàn không đạt chổi'
    HomNay['F3'] = 'Không hàn chấu'
    HomNay['G3'] = 'Hàn không đạt chấu'
    HomNay['H3'] = 'Bụi chì'
    HomNay['I3'] = 'Dị vật'
    HomNay['J3'] = 'Phế phẩm khác'
    HomNay['K3'] = 'Tổng số lượng PP'

    for i in range(67,77):
        HomNay[f'{str(chr(i))}3'].alignment = Alignment(horizontal='center')
        HomNay[f'{str(chr(i))}3'].font = Font(name= 'Calibri', size=12) 

    HomNay.column_dimensions['A'].width = 20
    HomNay.column_dimensions['B'].width = 20
    HomNay.column_dimensions['C'].width = 25
    HomNay.column_dimensions['D'].width = 18
    HomNay.column_dimensions['E'].width = 18
    HomNay.column_dimensions['F'].width = 18
    HomNay.column_dimensions['G'].width = 18
    HomNay.column_dimensions['H'].width = 18
    HomNay.column_dimensions['I'].width = 18
    HomNay.column_dimensions['J'].width = 18
    HomNay.column_dimensions['K'].width = 18
    # HomNay.column_dimensions['L'].width = 25
    # HomNay.column_dimensions['M'].width = 25

    wb.remove(wb['Sheet'])

    wb.save(f"excel/{mydate}_Dem.xlsx")
    try:
        shutil.copy(f"excel/{mydate}_Dem.xlsx", f"C:/excel/{mydate}_Dem.xlsx")
    except:
        pass
    #write_plc_keyence(1002,0) 

def change_label(model1): #Dung khi file train nham nhan wrong label names
    model1.names[0] = 'cacbon'
    model1.names[1] = 'buichi'
    model1.names[2] = 'divat'
    model1.names[3] = 'taychoi'
    model1.names[4] = 'chaudien'

def program_camera_1_FH(model,size,conf,regno,label_1):
    # read_D = read_plc_keyence(regno)  # doc thanh ghi D450
    # if read_D == 1:
        dir_path = 'C:/FH/camera1/**/*.jpg'
        window['result_cam1'].update(value= '-', text_color='green')
        filenames = glob(dir_path)
        if len(filenames) == 0:
            print('folder CAM1 empty')
        else:
            for filename1 in filenames:
                print(filename1)
                img1_orgin = cv2.imread(filename1)
                while type(img1_orgin) == type(None) or np.allclose(img1_orgin[1199,:],128,atol=0):
                    print('loading img 1...')
                    img1_orgin = cv2.imread(filename1)
                    
                t1 = time.time()
                img1_orgin = img1_orgin[0:1200,200:1400]
                # img1_orgin = img1_orgin1.copy()
                img1_save = img1_orgin
                print('CAM1 processing................')
                # ghi vao D450 gia tri 0
                #write_plc_keyence(regno, 0) 

                img1_orgin = cv2.cvtColor(img1_orgin, cv2.COLOR_BGR2RGB)
                result1 = model(img1_orgin,size= size,conf = conf) 

                table1 = result1.pandas().xyxy[0]
                area_remove1 = []

                myresult1 =0                
                for item in range(len(table1.index)):
                    width1 = table1['xmax'][item] - table1['xmin'][item]
                    height1 = table1['ymax'][item] - table1['ymin'][item]
                    conf1 = table1['confidence'][item] * 100
                    label_name = table1['name'][item]
                    for i1 in range(len(model1.names)):
                        if values[f'{model1.names[i1]}_1'] == True:
                            if label_name == model1.names[i1]:
                                if conf1 < int(values[f'{model1.names[i1]}_Conf_1']):
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif width1 < int(values[f'{model1.names[i1]}_Wn_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif width1 > int(values[f'{model1.names[i1]}_Wx_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif height1 < int(values[f'{model1.names[i1]}_Hn_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif height1 > int(values[f'{model1.names[i1]}_Hx_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                     
                        if values[f'{model1.names[i1]}_1'] == False:
                            if label_name == model1.names[i1]:
                                table1.drop(item, axis=0, inplace=True)
                                area_remove1.append(item)

                names1 = list(table1['name'])

                show1 = np.squeeze(result1.render(area_remove1))
                show_1 = cv2.resize(show1, (1000,1000), interpolation = cv2.INTER_AREA)
                show1 = cv2.resize(show1, (image_height_display,image_height_display), interpolation = cv2.INTER_AREA)
                show1 = cv2.cvtColor(show1, cv2.COLOR_BGR2RGB)
                show_1 = cv2.cvtColor(show_1, cv2.COLOR_BGR2RGB)
                hm=[]
                k=1
                for i1 in range(len(model1.names)):
                    register_ng = int(values[f'{model1.names[i1]}_PLC_1'])
                    if values[f'{model1.names[i1]}_1'] == True:

                        if values[f'{model1.names[i1]}_OK_1'] == True:
                            len_name1 = 0
                            for name1 in names1:
                                if name1 == model1.names[i1]:
                                    len_name1 +=1
                            if len_name1 != int(values[f'{model1.names[i1]}_Num_1']):
                                hm.append(model1.names[i1])   
                                #write_plc_keyence(register_ng,1)                            
                                cv2.putText(show1, 'NG',(result_height_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                cv2.putText(show1, model1.names[i1],(5,30*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                                print('ngc1:',model1.names[i1])
                                window['result_cam1'].update(value= 'NG', text_color='red')
                                myresult1 = 1
                                k+=1
                                

                        if values[f'{model1.names[i1]}_NG_1'] == True:
                            if model1.names[i1] in names1:
                                hm.append(model1.names[i1])
                                #write_plc_keyence(register_ng,1)
                                cv2.putText(show1, 'NG',(result_height_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                cv2.putText(show1, model1.names[i1],(5,30*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                                print('ngc1:',model1.names[i1])
                                window['result_cam1'].update(value= 'NG', text_color='red')    
                                myresult1 = 1   
                                k+=1

                fimg = time_to_name() + '-C1' #Name of save file'
                t2 = time.time() - t1
                if myresult1 == 0:
                    print('OK')                    
                    #write_plc_keyence(int(values['OK_PLC_1']),1)                              
                    cv2.putText(show1, 'OK',(result_height_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                    cv2.putText(show_1, 'OK',(1000,100),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                    window['result_cam1'].update(value= 'OK', text_color='green')
                    if values['have_save_OK_1']:
                        cv2.imwrite(values['save_OK_1']  + '/' + fimg + '.jpg',img1_save)
                else:
                    cv2.putText(show_1, 'NG',(1000,100),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                    if values['have_save_NG_1']:
                        cv2.imwrite(values['save_NG_1']  + '/' + fimg + '.jpg',img1_save)
                    hm1 = ', '.join(hm)
                    print('NG',hm1)
                    # pp.execute('INSERT INTO PHEPHAM (ANH, HANGMUC) VALUES (?,?)', (fimg, hm1))
                    # pp.commit()
                time_cam1 = str(int(t2*1000)) + 'ms'
                #Bao hoan tat CAM1
                #write_plc_keyence(17412,1)
                window['time_cam1'].update(value= time_cam1, text_color='black') 
            
                imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
                window['image1'].update(data= imgbytes1)

                imgbytes1 = cv2.imencode('.png',show_1)[1].tobytes()
                window['toan1'].update(data= imgbytes1)
                
                                    
                #Xoa thu muc
                fname=os.path.dirname(filename1)
                shutil.rmtree(fname)                                                 
                
                print('CAM1 Finished')

ktc = []               
mycount = 0 

def program_camera1_FH(model,size,conf,regno):

    global ktc
    global mycount
    global label_1

    dir_path = 'C:/FH/camera1/**/*.jpg'
    window['result_cam1'].update(value= '-', text_color='green')
    filenames = glob(dir_path)
    if len(filenames) == 0:
        # print('folder CAM1 empty')
        pass
    else:
        for filename1 in filenames:
            print(filename1)
            img1_orgin = cv2.imread(filename1)
            while type(img1_orgin) == type(None) or np.allclose(img1_orgin[1199,:],128,atol=0):
                # print('loading img 1...')
                img1_orgin = cv2.imread(filename1)
                
            t1 = time.time()
            # img1_orgin = img1_orgin[0:1200,200:1400]
            # img1_orgin = img1_orgin[0:1200,210:1410]
            # # img1_orgin = img1_orgin1.copy()
            # img1_orgin = cv2.rotate(img1_orgin, cv2.ROTATE_90_COUNTERCLOCKWISE)
            img1_save = img1_orgin
            print('CAM1 processing................')
            # ghi vao D450 gia tri 0

            #write_plc_keyence(regno, 0) 

            img1_orgin = cv2.cvtColor(img1_orgin, cv2.COLOR_BGR2RGB)
            result1 = model(img1_orgin,size= size,conf = conf) 

            table1 = result1.pandas().xyxy[0]
            area_remove1 = []
            ktr = []
            toadox = []
            toadoy = []
            _x_px = []
            _y_py = []
            myresult1 =0   
            # print('table1',table1)
            
            for item in range(len(table1.index)):
                width1 = table1['xmax'][item] - table1['xmin'][item]
                height1 = table1['ymax'][item] - table1['ymin'][item]
                conf1 = table1['confidence'][item] * 100
                label_name = table1['name'][item]
                
                if label_name == 'bui_chi' or label_name == 'hat_chi' or label_name == 'vun_chi' or label_name == 'di_vat'  :
                    ktr.append([table1['xmin'][item],table1['xmax'][item],table1['ymin'][item],table1['ymax'][item]])

                if label_name == 'bui_chi' :
                    toa_do_x = (int(table1['xmin'][item]) + int(table1['xmax'][item]))/2
                    toa_do_y = (int(table1['ymin'][item]) + int(table1['ymax'][item]))/2
                    _x_px.append(toa_do_x/1200)
                    _y_py.append(toa_do_y/1200)
                    print('width1:' ,width1)
                    print('height1:',height1)
                
                if label_name ==  'bui_chi' :
                    if conf1 > int(values['bui_chi_Conf_1']) :
                        toa_do_x = (int(table1['xmin'][item]) + int(table1['xmax'][item]))/2
                        toa_do_y = (int(table1['ymin'][item]) + int(table1['ymax'][item]))/2
                        toadox.append(toa_do_x)
                        toadoy.append(toa_do_y)
                        
                if label_name ==  'hat_chi' :
                    if conf1 > int(values['hat_chi_Conf_1']) :
                        toa_do_x = (int(table1['xmin'][item]) + int(table1['xmax'][item]))/2
                        toa_do_y = (int(table1['ymin'][item]) + int(table1['ymax'][item]))/2
                        toadox.append(toa_do_x)
                        toadoy.append(toa_do_y)
                
                if label_name ==  'vun_chi' :
                    if conf1 > int(values['vun_chi_Conf_1']) :
                        toa_do_x = (int(table1['xmin'][item]) + int(table1['xmax'][item]))/2
                        toa_do_y = (int(table1['ymin'][item]) + int(table1['ymax'][item]))/2
                        toadox.append(toa_do_x)
                        toadoy.append(toa_do_y)

                if label_name ==  'chi_dinh_thanh_de' :
                    if conf1 > int(values['chi_dinh_thanh_de_Conf_1']) :
                        toa_do_x = (int(table1['xmin'][item]) + int(table1['xmax'][item]))/2
                        toa_do_y = (int(table1['ymin'][item]) + int(table1['ymax'][item]))/2
                        toadox.append(toa_do_x)
                        toadoy.append(toa_do_y)

                for i1 in range(len(model1.names)):
                    if values[f'{model1.names[i1]}_1'] == True:
                        if label_name == model1.names[i1]:
                            if label_name ==  'bui_chi' :                                 
                                x,y,z,v,h,p = check_overlap(label_1)                              
                                if check_overload(x, y, z, v, h, p, _x_px,_y_py):    
                                    if width1 < int(values['W_min_bc_cc']) :
                                        table1.drop(item, axis=0, inplace=True)
                                        area_remove1.append(item)
                                    elif height1 < int(values['H_min_bc_cc']):
                                        table1.drop(item, axis=0, inplace=True)
                                        area_remove1.append(item)
                                    elif conf1 < int(values['bc_cc_Conf_1']):
                                        table1.drop(item, axis=0, inplace=True)
                                        area_remove1.append(item)
                                else:
                                    if conf1 < int(values['bui_chi_Conf_1']):
                                        table1.drop(item, axis=0, inplace=True)
                                        area_remove1.append(item)
                                    elif width1 < int(values['bui_chi_Wn_1']): 
                                        table1.drop(item, axis=0, inplace=True)
                                        area_remove1.append(item)
                                    elif width1 > int(values['bui_chi_Wx_1']): 
                                        table1.drop(item, axis=0, inplace=True)
                                        area_remove1.append(item)
                                    elif height1 < int(values['bui_chi_Hn_1']): 
                                        table1.drop(item, axis=0, inplace=True)
                                        area_remove1.append(item)
                                    elif height1 > int(values['bui_chi_Hx_1']): 
                                        table1.drop(item, axis=0, inplace=True)
                                        area_remove1.append(item)                                                                                          
                            else:                               
                                if conf1 < int(values[f'{model1.names[i1]}_Conf_1']):
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif width1 < int(values[f'{model1.names[i1]}_Wn_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif width1 > int(values[f'{model1.names[i1]}_Wx_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif height1 < int(values[f'{model1.names[i1]}_Hn_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif height1 > int(values[f'{model1.names[i1]}_Hx_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                                    
                    if values[f'{model1.names[i1]}_1'] == False:
                        if label_name == model1.names[i1]:
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)
        
            if ktr != []:
                ktc.append(ktr)
            else:
                ktc = []
                mycount = 0

            kco = 1
            if len(ktc) == 2:
                el1 = ktc[0]
                el2 = ktc[1]
                bk = 0 
                for e1 in el1:
                    for e2 in el2:
                        if abs(e1[0] - e2[0]) < 5 and abs(e1[1] - e2[1]) < 5 and abs(e1[2] - e2[2]) < 5 and abs(e1[3] - e2[3]) < 5:
                            mycount +=1
                            print('SO LAN BAT BUI CHI ==>>: ', mycount)
                            bk = 1
                            break
                    if bk == 1:
                        bk = 0
                        kco = 0
                        break
                if kco == 1:
                    mycount = 0

                ktc.pop(0)

            if mycount == 3:
                #write_plc_keyence(17510,1)
                print('HAY VE SINH KHUON CAMERA 1')
                ktc = []
                mycount = 0 
                break              
            # print('halo: ',_x_px[0],_y_py[0])
            names1 = list(table1['name'])
            print('area_remove1:',area_remove1)
            show1 = np.squeeze(result1.render(area_remove1))
            show_1 = cv2.resize(show1, (1000,1000), interpolation = cv2.INTER_AREA)
            show1 = cv2.resize(show1, (image_height_display,image_height_display), interpolation = cv2.INTER_AREA)
            show1 = cv2.cvtColor(show1, cv2.COLOR_BGR2RGB)
            show_1 = cv2.cvtColor(show_1, cv2.COLOR_BGR2RGB)
            hm=[]
            k=1
            flag = False
            for i1 in range(len(model1.names)):
                register_ng = int(values[f'{model1.names[i1]}_PLC_1'])
                if values[f'{model1.names[i1]}_1'] == True:
                    if values[f'{model1.names[i1]}_OK_1'] == True:
                        len_name1 = 0
                        for name1 in names1:
                            if name1 == model1.names[i1]:
                                len_name1 +=1
                        if len_name1 != int(values[f'{model1.names[i1]}_Num_1']):
                            hm.append(model1.names[i1])   
                            #write_plc_keyence(register_ng,1)                            
                            cv2.putText(show1, 'NG',(result_height_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                            cv2.putText(show1, model1.names[i1],(5,30*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)                              
                            window['result_cam1'].update(value= 'NG', text_color='red')
                            myresult1 = 1
                            k+=1
                            
                    if values[f'{model1.names[i1]}_NG_1'] == True:
                        if model1.names[i1] in names1:
                            if model1.names[i1] == 'bui_chi' or model1.names[i1] == 'hat_chi' or model1.names[i1] == 'vun_chi' or model1.names[i1] == 'chi_dinh_thanh_de' :   
                                for x, y in zip(toadox,toadoy):                                  
                                    if y < 300 and x < 1190:
                                        print('buichi tay choi tren', x, y)
                                        flag = True                                   
                                        #write_plc_keyence(register_ng,3)                                          
                                        cv2.putText(show1, 'NG',(result_height_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                        cv2.putText(show1, model1.names[i1],(5,30*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                                        window['result_cam1'].update(value= 'NG', text_color='red') 
                                        myresult1 = 1  
                                        k+=1   
                                        break   
                                        
                                    elif y > 900 and x < 1190:
                                        print('buichi tay choi duoi', x, y)
                                        flag = True                                   
                                        #write_plc_keyence(register_ng,4)                                           
                                        cv2.putText(show1, 'NG',(result_height_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                        cv2.putText(show1, model1.names[i1],(5,30*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                                        window['result_cam1'].update(value= 'NG', text_color='red') 
                                        myresult1 = 1  
                                        k+=1   
                                        break
                                                                            
                                    elif x < 300 and y < 1190:
                                        print('buichi chau dien trai', x, y)
                                        flag = True                                                 
                                        #write_plc_keyence(register_ng,1)                                            
                                        cv2.putText(show1, 'NG',(result_height_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                        cv2.putText(show1, model1.names[i1],(5,30*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                                        window['result_cam1'].update(value= 'NG', text_color='red') 
                                        myresult1 = 1  
                                        k+=1   
                                        break
                                        
                                    elif x > 900 and y < 1190:
                                        print('buichi chau dien phai', x, y)
                                        flag = True                                                 
                                        #write_plc_keyence(register_ng,2)                                            
                                        cv2.putText(show1, 'NG',(result_height_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                        cv2.putText(show1, model1.names[i1],(5,30*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                                        window['result_cam1'].update(value= 'NG', text_color='red') 
                                        myresult1 = 1  
                                        k+=1  
                                        break                                                                                                                        
                                if not flag : 
                                    print('buichi TH DAC BIET', x, y)
                                    cv2.putText(show1, 'NG',(result_height_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                    cv2.putText(show1, model1.names[i1],(5,30*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                                    window['result_cam1'].update(value= 'NG', text_color='red') 
                                    myresult1 = 1    
                                    k+=1          
                                    #write_plc_keyence(register_ng,3)
                                    toadox = []
                                    toadoy = [] 
                                    break 
                                    
                            else:
                                hm.append(model1.names[i1])
                                #write_plc_keyence(register_ng,1)
                                cv2.putText(show1, 'NG',(result_height_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                cv2.putText(show1, model1.names[i1],(5,30*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                                window['result_cam1'].update(value= 'NG', text_color='red')    
                                myresult1 = 1   
                                k+=1
            fimg = time_to_name() + '-C1' 
            t2 = time.time() - t1
            if myresult1 == 0:
                print('OK')                    
                #write_plc_keyence(int(values['OK_PLC_1']),1)                              
                cv2.putText(show1, 'OK',(result_height_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                cv2.putText(show_1, 'OK',(1000,100),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                window['result_cam1'].update(value= 'OK', text_color='green')
                if values['have_save_OK_1']:
                    cv2.imwrite(values['save_OK_1']  + '/' + fimg + '.jpg',img1_save)
            else:
                cv2.putText(show_1, 'NG',(1000,100),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                if values['have_save_NG_1']:
                    cv2.imwrite(values['save_NG_1']  + '/' + fimg + '.jpg',img1_save)
                hm1 = ', '.join(hm)
                print('NG',hm1)
                
            time_cam1 = str(int(t2*1000)) + 'ms'
            #write_plc_keyence(17412,1)
            window['time_cam1'].update(value= time_cam1, text_color='black') 
        
            imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
            window['image1'].update(data= imgbytes1)

            imgbytes1 = cv2.imencode('.png',show_1)[1].tobytes()
            window['toan1'].update(data= imgbytes1)

            fname=os.path.dirname(filename1)
            shutil.rmtree(fname)                                                 
            
            print('CAM1 Finished')

def program_camera2_FH(model,size,conf,file):
    global myitem 
    global time_all
    
    img2_orgin = cv2.imread(file)
    filename = os.path.dirname(file)
    while type(img2_orgin) == type(None):
        print('loading img 2...', file)
        for path2 in glob(filename + '/*'):
            img2_orgin = cv2.imread(path2)
        if not os.path.exists(file):
            continue
    img2_save = img2_orgin
    img2_orgin = cv2.cvtColor(img2_orgin, cv2.COLOR_BGR2RGB)
    t1 = time.time()

    result2 = model(img2_orgin,size= size,conf = conf) 
    table2 = result2.pandas().xyxy[0]
    area_remove2 = []
    
    myresult2 =0 
    for item in range(len(table2.index)):
        width2 = table2['xmax'][item] - table2['xmin'][item]
        height2 = table2['ymax'][item] - table2['ymin'][item]
        conf2 = table2['confidence'][item] * 100
        label_name = table2['name'][item]
        for i2 in range(len(model2.names)):
            if values[f'{model2.names[i2]}_2'] == True:
                if label_name == model2.names[i2]:
                    if conf2 < int(values[f'{model2.names[i2]}_Conf_2']):
                        table2.drop(item, axis=0, inplace=True)
                        area_remove2.append(item) 
                    elif width2 < int(values[f'{model2.names[i2]}_Wn_2']): 
                        table2.drop(item, axis=0, inplace=True)
                        area_remove2.append(item)
                    elif width2 > int(values[f'{model2.names[i2]}_Wx_2']): 
                        table2.drop(item, axis=0, inplace=True)
                        area_remove2.append(item)
                    elif height2 < int(values[f'{model2.names[i2]}_Hn_2']): 
                        table2.drop(item, axis=0, inplace=True)
                        area_remove2.append(item)
                    elif height2 > int(values[f'{model2.names[i2]}_Hx_2']): 
                        table2.drop(item, axis=0, inplace=True)
                        area_remove2.append(item)                   
            if values[f'{model2.names[i2]}_2'] == False:
                if label_name == model2.names[i2]:
                    table2.drop(item, axis=0, inplace=True)
                    area_remove2.append(item)

    names2 = list(table2['name'])
    print(names2)
    
    show2 = np.squeeze(result2.render(area_remove2))
    show2 = cv2.resize(show2, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
    show2 = cv2.cvtColor(show2, cv2.COLOR_BGR2RGB)
    ng2 = True
    hm=[]
    tag2 = ''
    k=1
    for i2 in range(len(model2.names)):
        error_number = int(values[f'{model2.names[i2]}_PLC_2'])
        if values[f'{model2.names[i2]}_2'] == True:

            if values[f'{model2.names[i2]}_NG_2'] == True:
                if model2.names[i2] in names2:
                    hm.append(model2.names[i2])                
                    all_error.append(error_number)
                    all_item_error.append(f'{model2.names[i2]}_CD')
                    cv2.putText(show2, 'NG',(result_width_display+20,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                    cv2.putText(show2, model2.names[i2],(5,30*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                    myresult2 = 1         
                    ng2 = False
                    k+=1
            
            if values[f'{model2.names[i2]}_OK_2'] == True:
                
                len_name2 = 0
                for name2 in names2:
                    if name2 == model2.names[i2]: 
                        len_name2 +=1
                if len_name2 != int(values[f'{model2.names[i2]}_Num_2']):
                    hm.append(model2.names[i2])  
                    
                    if ng2:
                        all_error.append(error_number)
                    cv2.putText(show2, 'NG',(result_width_display+20,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                    cv2.putText(show2, model2.names[i2],(5,30*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                    all_item_error.append(f'{model2.names[i2]}_CD')
                    myresult2 = 1
                    k+=1
    if cd==1:
        tag2 = '-CD4'
    else:
        tag2 = '-CD3'   
    
    fimg = c_hinh + tag2
    t2 = time.time() - t1    
    if myresult2 == 0:
        print('OK')        
        cv2.putText(show2, 'OK',(result_width_display+20,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)

        if values['have_save_OK_2']:
            cv2.imwrite(values['save_OK_2']  + '/' + fimg + '.jpg',img2_save)
        if cd==1:
            window['cd1'].update(value= 'OK', text_color='green')
            #write_plc_keyence(17480,1)
        if cd==2:
            window['cd2'].update(value= 'OK', text_color='green')
            #write_plc_keyence(17484,1)
    else:
        if cd==1:
            window['cd1'].update(value= 'NG', text_color='red')
            #write_plc_keyence(17480,0)
        if cd==2:
            window['cd2'].update(value= 'NG', text_color='red')
            #write_plc_keyence(17484,0)
        if values['have_save_NG_2']:
            cv2.imwrite(values['save_NG_2']  + '/' + fimg + '.jpg',img2_save)

    time_all += t2  
    time_cam2 = str(int(time_all*1000)) + 'ms'
    window['time_cam2'].update(value= time_cam2, text_color='black') 

    show_2_layout = cv2.resize(show2, (image_width_display_c2,image_height_display_c2))
    imgbytes2_layout = cv2.imencode('.png',show_2_layout)[1].tobytes()

    myitem +=1

    show_2 = cv2.resize(show2, (600,450))
    imgbytes2 = cv2.imencode('.png',show_2)[1].tobytes()

    if cd==1:
        window['cd_11'].update(data= imgbytes2_layout)
        window['cd_1'].update(data= imgbytes2)
    else:
        window['cd_22'].update(data= imgbytes2_layout)
        window['cd_2'].update(data= imgbytes2)

    shutil.rmtree(filename)
    
    #Xoa thu muc   
    #fname=os.path.dirname('C:/FH/CAM2/CHAU' + str(cd))
    # shutil.rmtree('C:/FH/camera2/CHAU' + str(cd))
    # print('CAM2-CD'+str(cd) + ' Finished')       


def program_camera3_FH(model,size,conf,file):
    global myitem 
    global time_all
    img3_orgin = cv2.imread(file)
    filename = os.path.dirname(file)
    while type(img3_orgin) == type(None):
        print('loading img 3...',file)
        for path3 in glob(filename + '/*'):
            img3_orgin = cv2.imread(path3)
        # if not os.path.exists(filename):
        #     print('1111')
        if not os.path.exists(file):
            continue
    img3_save = img3_orgin
    # print('CAM2-TC'+str(tc) + ' processing................')
    img3_orgin = cv2.cvtColor(img3_orgin, cv2.COLOR_BGR2RGB)
    t1 = time.time()   

    result3 = model(img3_orgin,size= size,conf = conf) 
    table3 = result3.pandas().xyxy[0]
    area_remove3 = []

    myresult3 =0        
    for item in range(len(table3.index)):
        width3 = table3['xmax'][item] - table3['xmin'][item]
        height3 = table3['ymax'][item] - table3['ymin'][item]
        conf3 = table3['confidence'][item] * 100
        label_name = table3['name'][item]
        for i3 in range(len(model3.names)):

            if values[f'{model3.names[i3]}_3'] == True:
                if label_name == model3.names[i3]:
                    if conf3 < int(values[f'{model3.names[i3]}_Conf_3']):
                        table3.drop(item, axis=0, inplace=True)
                        area_remove3.append(item)
                    elif width3 < int(values[f'{model3.names[i3]}_Wn_3']): 
                        table3.drop(item, axis=0, inplace=True)
                        area_remove3.append(item)
                    elif width3 > int(values[f'{model3.names[i3]}_Wx_3']): 
                        table3.drop(item, axis=0, inplace=True)
                        area_remove3.append(item)
                    elif height3 < int(values[f'{model3.names[i3]}_Hn_3']): 
                        table3.drop(item, axis=0, inplace=True)
                        area_remove3.append(item)
                    elif height3 > int(values[f'{model3.names[i3]}_Hx_3']): 
                        table3.drop(item, axis=0, inplace=True)
                        area_remove3.append(item)
                         

            if values[f'{model3.names[i3]}_3'] == False:
                if label_name == model3.names[i3]:
                    table3.drop(item, axis=0, inplace=True)
                    area_remove3.append(item)

    names3 = list(table3['name'])
    print(names3)

    show3 = np.squeeze(result3.render(area_remove3))
    show3 = cv2.resize(show3, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
    show3 = cv2.cvtColor(show3, cv2.COLOR_BGR2RGB)
    hm=[]
    tag3 = ''
    ng2 = True
    just_ng = 0
    k=1
    for i3 in range(len(model3.names)):
        error_number = int(values[f'{model3.names[i3]}_PLC_3'])
        if values[f'{model3.names[i3]}_3'] == True:


            if values[f'{model3.names[i3]}_NG_3'] == True:
                if model3.names[i3] in names3:
                    hm.append(model3.names[i3])   
                    if model3.names[i3] == 'buichi':
                        all_error.append(9)
                    else:
                        all_error.append(error_number)
                        all_item_error.append(f'{model3.names[i3]}_TC')

                    # all_error.append(error_number)
                    # print(error_number)
                    # print(all_error)
                    cv2.putText(show3, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                    cv2.putText(show3, model3.names[i3],(5,30*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                    all_item_error.append(f'{model3.names[i3]}_TC')
                    myresult3 = 1     
                    ng2 = False
                    just_ng = 1
                    k+=1

    # for i3 in range(len(model3.names)):
    #     error_number = int(values[f'{model3.names[i3]}_PLC_3'])

            if values[f'{model3.names[i3]}_OK_3'] == True:
                len_name3 = 0
                for name3 in names3:
                    if name3 == model3.names[i3]:
                        len_name3 +=1
                if len_name3 != int(values[f'{model3.names[i3]}_Num_3']):
                    hm.append(model3.names[i3])   
                    
                    if ng2:
                        all_error.append(error_number)
                    cv2.putText(show3, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                    cv2.putText(show3, model3.names[i3],(5,30*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                    all_item_error.append(f'{model3.names[i3]}_TC')
                    myresult3 = 1
                    k+=1
                    
    # if tc==1:
    #     tag3 = '-1TC'
    # else:
    #     tag3 = '-2TC'  
    if tc==1:
        tag3 = '-TC2'
    else:
        tag3 = '-TC1'            
    #Filename image
    fimg = c_hinh + tag3
    t2 = time.time() - t1
    if myresult3 == 0:
        print('OK')        
        cv2.putText(show3, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
        if values['have_save_OK_3']:
            cv2.imwrite(values['save_OK_3']  + '/' + fimg + '.jpg',img3_save)
        if tc==1:
            window['tc1'].update(value= 'OK', text_color='green')
            #write_plc_keyence(17482,1)
        if tc==2:
            window['tc2'].update(value= 'OK', text_color='green')
            #write_plc_keyence(17486,1)
    else:
        if tc==1:
            window['tc1'].update(value= 'NG', text_color='red')
            #write_plc_keyence(17482,0)
        if tc==2:
            window['tc2'].update(value= 'NG', text_color='red')
            #write_plc_keyence(17486,0)
        if values['have_save_NG_3']:
            cv2.imwrite(values['save_NG_3']  + '/' + fimg + '.jpg',img3_save)

    time_all +=t2  
    time_cam3 = str(int(time_all*1000)) + 'ms'

    window['time_cam2'].update(value= time_cam3, text_color='black') 

    myitem +=1

    show_3_layout = cv2.resize(show3, (image_width_display_c2,image_height_display_c2))
    imgbytes3_layout = cv2.imencode('.png',show_3_layout)[1].tobytes()

    show_3 = cv2.resize(show3, (600,450))
    imgbytes3 = cv2.imencode('.png',show_3)[1].tobytes()
    if tc==1:
        window['tc_11'].update(data= imgbytes3_layout)
        window['tc_1'].update(data= imgbytes3)
    else:
        window['tc_22'].update(data= imgbytes3_layout) 
        window['tc_2'].update(data= imgbytes3)

    #Xoa thu muc    
    #fname=os.path.dirname('C:/FH/CAM2/CHOI' + str(tc))
    
    shutil.rmtree(filename)

    # shutil.rmtree('C:/FH/camera2/CHOI' + str(tc))
    # print('CAM2-TC'+str(tc) + " Finished")
           

def make_window(theme):
    sg.theme(theme)

    file_weights = [('Weights (*.pt)', ('*.pt'))]

    right_click_menu = [[], ['Change Theme','Administrator']]

    layout_main = [

        [
        sg.Text('CAM 1',justification='center' ,font= ('Helvetica',30),text_color='red', expand_y=True),
        sg.Text('CAM 2',justification='center' ,font= ('Helvetica',30),text_color='red',expand_x=True),
        ],
        [
        #1
        sg.Frame('',[
            [sg.Text(' '*10,font=('Helvetica',20), justification='center'),
            sg.Image(filename='', size=(image_height_display,image_height_display),key='image1',background_color='black',)],
            [sg.Frame('',
            [
                [sg.Text('',font=('Helvetica',120), justification='center', key='result_cam1',expand_x=True)],
                [sg.Text('',font=('Helvetica',40), justification='center', key='time_cam1', expand_x=True)],
                
            ], vertical_alignment='top',size=(int(560*0.6),int(450*0.6))),
            sg.Frame('',[
                [sg.Button('Webcam', size=(8,1),  font=('Helvetica',14),disabled=True ,key= 'Webcam1')],
                [sg.Text('')],
                [sg.Button('Stop', size=(8,1), font=('Helvetica',14),disabled=True ,key= 'Stop1')],
                [sg.Text('')],
                [sg.Button('Snap', size=(8,1), font=('Helvetica',14),disabled=True ,key= 'Snap1')],
                [sg.Text('')],
                [sg.Checkbox('Check1',size=(6,1),font=('Helvetica',14), key='check_model1',enable_events=True,expand_x=True, expand_y=True)],
                ],element_justification='center', vertical_alignment='top', relief= sg.RELIEF_FLAT),
                
            sg.Frame('',[   
                [sg.Button('Change', size=(8,1), font=('Helvetica',14), disabled= True, key= 'Change1')],
                [sg.Text('')],
                [sg.Button('Pic', size=(8,1), font=('Helvetica',14),disabled=True,key= 'Pic1')],
                [sg.Text('')],
                [sg.Button('Detect', size=(8,1), font=('Helvetica',14),disabled=True,key= 'Detect1')],
                [sg.Text('',size=(4,1))],
                [sg.Combo(values=['1','3','4','5','6','7','8','9'], default_value='3',font=('Helvetica',20),size=(5, 100),text_color='navy',enable_events= True, key='choose_model'),],
                ],element_justification='center', vertical_alignment='top', relief= sg.RELIEF_FLAT),
            ],
            [sg.Text(' ',font=('Helvetica',30), justification='center', key='c1_1',expand_x=True),
            sg.Text(' ',font=('Helvetica',30), justification='center', key='c1_2',expand_x=True),
            sg.Text(' ',font=('Helvetica',30), justification='center', key='c1_3',expand_x=True),
            sg.Text(' ',font=('Helvetica',30), justification='center', key='c1_4',expand_x=True)],
                
        ],relief= sg.RELIEF_FLAT),
    
        # 2
        sg.Frame('',[
            [sg.Text(' ',font= ('Helvetica',5))],
            [sg.Image(filename='', size=(image_width_display_c2,image_height_display_c2),key='cd_11',background_color='black'),
             sg.Image(filename='',  size=(image_width_display_c2,image_height_display_c2),key='tc_11',background_color='black')],
            
            [sg.Image(filename='',  size=(image_width_display_c2,image_height_display_c2),key='cd_22',background_color='black'),
            sg.Image(filename='',  size=(image_width_display_c2,image_height_display_c2),key='tc_22',background_color='black')],
            [sg.Frame('',
            [
                [sg.Text('',font=('Helvetica',120), justification='center', key='result_cam2',expand_x=True)],
                [sg.Text('',font=('Helvetica',40),justification='center', key='time_cam2',expand_x=True)],
            ], vertical_alignment='top',size=(int(560*0.6),int(450*0.6))),
            sg.Frame('',[
                [sg.Button('Webcam', size=(8,1),  font=('Helvetica',14),disabled=True,key= 'Webcam2',auto_size_button=True)],
                [sg.Text('')],
                [sg.Button('Stop', size=(8,1), font=('Helvetica',14),disabled=True  ,key='Stop2')],
                [sg.Text('')],
                [sg.Button('Snap', size=(8,1), font=('Helvetica',14),disabled=True  ,key='Snap2')],
                [sg.Text('')],
                [sg.Checkbox('Taychoi',size=(6,1),font=('Helvetica',14), key='Tay_choi',enable_events=True,expand_x=True,expand_y=True)],
                [sg.Checkbox('Check2',size=(6,1),font=('Helvetica',14), key='check_model2',enable_events=True,expand_y=True)],
                ],element_justification='center', vertical_alignment='top', relief= sg.RELIEF_FLAT),

            sg.Frame('',[   
                [sg.Button('Change', size=(8,1), font=('Helvetica',14), disabled= True, key= 'Change2')],
                [sg.Text('')],
                [sg.Button('Pic', size=(8,1), font=('Helvetica',14),disabled=True,key='Pic2')],
                [sg.Text('')],
                [sg.Button('Detect', size=(8,1), font=('Helvetica',14),disabled=True ,key= 'Detect2')],
                [sg.Text('',size=(4,2))],
                [sg.Text('',size=(4,1))],
                ],element_justification='center', vertical_alignment='top', relief= sg.RELIEF_FLAT),
            ],
            [sg.Text('',font=('Helvetica',30), justification='center', key='cd1',expand_x=True),
            sg.Text('',font=('Helvetica',30), justification='center', key='tc1',expand_x=True),
            sg.Text('',font=('Helvetica',30), justification='center', key='cd2',expand_x=True),
            sg.Text('',font=('Helvetica',30), justification='center', key='tc2',expand_x=True)],          
        ],relief= sg.RELIEF_FLAT, expand_y= True),

    ],
    [
        sg.Frame('',[
            [sg.Text('  ',font=('Helvetica',48), justification='center', key='choosemodel_running')],
        ],element_justification='center',expand_x= True, expand_y= True),
    ] 
    ]

    layout_cam1 = [
        [sg.Image(filename='', key='toan1'),],
    ]

    layout_cam2 = [

        [
        #1
        sg.Frame('CHAU DIEN',[
            [sg.Image(filename='', size=(600,450),key='cd_1',background_color='black')],
            [sg.Image(filename='', size=(600,450),key='cd_2',background_color='black')],
        ],relief= sg.RELIEF_FLAT, expand_x= True),

        # 2
        sg.Frame('TAY CHOI',[
            [sg.Image(filename='', size=(600,450),key='tc_1',background_color='black')],
            [sg.Image(filename='', size=(600,450),key='tc_2',background_color='black')],
                  
        ],relief= sg.RELIEF_FLAT, expand_x= True),

    ],
    
    ]

    layout_option1 = [
        [sg.Frame('',[
        [sg.Frame('',
        [   
            [sg.Text('Weights', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Input(size=(60,1), font=('Helvetica',12), key='file_weights1',readonly= True, text_color='navy',enable_events= True),
            sg.Frame('',[
                [sg.FileBrowse(file_types= file_weights, size=(12,1), font=('Helvetica',10),key= 'file_browse1',enable_events=True, disabled=True)]
            ], relief= sg.RELIEF_FLAT),
            sg.Frame('',[
                [sg.Button('Change Model', size=(14,1), font=('Helvetica',10), disabled= True, key= 'Change_1'), sg.Checkbox('',size=(3,1),default=False,font=('Helvetica',15),  key='checkpass',enable_events=True), ]
            ], relief= sg.RELIEF_FLAT),],
            [sg.Text('Confidence',size=(12,1),font=('Helvetica',15), text_color='red'), sg.Slider(range=(1,100),default_value=30,orientation='h',size=(60,20),font=('Helvetica',11),disabled=True, key= 'conf_thres1'),],

            [sg.Text('')],
            
            [sg.Text('Thiết định thông số bụi chì [ cuộn cảm ] : ',size=(35,1),font=('Helvetica',15), text_color='red'),
            sg.Text('With min', size=(7,1), font=('Helvetica',15),text_color='red'),
            sg.InputText(default_text=load_values()[0], size=(7, 1), font=('Helvetica', 15), key='W_min_bc_cc', text_color='navy',enable_events=True, disabled=True),
            sg.Text('Height min', size=(8,1), font=('Helvetica',15),text_color='red'),
            sg.InputText(default_text=load_values()[1], size=(7, 1), font=('Helvetica', 15), key='H_min_bc_cc', text_color='navy',enable_events=True, disabled=True),sg.Text('Confidence', size=(9,1), font=('Helvetica',15),text_color='red'),
            sg.Slider(range=(1,100),default_value=load_values()[2],orientation='h',size=(28,15),font=('Helvetica',10), key= 'bc_cc_Conf_1',enable_events=True,disabled=True),],
            [sg.Text('')],

            [sg.Text('Thiết định thông số bụi chì [ phân vùng 1 ] : ',size=(35,1),font=('Helvetica',15), text_color='red'),
            sg.Text('With min', size=(7,1), font=('Helvetica',15),text_color='red'),
            sg.InputText(default_text=load_values()[0], size=(7, 1), font=('Helvetica', 15), key='W_min_bc_cc', text_color='navy',enable_events=True, disabled=True),
            sg.Text('Height min', size=(8,1), font=('Helvetica',15),text_color='red'),
            sg.InputText(default_text=load_values()[1], size=(7, 1), font=('Helvetica', 15), key='H_min_bc_cc', text_color='navy',enable_events=True, disabled=True),sg.Text('Confidence', size=(9,1), font=('Helvetica',15),text_color='red'),
            sg.Slider(range=(1,100),default_value=load_values()[2],orientation='h',size=(28,15),font=('Helvetica',10), key= 'bc_cc_Conf_1',enable_events=True,disabled=True),],
            [sg.Text('')],

            [sg.Text('Thiết định thông số bụi chì [ phân vùng 2 ] : ',size=(35,1),font=('Helvetica',15), text_color='red'),
            sg.Text('With min', size=(7,1), font=('Helvetica',15),text_color='red'),
            sg.InputText(default_text=load_values()[0], size=(7, 1), font=('Helvetica', 15), key='W_min_bc_cc', text_color='navy',enable_events=True, disabled=True),
            sg.Text('Height min', size=(8,1), font=('Helvetica',15),text_color='red'),
            sg.InputText(default_text=load_values()[1], size=(7, 1), font=('Helvetica', 15), key='H_min_bc_cc', text_color='navy',enable_events=True, disabled=True),sg.Text('Confidence', size=(9,1), font=('Helvetica',15),text_color='red'),
            sg.Slider(range=(1,100),default_value=load_values()[2],orientation='h',size=(28,15),font=('Helvetica',10), key= 'bc_cc_Conf_1',enable_events=True,disabled=True),],
            [sg.Text('')],

            
        ], relief=sg.RELIEF_FLAT),
        ],
        [sg.Frame('',[
            [sg.Text('Name',size=(12,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Join',size=(5,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('OK',size=(5,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Num',size=(5,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('NG',size=(5,1),font=('Helvetica',15), text_color='red'),  
            sg.Text('Width Min',size=(9,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Width Max',size=(9,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Min',size=(9,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Max',size=(12,1),font=('Helvetica',15), text_color='red'),
            sg.Text('PLC',size=(9,1),font=('Helvetica',15), text_color='red'),
            sg.Text('Private Conf',size=(10,1),font=('Helvetica',15), text_color='red')],
        ], relief=sg.RELIEF_FLAT)],

        [sg.Frame('',[
            [
                sg.Text(f'{model1.names[i1]}_1',size=(12,1),font=('Helvetica',15), text_color='yellow'), 
                sg.Checkbox('',size=(3,1),default=True,font=('Helvetica',15),  key=f'{model1.names[i1]}_1',enable_events=True, disabled=True), 
                sg.Radio('',group_id=f'Cam1 {i1}',size=(3,1),default=False,font=('Helvetica',15),  key=f'{model1.names[i1]}_OK_1',enable_events=True, disabled=True), 
                sg.Input('1',size=(2,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Num_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Radio('',group_id=f'Cam1 {i1}',size=(2,1),default=False,font=('Helvetica',15),  key=f'{model1.names[i1]}_NG_1',enable_events=True, disabled=True), 
                sg.Input('0',size=(7,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Wn_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(1,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('1600',size=(7,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Wx_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(1,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(7,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Hn_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(1,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('1200',size=(7,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Hx_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(7,1),font=('Helvetica',15),key= f'{model1.names[i1]}_PLC_1',text_color='navy',enable_events=True, disabled=True),
                sg.Slider(range=(1,100),default_value=30,orientation='h',size=(28,9),font=('Helvetica',10), key= f'{model1.names[i1]}_Conf_1',disabled=True),
            ] for i1 in range(len(model1.names))
        ], relief=sg.RELIEF_FLAT)],
        
        [sg.Text('  OK',size=(15,1),font=('Helvetica',15), text_color='yellow'),
        sg.Text(' '*178), 
        sg.Input('420',size=(7,1),font=('Helvetica',15),key= 'OK_PLC_1',text_color='navy',enable_events=True,disabled=True)],
        [sg.Text(' ')],
        [sg.Text(' '*250), sg.Button('Save Data', size=(12,1),  font=('Helvetica',12),key='SaveData1',enable_events=True,disabled=True)] 
        ])]
    ]
    
    layout_option2 = [
        [sg.Frame('',[
        [sg.Frame('',[
            [sg.Text('Weights', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Input(size=(60,1), font=('Helvetica',12), key='file_weights2',readonly= True, text_color='navy',enable_events= True),
            sg.Frame('',[
                [sg.FileBrowse(file_types= file_weights, size=(12,1), font=('Helvetica',10),key= 'file_browse2',enable_events=True, disabled=True)]
            ], relief= sg.RELIEF_FLAT),
            sg.Frame('',[
                [sg.Button('Change Model', size=(14,1), font=('Helvetica',10), disabled= True, key= 'Change_2')]
            ], relief= sg.RELIEF_FLAT),],
            [sg.Text('Confidence',size=(12,1),font=('Helvetica',15), text_color='red'), sg.Slider(range=(1,100),default_value=30,orientation='h',size=(60,20),font=('Helvetica',11),disabled=True, key= 'conf_thres2'),sg.Button('Save Data', size=(12,1),  font=('Helvetica',12),key='SaveData2', enable_events=True, disabled=True)],

        ], relief=sg.RELIEF_FLAT, expand_y= True),],
        [sg.Frame('',[
            [sg.Text('Name',size=(12,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Join',size=(5,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('OK',size=(5,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Num',size=(5,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('NG',size=(5,1),font=('Helvetica',15), text_color='red'),  
            sg.Text('Width Min',size=(9,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Width Max',size=(9,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Min',size=(9,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Max',size=(12,1),font=('Helvetica',15), text_color='red'),
            sg.Text('PLC',size=(9,1),font=('Helvetica',15), text_color='red'),
            sg.Text('Private Conf',size=(10,1),font=('Helvetica',15), text_color='red')],
        ], relief=sg.RELIEF_FLAT)],

        [sg.Frame('',[
            [
                sg.Text(f'{model2.names[i2]}_2',size=(12,1),font=('Helvetica',15), text_color='yellow'), 
                sg.Checkbox('',size=(3,1),default=True,font=('Helvetica',15),  key=f'{model2.names[i2]}_2',enable_events=True, disabled=True), 
                sg.Radio('',group_id=f'Cam2 {i2}',size=(3,1),font=('Helvetica',15),  key=f'{model2.names[i2]}_OK_2',enable_events=True, disabled=True), 
                sg.Input('1',size=(2,1),font=('Helvetica',15),key= f'{model2.names[i2]}_Num_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Radio('',group_id=f'Cam2 {i2}',size=(2,1),font=('Helvetica',15),  key=f'{model2.names[i2]}_NG_2',enable_events=True, disabled=True), 
                sg.Input('0',size=(7,1),font=('Helvetica',15),key= f'{model2.names[i2]}_Wn_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(1,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('1600',size=(7,1),font=('Helvetica',15),key= f'{model2.names[i2]}_Wx_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(1,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(7,1),font=('Helvetica',15),key= f'{model2.names[i2]}_Hn_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(1,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('1200',size=(7,1),font=('Helvetica',15),key= f'{model2.names[i2]}_Hx_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(7,1),font=('Helvetica',15),key= f'{model2.names[i2]}_PLC_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Slider(range=(1,100),default_value=30,orientation='h',size=(28,9),font=('Helvetica',11), key= f'{model2.names[i2]}_Conf_2', disabled=True),
            ] for i2 in range(len(model2.names))
        ], relief=sg.RELIEF_FLAT)],

        [sg.Text('  OK',size=(15,1),font=('Helvetica',15), text_color='yellow'),
        sg.Text(' '*178), 
        sg.Input('1',size=(7,1),font=('Helvetica',15),key= 'OK_PLC_2',text_color='navy',enable_events=True, disabled=True)],
        [sg.Text(' ')],
        ])]
    ]


    layout_option3 = [
        [sg.Frame('',[
        [sg.Frame('',
        [   
            [sg.Text('Weights', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Input(size=(60,1), font=('Helvetica',12), key='file_weights3',readonly= True, text_color='navy',enable_events= True),
            sg.Frame('',[
                [sg.FileBrowse(file_types= file_weights, size=(12,1), font=('Helvetica',10),key= 'file_browse3',enable_events=True, disabled=True)]
            ], relief= sg.RELIEF_FLAT),
            sg.Frame('',[
                [sg.Button('Change Model', size=(14,1), font=('Helvetica',10), disabled= True, key= 'Change_3')]
            ], relief= sg.RELIEF_FLAT),],
            [sg.Text('Confidence',size=(12,1),font=('Helvetica',15), text_color='red'), sg.Slider(range=(1,100),default_value=30,orientation='h',size=(60,20),font=('Helvetica',11),disabled=True, key= 'conf_thres3'),sg.Button('Save Data', size=(12,1),  font=('Helvetica',12),key='SaveData3', enable_events=True, disabled=True)]
        ], relief=sg.RELIEF_FLAT),
        ],
        [sg.Frame('',[
            [sg.Text('Name',size=(12,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Join',size=(5,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('OK',size=(5,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Num',size=(5,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('NG',size=(5,1),font=('Helvetica',15), text_color='red'),  
            sg.Text('Width Min',size=(9,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Width Max',size=(9,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Min',size=(9,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Max',size=(12,1),font=('Helvetica',15), text_color='red'),
            sg.Text('PLC',size=(9,1),font=('Helvetica',15), text_color='red'),
            sg.Text('Private Conf',size=(10,1),font=('Helvetica',15), text_color='red')],
        ], relief=sg.RELIEF_FLAT)],

        [sg.Frame('',[
            [
                sg.Text(f'{model3.names[i3]}_3',size=(12,1),font=('Helvetica',15), text_color='yellow'), 
                sg.Checkbox('',size=(3,1),default=True,font=('Helvetica',15),  key=f'{model3.names[i3]}_3',enable_events=True, disabled=True), 
                sg.Radio('',group_id=f'Cam3 {i3}',size=(3,1),font=('Helvetica',15),  key=f'{model3.names[i3]}_OK_3',enable_events=True, disabled=True), 
                sg.Input('1',size=(2,1),font=('Helvetica',15),key= f'{model3.names[i3]}_Num_3',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Radio('',group_id=f'Cam3 {i3}',size=(2,1),font=('Helvetica',15),  key=f'{model3.names[i3]}_NG_3',enable_events=True, disabled=True), 
                sg.Input('0',size=(7,1),font=('Helvetica',15),key= f'{model3.names[i3]}_Wn_3',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(1,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('1600',size=(7,1),font=('Helvetica',15),key= f'{model3.names[i3]}_Wx_3',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(1,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(7,1),font=('Helvetica',15),key= f'{model3.names[i3]}_Hn_3',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(1,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('1200',size=(7,1),font=('Helvetica',15),key= f'{model3.names[i3]}_Hx_3',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(7,1),font=('Helvetica',15),key= f'{model3.names[i3]}_PLC_3',text_color='navy',enable_events=True, disabled=True), 
                sg.Slider(range=(1,100),default_value=30,orientation='h',size=(28,9),font=('Helvetica',11), key= f'{model3.names[i3]}_Conf_3', disabled=True),
            ] for i3 in range(len(model3.names))
        ], relief=sg.RELIEF_FLAT)],

        [sg.Text('  OK',size=(15,1),font=('Helvetica',15), text_color='yellow'),
        sg.Text(' '*178), 
        sg.Input('1',size=(7,1),font=('Helvetica',15),key= 'OK_PLC_3',text_color='navy',enable_events=True, disabled=True)],
        [sg.Text(' ')],
        [sg.Text(' '*250), sg.Button('Save Data', size=(12,1),  font=('Helvetica',12),key='SaveData3', enable_events=True, disabled=True)] 
        ])]
    ]

    layout_savimg = [
        [sg.Frame('',[
                [sg.Text('Have save folder image OK for camera 1',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=False,font=('Helvetica',15),  key='have_save_OK_1',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image OK for camera 1', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='D:/X75/' + maude + '/OK/OPT1' ,font=('Helvetica',12), key='save_OK_1',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_OK_1',enable_events=True) ],
                [sg.Text('')],
                [sg.Text('Have save folder image OK for camera 2',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=False,font=('Helvetica',15),  key='have_save_OK_2',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image OK for camera 2', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='D:/X75/' + maude + '/OK/OPT2' , font=('Helvetica',12), key='save_OK_2',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_OK_2',enable_events=True) ],
                [sg.Text('')],
                [sg.Text('Have save folder image NG for camera 1',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key='have_save_NG_1',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image NG for camera 1', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='D:/X75/' + maude + '/NG/OPT1' , font=('Helvetica',12), key='save_NG_1',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_NG_1',enable_events=True) ],
                [sg.Text('')],
                [sg.Text('Have save folder image NG for camera 2',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key='have_save_NG_2',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image NG for camera 2', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='D:/X75/' + maude + '/NG/OPT2' , font=('Helvetica',12), key='save_NG_2',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_NG_2',enable_events=True) ],
        ], relief=sg.RELIEF_FLAT),
        sg.Frame('',[
                [sg.Text('Have save folder image OK for camera 3',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=False,font=('Helvetica',15),  key='have_save_OK_3',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image OK for camera 3', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='D:/X75/' + maude + '/OK/OPT3' ,font=('Helvetica',12), key='save_OK_3',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_OK_3',enable_events=True) ],
                [sg.Text('')],
                [sg.Text('Have save folder image NG for camera 3',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key='have_save_NG_3',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image NG for camera 3', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='D:/X75/' + maude + '/NG/OPT3' , font=('Helvetica',12), key='save_NG_3',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_NG_3',enable_events=True) ],
                [sg.Text('')],
                [sg.T('Choose folder save File CSV log', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='D:/CSV' , font=('Helvetica',12), key='save_CSV',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_CSV',enable_events=True) ],
                [sg.Text('')],
                [sg.Button('Out CSV', size=(12,1),  font=('Helvetica',12),key='OutCSV',enable_events=True),sg.Checkbox('',size=(3,1),default=False,font=('Helvetica',15),  key='checkpass',enable_events=True),], 
        ], relief=sg.RELIEF_FLAT)],
        ]
    # layout_terminal = [[sg.Text("Anything printed will display here!")],
    #                   [sg.Multiline( font=('Helvetica',14), write_only=True, autoscroll=True, auto_refresh=True,reroute_stdout=True, reroute_stderr=True, echo_stdout_stderr=True,expand_x=True,expand_y=True)]
    #                   ]
    layout_option1 = [[sg.Column(layout_option1, scrollable = True, expand_x = True, expand_y = True)]]
    layout_option2 = [[sg.Column(layout_option2, scrollable = True, expand_x = True, expand_y = True)]]
    layout_option3 = [[sg.Column(layout_option3, scrollable = True, expand_x = True, expand_y = True)]]

    
    layout = [[sg.TabGroup([[  sg.Tab('Main', layout_main),
                               sg.Tab('Cam1', layout_cam1),
                               sg.Tab('Cam2', layout_cam2),
                               sg.Tab('Option for model 1', layout_option1),
                               sg.Tab('Option for model 2', layout_option2),
                               sg.Tab('Option for model 3', layout_option3),
                               sg.Tab('Save Image', layout_savimg)]])
                               #sg.Tab('Output', layout_terminal)]])
               ]]

    window = sg.Window('HuynhLeVu', layout, location=(0,0),right_click_menu=right_click_menu,resizable=True).Finalize()
    #window.bind('<Configure>',"Configure")
    window.Maximize()

    return window


image_width_display = int(750*0.8)
image_height_display = int(480*0.8)

image_width_display_c2 = int(image_width_display/2)
image_height_display_c2 = int(image_height_display/2)

result_width_display = 400
result_height_display = 100 


file_name_img = [("Img(*.jpg,*.png)",("*jpg","*.png"))]


recording1 = False
recording2 = False 

error_cam1 = True
error_cam2 = True

recording3 = False

error_cam3 = True
time_all=0
myitem = 0 
all_error = []
all_item_error = []
all_item_error1 = []
changed = 0
c_hinh = ''

os.system('shutdown -a') #Cancel order shutdown if exist

# connected = False
# while connected == False:
#     print('connecting....')
#     connected = plc.socket_connect('192.168.250.20', 8000)
# print("connected plc")  

# run_plc_keyence("192.168.1.1", 8501)

# #write_plc_keyence(17406,1)
# print(read_plc_keyence(17406))


mypath1 = load_model(1)
model1 = torch.hub.load('./levu','custom', path= mypath1, source='local',force_reload =False)
if mypath1[-7:] == 'edit.pt': 
    change_label(model1)


img1_test = os.path.join(os.getcwd(), 'img/imgtest.jpg')
result1 = model1(img1_test,608,0.25) 
print('model1 already')


mypath2 = load_model(2)
model2 = torch.hub.load('./levu','custom', path= mypath2, source='local',force_reload =False)

img2_test = os.path.join(os.getcwd(), 'img/imgtest.jpg')
result2 = model2(img2_test,608,0.25) 

print('model2 already')

mypath3 = load_model(3)
model3 = torch.hub.load('./levu','custom', path= mypath3, source='local',force_reload =False)

img1_test = os.path.join(os.getcwd(), 'img/imgtest.jpg')
result3 = model3(img1_test,608,0.25) 
print('model3 already')
# print(read_plc_keyence(400))

themes = load_theme()
theme = themes[0]
choose_model = load_choosemodel()

if choose_model == '1':
    maude = 'TRANG'
    # window = make_window(theme)
    # window['choosemodel_running'].update(value = 'Đế TRẮNG')

    # size_model = 608

if choose_model == '3':

    maude = 'DEN'
    # window = make_window(theme)
    # window['choosemodel_running'].update(value = 'Đế ĐEN')

# #     # size_model = 6
window = make_window(theme)


# if read_plc_keyence(17402) ==1:
#     choose_model = '3'
#     maude = 'DEN'
#     window['choosemodel_running'].update(value = 'Đế ĐEN')
#     print('Đế ĐEN')
# if read_plc_keyence(17400) ==1:
#     choose_model ='1'
#     maude = 'TRANG'
#     window['choosemodel_running'].update(value = 'Đế TRẮNG')
#     print('Đế TRẮNG')
# window['choose_model'].update(value=choose_model)
# time.sleep(4)
started = 1

# try:
#     load_all_sql(1,choose_model)
# except:
#     pass
#     # window['time_cam1'].update(value= "Error data") 


# try:
#     load_all_sql(2,choose_model)
# except:
#     pass
#     # window['time_cam2'].update(value= "Error data") 

# try:
#     load_all_sql(3,choose_model)
# except:
#     pass
#     # window['time_cam2'].update(value= "Error data") 



connect_camera1 = False
connect_camera2 = False
connect_camera3 = False

connect_total = False


if connect_camera1 == True and connect_total == True:
    window['result_cam1'].update(value= 'Done', text_color='blue')
if connect_camera2 == True and connect_total == True:
    window['result_cam2'].update(value= 'Done', text_color='blue')

# #Reset 
# #write_plc_keyence(450,0)
# #write_plc_keyence(460,0)
try:
    removefile()
    time.sleep(2)
except:
    pass
#Bao cho PLC reset all
# #write_plc_keyence(600,1)

#ket noi dabase detect luu phe pham
pp = sqlite3.connect('defect.db')

try:
    while True:
        event, values = window.read(timeout=20)
       
        if event =='Exit' or event == sg.WINDOW_CLOSED :
            #write_plc_keyence(600,0) 
            break

        if event == 'OutCSV':
            df = pd.read_sql_query("select * from PHEPHAM", pp)
            fcsv = values['save_CSV'] +'/' + time_to_name() + '.csv'
            df.to_csv(fcsv)
            sg.popup('Duong dan:\n' + fcsv, title="Thong bao da xuat CSV")

        if event =='checkpass' and values['checkpass'] == True:
              
            window['conf_thres2'].update(disabled= False)
            window['conf_thres1'].update(disabled= False)

            window['file_browse2'].update(disabled= False,button_color='turquoise')
            window['file_browse1'].update(disabled= False,button_color='turquoise')

            window['SaveData1'].update(disabled= False,button_color='turquoise')
            window['SaveData2'].update(disabled= False,button_color='turquoise')

            window['Webcam1'].update(disabled= True,button_color='turquoise')
            window['Webcam2'].update(disabled= True,button_color='turquoise')
            window['Stop1'].update(disabled= False,button_color='turquoise')
            window['Stop2'].update(disabled= False,button_color='turquoise')
            window['Pic1'].update(disabled= False,button_color='turquoise')
            window['Pic2'].update(disabled= False,button_color='turquoise')
            window['Snap1'].update(disabled= True,button_color='turquoise')
            window['Snap2'].update(disabled= True,button_color='turquoise')
            window['Change1'].update(button_color='turquoise')
            window['Change2'].update(button_color='turquoise')
            window['Change_1'].update(button_color='turquoise')
            window['Change_2'].update(button_color='turquoise')
            window['Detect1'].update(button_color='turquoise')
            window['Detect2'].update(button_color='turquoise')


            window['have_save_OK_1'].update(disabled=False)
            window['have_save_NG_1'].update(disabled=False)
            window['have_save_OK_2'].update(disabled=False)
            window['have_save_NG_2'].update(disabled=False)

            window['save_OK_1'].update(disabled=False)
            window['save_NG_1'].update(disabled=False)
            window['save_OK_2'].update(disabled=False)
            window['save_NG_2'].update(disabled=False)

            window['save_folder_OK_1'].update(disabled= False,button_color='turquoise')
            window['save_folder_NG_1'].update(disabled= False,button_color='turquoise')
            window['save_folder_OK_2'].update(disabled= False,button_color='turquoise')
            window['save_folder_NG_2'].update(disabled= False,button_color='turquoise')

            window['OK_PLC_1'].update(disabled=False)
            window['OK_PLC_2'].update(disabled=False)
            window['OK_PLC_3'].update(disabled=False)

            for i1 in range(len(model1.names)):
                window[f'{model1.names[i1]}_1'].update(disabled=False)
                window[f'{model1.names[i1]}_OK_1'].update(disabled=False)
                window[f'{model1.names[i1]}_Num_1'].update(disabled=False)
                window[f'{model1.names[i1]}_NG_1'].update(disabled=False)
                window[f'{model1.names[i1]}_Wn_1'].update(disabled=False)
                window[f'{model1.names[i1]}_Wx_1'].update(disabled=False)
                window[f'{model1.names[i1]}_Hn_1'].update(disabled=False)
                window[f'{model1.names[i1]}_Hx_1'].update(disabled=False)
                window[f'{model1.names[i1]}_PLC_1'].update(disabled=False)
                window[f'{model1.names[i1]}_Conf_1'].update(disabled=False)
                window['W_min_bc_cc'].update(disabled=False)
                window['H_min_bc_cc'].update(disabled=False)
                window['bc_cc_Conf_1'].update(disabled=False)
                # window['W_min_bc_cb'].update(disabled=False)
                # window['H_min_bc_cb'].update(disabled=False)
                

            for i2 in range(len(model2.names)):
                window[f'{model2.names[i2]}_2'].update(disabled=False)
                window[f'{model2.names[i2]}_OK_2'].update(disabled=False)
                window[f'{model2.names[i2]}_Num_2'].update(disabled=False)
                window[f'{model2.names[i2]}_NG_2'].update(disabled=False)
                window[f'{model2.names[i2]}_Wn_2'].update(disabled=False)
                window[f'{model2.names[i2]}_Wx_2'].update(disabled=False)
                window[f'{model2.names[i2]}_Hn_2'].update(disabled=False)
                window[f'{model2.names[i2]}_Hx_2'].update(disabled=False)
                window[f'{model2.names[i2]}_PLC_2'].update(disabled=False)
                window[f'{model2.names[i2]}_Conf_2'].update(disabled=False)

            window['conf_thres3'].update(disabled= False)

            window['file_browse3'].update(disabled= False,button_color='turquoise')

            window['SaveData3'].update(disabled= False,button_color='turquoise')

            window['have_save_OK_3'].update(disabled=False)
            window['have_save_NG_3'].update(disabled=False)
            
            window['save_OK_3'].update(disabled=False)
            window['save_NG_3'].update(disabled=False)

            window['save_folder_OK_3'].update(disabled= False,button_color='turquoise')
            window['save_folder_NG_3'].update(disabled= False,button_color='turquoise')

            for i3 in range(len(model3.names)):
                window[f'{model3.names[i3]}_3'].update(disabled=False)
                window[f'{model3.names[i3]}_OK_3'].update(disabled=False)
                window[f'{model3.names[i3]}_Num_3'].update(disabled=False)
                window[f'{model3.names[i3]}_NG_3'].update(disabled=False)
                window[f'{model3.names[i3]}_Wn_3'].update(disabled=False)
                window[f'{model3.names[i3]}_Wx_3'].update(disabled=False)
                window[f'{model3.names[i3]}_Hn_3'].update(disabled=False)
                window[f'{model3.names[i3]}_Hx_3'].update(disabled=False)
                window[f'{model3.names[i3]}_PLC_3'].update(disabled=False)
                window[f'{model3.names[i3]}_Conf_3'].update(disabled=False)

        if event =='checkpass' and values['checkpass'] == False:
            
            window['conf_thres2'].update(disabled= True)
            window['conf_thres1'].update(disabled= True)

            window['file_browse2'].update(disabled= True,button_color='turquoise')
            window['file_browse1'].update(disabled= True,button_color='turquoise')

            window['SaveData1'].update(disabled= True,button_color='turquoise')
            window['SaveData2'].update(disabled= True,button_color='turquoise')

            window['Webcam1'].update(disabled= True,button_color='turquoise')
            window['Webcam2'].update(disabled= True,button_color='turquoise')
            window['Stop1'].update(disabled= True,button_color='turquoise')
            window['Stop2'].update(disabled= True,button_color='turquoise')
            window['Pic1'].update(disabled= True,button_color='turquoise')
            window['Pic2'].update(disabled= True,button_color='turquoise')
            window['Snap1'].update(disabled= True,button_color='turquoise')
            window['Snap2'].update(disabled= True,button_color='turquoise')
            window['Change1'].update(button_color='turquoise')
            window['Change2'].update(button_color='turquoise')
            window['Change_1'].update(button_color='turquoise')
            window['Change_2'].update(button_color='turquoise')
            window['Detect1'].update(button_color='turquoise')
            window['Detect2'].update(button_color='turquoise')


            window['have_save_OK_1'].update(disabled=True)
            window['have_save_NG_1'].update(disabled=True)
            window['have_save_OK_2'].update(disabled=True)
            window['have_save_NG_2'].update(disabled=True)

            window['save_OK_1'].update(disabled=True)
            window['save_NG_1'].update(disabled=True)
            window['save_OK_2'].update(disabled=True)
            window['save_NG_2'].update(disabled=True)

            window['save_folder_OK_1'].update(disabled= True,button_color='turquoise')
            window['save_folder_NG_1'].update(disabled= True,button_color='turquoise')
            window['save_folder_OK_2'].update(disabled= True,button_color='turquoise')
            window['save_folder_NG_2'].update(disabled= True,button_color='turquoise')

            window['OK_PLC_1'].update(disabled=True)
            window['OK_PLC_2'].update(disabled=True)
            window['OK_PLC_3'].update(disabled=True)

            for i1 in range(len(model1.names)):
                window[f'{model1.names[i1]}_1'].update(disabled=True)
                window[f'{model1.names[i1]}_OK_1'].update(disabled=True)
                window[f'{model1.names[i1]}_Num_1'].update(disabled=True)
                window[f'{model1.names[i1]}_NG_1'].update(disabled=True)
                window[f'{model1.names[i1]}_Wn_1'].update(disabled=True)
                window[f'{model1.names[i1]}_Wx_1'].update(disabled=True)
                window[f'{model1.names[i1]}_Hn_1'].update(disabled=True)
                window[f'{model1.names[i1]}_Hx_1'].update(disabled=True)
                window[f'{model1.names[i1]}_PLC_1'].update(disabled=True)
                window[f'{model1.names[i1]}_Conf_1'].update(disabled=True)
                

            for i2 in range(len(model2.names)):
                window[f'{model2.names[i2]}_2'].update(disabled=True)
                window[f'{model2.names[i2]}_OK_2'].update(disabled=True)
                window[f'{model2.names[i2]}_Num_2'].update(disabled=True)
                window[f'{model2.names[i2]}_NG_2'].update(disabled=True)
                window[f'{model2.names[i2]}_Wn_2'].update(disabled=True)
                window[f'{model2.names[i2]}_Wx_2'].update(disabled=True)
                window[f'{model2.names[i2]}_Hn_2'].update(disabled=True)
                window[f'{model2.names[i2]}_Hx_2'].update(disabled=True)
                window[f'{model2.names[i2]}_PLC_2'].update(disabled=True)
                window[f'{model2.names[i2]}_Conf_2'].update(disabled=True)

            window['conf_thres3'].update(disabled= True)

            window['file_browse3'].update(disabled= True,button_color='turquoise')

            window['SaveData3'].update(disabled= True,button_color='turquoise')

            window['have_save_OK_3'].update(disabled=True)
            window['have_save_NG_3'].update(disabled=True)
            
            window['save_OK_3'].update(disabled=True)
            window['save_NG_3'].update(disabled=True)

            window['save_folder_OK_3'].update(disabled= True,button_color='turquoise')
            window['save_folder_NG_3'].update(disabled= True,button_color='turquoise')

            for i3 in range(len(model3.names)):
                window[f'{model3.names[i3]}_3'].update(disabled=True)
                window[f'{model3.names[i3]}_OK_3'].update(disabled=True)
                window[f'{model3.names[i3]}_Num_3'].update(disabled=True)
                window[f'{model3.names[i3]}_NG_3'].update(disabled=True)
                window[f'{model3.names[i3]}_Wn_3'].update(disabled=True)
                window[f'{model3.names[i3]}_Wx_3'].update(disabled=True)
                window[f'{model3.names[i3]}_Hn_3'].update(disabled=True)
                window[f'{model3.names[i3]}_Hx_3'].update(disabled=True)
                window[f'{model3.names[i3]}_PLC_3'].update(disabled=True)
                window[f'{model3.names[i3]}_Conf_3'].update(disabled=True)            

        if event =='Administrator':
            login_password = 'vu123ppp'  # helloworld
            password = sg.popup_get_text(
                'Enter PassworE: ', password_char='*') 
            if password == login_password:
                sg.popup_ok('Login Successed!!! ',text_color='green', font=('Helvetica',14))  

                window['conf_thres2'].update(disabled= False)
                window['conf_thres1'].update(disabled= False)

                window['file_browse2'].update(disabled= False,button_color='turquoise')
                window['file_browse1'].update(disabled= False,button_color='turquoise')

                window['SaveData1'].update(disabled= False,button_color='turquoise')
                window['SaveData2'].update(disabled= False,button_color='turquoise')

                window['Webcam1'].update(disabled= True,button_color='turquoise')
                window['Webcam2'].update(disabled= True,button_color='turquoise')
                window['Stop1'].update(disabled= False,button_color='turquoise')
                window['Stop2'].update(disabled= False,button_color='turquoise')
                window['Pic1'].update(disabled= False,button_color='turquoise')
                window['Pic2'].update(disabled= False,button_color='turquoise')
                window['Snap1'].update(disabled= True,button_color='turquoise')
                window['Snap2'].update(disabled= True,button_color='turquoise')
                window['Change1'].update(button_color='turquoise')
                window['Change2'].update(button_color='turquoise')
                window['Change_1'].update(button_color='turquoise')
                window['Change_2'].update(button_color='turquoise')
                window['Detect1'].update(button_color='turquoise')
                window['Detect2'].update(button_color='turquoise')


                window['have_save_OK_1'].update(disabled=False)
                window['have_save_NG_1'].update(disabled=False)
                window['have_save_OK_2'].update(disabled=False)
                window['have_save_NG_2'].update(disabled=False)

                window['save_OK_1'].update(disabled=False)
                window['save_NG_1'].update(disabled=False)
                window['save_OK_2'].update(disabled=False)
                window['save_NG_2'].update(disabled=False)

                window['save_folder_OK_1'].update(disabled= False,button_color='turquoise')
                window['save_folder_NG_1'].update(disabled= False,button_color='turquoise')
                window['save_folder_OK_2'].update(disabled= False,button_color='turquoise')
                window['save_folder_NG_2'].update(disabled= False,button_color='turquoise')

                window['OK_PLC_1'].update(disabled=False)
                window['OK_PLC_2'].update(disabled=False)
                window['OK_PLC_3'].update(disabled=False)

                for i1 in range(len(model1.names)):
                    window[f'{model1.names[i1]}_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_OK_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Num_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_NG_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Wn_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Wx_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Hn_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Hx_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_PLC_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Conf_1'].update(disabled=False)
                    

                for i2 in range(len(model2.names)):
                    window[f'{model2.names[i2]}_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_OK_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Num_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_NG_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Wn_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Wx_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Hn_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Hx_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_PLC_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Conf_2'].update(disabled=False)

                window['conf_thres3'].update(disabled= False)

                window['file_browse3'].update(disabled= False,button_color='turquoise')

                window['SaveData3'].update(disabled= False,button_color='turquoise')

                window['have_save_OK_3'].update(disabled=False)
                window['have_save_NG_3'].update(disabled=False)
                
                window['save_OK_3'].update(disabled=False)
                window['save_NG_3'].update(disabled=False)

                window['save_folder_OK_3'].update(disabled= False,button_color='turquoise')
                window['save_folder_NG_3'].update(disabled= False,button_color='turquoise')

                for i3 in range(len(model3.names)):
                    window[f'{model3.names[i3]}_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_OK_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_Num_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_NG_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_Wn_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_Wx_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_Hn_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_Hx_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_PLC_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_Conf_3'].update(disabled=False)

            else:
                sg.popup_cancel('Wrong Password!!!',text_color='red', font=('Helvetica',14))

        if event == 'Change Theme':
            layout_theme = [
                [sg.Listbox(values= sg.theme_list(), size = (30,20),auto_size_text=18,default_values='Dark',key='theme', enable_events=True)],
                [
                    [sg.Button('Apply'),
                    sg.Button('Cancel')]
                ]
            ] 
            window_theme = sg.Window('Change Theme', layout_theme, location=(50,50),keep_on_top=True).Finalize()
            window_theme.set_min_size((300,400))   

            while True:
                event_theme, values_theme = window_theme.read(timeout=20)
                if event_theme == sg.WINDOW_CLOSED:
                    break

                if event_theme == 'Apply':
                    theme_choose = values_theme['theme'][0]
                    if theme_choose == 'Default':
                        continue
                    window.close()
                    window = make_window(theme_choose)
                    save_theme(theme_choose)
                    #print(theme_choose)
                if event_theme == 'Cancel':
                    answer = sg.popup_yes_no('Do you want to exit?')
                    if answer == 'Yes':
                        break
                    if answer == 'No':
                        continue
            window_theme.close()

        if event == 'file_browse1': 
            window['file_weights1'].update(value=values['file_browse1'])
            if values['file_browse1']:
                window['Change1'].update(disabled=False)
                window['Change_1'].update(disabled=False)

        if event == 'file_browse2':
            window['file_weights2'].update(value=values['file_browse2'])
            if values['file_browse2']:
                window['Change2'].update(disabled=False)
                window['Change_2'].update(disabled=False)

        if event == 'file_browse3': 
            window['file_weights3'].update(value=values['file_browse3'])
            if values['file_browse3']:
                window['Change_3'].update(disabled=False)

        # change_chooose_model_trang = read_plc_keyence(17400)
        # change_chooose_model_den = read_plc_keyence(17402)
        # change_chooose_model_sub = read_plc_keyence(17404)

        # if change_chooose_model_trang == 1  and change_chooose_model_sub ==1:
        #     window['choose_model'].update(value = '1')
        #     values['choose_model'] = '1'
        #     print('Trang',values['choose_model'])
        #     maude = 'TRANG'
        #     # #write_plc_keyence(620,0)
        #     #write_plc_keyence(17404,0)

        #     # size_model = 608
        #     # print('size model DE TRANG la: ',size_model) 
        #     changed = 1

        # if change_chooose_model_den == 1 and change_chooose_model_sub ==2:
        #     window['choose_model'].update(value = '3')
        #     values['choose_model'] = '3'
        #     print('Den',values['choose_model'])
        #     maude = 'DEN'
        #     # #write_plc_keyence(620,0)
        #     #write_plc_keyence(17404,0)

        #     # size_model = 416
        #     # print('size model DE DEN la: ',size_model) 
        #     changed = 1

        if changed==1 or event == 'choose_model' or started == 1:
            
            # mychoose = values['choose_model']
            print('Changed')
            weight1 = ''
            conf_thres1 = 1
            weight2 = ''
            conf_thres2 = 1

            OK_Cam1 = False
            OK_Cam2 = False
            NG_Cam1 = True
            NG_Cam2 = True
            Folder_OK_Cam1 = 'C:/Cam1/OK'
            Folder_OK_Cam2 = 'C:/Cam2/OK'
            Folder_NG_Cam1 = 'C:/Cam1/NG'
            Folder_NG_Cam2 = 'C:/Cam2/NG'

            weight3 = ''
            conf_thres3 = 1

            OK_Cam3 = False
            NG_Cam3 = True

            Folder_OK_Cam3 = 'C:/Cam3/OK'
            Folder_ = 'C:/OK'
            Folder_NG_Cam3 = 'C:/Cam3/NG'
            Folder_NG_ = 'C:/NG'
            try:
                conn = sqlite3.connect('2cam_3model.db')
                cursor = conn.execute("SELECT * from MYMODEL")
            except:
                print('loi ket noi csdl')
            for row in cursor:
                if row[0] == values['choose_model']:
 
                    mychoose = values['choose_model']
                    row1_a, row1_b = row[1].strip().split('_')
                    if row1_a == '1' and row1_b == '0':
                        weight1 = row[2]
                        conf_thres1 = row[3]
                        OK_Cam1 = str2bool(row[4])
                        OK_Cam2 = str2bool(row[5])
                        OK_Cam3 = str2bool(row[6])
                        NG_Cam1 = str2bool(row[7])
                        NG_Cam2 = str2bool(row[8])
                        NG_Cam3 = str2bool(row[9])
                        Folder_OK_Cam1 = row[10]
                        Folder_OK_Cam2 = row[11]
                        Folder_OK_Cam3 = row[12]
                        Folder_NG_Cam1 = row[13]
                        Folder_NG_Cam2 = row[14]
                        Folder_NG_Cam3 = row[15]

                        model1 = torch.hub.load('./levu','custom', path= row[2], source='local',force_reload =False)
                        if row[2][-7:] == 'edit.pt': 
                            change_label(model1)
                    if row1_a == '2' and row1_b == '0':
                        weight2 = row[2]
                        conf_thres2 = row[3]
                        OK_Cam1 = str2bool(row[4])
                        OK_Cam2 = str2bool(row[5])
                        OK_Cam3 = str2bool(row[6])

                        NG_Cam1 = str2bool(row[7])
                        NG_Cam2 = str2bool(row[8])
                        NG_Cam3 = str2bool(row[9])
            
                        Folder_OK_Cam1 = row[10]
                        Folder_OK_Cam2 = row[11]
                        Folder_OK_Cam3 = row[12]
             
                        Folder_NG_Cam1 = row[13]
                        Folder_NG_Cam2 = row[14]
                        Folder_NG_Cam3 = row[15]
            
                        model2 = torch.hub.load('./levu','custom', path= row[2], source='local',force_reload =False)

                    if row1_a == '3' and row1_b == '0':
                        weight3 = row[2]
                        conf_thres3 = row[3]
                        OK_Cam1 = str2bool(row[4])
                        OK_Cam2 = str2bool(row[5])
                        OK_Cam3 = str2bool(row[6])
               
                        NG_Cam1 = str2bool(row[7])
                        NG_Cam2 = str2bool(row[8])
                        NG_Cam3 = str2bool(row[9])
                      
                        Folder_OK_Cam1 = row[10]
                        Folder_OK_Cam2 = row[11]
                        Folder_OK_Cam3 = row[12]
           
                        Folder_NG_Cam1 = row[13]
                        Folder_NG_Cam2 = row[14]
                        Folder_NG_Cam3 = row[15]
                   
                        model3 = torch.hub.load('./levu','custom', path= row[2], source='local',force_reload =False)

                    if row1_a == '4' and row1_b == '0':
                        weight4 = row[2]
                        conf_thres4 = row[3]
                        OK_Cam1 = str2bool(row[4])
                        OK_Cam2 = str2bool(row[5])
                        OK_Cam3 = str2bool(row[6])
                        
                        NG_Cam1 = str2bool(row[7])
                        NG_Cam2 = str2bool(row[8])
                        NG_Cam3 = str2bool(row[9])
                  
                        Folder_OK_Cam1 = row[10]
                        Folder_OK_Cam2 = row[11]
                        Folder_OK_Cam3 = row[12]
                    
                        Folder_NG_Cam1 = row[13]
                        Folder_NG_Cam2 = row[14]
                        Folder_NG_Cam3 = row[15]
               
                        model4 = torch.hub.load('./levu','custom', path= row[2], source='local',force_reload =False)

            changed=0        
            started = 0
            window.close() 
            window = make_window(theme)
            if mychoose=='3':
                window['choosemodel_running'].update(value = 'Đế ĐEN')
            if mychoose=='1':
                window['choosemodel_running'].update(value = 'Đế TRẮNG')
            #print(weight1)
            window['file_weights1'].update(value=weight1)
            window['conf_thres1'].update(value=conf_thres1)
            window['file_weights2'].update(value=weight2)
            window['conf_thres2'].update(value=conf_thres2)
            window['choose_model'].update(value=mychoose)

            window['have_save_OK_1'].update(value=OK_Cam1)
            window['have_save_OK_2'].update(value=OK_Cam2)
            window['have_save_NG_1'].update(value=NG_Cam1)
            window['have_save_NG_2'].update(value=NG_Cam2)

            window['save_OK_1'].update(value=Folder_OK_Cam1)
            window['save_OK_2'].update(value=Folder_OK_Cam2)
            window['save_NG_1'].update(value=Folder_NG_Cam1)
            window['save_NG_2'].update(value=Folder_NG_Cam2)

            window['file_weights3'].update(value=weight3)
            window['conf_thres3'].update(value=conf_thres3)

            window['choose_model'].update(value=mychoose)

            window['have_save_OK_3'].update(value=OK_Cam3)
  
            window['have_save_NG_3'].update(value=NG_Cam3)


            window['save_OK_3'].update(value=Folder_OK_Cam3)

            window['save_NG_3'].update(value=Folder_NG_Cam3)
   


            cursor = conn.execute("SELECT ChooseModel,Camera,Weights,Confidence,OK_Cam1,OK_Cam2,OK_Cam3,NG_Cam1,NG_Cam2,NG_Cam3,Folder_OK_Cam1,Folder_OK_Cam2,Folder_OK_Cam3,Folder_NG_Cam1,Folder_NG_Cam2,Folder_NG_Cam3,Joined,Ok,Num,NG,WidthMin,WidthMax,HeightMin,HeightMax,PLC_NG,PLC_OK,Conf from MYMODEL")
            for row in cursor:
                if row[0] == values['choose_model']:
                    row1_a, row1_b = row[1].strip().split('_')
                    if row1_a == '1':
                        for item in range(len(model1.names)):
                            if int(row1_b) == item:
                                window[f'{model1.names[item]}_1'].update(value=str2bool(row[16]))
                                window[f'{model1.names[item]}_OK_1'].update(value=str2bool(row[17]))
                                window[f'{model1.names[item]}_Num_1'].update(value=str(row[18]))
                                window[f'{model1.names[item]}_NG_1'].update(value=str2bool(row[19]))
                                window[f'{model1.names[item]}_Wn_1'].update(value=str(row[20]))
                                window[f'{model1.names[item]}_Wx_1'].update(value=str(row[21]))
                                window[f'{model1.names[item]}_Hn_1'].update(value=str(row[22]))
                                window[f'{model1.names[item]}_Hx_1'].update(value=str(row[23]))
                                window[f'{model1.names[item]}_PLC_1'].update(value=str(row[24]))
                                window['OK_PLC_1'].update(value=str(row[25]))
                                window[f'{model1.names[item]}_Conf_1'].update(value=str(row[26]))

                    if row1_a == '2':
                        for item in range(len(model2.names)):
                            if int(row1_b) == item:
                                window[f'{model2.names[item]}_2'].update(value=str2bool(row[16]))
                                window[f'{model2.names[item]}_OK_2'].update(value=str2bool(row[17]))
                                window[f'{model2.names[item]}_Num_2'].update(value=str(row[18]))
                                window[f'{model2.names[item]}_NG_2'].update(value=str2bool(row[19]))
                                window[f'{model2.names[item]}_Wn_2'].update(value=str(row[20]))
                                window[f'{model2.names[item]}_Wx_2'].update(value=str(row[21]))
                                window[f'{model2.names[item]}_Hn_2'].update(value=str(row[22]))
                                window[f'{model2.names[item]}_Hx_2'].update(value=str(row[23]))
                                window[f'{model2.names[item]}_PLC_2'].update(value=str(row[24]))
                                window['OK_PLC_2'].update(value=str(row[25]))
                                window[f'{model2.names[item]}_Conf_2'].update(value=str(row[26]))
                    if row1_a == '3':
                        for item in range(len(model3.names)):
                            if int(row1_b) == item:
                                window[f'{model3.names[item]}_3'].update(value=str2bool(row[16]))
                                window[f'{model3.names[item]}_OK_3'].update(value=str2bool(row[17]))
                                window[f'{model3.names[item]}_Num_3'].update(value=str(row[18]))
                                window[f'{model3.names[item]}_NG_3'].update(value=str2bool(row[19]))
                                window[f'{model3.names[item]}_Wn_3'].update(value=str(row[20]))
                                window[f'{model3.names[item]}_Wx_3'].update(value=str(row[21]))
                                window[f'{model3.names[item]}_Hn_3'].update(value=str(row[22]))
                                window[f'{model3.names[item]}_Hx_3'].update(value=str(row[23]))
                                window[f'{model3.names[item]}_PLC_3'].update(value=str(row[24]))
                                window['OK_PLC_3'].update(value=str(row[25]))
                                window[f'{model3.names[item]}_Conf_3'].update(value=str(row[26]))


            conn.close()


        if event == 'SaveData1':
            save_all_sql(model1,1,str(values['choose_model']))
            save_choosemodel(str(values['choose_model']))
            save_model(1,values['file_weights1'])
            save_values(str(values['W_min_bc_cc']),str(values['H_min_bc_cc']),int(values['bc_cc_Conf_1']))
            sg.popup('Saved param model 1 successed',font=('Helvetica',15), text_color='green',keep_on_top= True)


        if event == 'SaveData2':
            save_all_sql(model2,2,str(values['choose_model']))
            save_choosemodel(str(values['choose_model']))
            save_model(2,values['file_weights2'])
            sg.popup('Saved param model 2 successed',font=('Helvetica',15), text_color='green',keep_on_top= True)


        if event == 'SaveData3':
            save_all_sql(model3,3,str(values['choose_model']))
            save_choosemodel(str(values['choose_model']))
            save_model(3,values['file_weights3'])
            sg.popup('Saved param model 3 successed',font=('Helvetica',15), text_color='green',keep_on_top= True)
        # try:
        #     # if read_plc_keyence(17408) == 1:
        #         removefile_cam1()    

        #         removefile_cam2()    
        #         myitem =0
        #         all_error = []
        #         c_hinh = ''  
        #         # #write_plc_keyence(17408,0)
        #         # #write_plc_keyence(460,0)

        # except:
        #     pass


        # try:
        #     if read_plc_keyence(17409) == 1:
        #         removefile_cam2()    
        #         myitem =0
        #         all_error = []
        #         c_hinh = ''  
        #         #write_plc_keyence(17409,0)
        #         # #write_plc_keyence(460,0)

        # except:
        #     pass
        try:
            #    HomNay
            if  read_plc_keyence(1002) == 1:
                print('Tao excel')
                excel_sang()

        #Dem
            if  read_plc_keyence(1002) == 2:
                excel_dem()


            if  read_plc_keyence(1000) == 1:
                today = date.today()
                d1 = today.strftime("%d/%m/%Y")
                now = datetime.datetime.now()
                t1 = now.strftime("%H:%M:%S")

                mydate = today.strftime("%Y_%m_%d")
            
                hour = int(now.strftime("%H"))
        
                if 7<= hour <=18:
                    if not os.path.isfile(f"excel/{mydate}_Ngay.xlsx"):
                        excel_sang()
                    wb = openpyxl.load_workbook(f"excel/{mydate}_Ngay.xlsx")

                if 19 <= hour <= 23:
                    if not os.path.isfile(f"excel/{mydate}_Dem.xlsx"):
                        excel_dem()
                    wb = openpyxl.load_workbook(f"excel/{mydate}_Dem.xlsx")

                if 0 <= hour <= 6:
                    Previous_Date = datetime.datetime.today() - datetime.timedelta(days=1)
                    Previous_Date = Previous_Date.strftime("%Y_%m_%d")
                    if not os.path.isfile(f"excel/{Previous_Date}_Dem.xlsx"):
                        excel_dem()


                    wb = openpyxl.load_workbook(f"excel/{Previous_Date}_Dem.xlsx")

                ws = wb.active
                col1 = d1
                col2 = t1
                col3 = int(read_plc_keyence(1024))
                col4 = int(read_plc_keyence(1010))
                col5 = int(read_plc_keyence(1012))
                col6 = int(read_plc_keyence(1014))
                col7 = int(read_plc_keyence(1016))
                col8 = int(read_plc_keyence(1018))
                col9 = int(read_plc_keyence(1020))
                col10 = int(read_plc_keyence(1022))
                col11 = int(read_plc_keyence(1026))
                col12 = int(read_plc_keyence(6028))
                col13 = int(read_plc_keyence(6042))
                ws.append([col1, col2,col3, col4,col5, col6,col7, col8,col9, col10,col11])

                for row in range(ws.max_row, ws.max_row+1):
                    for col in range(1, ws.max_column+1):
                        #print(col[row].value)
                        d = ws.cell(row = row, column = col)
                        #currentCell = ws.cell(col[row])
                        d.alignment = Alignment(horizontal='center')
                        #d.style.alignment.horizontal = 'center'
                        d.font = Font(name= 'Calibri', size=12)

                if 7<= hour <=18:
                    
                    wb.save(f"excel/{mydate}_Ngay.xlsx")
                    try:
                        shutil.copy(f"excel/{mydate}_Ngay.xlsx", f"C:/excel/{mydate}_Ngay.xlsx")
                    except:
                        pass
                    
                    try:
                        wb1 = openpyxl.load_workbook(f"C:/excel/{mydate}_Ngay.xlsx")

                        ws1 = wb1.active
                        all_cell = []


                        for row in range(1, ws1.max_row+1):
                            all_col = []    
                            for col in range(1, 12):
                                cell_obj = ws1.cell(row = row, column = col)
                                all_col.append(cell_obj.value)
                            all_cell.append(all_col)




                        wb2 = openpyxl.load_workbook("C:/excel/Now.xlsx")
                        ws2 = wb2.active

                        index_row = ws2.max_row


                        if ws1.max_row > ws2.max_row:
                            for row in range(4,ws1.max_row+1):
                                for col in range(1, 12):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                    
                        else:
                            for row in range(4,ws2.max_row+1):
                                for col in range(1, 12):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                                    else:
                                        ws2.cell(row = row, column = col).value = None
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)   

                        wb2.save("C:/excel/Now.xlsx")



                    except:
                        pass
                if 19 <= hour <= 23:
                    wb.save(f"excel/{mydate}_Dem.xlsx")
                    try:
                        shutil.copy(f"excel/{mydate}_Dem.xlsx", f"C:/excel/{mydate}_Dem.xlsx")
                    except:
                        pass

                    try:
                        wb1 = openpyxl.load_workbook(f"C:/excel/{mydate}_Dem.xlsx")

                        ws1 = wb1.active
                        all_cell = []


                        for row in range(1, ws1.max_row+1):
                            all_col = []    
                            for col in range(1, 12):
                                cell_obj = ws1.cell(row = row, column = col)
                                all_col.append(cell_obj.value)
                            all_cell.append(all_col)




                        wb2 = openpyxl.load_workbook("C:/excel/Now.xlsx")
                        ws2 = wb2.active

                        index_row = ws2.max_row


                        if ws1.max_row > ws2.max_row:
                            for row in range(4,ws1.max_row+1):
                                for col in range(1, 12):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                    
                        else:
                            for row in range(4,ws2.max_row+1):
                                for col in range(1, 12):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                                    else:
                                        ws2.cell(row = row, column = col).value = None
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)   

                        wb2.save("C:/excel/Now.xlsx")



                    except:
                        pass


                if 0 <= hour <= 6:
                    Previous_Date = datetime.datetime.today() - datetime.timedelta(days=1)
                    Previous_Date = Previous_Date.strftime("%Y_%m_%d")
                    wb.save(f"excel/{Previous_Date}_Dem.xlsx")
                    try:
                        shutil.copy(f"excel/{Previous_Date}_Dem.xlsx", f"C:/excel/{Previous_Date}_Dem.xlsx")
                    except:
                        pass

                    try:
                        wb1 = openpyxl.load_workbook(f"C:/excel/{Previous_Date}_Dem.xlsx")

                        ws1 = wb1.active
                        all_cell = []


                        for row in range(1, ws1.max_row+1):
                            all_col = []    
                            for col in range(1, 12):
                                cell_obj = ws1.cell(row = row, column = col)
                                all_col.append(cell_obj.value)
                            all_cell.append(all_col)




                        wb2 = openpyxl.load_workbook("C:/excel/Now.xlsx")
                        ws2 = wb2.active

                        index_row = ws2.max_row


                        if ws1.max_row > ws2.max_row:
                            for row in range(4,ws1.max_row+1):
                                for col in range(1, 12):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                    
                        else:
                            for row in range(4,ws2.max_row+1):
                                for col in range(1, 12):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                                    else:
                                        ws2.cell(row = row, column = col).value = None
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)   

                        wb2.save("C:/excel/Now.xlsx")



                    except:
                        pass


                #write_plc_keyence(1000,0) 
            

            if  read_plc_keyence(1004) == 1:
                today = date.today()
                d1 = today.strftime("%d/%m/%Y")
                now = datetime.datetime.now()
                t1 = now.strftime("%H:%M:%S")

                wb = openpyxl.load_workbook("excel/All.xlsx")
                ws = wb.active
                col1 = d1
                col2 = t1
                col3 = int(read_plc_keyence(1024))
                col4 = int(read_plc_keyence(1010))
                col5 = int(read_plc_keyence(1012))
                col6 = int(read_plc_keyence(1014))
                col7 = int(read_plc_keyence(1016))
                col8 = int(read_plc_keyence(1018))
                col9 = int(read_plc_keyence(1020))
                col10 = int(read_plc_keyence(1022))
                col11 = int(read_plc_keyence(1026))
                ws.append([col1, col2,col3, col4,col5, col6,col7, col8,col9, col10,col11])

                for row in range(ws.max_row, ws.max_row+1):
                    for col in range(1, ws.max_column+1):
                        #print(col[row].value)
                        d = ws.cell(row = row, column = col)
                        #currentCell = ws.cell(col[row])
                        d.alignment = Alignment(horizontal='center')
                        #d.style.alignment.horizontal = 'center'
                        d.font = Font(name= 'Calibri', size=12)


                wb.save("excel/All.xlsx")
                try:
                    shutil.copy("excel/All.xlsx", "C:/excel/A.xlsx")



                    wb1 = openpyxl.load_workbook("C:/excel/A.xlsx")

                    ws1 = wb1.active
                    all_cell = []
                    all_cell_date = []
                    all_cell_time = []

                    for row in range(1, ws1.max_row+1):
                        all_col = []    
                        for col in range(1, 12):
                            cell_obj = ws1.cell(row = row, column = col)
                            all_col.append(cell_obj.value)
                        all_cell.append(all_col)
                        all_cell_date.append(all_col[0])
                        all_cell_time.append(all_col[1])



                    wb2 = openpyxl.load_workbook("C:/excel/All.xlsx")
                    ws2 = wb2.active

                    index_row = ws2.max_row
                    for row in range(4, ws2.max_row+1):
                        if ws2.cell(row=row,column=1).value == None:
                            index_row = row
                            break

                    for row in range(index_row,ws1.max_row+1):
                        for col in range(1, 12):
                            ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                            d = ws2.cell(row = row, column = col)
                            d.alignment = Alignment(horizontal='center')
                            d.font = Font(name= 'Calibri', size=12)
                    wb2.save("C:/excel/All.xlsx")


                except:
                    pass
                
                #write_plc_keyence(1004,0) 
            
        except:
            pass


        # try:
        #Xuly CAM1
        # read_D = read_plc_keyence(17449)  # doc thanh ghi D450
        # if read_D == 1:
        program_camera1_FH(model=model1,size=768,conf= values['conf_thres1']/100, regno=17410)
        
        #Xuly CAM2  
        # read_D = read_plc_keyence(17450) 
        # if read_D == 1:   
        already = False
        window['result_cam2'].update(value= ' ', text_color='green')
        if myitem==0:
            window['cd1'].update(value= ' ', text_color='green')
            window['tc1'].update(value= ' ', text_color='green')
            window['cd2'].update(value= ' ', text_color='green')
            window['tc2'].update(value= ' ', text_color='green')
            
            if len(c_hinh)==0:
                c_hinh = time_to_name()
            
            imgbytesa = np.zeros([100,100,3],dtype=np.uint8)
            imgbytesa = cv2.resize(imgbytesa, (600,450), interpolation = cv2.INTER_AREA)
            imgbytesa = cv2.imencode('.png',imgbytesa)[1].tobytes()
            window['cd_1'].update(data=imgbytesa)
            window['cd_2'].update(data=imgbytesa)
            window['tc_1'].update(data=imgbytesa)
            window['tc_2'].update(data=imgbytesa)



            folder1 = glob('C:/FH/camera2/CHAU1/**/*.jpg')
            folder2 = glob('C:/FH/camera2/CHOI1/**/*.jpg')
            folder3 = glob('C:/FH/camera2/CHAU2/**/*.jpg')
            folder4 = glob('C:/FH/camera2/CHOI2/**/*.jpg')

            
        
            for f1 in folder1:
                cd =1
                program_camera2_FH(model=model2,size=608,conf=values['conf_thres2']/100, file = f1)     
                
            for f2 in folder2:
                tc =1
                program_camera3_FH(model=model3,size=608,conf=values['conf_thres3']/100, file = f2)                 
                
            for f3 in folder3:
                cd =2
                program_camera2_FH(model=model2,size=608,conf=values['conf_thres2']/100, file = f3)     
                
            for f4 in folder4:
                tc=2
                program_camera3_FH(model=model3,size=608,conf=values['conf_thres3']/100, file = f4)          
            # print("myitem: ", myitem)
            if myitem >= 4:
                #write_plc_keyence(17450,0)
                all_error = set(all_error)
                all_error = list(all_error)
                all_item_error = set(all_item_error)
                all_item_error = list(all_item_error)

                if len(all_error) == 0:
                    write_plc_keyence(int(values['OK_PLC_2']),1)    
                    window['result_cam2'].update(value= 'OK', text_color='green')
                if len(all_error) == 1:
                    if int(all_error[0]) == 9:
                        #write_plc_keyence(17470,2)
                        print('er1:',all_error)                       
                    else:
                        #write_plc_keyence(int(all_error[0]),1)
                    #write_plc_keyence(int(all_error[0]),1) 
                        print('er1:',all_error)                       
                    window['result_cam2'].update(value= 'NG', text_color='red')
                if len(all_error) >= 2:
                    for error in all_error:
                        if error == 9:
                            #write_plc_keyence(17470,2)
                            print('er2:',all_error)
                        else:
                            print('er2:',all_error)
                            #write_plc_keyence(error,1) 
                        #write_plc_keyence(error,1)           
                    window['result_cam2'].update(value= 'NG*', text_color='red')
                    print('all_error', all_error)
                if len(all_item_error) >=25: 
                    file_path = r'C:\combine\tenfile.txt'
                    with open(file_path, 'r+',encoding='utf-8') as file:
                        line = file.readline()
                        if not line:
                            current_time = datetime.datetime.now() 
                            file.write(f'       DATE TIME             |              HẠNG MỤC NG*')
                            file.write('\n'+ f'{str(current_time)}   |   {all_item_error}')
                        else:    
                            current_time = datetime.datetime.now() 
                            file.write('\n'+ f'{str(current_time)}   |   {all_item_error}')
                #write_plc_keyence(17452,1)
                myitem =0
                time_all=0
                all_error = []
                all_item_error=[]
                c_hinh = ''
                print('CAM2 FINISHED---------------------------------------------------------------------------------------------')
    
    if event == 'check_model1' and values['check_model1'] == True or read_plc_keyence(17500) :
        directory1 = 'D:/CHECK/' + maude + '/Cam1/NG/'        
        print(directory1)
        if os.listdir(directory1) == []:
            print('folder 1 empty')
        else:
            print('received folder 1')
            bomau = glob(directory1 + '*.jpg')
            cnt=0
            allng1 =0
            for path1 in bomau:
                ten = os.path.basename(path1)
                img1_orgin = cv2.imread(path1)
                img1_orgin = cv2.cvtColor(img1_orgin, cv2.COLOR_BGR2RGB)     
                result1 = model1(img1_orgin,size=768,conf = values['conf_thres1']/100)
                table1 = result1.pandas().xyxy[0]
                area_remove1 = []
                myresult1 =0 
                for item in range(len(table1.index)):
                    width1 = table1['xmax'][item] - table1['xmin'][item]
                    height1 = table1['ymax'][item] - table1['ymin'][item]
                    label_name = table1['name'][item]
                    conf1 = table1['confidence'][item] * 100
                    for i1 in range(len(model1.names)):
                        if values[f'{model1.names[i1]}_1'] == True:
                            if label_name == model1.names[i1]:
                                if width1 < int(values[f'{model1.names[i1]}_Wn_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif width1 > int(values[f'{model1.names[i1]}_Wx_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif height1 < int(values[f'{model1.names[i1]}_Hn_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif height1 > int(values[f'{model1.names[i1]}_Hx_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif conf1 < int(values[f'{model1.names[i1]}_Conf_1']):
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item) 
                    if values[f'{model1.names[i1]}_1'] == False:
                        if label_name == model1.names[i1]:
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)

                names1 = list(table1['name'])

                show1 = np.squeeze(result1.render(area_remove1))
                show1 = cv2.resize(show1, (image_height_display,image_height_display), interpolation = cv2.INTER_AREA)
                show1 = cv2.cvtColor(show1, cv2.COLOR_BGR2RGB) 
                for i1 in range(len(model1.names)):
                    if values[f'{model1.names[i1]}_OK_1'] == True:
                        len_name1 = 0
                        for name1 in names1:
                            if name1 == model1.names[i1]:
                                len_name1 +=1
                        if len_name1 != int(values[f'{model1.names[i1]}_Num_1']):
                            print('NG1')
                            myresult1 = 1                              
                    elif values[f'{model1.names[i1]}_NG_1'] == True:
                        if model1.names[i1] in names1:
                            print('NG2')
                            myresult1 = 1         
                cv2.putText(show1, ten,(20,20),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),1)
                imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
                window['image1'].update(data= imgbytes1)
                if myresult1 == 0:
                    allng1 =1
                    print('OK')                                                                       
                    window['result_cam1'].update(value= 'OK', text_color='green')
                    # sg.popup('Test sample NG fail')
                    #write_plc_keyence(17500,0)
                    #write_plc_keyence(17492,2)
                    # #write_plc_keyence(17492,1)
                    break
                else:                        
                    window['result_cam1'].update(value= 'NG', text_color='red')    
                cnt += 1
            if allng1 ==0:
                #write_plc_keyence(17500,0)
                #write_plc_keyence(17492,1)   
            
        if event == 'check_model2' and values['check_model2'] == True and values['Tay_choi'] == False or read_plc_keyence(17502) ==1:
            directory2 = 'D:/CHECK/' + maude + '/Cam2/CD/NG/'
            if os.listdir(directory2) == []:
                print('folder 2 empty')
            else:
                print('received folder 2')
                bomau = glob('D:/CHECK/' + maude + '/Cam2/CD/NG/*.jpg')
                cnt=0
                allng2 =0
                for path2 in bomau:
                    ten = os.path.basename(path2)
                    img2_orgin = cv2.imread(path2)
                    img2_orgin = cv2.cvtColor(img2_orgin, cv2.COLOR_BGR2RGB) 
                    result2 = model2(img2_orgin,size= 608,conf = values['conf_thres2']/100)
                    table2 = result2.pandas().xyxy[0]
                    area_remove2 = []
                    myresult2 =0 
                    for item in range(len(table2.index)):
                        width2 = table2['xmax'][item] - table2['xmin'][item]
                        height2 = table2['ymax'][item] - table2['ymin'][item]
                        label_name = table2['name'][item]
                        conf2 = table2['confidence'][item] * 100
                        for i2 in range(len(model2.names)):
                            if values[f'{model2.names[i2]}_2'] == True:
                                if label_name == model2.names[i2]:
                                    if width2 < int(values[f'{model2.names[i2]}_Wn_2']): 
                                        table2.drop(item, axis=0, inplace=True)
                                        area_remove2.append(item)
                                    elif width2 > int(values[f'{model2.names[i2]}_Wx_2']): 
                                        table2.drop(item, axis=0, inplace=True)
                                        area_remove2.append(item)
                                    elif height2 < int(values[f'{model2.names[i2]}_Hn_2']): 
                                        table2.drop(item, axis=0, inplace=True)
                                        area_remove2.append(item)
                                    elif height2 > int(values[f'{model2.names[i2]}_Hx_2']): 
                                        table2.drop(item, axis=0, inplace=True)
                                        area_remove2.append(item)
                                    elif conf2 < int(values[f'{model2.names[i2]}_Conf_2']):
                                        table2.drop(item, axis=0, inplace=True)
                                        area_remove2.append(item) 
                        if values[f'{model2.names[i2]}_2'] == False:
                            if label_name == model2.names[i2]:
                                table2.drop(item, axis=0, inplace=True)
                                area_remove2.append(item)

                    names2 = list(table2['name'])

                    show2 = np.squeeze(result2.render(area_remove2))
                    show2 = cv2.resize(show2, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
                    show2 = cv2.cvtColor(show2, cv2.COLOR_BGR2RGB) 
                    for i2 in range(len(model2.names)):
                        if values[f'{model2.names[i2]}_OK_2'] == True:
                            len_name2 = 0
                            for name2 in names2:
                                if name2 == model2.names[i2]:
                                    len_name2 +=1
                            if len_name2 != int(values[f'{model2.names[i2]}_Num_2']):
                                print('NG1')                                                                                                
                                myresult2 = 1
                        if values[f'{model2.names[i2]}_NG_2'] == True:
                            if model2.names[i2] in names2:
                                print('NG2')                                
                                myresult2 = 1      
    
                    cv2.putText(show2, ten,(20,20),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),1)
                    imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()
                    window['cd_11'].update(data= imgbytes2)
                    if myresult2 == 0:
                        allng2 =1
                        print('OK')
                        window['result_cam2'].update(value= 'OK', text_color='green')
                        # sg.popup('Test sample NG fail')
                        # break
                        #write_plc_keyence(17502,0)
                        #write_plc_keyence(17494,2)
                        break
                    else:
                        window['result_cam2'].update(value= 'NG', text_color='red')                  
                    cnt += 1
                if allng2 ==0:
                    #write_plc_keyence(17502,0)
                    #write_plc_keyence(17494,1)   
               
        # thu mau Tay choi
        if event == 'check_model2' and values['check_model2'] == True and values['Tay_choi'] == True or read_plc_keyence(17504) ==1:
            directory2 = 'D:/CHECK/' + maude + '/Cam2/TC/NG/'
            if os.listdir(directory2) == []:
                print('folder 2 empty')
            else:
                print('received folder 2')
                bomau = glob('D:/CHECK/' + maude + '/Cam2/TC/NG/*.jpg')
                cnt=0
                for path2 in bomau:
                    ten = os.path.basename(path2)
                    img2_orgin = cv2.imread(path2)                    
                    img2_orgin = cv2.cvtColor(img2_orgin, cv2.COLOR_BGR2RGB)
                    result2 = model3(img2_orgin,size= 608,conf = values['conf_thres3']/100)
                    table2 = result2.pandas().xyxy[0]
                    area_remove2 = []
                    myresult2 =0 
                    allng3 =0

                    for item in range(len(table2.index)):
                        width2 = table2['xmax'][item] - table2['xmin'][item]
                        height2 = table2['ymax'][item] - table2['ymin'][item]
                        label_name = table2['name'][item]
                        conf2 = table2['confidence'][item] * 100
                        for i2 in range(len(model3.names)):
                            if values[f'{model3.names[i2]}_3'] == True:
                                if label_name == model3.names[i2]:
                                    if width2 < int(values[f'{model3.names[i2]}_Wn_3']): 
                                        table2.drop(item, axis=0, inplace=True)
                                        area_remove2.append(item)
                                    elif width2 > int(values[f'{model3.names[i2]}_Wx_3']): 
                                        table2.drop(item, axis=0, inplace=True)
                                        area_remove2.append(item)
                                    elif height2 < int(values[f'{model3.names[i2]}_Hn_3']): 
                                        table2.drop(item, axis=0, inplace=True)
                                        area_remove2.append(item)
                                    elif height2 > int(values[f'{model3.names[i2]}_Hx_3']): 
                                        table2.drop(item, axis=0, inplace=True)
                                        area_remove2.append(item)
                                    elif conf2 < int(values[f'{model3.names[i2]}_Conf_3']):
                                        table2.drop(item, axis=0, inplace=True)
                                        area_remove2.append(item) 
                        if values[f'{model3.names[i2]}_3'] == False:
                            if label_name == model3.names[i2]:
                                table2.drop(item, axis=0, inplace=True)
                                area_remove2.append(item)

                    names2 = list(table2['name'])

                    show2 = np.squeeze(result2.render(area_remove2))
                    show2 = cv2.resize(show2, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
                    show2 = cv2.cvtColor(show2, cv2.COLOR_BGR2RGB) 
                    for i2 in range(len(model3.names)):
                        if values[f'{model3.names[i2]}_OK_3'] == True:
                            len_name2 = 0
                            for name2 in names2:
                                if name2 == model3.names[i2]:
                                    len_name2 +=1
                            if len_name2 != int(values[f'{model3.names[i2]}_Num_3']):
                                print('NG')                            
                                myresult2 = 1
                                
                        if values[f'{model3.names[i2]}_NG_3'] == True:
                            if model3.names[i2] in names2:
                                print('NG')
                                myresult2 = 1      
                                
                    cv2.putText(show2, ten,(20,20),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),1)
                    imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()                  
                    window['cd_11'].update(data= imgbytes2)

                    if myresult2 == 0:
                        allng3 =1
                        print('OK')            
                        window['result_cam2'].update(value= 'OK', text_color='green')
                        # sg.popup('Test sample NG fail')
                        #write_plc_keyence(17504,0)
                        #write_plc_keyence(17496,2)
                        break
                    else:
                        window['result_cam2'].update(value= 'NG', text_color='red')

                    cnt += 1
                if allng3 ==0:
                    #write_plc_keyence(17504,0)
                    #write_plc_keyence(17496,1)   
                     

        if event == 'Webcam1':
            recording1 = True

        elif event == 'Stop1':
            recording1 = False 
            imgbytes1 = np.zeros([100,100,3],dtype=np.uint8)
            imgbytes1 = cv2.resize(imgbytes1, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
            imgbytes1 = cv2.imencode('.png',imgbytes1)[1].tobytes()
            window['image1'].update(data=imgbytes1)
            window['result_cam1'].update(value='')

        if event == 'Webcam2':
            recording2 = True

        elif event == 'Stop2':
            recording2 = False 
            imgbytes2 = np.zeros([100,100,3],dtype=np.uint8)
            imgbytes2 = cv2.resize(imgbytes2, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
            imgbytes2 = cv2.imencode('.png',imgbytes2)[1].tobytes()
            window['cd_11'].update(data=imgbytes2)
            window['result_cam2'].update(value='')

        if event == 'Pic1':
            dir_img1 = sg.popup_get_file('Choose your image 1',file_types=file_name_img,keep_on_top= True)
            if dir_img1 not in ('',None):
                pic1 = Image.open(dir_img1)
                img1_resize = pic1.resize((image_width_display,image_height_display))
                
                imgbytes1 = ImageTk.PhotoImage(img1_resize)
                window['image1'].update(data= imgbytes1)
                window['Detect1'].update(disabled= False)         

        if event == 'Pic2':
            dir_img2 = sg.popup_get_file('Choose your image 2',file_types=file_name_img,keep_on_top= True)
            if dir_img2 not in ('',None):
                pic2 = Image.open(dir_img2)
                
                img2_resize = pic2.resize((image_width_display,image_height_display))
                imgbytes2 = ImageTk.PhotoImage(img2_resize)
                window['cd_11'].update(data=imgbytes2)
                window['Detect2'].update(disabled= False)


        if event == 'Change1' or event == 'Change_1':
            mypath1 = values['file_weights1']
            model1= torch.hub.load('./levu','custom',path=mypath1,source='local',force_reload=False)
            if mypath1[-7:] == 'edit.pt': #Dung khi bi sai label
                change_label(model1)
            mychoose = values['choose_model']
            weight1 = values['file_weights1']
            conf_thres1 = values['conf_thres1'] 
            weight2 = values['file_weights2']
            conf_thres2 = values['conf_thres2'] 

            OK_Cam1 = values['have_save_OK_1']
            OK_Cam2 = values['have_save_OK_2']
            NG_Cam1 = values['have_save_NG_1']
            NG_Cam2 = values['have_save_NG_2']
            Folder_OK_Cam1 = values['save_OK_1']
            Folder_OK_Cam2 = values['save_OK_2']
            Folder_NG_Cam1 = values['save_NG_1']
            Folder_NG_Cam2 = values['save_NG_2']

            weight3 = values['file_weights3']
            conf_thres3 = values['conf_thres3'] 

            OK_Cam3 = values['have_save_OK_3']
         
            NG_Cam3 = values['have_save_NG_3']
          
            Folder_OK_Cam3 = values['save_OK_3']
           
            Folder_NG_Cam3 = values['save_NG_3']
        
            window.close() 
            window = make_window(theme)

            window['choose_model'].update(value=mychoose)
            window['file_weights1'].update(value=weight1)
            window['conf_thres1'].update(value=conf_thres1)
            window['file_weights2'].update(value=weight2)
            window['conf_thres2'].update(value=conf_thres2)

            window['have_save_OK_1'].update(value=OK_Cam1)
            window['have_save_OK_2'].update(value=OK_Cam2)
            window['have_save_NG_1'].update(value=NG_Cam1)
            window['have_save_NG_2'].update(value=NG_Cam2)

            window['save_OK_1'].update(value=Folder_OK_Cam1)
            window['save_OK_2'].update(value=Folder_OK_Cam2)
            window['save_NG_1'].update(value=Folder_NG_Cam1)
            window['save_NG_2'].update(value=Folder_NG_Cam2)

            window['choose_model'].update(value=mychoose)
            window['file_weights3'].update(value=weight3)
            window['conf_thres3'].update(value=conf_thres3)

            window['have_save_OK_3'].update(value=OK_Cam3)
            window['have_save_NG_3'].update(value=NG_Cam3)
            window['save_OK_3'].update(value=Folder_OK_Cam3)
            window['save_NG_3'].update(value=Folder_NG_Cam3)
    
        if event == 'Change2' or event == 'Change_2':
            mypath2 = values['file_weights2']
            model2= torch.hub.load('./levu','custom',path=mypath2,source='local',force_reload=False)
            mychoose = values['choose_model']
            weight1 = values['file_weights1']
            conf_thres1 = values['conf_thres1'] 
            weight2 = values['file_weights2']
            conf_thres2 = values['conf_thres2'] 

            OK_Cam1 = values['have_save_OK_1']
            OK_Cam2 = values['have_save_OK_2']
            NG_Cam1 = values['have_save_NG_1']
            NG_Cam2 = values['have_save_NG_2']
            Folder_OK_Cam1 = values['save_OK_1']
            Folder_OK_Cam2 = values['save_OK_2']
            Folder_NG_Cam1 = values['save_NG_1']
            Folder_NG_Cam2 = values['save_NG_2']

            weight3 = values['file_weights3']
            conf_thres3 = values['conf_thres3'] 
   
            OK_Cam3 = values['have_save_OK_3']        
            NG_Cam3 = values['have_save_NG_3']       
            Folder_OK_Cam3 = values['save_OK_3']         
            Folder_NG_Cam3 = values['save_NG_3']    

            window.close() 
            window = make_window(theme)

            window['choose_model'].update(value=mychoose)
            window['file_weights1'].update(value=weight1)
            window['conf_thres1'].update(value=conf_thres1)
            window['file_weights2'].update(value=weight2)
            window['conf_thres2'].update(value=conf_thres2)

            window['have_save_OK_1'].update(value=OK_Cam1)
            window['have_save_OK_2'].update(value=OK_Cam2)
            window['have_save_NG_1'].update(value=NG_Cam1)
            window['have_save_NG_2'].update(value=NG_Cam2)

            window['save_OK_1'].update(value=Folder_OK_Cam1)
            window['save_OK_2'].update(value=Folder_OK_Cam2)
            window['save_NG_1'].update(value=Folder_NG_Cam1)
            window['save_NG_2'].update(value=Folder_NG_Cam2)

            window['choose_model'].update(value=mychoose)
            window['file_weights3'].update(value=weight3)
            window['conf_thres3'].update(value=conf_thres3)

            window['have_save_OK_3'].update(value=OK_Cam3)
            window['have_save_NG_3'].update(value=NG_Cam3)
            window['save_OK_3'].update(value=Folder_OK_Cam3)
            window['save_NG_3'].update(value=Folder_NG_Cam3)

        if event == 'Change_3':
            mypath3 = values['file_weights3']
            model3= torch.hub.load('./levu','custom',path=mypath3,source='local',force_reload=False)
            mychoose = values['choose_model']
            weight1 = values['file_weights1']
            conf_thres1 = values['conf_thres1'] 
            weight2 = values['file_weights2']
            conf_thres2 = values['conf_thres2'] 

            OK_Cam1 = values['have_save_OK_1']
            OK_Cam2 = values['have_save_OK_2']
            NG_Cam1 = values['have_save_NG_1']
            NG_Cam2 = values['have_save_NG_2']
            Folder_OK_Cam1 = values['save_OK_1']
            Folder_OK_Cam2 = values['save_OK_2']
            Folder_NG_Cam1 = values['save_NG_1']
            Folder_NG_Cam2 = values['save_NG_2']

            weight3 = values['file_weights3']
            conf_thres3 = values['conf_thres3']    

            OK_Cam3 = values['have_save_OK_3']  
            NG_Cam3 = values['have_save_NG_3']          
            Folder_OK_Cam3 = values['save_OK_3']            
            Folder_NG_Cam3 = values['save_NG_3']           

            window.close() 
            window = make_window(theme)

            window['choose_model'].update(value=mychoose)
            window['file_weights1'].update(value=weight1)
            window['conf_thres1'].update(value=conf_thres1)
            window['file_weights2'].update(value=weight2)
            window['conf_thres2'].update(value=conf_thres2)

            window['have_save_OK_1'].update(value=OK_Cam1)
            window['have_save_OK_2'].update(value=OK_Cam2)
            window['have_save_NG_1'].update(value=NG_Cam1)
            window['have_save_NG_2'].update(value=NG_Cam2)

            window['save_OK_1'].update(value=Folder_OK_Cam1)
            window['save_OK_2'].update(value=Folder_OK_Cam2)
            window['save_NG_1'].update(value=Folder_NG_Cam1)
            window['save_NG_2'].update(value=Folder_NG_Cam2)

            window['choose_model'].update(value=mychoose)
            window['file_weights3'].update(value=weight3)
            window['conf_thres3'].update(value=conf_thres3)


            window['have_save_OK_3'].update(value=OK_Cam3)
            window['have_save_NG_3'].update(value=NG_Cam3)
            window['save_OK_3'].update(value=Folder_OK_Cam3)  
            window['save_NG_3'].update(value=Folder_NG_Cam3)

        if event == 'Detect1':
            print('CAM 1 DETECT')
            t1 = time.time()
            try:               
                result1 = model1(pic1,size= 768,conf = values['conf_thres1']/100)
                table1 = result1.pandas().xyxy[0]
                print(table1)
                area_remove1 = []
                myresult1 =0 

                for item in range(len(table1.index)):
                    width1 = table1['xmax'][item] - table1['xmin'][item]
                    height1 = table1['ymax'][item] - table1['ymin'][item]
                    conf1 = table1['confidence'][item] * 100
                    label_name = table1['name'][item]
                    for i1 in range(len(model1.names)):
                        if values[f'{model1.names[i1]}_1'] == True:
                            if label_name == model1.names[i1]:
                                if width1 < int(values[f'{model1.names[i1]}_Wn_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif width1 > int(values[f'{model1.names[i1]}_Wx_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif height1 < int(values[f'{model1.names[i1]}_Hn_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif height1 > int(values[f'{model1.names[i1]}_Hx_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif conf1 < int(values[f'{model1.names[i1]}_Conf_1']):
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)     
                        if values[f'{model1.names[i1]}_1'] == False:
                            if label_name == model1.names[i1]:
                                table1.drop(item, axis=0, inplace=True)
                                area_remove1.append(item)

                names1 = list(table1['name'])
                hm=[]
                k=1
                show1 = np.squeeze(result1.render(area_remove1))
                show1 = cv2.resize(show1, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
                show1 = cv2.cvtColor(show1, cv2.COLOR_BGR2RGB)
                for i1 in range(len(model1.names)):
                    if values[f'{model1.names[i1]}_1'] == True:
                        if values[f'{model1.names[i1]}_OK_1'] == True:
                            len_name1 = 0
                            for name1 in names1:
                                if name1 == model1.names[i1]:
                                    len_name1 +=1
                            if len_name1 != int(values[f'{model1.names[i1]}_Num_1']):
                                print('NG')
                                cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                hm.append(model1.names[i1])  
                                cv2.putText(show1, model1.names[i1],(5,30*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                                print('ngc1:',model1.names[i1])
                                window['result_cam1'].update(value= 'NG', text_color='red')
                                myresult1 = 1
                                k+=1

                        if values[f'{model1.names[i1]}_NG_1'] == True:
                            if model1.names[i1] in names1:
                                print('NG')
                                cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                window['result_cam1'].update(value= 'NG', text_color='red')  
                                hm.append(model1.names[i1]) 
                                cv2.putText(show1, model1.names[i1],(5,30*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0) 
                                print('ngc1:',model1.names[i1])  
                                myresult1 = 1         
                                k+=1    

                if myresult1 == 0:
                    print('OK')
                    cv2.putText(show1, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                    window['result_cam1'].update(value= 'OK', text_color='green')

                imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
                window['image1'].update(data= imgbytes1)
            
            except:
                print(traceback.format_exc())
                sg.popup_annoying("Don't have image or parameter wrong", font=('Helvetica',14),text_color='red')
            
            t2 = time.time() - t1
            print(t2)
            time_cam1 = str(int(t2*1000)) + 'ms'
            window['time_cam1'].update(value= time_cam1, text_color='black') 
            print('---------------------------------------------') 
            
        if event == 'Detect2' and values['Tay_choi'] == False:
            print('Chau dien')
            t1 = time.time()
            try:
                
                result2 = model2(pic2,size= 608,conf = values['conf_thres2']/100)
                table2 = result2.pandas().xyxy[0]
                area_remove2 = []

                myresult2 =0 

                for item in range(len(table2.index)):
                    width2 = table2['xmax'][item] - table2['xmin'][item]
                    height2 = table2['ymax'][item] - table2['ymin'][item]
                    conf2 = table2['confidence'][item] * 100
                    label_name = table2['name'][item]
                    for i2 in range(len(model2.names)):

                        if values[f'{model2.names[i2]}_2'] == True:
                            if label_name == model2.names[i2]:
                                if width2 < int(values[f'{model2.names[i2]}_Wn_2']): 
                                    table2.drop(item, axis=0, inplace=True)
                                    area_remove2.append(item)
                                elif width2 > int(values[f'{model2.names[i2]}_Wx_2']): 
                                    table2.drop(item, axis=0, inplace=True)
                                    area_remove2.append(item)
                                elif height2 < int(values[f'{model2.names[i2]}_Hn_2']): 
                                    table2.drop(item, axis=0, inplace=True)
                                    area_remove2.append(item)
                                elif height2 > int(values[f'{model2.names[i2]}_Hx_2']): 
                                    table2.drop(item, axis=0, inplace=True)
                                    area_remove2.append(item)
                                elif conf2 < int(values[f'{model2.names[i2]}_Conf_2']):
                                    table2.drop(item, axis=0, inplace=True)
                                    area_remove2.append(item) 

                        if values[f'{model2.names[i2]}_2'] == False:
                            if label_name == model2.names[i2]:
                                table2.drop(item, axis=0, inplace=True)
                                area_remove2.append(item)

                names2 = list(table2['name'])

                show2 = np.squeeze(result2.render(area_remove2))
                show2 = cv2.resize(show2, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
                show2 = cv2.cvtColor(show2, cv2.COLOR_BGR2RGB)
                for i2 in range(len(model2.names)):
                    if values[f'{model2.names[i2]}_OK_2'] == True:
                        len_name2 = 0
                        for name2 in names2:
                            if name2 == model2.names[i2]:
                                len_name2 +=1
                        if len_name2 != int(values[f'{model2.names[i2]}_Num_2']):
                            print('NG')
                            cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                            window['result_cam2'].update(value= 'NG', text_color='red')
                            myresult2 = 1
                            break

                    if values[f'{model2.names[i2]}_NG_2'] == True:
                        if model2.names[i2] in names2:
                            print('NG')
                            cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                            window['result_cam2'].update(value= 'NG', text_color='red')    
                            myresult2 = 1      
                            break    

                if myresult2 == 0:
                    print('OK')
                    cv2.putText(show2, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                    window['result_cam2'].update(value= 'OK', text_color='green')

                imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()
                window['cd_11'].update(data= imgbytes2)
            
            except:
                print(traceback.format_exc())
                sg.popup_annoying("Don't have image or parameter wrong", font=('Helvetica',24),text_color='red')
            
            t2 = time.time() - t1
            
            time_cam2 = str(int(t2*1000)) + 'ms'
            window['time_cam2'].update(value= time_cam2, text_color='black') 
            print('---------------------------------------------') 

        if event == 'Detect2' and values['Tay_choi'] == True:
            print('Tay choi')
            t1 = time.time()
            try:
                
                result3 = model3(pic2,size= 608,conf = values['conf_thres3']/100)
                table3 = result3.pandas().xyxy[0]
                area_remove3 = []
                myresult3 =0 

                for item in range(len(table3.index)):
                    width3 = table3['xmax'][item] - table3['xmin'][item]
                    height3 = table3['ymax'][item] - table3['ymin'][item]
                    conf3 = table3['confidence'][item] * 100
                    label_name = table3['name'][item]
                    for i3 in range(len(model3.names)):

                        if values[f'{model3.names[i3]}_3'] == True:
                            if label_name == model3.names[i3]:
                                if width3 < int(values[f'{model3.names[i3]}_Wn_3']): 
                                    table3.drop(item, axis=0, inplace=True)
                                    area_remove3.append(item)
                                elif width3 > int(values[f'{model3.names[i3]}_Wx_3']): 
                                    table3.drop(item, axis=0, inplace=True)
                                    area_remove3.append(item)
                                elif height3 < int(values[f'{model3.names[i3]}_Hn_3']): 
                                    table3.drop(item, axis=0, inplace=True)
                                    area_remove3.append(item)
                                elif height3 > int(values[f'{model3.names[i3]}_Hx_3']): 
                                    table3.drop(item, axis=0, inplace=True)
                                    area_remove3.append(item)
                                elif conf3 < int(values[f'{model3.names[i3]}_Conf_3']):
                                    table3.drop(item, axis=0, inplace=True)
                                    area_remove3.append(item) 

                        if values[f'{model3.names[i3]}_3'] == False:
                            if label_name == model3.names[i3]:
                                table3.drop(item, axis=0, inplace=True)
                                area_remove3.append(item)

                names3 = list(table3['name'])

                show3 = np.squeeze(result3.render(area_remove3))
                show3 = cv2.resize(show3, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
                show3 = cv2.cvtColor(show3, cv2.COLOR_BGR2RGB)
                for i3 in range(len(model3.names)):
                    if values[f'{model3.names[i3]}_OK_3'] == True:
                        len_name3 = 0
                        for name3 in names3:
                            if name3 == model3.names[i3]:
                                len_name3 +=1
                        if len_name3 != int(values[f'{model3.names[i3]}_Num_3']):
                            print('NG')
                            cv2.putText(show3, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                            window['result_cam2'].update(value= 'NG', text_color='red')
                            myresult3 = 1
                            break

                    if values[f'{model3.names[i3]}_NG_3'] == True:
                        if model3.names[i3] in names3:
                            print('NG')
                            cv2.putText(show3, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                            window['result_cam2'].update(value= 'NG', text_color='red')    
                            myresult3 = 1         
                            break    

                if myresult3 == 0:
                    print('OK')
                    cv2.putText(show3, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                    window['result_cam2'].update(value= 'OK', text_color='green')

                imgbytes3 = cv2.imencode('.png',show3)[1].tobytes()
                window['cd_11'].update(data= imgbytes3)

            
            except:
                print(traceback.format_exc())
                sg.popup_annoying("Don't have image or parameter wrong", font=('Helvetica',34),text_color='red')
            
            t2 = time.time() - t1
            
            time_cam3 = str(int(t2*1000)) + 'ms'
            window['time_cam2'].update(value= time_cam3, text_color='black') 
            print('---------------------------------------------') 




    window.close() 

except Exception as e:
    #write_plc_keyence(17406,0)

    # #write_plc_keyence(600,0) 
    print(traceback.print_exc())
    print(e)
    #shutdown this PC after 10minutes, if restart program shutdown will be cancelled.
    os.system('shutdown -s -t 300')