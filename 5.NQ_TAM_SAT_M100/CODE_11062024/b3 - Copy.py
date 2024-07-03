from copyreg import remove_extension
from faulthandler import disable
import glob
import os
from tkinter.tix import Tree
import cv2
import threading
import torch
import numpy as np 
import time
from time import sleep

import PySimpleGUI as sg

from PIL import Image,ImageTk
import os
import datetime 
import shutil

from PIL import Image
from yaml import load
import keyboard
from udp import UDPFinsConnection
from initialization import FinsPLCMemoryAreas

import traceback

import sqlite3

import stapipy as st
import multiprocessing
import socket
import openpyxl

from openpyxl.styles import Alignment
from openpyxl import Workbook
from datetime import date
from openpyxl.styles import Font
#CAM 1   2048x1536              192.168.25.2
#CAM 2   1440x1080              169.254.139.50


# SCALE_X_CAM1 = 1/3.2
# SCALE_Y_CAM1 = 1/3.2

# SCALE_X_CAM2 = 1/2.25
# SCALE_Y_CAM2 = 1/2.25


# SCALE_X_CAM1 = 640/2048
# SCALE_Y_CAM1 = 480/1536

mysleep = 0.1

SCALE_X_CAM1 = 640*1.2/2048
SCALE_Y_CAM1 = 480*1.2/1536

SCALE_X_CAM2 = 640/1440
SCALE_Y_CAM2 = 480/1080



def excel_sang():
    today = date.today()
    mydate = today.strftime("%Y_%m_%d")
    wb = openpyxl.Workbook()

    HomNay = wb.create_sheet("Data")


    HomNay.merge_cells('A1:M1')
    HomNay.merge_cells('A2:A3')
    HomNay.merge_cells('B2:B3')
    HomNay.merge_cells('C2:M2')
    #HomNay.unmerge_cells('A2:D2')
    HomNay['A1'] = 'DỮ LIỆU MÁY NQVNHT RS656 A17'
    HomNay['A1'].alignment = Alignment(horizontal='center')
    HomNay['A1'].font = Font(name= 'Calibri', size=20)

    HomNay['A2'] = 'Ngày sản xuất'
    HomNay['A2'].alignment = Alignment(horizontal='center')
    HomNay['A2'].font = Font(name= 'Calibri', size=12)
    HomNay['B2'] = 'Giờ lưu dữ liệu'
    HomNay['B2'].alignment = Alignment(horizontal='center')
    HomNay['B2'].font = Font(name= 'Calibri', size=12)
    HomNay['C2'] = 'HẠNG MỤC PHẾ PHẨM'
    HomNay['C2'].alignment = Alignment(horizontal='center')
    HomNay['C2'].font = Font(name= 'Calibri', size=12)
    HomNay['C3'] = 'Tổng số lượng sản xuất'
    HomNay['D3'] = 'Hàn chì'
    HomNay['E3'] = 'Bụi chì'
    HomNay['F3'] = 'PP DV'
    HomNay['G3'] = 'Lệch DV'
    HomNay['H3'] = 'Mối hàn'
    HomNay['I3'] = 'PP dây quấn'
    HomNay['J3'] = 'Nhổ tâm comi'
    HomNay['K3'] = 'PP khác'
    HomNay['L3'] = 'Tổng số lượng PP'
    #HomNay['M3'] = 'Tổng số lượng PP'

    for i in range(67,77):
        HomNay[f'{str(chr(i))}3'].alignment = Alignment(horizontal='center')
        HomNay[f'{str(chr(i))}3'].font = Font(name= 'Calibri', size=12) 

    HomNay.column_dimensions['A'].width = 20
    HomNay.column_dimensions['B'].width = 20
    HomNay.column_dimensions['C'].width = 25
    HomNay.column_dimensions['D'].width = 18
    HomNay.column_dimensions['E'].width = 18
    HomNay.column_dimensions['F'].width = 18
    HomNay.column_dimensions['G'].width = 18
    HomNay.column_dimensions['H'].width = 18
    HomNay.column_dimensions['I'].width = 18
    HomNay.column_dimensions['J'].width = 18
    HomNay.column_dimensions['K'].width = 18
    HomNay.column_dimensions['L'].width = 25
    #HomNay.column_dimensions['M'].width = 25

    wb.remove(wb['Sheet'])

    wb.save(f"excel/{mydate}_Ngay.xlsx")
    try:
        shutil.copy(f"excel/{mydate}_Ngay.xlsx", f"C:/excel/{mydate}_Ngay.xlsx")
    except:
        pass
    writedata('DM6050.U',0) 

def excel_dem():
    today = date.today()
    mydate = today.strftime("%Y_%m_%d")
    wb = openpyxl.Workbook()

    HomNay = wb.create_sheet("Data")


    HomNay.merge_cells('A1:L1')
    HomNay.merge_cells('A2:A3')
    HomNay.merge_cells('B2:B3')
    HomNay.merge_cells('C2:L2')
    #HomNay.unmerge_cells('A2:D2')
    HomNay['A1'] = 'DỮ LIỆU MÁY NQVNHT RS656 A17'
    HomNay['A1'].alignment = Alignment(horizontal='center')
    HomNay['A1'].font = Font(name= 'Calibri', size=20)

    HomNay['A2'] = 'Ngày sản xuất'
    HomNay['A2'].alignment = Alignment(horizontal='center')
    HomNay['A2'].font = Font(name= 'Calibri', size=12)
    HomNay['B2'] = 'Giờ lưu dữ liệu'
    HomNay['B2'].alignment = Alignment(horizontal='center')
    HomNay['B2'].font = Font(name= 'Calibri', size=12)
    HomNay['C2'] = 'HẠNG MỤC PHẾ PHẨM'
    HomNay['C2'].alignment = Alignment(horizontal='center')
    HomNay['C2'].font = Font(name= 'Calibri', size=12)
    HomNay['C3'] = 'Tổng số lượng sản xuất'
    HomNay['D3'] = 'Hàn chì'
    HomNay['E3'] = 'Bụi chì'
    HomNay['F3'] = 'PP DV'
    HomNay['G3'] = 'Lệch DV'
    HomNay['H3'] = 'Mối hàn'
    HomNay['I3'] = 'PP dây quấn'
    HomNay['J3'] = 'Nhổ tâm comi'
    HomNay['K3'] = 'PP khác'
    HomNay['L3'] = 'Tổng số lượng PP'

    for i in range(67,77):
        HomNay[f'{str(chr(i))}3'].alignment = Alignment(horizontal='center')
        HomNay[f'{str(chr(i))}3'].font = Font(name= 'Calibri', size=12) 

    HomNay.column_dimensions['A'].width = 20
    HomNay.column_dimensions['B'].width = 20
    HomNay.column_dimensions['C'].width = 25
    HomNay.column_dimensions['D'].width = 18
    HomNay.column_dimensions['E'].width = 18
    HomNay.column_dimensions['F'].width = 18
    HomNay.column_dimensions['G'].width = 18
    HomNay.column_dimensions['H'].width = 18
    HomNay.column_dimensions['I'].width = 18
    HomNay.column_dimensions['J'].width = 18
    HomNay.column_dimensions['K'].width = 18
    HomNay.column_dimensions['L'].width = 25
    #HomNay.column_dimensions['M'].width = 25

    wb.remove(wb['Sheet'])

    wb.save(f"excel/{mydate}_Dem.xlsx")
    try:
        shutil.copy(f"excel/{mydate}_Dem.xlsx", f"C:/excel/{mydate}_Dem.xlsx")
    except:
        pass
    writedata('DM6050.U',0) 
  


def removefile():
    directory1 = 'C:/FH/camera1/'

    if os.listdir(directory1) != []:
        for i in glob.glob(directory1+'*'):
            for j in glob.glob(i+'/*'):
                os.remove(j)
            os.rmdir(i)


    print('already delete folder')



class CMyCallback:
    """
    Class that contains a callback function.
    """

    def __init__(self):
        self._image = None
        self._lock = threading.Lock()

    @property
    def image(self):
        """Property: return PyIStImage of the grabbed image."""
        duplicate = None
        self._lock.acquire()
        if self._image is not None:
            duplicate = self._image.copy()
        self._lock.release()
        return duplicate

    def datastream_callback1(self, handle=None, context=None):
        """
        Callback to handle events from DataStream.

        :param handle: handle that `trigger` the callback.
        :param context: user data passed on during callback registration.
        """
        st_datastream = handle.module
        if st_datastream:
            with st_datastream.retrieve_buffer() as st_buffer:
                # Check if the acquired data contains image data.
                if st_buffer.info.is_image_present:
                    # Create an image object.
                    st_image = st_buffer.get_image()

                    # Check the pixelformat of the input image.
                    pixel_format = st_image.pixel_format
                    pixel_format_info = st.get_pixel_format_info(pixel_format)

                    # Only mono or bayer is processed.
                    if not(pixel_format_info.is_mono or pixel_format_info.is_bayer):
                        return

                    # Get raw image data.
                    data = st_image.get_image_data()

                    # Perform pixel value scaling if each pixel component is
                    # larger than 8bit. Example: 10bit Bayer/Mono, 12bit, etc.
                    if pixel_format_info.each_component_total_bit_count > 8:
                        nparr = np.frombuffer(data, np.uint16)
                        division = pow(2, pixel_format_info
                                       .each_component_valid_bit_count - 8)
                        nparr = (nparr / division).astype('uint8')
                    else:
                        nparr = np.frombuffer(data, np.uint8)

                    # Process image for display.
                    nparr = nparr.reshape(st_image.height, st_image.width, 1)

                    # Perform color conversion for Bayer.
                    if pixel_format_info.is_bayer:
                        bayer_type = pixel_format_info.get_pixel_color_filter()
                        if bayer_type == st.EStPixelColorFilter.BayerRG:
                            nparr = cv2.cvtColor(nparr, cv2.COLOR_BAYER_RG2RGB)
                        elif bayer_type == st.EStPixelColorFilter.BayerGR:
                            nparr = cv2.cvtColor(nparr, cv2.COLOR_BAYER_GR2RGB)
                        elif bayer_type == st.EStPixelColorFilter.BayerGB:
                            nparr = cv2.cvtColor(nparr, cv2.COLOR_BAYER_GB2RGB)
                        elif bayer_type == st.EStPixelColorFilter.BayerBG:
                            nparr = cv2.cvtColor(nparr, cv2.COLOR_BAYER_BG2RGB)

                    # Resize image and store to self._image.
                    nparr = cv2.resize(nparr, None,
                                       fx=SCALE_X_CAM1,
                                       fy=SCALE_Y_CAM1)
                    self._lock.acquire()
                    self._image = nparr
                    self._lock.release()


    def datastream_callback2(self, handle=None, context=None):
        """
        Callback to handle events from DataStream.

        :param handle: handle that trigger the callback.
        :param context: user data passed on during callback registration.
        """
        st_datastream = handle.module
        if st_datastream:
            with st_datastream.retrieve_buffer() as st_buffer:
                # Check if the acquired data contains image data.
                if st_buffer.info.is_image_present:
                    # Create an image object.
                    st_image = st_buffer.get_image()

                    # Check the pixelformat of the input image.
                    pixel_format = st_image.pixel_format
                    pixel_format_info = st.get_pixel_format_info(pixel_format)

                    # Only mono or bayer is processed.
                    if not(pixel_format_info.is_mono or pixel_format_info.is_bayer):
                        return

                    # Get raw image data.
                    data = st_image.get_image_data()

                    # Perform pixel value scaling if each pixel component is
                    # larger than 8bit. Example: 10bit Bayer/Mono, 12bit, etc.
                    if pixel_format_info.each_component_total_bit_count > 8:
                        nparr = np.frombuffer(data, np.uint16)
                        division = pow(2, pixel_format_info
                                       .each_component_valid_bit_count - 8)
                        nparr = (nparr / division).astype('uint8')
                    else:
                        nparr = np.frombuffer(data, np.uint8)

                    # Process image for display.
                    nparr = nparr.reshape(st_image.height, st_image.width, 1)

                    # Perform color conversion for Bayer.
                    if pixel_format_info.is_bayer:
                        bayer_type = pixel_format_info.get_pixel_color_filter()
                        if bayer_type == st.EStPixelColorFilter.BayerRG:
                            nparr = cv2.cvtColor(nparr, cv2.COLOR_BAYER_RG2RGB)
                        elif bayer_type == st.EStPixelColorFilter.BayerGR:
                            nparr = cv2.cvtColor(nparr, cv2.COLOR_BAYER_GR2RGB)
                        elif bayer_type == st.EStPixelColorFilter.BayerGB:
                            nparr = cv2.cvtColor(nparr, cv2.COLOR_BAYER_GB2RGB)
                        elif bayer_type == st.EStPixelColorFilter.BayerBG:
                            nparr = cv2.cvtColor(nparr, cv2.COLOR_BAYER_BG2RGB)

                    # Resize image and store to self._image.
                    nparr = cv2.resize(nparr, None,
                                       fx=SCALE_X_CAM2,
                                       fy=SCALE_Y_CAM2)
                    self._lock.acquire()
                    self._image = nparr
                    self._lock.release()


def set_enumeration(nodemap, enum_name, entry_name):
    enum_node = st.PyIEnumeration(nodemap.get_node(enum_name))
    entry_node = st.PyIEnumEntry(enum_node[entry_name])
    enum_node.set_entry_value(entry_node)



def setup_camera1_stc():
    #lobal error_cam1
    #while error_cam1 == True:
    try:
        st_device1 = st_system.create_first_device()
        print('Device1=', st_device1.info.display_name)
        st_datastream1 = st_device1.create_datastream()
        callback1 = st_datastream1.register_callback(cb_func1)
        st_datastream1.start_acquisition()
        st_device1.acquisition_start()
        remote_nodemap1 = st_device1.remote_port.nodemap
        set_enumeration(remote_nodemap1,"TriggerMode", "Off")
        error_cam1 = False
        return  st_datastream1, st_device1,remote_nodemap1

    except Exception as exception:
        print(' Error Cam 1:', exception)
        str_error = "Error"
        window['result_cam1'].update(value= str_error, text_color='red',)



def setup_camera2_stc():
    #global error_cam2
    #while error_cam2 == True:
    try:
        st_device2 = st_system.create_first_device()
        print('Device2=', st_device2.info.display_name)
        st_datastream2 = st_device2.create_datastream()
        callback2 = st_datastream2.register_callback(cb_func2)
        st_datastream2.start_acquisition()
        st_device2.acquisition_start()
        remote_nodemap2 = st_device2.remote_port.nodemap
        set_enumeration(remote_nodemap2,"TriggerMode", "Off")
        error_cam2 = False
        return  st_datastream2, st_device2,remote_nodemap2
    except Exception as exception:     
        print('Error Cam 2:', exception)
        str_error = "Error"
        #sg.popup(str_error,font=('Helvetica',15), text_color='red',keep_on_top= True)
        window['result_cam2'].update(value= str_error, text_color='red')


soc = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

def socket_connect(host, port):
    try:
        soc.connect((host, port))
        return True
    except OSError:
        print("Can't connect to PLC")
        sleep(3)
        print("Reconnecting....")
        return False

def readdata(data):
    a = 'RD '
    c = '\x0D'
    d = a+ data +c
    datasend = d.encode("UTF-8")
    soc.sendall(datasend)
    data = soc.recv(1024)
    datadeco = data.decode("UTF-8")
    data1 = int(datadeco)

    return data1

#Write data
def writedata(register, data):
    a = 'WR '
    b = ' '
    c = '\x0D'
    d = a+ register + b + str(data) + c
    datasend  = d.encode("UTF-8")
    soc.sendall(datasend)
    datares = soc.recv(1024)
    #print(datares)


def time_to_name():
    current_time = datetime.datetime.now() 
    name_folder = str(current_time)
    name_folder = list(name_folder)
    for i in range(len(name_folder)):
        if name_folder[i] == ':':
            name_folder[i] = '-'
        if name_folder[i] == ' ':
            name_folder[i] ='_'
        if name_folder[i] == '.':
            name_folder[i] ='-'
    name_folder = ''.join(name_folder)
    return name_folder




def load_theme():
    name_themes = []
    with open('static/theme.txt') as lines:
        for line in lines:
            _, name_theme = line.strip().split(':')
            name_themes.append(name_theme)
    return name_themes

def load_choosemodel():
    with open('static/choose_model.txt') as lines:
        for line in lines:
            _, name_model = line.strip().split('=')
    return name_model

def save_theme(name_theme):
    line = 'theme:' + name_theme
    with open('static/theme.txt','w') as f:
        f.write(line)


def save_choosemodel(name_model):
    line = 'choose_model=' + name_model
    with open('static/choose_model.txt','w') as f:
        f.write(line)

def load_model(i):
    with open('static/model'+ str(i) + '.txt','r') as lines:
        for line in lines:
            _, name_model = line.strip().split('=')
    return name_model

def save_model(i,name_model):
    line = 'model' + str(i) + '=' + name_model
    with open('static/model' + str(i) + '.txt','w') as f:
        f.write(line)


def str2bool(v):
    return v.lower() in ("yes", "true", "t", "1")

def load_all(model,i):
    values_all = []
    with open('static/all'+ str(i) + '.txt','r') as lines:
        for line in lines:
            _, name_all = line.strip().split('=')
            values_all.append(name_all)
    window['file_weights' + str(i)].update(value=values_all[0])
    window['conf_thres' + str(i)].update(value=values_all[1])
    a=1
    for item in range(len(model.names)):
        window[f'{model.names[item]}_' + str(i)].update(value=str2bool(values_all[a+1]))
        window[f'{model.names[item]}_OK_' + str(i)].update(value=str2bool(values_all[a+2]))
        window[f'{model.names[item]}_Num_' + str(i)].update(value=str(values_all[a+3]))
        window[f'{model.names[item]}_NG_' + str(i)].update(value=str2bool(values_all[a+4]))
        window[f'{model.names[item]}_Wn_' + str(i)].update(value=str(values_all[a+5]))
        window[f'{model.names[item]}_Wx_' + str(i)].update(value=str(values_all[a+6]))
        window[f'{model.names[item]}_Hn_' + str(i)].update(value=str(values_all[a+7]))
        window[f'{model.names[item]}_Hx_' + str(i)].update(value=str(values_all[a+8]))
        a += 8


def save_all(model,i):
    with open('static/all'+ str(i) + '.txt','w') as f:
        f.write('weights' + str(i) + '=' + str(values['file_weights' + str(i)]))
        f.write('\n')
        f.write('conf' + str(i) + '=' + str(values['conf_thres' + str(i)]))
        f.write('\n')

        for item in range(len(model.names)):
            f.write(str(f'{model.names[item]}_' + str(i)) + '=' + str(values[f'{model.names[item]}_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_OK_' + str(i)) + '=' + str(values[f'{model.names[item]}_OK_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_Num_' + str(i)) + '=' + str(values[f'{model.names[item]}_Num_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_NG_' + str(i)) + '=' + str(values[f'{model.names[item]}_NG_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_Wn_' + str(i)) + '=' + str(values[f'{model.names[item]}_Wn_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_Wx_' + str(i)) + '=' + str(values[f'{model.names[item]}_Wx_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_Hn_' + str(i)) + '=' + str(values[f'{model.names[item]}_Hn_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_Hx_' + str(i)) + '=' + str(values[f'{model.names[item]}_Hx_' + str(i)]))
            if item != len(model.names)-1:
                f.write('\n')



def load_all_sql(i,choose_model):
    conn = sqlite3.connect('modeldb_2_PLC_conf.db')
    cursor = conn.execute("SELECT ChooseModel,Camera,Weights,Confidence,OK_Cam1,OK_Cam2,NG_Cam1,NG_Cam2,Folder_OK_Cam1,Folder_OK_Cam2,Folder_NG_Cam1,Folder_NG_Cam2,Joined,Ok,Num,NG,WidthMin,WidthMax,HeightMin,HeightMax,PLC_NG,PLC_OK,Conf from MYMODEL")
    for row in cursor:
        #if row[0] == values['choose_model']:
        if row[0] == choose_model:
            row1_a, row1_b = row[1].strip().split('_')
            if row1_a == str(i) and row1_b == '0':
                window['file_weights' + str(i)].update(value=row[2])
                window['conf_thres' + str(i)].update(value=row[3])
                window['have_save_OK_1'].update(value=str2bool(row[4]))
                window['have_save_OK_2'].update(value=str2bool(row[5]))
    
                window['have_save_NG_1'].update(value=str2bool(row[6]))
                window['have_save_NG_2'].update(value=str2bool(row[7]))


                window['save_OK_1'].update(value=row[8])
                window['save_OK_2'].update(value=row[9])

        
                window['save_NG_1'].update(value=row[10])
                window['save_NG_2'].update(value=row[11])



                model = torch.hub.load('./levu','custom', path= row[2], source='local',force_reload =False)
            if row1_a == str(i):
                for item in range(len(model.names)):
                    if int(row1_b) == item:
                        window[f'{model.names[item]}_' + str(i)].update(value=str2bool(row[12]))
                        window[f'{model.names[item]}_OK_' + str(i)].update(value=str2bool(row[13]))
                        window[f'{model.names[item]}_Num_' + str(i)].update(value=str(row[14]))
                        window[f'{model.names[item]}_NG_' + str(i)].update(value=str2bool(row[15]))
                        window[f'{model.names[item]}_Wn_' + str(i)].update(value=str(row[16]))
                        window[f'{model.names[item]}_Wx_' + str(i)].update(value=str(row[17]))
                        window[f'{model.names[item]}_Hn_' + str(i)].update(value=str(row[18]))
                        window[f'{model.names[item]}_Hx_' + str(i)].update(value=str(row[19]))
                        window[f'{model.names[item]}_PLC_' + str(i)].update(value=str(row[20]))
                        window[f'OK_PLC_' + str(i)].update(value=str(row[21]))
                        window[f'{model.names[item]}_Conf_' + str(i)].update(value=str(row[22]))

                    

    conn.close()


def save_all_sql(model,i,choose_model):
    conn = sqlite3.connect('modeldb_2_PLC_conf.db')
    cursor = conn.execute("SELECT ChooseModel,Camera,Weights,Confidence,OK_Cam1,OK_Cam2,NG_Cam1,NG_Cam2,Folder_OK_Cam1,Folder_OK_Cam2,Folder_NG_Cam1,Folder_NG_Cam2,Joined,Ok,Num,NG,WidthMin,WidthMax,HeightMin,HeightMax,PLC_NG,PLC_OK,Conf from MYMODEL")
    update = 0 

    for row in cursor:
        if row[0] == choose_model:            
            row1_a, _ = row[1].strip().split('_')
            if row1_a == str(i):
                conn.execute("DELETE FROM MYMODEL WHERE (ChooseModel = ? AND Camera LIKE ?)", (choose_model,str(i) + '%'))
                for item in range(len(model.names)):
                    #conn.execute("UPDATE MYMODEL SET ChooseModel = ? , Camera = ?, Weights = ?,Confidence = ?, Joined = ?, Ok = ?, Num = ?, NG = ?, WidthMin = ?, WidthMax = ?, HeightMin = ?, HeightMax = ? WHERE (ChooseModel = ? AND Camera = ?)",(str(values['choose_model']),str(i)+ '_' +str(item) ,str(values['file_weights' + str(i)]),int(values['conf_thres' + str(i)]), str(values[model.names[item] + '_' + str(i)]), str(values[model.names[item]+ '_OK_' + str(i)]), int(values[model.names[item]+ '_Num_' + str(i)]), str(values[model.names[item]+ '_NG_' + str(i)]), int(values[model.names[item] + '_Wn_' + str(i)]), int(values[model.names[item] + '_Wx_' + str(i)]), int(values[model.names[item]+ '_Hn_' + str(i)]), int(values[model.names[item] + '_Hx_' + str(i)]), choose_model,str(i) + '_' + str(item)))
                    #conn.execute("DELETE FROM MYMODEL WHERE (ChooseModel = ? AND Camera = ?)", (choose_model,str(i) + '_' + str(item)))
                    conn.execute("INSERT INTO MYMODEL (ChooseModel,Camera, Weights,Confidence,OK_Cam1,OK_Cam2,NG_Cam1,NG_Cam2,Folder_OK_Cam1,Folder_OK_Cam2,Folder_NG_Cam1,Folder_NG_Cam2,Joined,Ok,Num,NG,WidthMin, WidthMax,HeightMin,HeightMax,PLC_NG,PLC_OK,Conf) \
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (str(values['choose_model']),str(i)+ '_' +str(item) ,str(values['file_weights' + str(i)]), int(values['conf_thres' + str(i)]),str(values['have_save_OK_1']),str(values['have_save_OK_2']),str(values['have_save_NG_1']),str(values['have_save_NG_2']),str(values['save_OK_1']),str(values['save_OK_2']),str(values['save_NG_1']),str(values['save_NG_2']),str(values[f'{model.names[item]}_' + str(i)]), str(values[f'{model.names[item]}_OK_' + str(i)]), int(values[f'{model.names[item]}_Num_' + str(i)]), str(values[f'{model.names[item]}_NG_' + str(i)]), int(values[f'{model.names[item]}_Wn_' + str(i)]), int(values[f'{model.names[item]}_Wx_' + str(i)]), int(values[f'{model.names[item]}_Hn_' + str(i)]), int(values[f'{model.names[item]}_Hx_' + str(i)]), int(values[f'{model.names[item]}_PLC_' + str(i)]), int(values['OK_PLC_' + str(i)]),int(values[f'{model.names[item]}_Conf_' + str(i)])))           
                    #conn.execute("INSERT INTO MYMODEL (ChooseModel,Camera, Weights,Confidence,Joined,Ok,Num,NG,WidthMin, WidthMax,HeightMin,HeightMax) \
                    #    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)", (str(values['choose_model']),str(i)+ '_' +str(item) ,str(values['file_weights' + str(i)]), int(values['conf_thres' + str(i)]),str(values[f'{model.names[item]}_' + str(i)]), str(values[f'{model.names[item]}_OK_' + str(i)]), int(values[f'{model.names[item]}_Num_' + str(i)]), str(values[f'{model.names[item]}_NG_' + str(i)]), int(values[f'{model.names[item]}_Wn_' + str(i)]), int(values[f'{model.names[item]}_Wx_' + str(i)]), int(values[f'{model.names[item]}_Hn_' + str(i)]), int(values[f'{model.names[item]}_Hx_' + str(i)])))           
                    update = 1
                break

    if update == 0:
        for item in range(len(model.names)):
            conn.execute("INSERT INTO MYMODEL (ChooseModel,Camera, Weights,Confidence,OK_Cam1,OK_Cam2,NG_Cam1,NG_Cam2,Folder_OK_Cam1,Folder_OK_Cam2,Folder_NG_Cam1,Folder_NG_Cam2,Joined,Ok,Num,NG,WidthMin, WidthMax,HeightMin,HeightMax,PLC_NG,PLC_OK,Conf) \
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (str(values['choose_model']),str(i)+ '_' +str(item) ,str(values['file_weights' + str(i)]), int(values['conf_thres' + str(i)]),str(values['have_save_OK_1']),str(values['have_save_OK_2']),str(values['have_save_NG_1']),str(values['have_save_NG_2']),str(values['save_OK_1']),str(values['save_OK_2']),str(values['save_NG_1']),str(values['save_NG_2']),str(values[f'{model.names[item]}_' + str(i)]), str(values[f'{model.names[item]}_OK_' + str(i)]), int(values[f'{model.names[item]}_Num_' + str(i)]), str(values[f'{model.names[item]}_NG_' + str(i)]), int(values[f'{model.names[item]}_Wn_' + str(i)]), int(values[f'{model.names[item]}_Wx_' + str(i)]), int(values[f'{model.names[item]}_Hn_' + str(i)]), int(values[f'{model.names[item]}_Hx_' + str(i)]),int(values[f'{model.names[item]}_PLC_' + str(i)]), int(values['OK_PLC_' + str(i)]),int(values[f'{model.names[item]}_Conf_' + str(i)])))
            #conn.execute("INSERT INTO MYMODEL (ChooseModel,Camera, Weights,Confidence,Joined,Ok,Num,NG,WidthMin, WidthMax,HeightMin,HeightMax) \
            #    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)", (str(values['choose_model']),str(i)+ '_' +str(item) ,str(values['file_weights' + str(i)]), int(values['conf_thres' + str(i)]),str(values[f'{model.names[item]}_' + str(i)]), str(values[f'{model.names[item]}_OK_' + str(i)]), int(values[f'{model.names[item]}_Num_' + str(i)]), str(values[f'{model.names[item]}_NG_' + str(i)]), int(values[f'{model.names[item]}_Wn_' + str(i)]), int(values[f'{model.names[item]}_Wx_' + str(i)]), int(values[f'{model.names[item]}_Hn_' + str(i)]), int(values[f'{model.names[item]}_Hx_' + str(i)])))
        
    for row in cursor:
        if row[0] == choose_model:
            conn.execute("UPDATE MYMODEL SET OK_Cam1 = ? , OK_Cam2 = ?, NG_Cam1 = ?,NG_Cam2 = ?, Folder_OK_Cam1 = ?, Folder_OK_Cam2 = ?,Folder_NG_Cam1 = ?, Folder_NG_Cam2 = ? WHERE ChooseModel = ? ",(str(values['have_save_OK_1']),str(values['have_save_OK_2']),str(values['have_save_NG_1']),str(values['have_save_NG_2']),str(values['save_OK_1']),str(values['save_OK_2']),str(values['save_NG_1']),str(values['save_NG_2']),choose_model))


    conn.commit()
    conn.close()



def program_camera1_test(model,size,conf):
    global img1_orgin
    global myindex1
    if keyboard.is_pressed('1'): 

        img1_orgin = my_callback1.image 
        img1_save = img1_orgin
        print('MODEL 1')


        t1 = time.time()


        img1_orgin = cv2.cvtColor(img1_orgin, cv2.COLOR_BGR2RGB)

        result1 = model(img1_orgin,size= size,conf = conf) 
        table1 = result1.pandas().xyxy[0]
        area_remove1 = []

        myresult1 =0 

        for item in range(len(table1.index)):
            width1 = table1['xmax'][item] - table1['xmin'][item]
            height1 = table1['ymax'][item] - table1['ymin'][item]
            conf1 = table1['confidence'][item] * 100

            #area1 = width1*height1
            label_name = table1['name'][item]
            for i1 in range(len(model1.names)):
                if values[f'{model1.names[i1]}_1'] == True:
                    #if values[f'{model1.names[i1]}_WH'] == True:
                    if label_name == model1.names[i1]:
                        if width1 < int(values[f'{model1.names[i1]}_Wn_1']): 
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)
                        elif width1 > int(values[f'{model1.names[i1]}_Wx_1']): 
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)
                        elif height1 < int(values[f'{model1.names[i1]}_Hn_1']): 
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)
                        elif height1 > int(values[f'{model1.names[i1]}_Hx_1']): 
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)
                        elif conf1 < int(values[f'{model1.names[i1]}_Conf_1']):
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)                                             
                if values[f'{model1.names[i1]}_1'] == False:
                    if label_name == model1.names[i1]:
                        table1.drop(item, axis=0, inplace=True)
                        area_remove1.append(item)

        names1 = list(table1['name'])

        show1 = np.squeeze(result1.render(area_remove1))
        show1 = cv2.resize(show1, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
        show1 = cv2.cvtColor(show1, cv2.COLOR_BGR2RGB)

        #ta = time.time()
        for i1 in range(len(model1.names)):
            register_ng = str(values[f'{model1.names[i1]}_PLC_1'])

            if values[f'{model1.names[i1]}_OK_1'] == True:
                len_name1 = 0
                for name1 in names1:
                    if name1 == model1.names[i1]:
                        len_name1 +=1
                if len_name1 != int(values[f'{model1.names[i1]}_Num_1']):
                    print('NG')
                    writedata('DM'+ register_ng +'.U',1) 

                    #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register_ng,b'\x00\x01',1)
                    t2 = time.time() - t1
                    print(t2) 
                    cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                    window['result_cam1'].update(value= 'NG', text_color='red')
                    
                    myresult1 = 1
      

            if values[f'{model1.names[i1]}_NG_1'] == True:
                if model1.names[i1] in names1:
                    print('NG')
                    writedata('DM'+ register_ng +'.U',1) 

                    t2 = time.time() - t1
                    print(t2) 
                    cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                    window['result_cam1'].update(value= 'NG', text_color='red')    
                    
                    myresult1 = 1         
  

        if myresult1 == 0:
            print('OK')
            writedata('DM'+ str(values['OK_PLC_1']) +'.U',1) 

            #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,(3000).to_bytes(2, byteorder='big') + b'\x00',b'\x00\x01',1)
            t2 = time.time() - t1
            print(t2) 
            cv2.putText(show1, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
            window['result_cam1'].update(value= 'OK', text_color='green')
            if values['have_save_OK_1']:
                name_folder_ng = time_to_name()
                cv2.imwrite(values['save_OK_1']  + '/' + name_folder_ng + '.jpg',img1_save)
        else:
            if values['have_save_NG_1']:
                name_folder_ng = time_to_name()
                cv2.imwrite(values['save_NG_1']  + '/' + name_folder_ng + '.jpg',img1_save)

        time_cam1 = str(int(t2*1000)) + 'ms'
        window['time_cam1'].update(value= time_cam1, text_color='black') 
        cv2.putText(show1, ' M1',(100,100),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)

        imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
        window['image1'].update(data= imgbytes1)
        if myindex1 == 6:
            myindex1 = 0
        myindex1 +=1

        if myindex1 == 1:
            imgbytes = np.zeros([100,100,3],dtype=np.uint8)
            imgbytes = cv2.resize(imgbytes, (int(image_width_display/2.2),int(image_height_display/1.4)), interpolation = cv2.INTER_AREA)
            imgbytes = cv2.imencode('.png',imgbytes)[1].tobytes()
            window['image1.1'].update(data= imgbytes)
            window['image1.2'].update(data= imgbytes)
            window['image1.3'].update(data= imgbytes)
            window['image1.4'].update(data= imgbytes)
            window['image1.5'].update(data= imgbytes)
            window['image1.6'].update(data= imgbytes)
            show1 = cv2.resize(show1, (int(image_width_display/2.2),int(image_height_display/1.4)), interpolation = cv2.INTER_AREA)
            cv2.putText(show1, '1',(30,70),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),1)

            imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()

            window['image1.1'].update(data= imgbytes1)

        if myindex1 == 3:
            show1 = cv2.resize(show1, (int(image_width_display/2.2),int(image_height_display/1.4)), interpolation = cv2.INTER_AREA)
            cv2.putText(show1, '3',(30,70),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),1)
            
            imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()

            window['image1.3'].update(data= imgbytes1)

        if myindex1 == 5:
            show1 = cv2.resize(show1, (int(image_width_display/2.2),int(image_height_display/1.4)), interpolation = cv2.INTER_AREA)
            cv2.putText(show1, '5',(30,70),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),1)
            
            imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()

            window['image1.5'].update(data= imgbytes1)


        print('---------------------------------------------')



def program_camera2_test(model,size,conf):

    global img1_orgin
    global myindex1

    if keyboard.is_pressed('2'): 

        img2_orgin = my_callback1.image 
        img2_save = img2_orgin
        #edit
        print('MODEL 2')


        t1 = time.time()

        img2_orgin = cv2.cvtColor(img2_orgin, cv2.COLOR_BGR2RGB)

        result2 = model(img2_orgin,size= size,conf = conf) 
        table2 = result2.pandas().xyxy[0]
        area_remove2 = []

        myresult2 =0 

        for item in range(len(table2.index)):
            width2 = table2['xmax'][item] - table2['xmin'][item]
            height2 = table2['ymax'][item] - table2['ymin'][item]
            conf2 = table2['confidence'][item] * 100

            label_name = table2['name'][item]
            for i2 in range(len(model2.names)):
                if values[f'{model2.names[i2]}_2'] == True:
                    if label_name == model2.names[i2]:
                        if width2 < int(values[f'{model2.names[i2]}_Wn_2']): 
                            table2.drop(item, axis=0, inplace=True)
                            area_remove2.append(item)
                        elif width2 > int(values[f'{model2.names[i2]}_Wx_2']): 
                            table2.drop(item, axis=0, inplace=True)
                            area_remove2.append(item)
                        elif height2 < int(values[f'{model2.names[i2]}_Hn_2']): 
                            table2.drop(item, axis=0, inplace=True)
                            area_remove2.append(item)
                        elif height2 > int(values[f'{model2.names[i2]}_Hx_2']): 
                            table2.drop(item, axis=0, inplace=True)
                            area_remove2.append(item)
                        elif conf2 < int(values[f'{model2.names[i2]}_Conf_2']):
                            table2.drop(item, axis=0, inplace=True)
                            area_remove2.append(item)    

                if values[f'{model2.names[i2]}_2'] == False:
                    if label_name == model2.names[i2]:
                        table2.drop(item, axis=0, inplace=True)
                        area_remove2.append(item)

        names2 = list(table2['name'])

        show2 = np.squeeze(result2.render(area_remove2))
        show2 = cv2.resize(show2, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
        show2 = cv2.cvtColor(show2, cv2.COLOR_BGR2RGB)
        #ta = time.time()

        for i2 in range(len(model2.names)):
            register_ng = str(values[f'{model2.names[i2]}_PLC_2'])

            if values[f'{model2.names[i2]}_OK_2'] == True:
                len_name2 = 0
                for name2 in names2:
                    if name2 == model2.names[i2]:
                        len_name2 +=1
                if len_name2 != int(values[f'{model2.names[i2]}_Num_2']):
                    print('NG')
                    writedata('DM'+ register_ng +'.U',1) 
                    t2 = time.time() - t1
                    print(t2) 
                    cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                    window['result_cam1'].update(value= 'NG', text_color='red')
                    
                    myresult2 = 1
       

            if values[f'{model2.names[i2]}_NG_2'] == True:
                if model2.names[i2] in names2:
                    print('NG')
                    writedata('DM'+ register_ng +'.U',1) 

                    t2 = time.time() - t1
                    print(t2) 
                    cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                    window['result_cam1'].update(value= 'NG', text_color='red')    
                    
                    myresult2 = 1         
                  

        if myresult2 == 0:
            print('OK')
            writedata('DM'+ str(values['OK_PLC_2']) +'.U',1) 

            t2 = time.time() - t1
            print(t2) 
            cv2.putText(show2, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
            window['result_cam1'].update(value= 'OK', text_color='green')
            if values['have_save_OK_2']:
                name_folder_ng = time_to_name()
                cv2.imwrite(values['save_OK_2']  + '/' + name_folder_ng + '.jpg',img2_save)
        else:
            if values['have_save_NG_2']:
                name_folder_ng = time_to_name()
                cv2.imwrite(values['save_NG_2']  + '/' + name_folder_ng + '.jpg',img2_save)



        time_cam2 = str(int(t2*1000)) + 'ms'
        window['time_cam1'].update(value= time_cam2, text_color='black') 
        cv2.putText(show2, 'M2',(120,100),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)

        imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()
        window['image1'].update(data= imgbytes2)

        myindex1 +=1


        if myindex1 == 2:
            show2 = cv2.resize(show2, (int(image_width_display/2.2),int(image_height_display/1.4)), interpolation = cv2.INTER_AREA)
            cv2.putText(show2, '2',(30,70),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),1)
            
            imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()

            window['image1.2'].update(data= imgbytes2)

        if myindex1 == 4:
            show2 = cv2.resize(show2, (int(image_width_display/2.2),int(image_height_display/1.4)), interpolation = cv2.INTER_AREA)
            cv2.putText(show2, '4',(30,70),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),1)
            
            imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()

            window['image1.4'].update(data= imgbytes2)

        if myindex1 == 6:
            show2 = cv2.resize(show2, (int(image_width_display/2.2),int(image_height_display/1.4)), interpolation = cv2.INTER_AREA)
            cv2.putText(show2, '6',(30,70),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),1)
            
            imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()

            window['image1.6'].update(data= imgbytes2)  


        print('---------------------------------------------')



def program_camera1_1():
    read_7102 = readdata('DM7102') 
    read_7100 = readdata('DM7100')


    if read_7102 == 1 and read_7100  == 1: 

        writedata('DM7002.U',1)

        writedata('DM7102.U',0) 
        writedata('DM7100.U',0)


        print('1---------------------------------------------')



def program_camera2_1():
    read_7102 = readdata('DM7102') 
    read_7100 = readdata('DM7100')


    if read_7102 == 1 and read_7100  == 2: 

        writedata('DM7002.U',1)

        writedata('DM7102.U',0) 
        writedata('DM7100.U',0)


        print('2---------------------------------------------')




def program_camera1(model,size,conf):
    read_7102 = readdata('DM7102') 
    read_7100 = readdata('DM7100')
    global img1_orgin
    global myindex1


    if read_7102 == 1 and read_7100  == 1: 

        writedata('DM7002.U',1)

        img1_orgin = my_callback1.image 
        img1_save = img1_orgin
        print('MODEL 1')

        writedata('DM7102.U',0) 
        writedata('DM7100.U',0)

        t1 = time.time()


        img1_orgin = cv2.cvtColor(img1_orgin, cv2.COLOR_BGR2RGB)

        result1 = model(img1_orgin,size= size,conf = conf) 
        table1 = result1.pandas().xyxy[0]
        area_remove1 = []

        myresult1 =0 
        writedata('DM7002.U',0)

        for item in range(len(table1.index)):
            width1 = table1['xmax'][item] - table1['xmin'][item]
            height1 = table1['ymax'][item] - table1['ymin'][item]
            conf1 = table1['confidence'][item] * 100
            # if item ==0:
            #     center_x = int(float(table1['xmin'][item]) + float(table1['xmax'][item] - table1['xmin'][item])/2)
            #     center_y = int(float(table1['ymin'][item]) + floloat(table1['ymax'][item] - table1['ymin'][item])/2)
            #     print('x label 1: ',center_x)
            #     print('y label 1: ',center_y)



            #area1 = width1*height1
            label_name = table1['name'][item]
            for i1 in range(len(model1.names)):
                if values[f'{model1.names[i1]}_1'] == True:
                    #if values[f'{model1.names[i1]}_WH'] == True:
                    if label_name == model1.names[i1]:
                        if width1 < int(values[f'{model1.names[i1]}_Wn_1']): 
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)
                        elif width1 > int(values[f'{model1.names[i1]}_Wx_1']): 
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)
                        elif height1 < int(values[f'{model1.names[i1]}_Hn_1']): 
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)
                        elif height1 > int(values[f'{model1.names[i1]}_Hx_1']): 
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)
                        elif conf1 < int(values[f'{model1.names[i1]}_Conf_1']):
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)                                             
                if values[f'{model1.names[i1]}_1'] == False:
                    if label_name == model1.names[i1]:
                        table1.drop(item, axis=0, inplace=True)
                        area_remove1.append(item)

        names1 = list(table1['name'])

        show1 = np.squeeze(result1.render(area_remove1))
        show1 = cv2.resize(show1, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
        show1 = cv2.cvtColor(show1, cv2.COLOR_BGR2RGB)

        #ta = time.time()
        for i1 in range(len(model1.names)):
            register_ng = str(values[f'{model1.names[i1]}_PLC_1'])

            if values[f'{model1.names[i1]}_OK_1'] == True:
                len_name1 = 0
                for name1 in names1:
                    if name1 == model1.names[i1]:
                        len_name1 +=1
                if len_name1 != int(values[f'{model1.names[i1]}_Num_1']):
                    print('NG')
                    writedata('DM'+ register_ng +'.U',1) 
                    print('DM'+ register_ng +'.U')

                    #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register_ng,b'\x00\x01',1)
                    t2 = time.time() - t1
                    print(t2) 
                    cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                    window['result_cam1'].update(value= 'NG', text_color='red')
                    
                    myresult1 = 1
           

            if values[f'{model1.names[i1]}_NG_1'] == True:
                if model1.names[i1] in names1:
                    print('NG')
                    writedata('DM'+ register_ng +'.U',1) 
                    print('DM'+ register_ng +'.U')

                    t2 = time.time() - t1
                    print(t2) 
                    cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                    window['result_cam1'].update(value= 'NG', text_color='red')    
                    
                    myresult1 = 1         
        

        if myresult1 == 0:
            print('OK')
            writedata('DM'+ str(values['OK_PLC_1']) +'.U',1) 

            #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,(3000).to_bytes(2, byteorder='big') + b'\x00',b'\x00\x01',1)
            t2 = time.time() - t1
            print(t2) 
            cv2.putText(show1, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
            window['result_cam1'].update(value= 'OK', text_color='green')
            if values['have_save_OK_1']:
                # name_folder_ok = time_to_name()
                # cv2.imwrite(values['save_OK_1']  + '/' + name_folder_ok + '.jpg',img1_save)

                name_folder_ng = time_to_name()
                today = datetime.date.today()
                if not os.path.isdir(values['save_OK_1'] + "/" + str(today)):
                    os.mkdir(values['save_OK_1']  + "/" + str(today))


                cv2.imwrite(values['save_OK_1']  + '/' + name_folder_ng + '.jpg',img1_save)


        else:
            if values['have_save_NG_1']:
                # name_folder_ng = time_to_name()
                # cv2.imwrite(values['save_NG_1']  + '/' + name_folder_ng + '.jpg',img1_save)
                # cv2.imwrite('D:/1' + '/' + name_folder_ng + '.jpg',img1_save)

                name_folder_ng = time_to_name()
                today = datetime.date.today()
                if not os.path.isdir(values['save_NG_1'] + "/" + str(today)):
                    os.mkdir(values['save_NG_1'] + "/" + str(today))

                cv2.imwrite(values['save_NG_1']  + "/" + str(today) + '/' + name_folder_ng + '.jpg',img1_save)
        



        writedata('DM7520.U',1)
        

        time_cam1 = str(int(t2*1000)) + 'ms'
        window['time_cam1'].update(value= time_cam1, text_color='black') 
        cv2.putText(show1, ' M1',(100,100),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)

        imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
        window['image1'].update(data= imgbytes1)

        if myindex1 == 6:
            myindex1 = 0
        myindex1 +=1


        if myindex1 == 1:
            imgbytes = np.zeros([100,100,3],dtype=np.uint8)
            imgbytes = cv2.resize(imgbytes, (int(image_width_display/2.2),int(image_height_display/1.4)), interpolation = cv2.INTER_AREA)
            imgbytes = cv2.imencode('.png',imgbytes)[1].tobytes()
            window['image1.1'].update(data= imgbytes)
            window['image1.2'].update(data= imgbytes)
            window['image1.3'].update(data= imgbytes)
            window['image1.4'].update(data= imgbytes)
            window['image1.5'].update(data= imgbytes)
            window['image1.6'].update(data= imgbytes)
            show1 = cv2.resize(show1, (int(image_width_display/2.2),int(image_height_display/1.4)), interpolation = cv2.INTER_AREA)
            cv2.putText(show1, '1',(30,70),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),1)

            imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()

            window['image1.1'].update(data= imgbytes1)

        if myindex1 == 3:
            show1 = cv2.resize(show1, (int(image_width_display/2.2),int(image_height_display/1.4)), interpolation = cv2.INTER_AREA)
            cv2.putText(show1, '3',(30,70),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),1)
            
            imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()

            window['image1.3'].update(data= imgbytes1)

        if myindex1 == 5:
            show1 = cv2.resize(show1, (int(image_width_display/2.2),int(image_height_display/1.4)), interpolation = cv2.INTER_AREA)
            cv2.putText(show1, '5',(30,70),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),1)
            
            imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()

            window['image1.5'].update(data= imgbytes1)



        print('---------------------------------------------')



def program_camera2(model,size,conf):
    read_7102 = readdata('DM7102') 
    read_7100 = readdata('DM7100')
    global img1_orgin
    global myindex1

    if read_7102 == 1 and read_7100  == 2: 

        writedata('DM7002.U',1)
        img2_orgin = my_callback1.image 
        img2_save = img2_orgin
        #edit
        print('MODEL 2')


        writedata('DM7102.U',0) 
        writedata('DM7100.U',0)
        t1 = time.time()

        img2_orgin = cv2.cvtColor(img2_orgin, cv2.COLOR_BGR2RGB)

        result2 = model(img2_orgin,size= size,conf = conf) 
        table2 = result2.pandas().xyxy[0]
        area_remove2 = []

        myresult2 =0 
        writedata('DM7002.U',0)

        for item in range(len(table2.index)):
            width2 = table2['xmax'][item] - table2['xmin'][item]
            height2 = table2['ymax'][item] - table2['ymin'][item]
            conf2 = table2['confidence'][item] * 100

            label_name = table2['name'][item]
            for i2 in range(len(model2.names)):
                if values[f'{model2.names[i2]}_2'] == True:
                    if label_name == model2.names[i2]:
                        if width2 < int(values[f'{model2.names[i2]}_Wn_2']): 
                            table2.drop(item, axis=0, inplace=True)
                            area_remove2.append(item)
                        elif width2 > int(values[f'{model2.names[i2]}_Wx_2']): 
                            table2.drop(item, axis=0, inplace=True)
                            area_remove2.append(item)
                        elif height2 < int(values[f'{model2.names[i2]}_Hn_2']): 
                            table2.drop(item, axis=0, inplace=True)
                            area_remove2.append(item)
                        elif height2 > int(values[f'{model2.names[i2]}_Hx_2']): 
                            table2.drop(item, axis=0, inplace=True)
                            area_remove2.append(item)
                        elif conf2 < int(values[f'{model2.names[i2]}_Conf_2']):
                            table2.drop(item, axis=0, inplace=True)
                            area_remove2.append(item)    

                if values[f'{model2.names[i2]}_2'] == False:
                    if label_name == model2.names[i2]:
                        table2.drop(item, axis=0, inplace=True)
                        area_remove2.append(item)

        names2 = list(table2['name'])

        show2 = np.squeeze(result2.render(area_remove2))
        show2 = cv2.resize(show2, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
        show2 = cv2.cvtColor(show2, cv2.COLOR_BGR2RGB)
        #ta = time.time()

        for i2 in range(len(model2.names)):
            register_ng = str(values[f'{model2.names[i2]}_PLC_2'])

            if values[f'{model2.names[i2]}_OK_2'] == True:
                len_name2 = 0
                for name2 in names2:
                    if name2 == model2.names[i2]:
                        len_name2 +=1
                if len_name2 != int(values[f'{model2.names[i2]}_Num_2']):
                    print('NG')
                    writedata('DM'+ register_ng +'.U',1) 
                    print('DM'+ register_ng +'.U')

                    t2 = time.time() - t1
                    print(t2) 
                    cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                    window['result_cam1'].update(value= 'NG', text_color='red')
                    
                    myresult2 = 1
       

            if values[f'{model2.names[i2]}_NG_2'] == True:
                if model2.names[i2] in names2:
                    print('NG')
                    writedata('DM'+ register_ng +'.U',1) 
                    print('DM'+ register_ng +'.U')


                    t2 = time.time() - t1
                    print(t2) 
                    cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                    window['result_cam1'].update(value= 'NG', text_color='red')    
                    
                    myresult2 = 1         
                  

        if myresult2 == 0:
            print('OK')
            writedata('DM'+ str(values['OK_PLC_2']) +'.U',1) 

            t2 = time.time() - t1
            print(t2) 
            cv2.putText(show2, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
            window['result_cam1'].update(value= 'OK', text_color='green')
            if values['have_save_OK_2']:
                # name_folder_ok = time_to_name()
                # cv2.imwrite(values['save_OK_2']  + '/' + name_folder_ok + '.jpg',img2_save)

                name_folder_ng = time_to_name()
                today = datetime.date.today()
                if not os.path.isdir(values['save_OK_2']  + "/" + str(today)):
                    os.mkdir(values['save_OK_2']  + "/" + str(today))
                cv2.imwrite(values['save_OK_2']  + "/" + str(today) + '/' + name_folder_ng + '.jpg',img2_save)

        else:
            if values['have_save_NG_2']:
                # name_folder_ng = time_to_name()
                # cv2.imwrite(values['save_NG_2']  + '/' + name_folder_ng + '.jpg',img2_save)
                # cv2.imwrite('D:/2' + '/' + name_folder_ng + '.jpg',img2_save)
                name_folder_ng = time_to_name()
                today = datetime.date.today()
                if not os.path.isdir(values['save_NG_2']  + "/" + str(today)):
                    os.mkdir(values['save_NG_2']  + "/" + str(today))

                cv2.imwrite(values['save_NG_2']  + "/" + str(today) + '/' + name_folder_ng + '.jpg',img2_save)



        writedata('DM7522.U',1)
        

        time_cam2 = str(int(t2*1000)) + 'ms'
        window['time_cam1'].update(value= time_cam2, text_color='black') 
        cv2.putText(show2, 'M2',(120,100),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)

        imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()
        window['image1'].update(data= imgbytes2)

        myindex1 +=1


        if myindex1 == 2:
            show2 = cv2.resize(show2, (int(image_width_display/2.2),int(image_height_display/1.4)), interpolation = cv2.INTER_AREA)
            cv2.putText(show2, '2',(30,70),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),1)
            
            imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()

            window['image1.2'].update(data= imgbytes2)

        if myindex1 == 4:
            show2 = cv2.resize(show2, (int(image_width_display/2.2),int(image_height_display/1.4)), interpolation = cv2.INTER_AREA)
            cv2.putText(show2, '4',(30,70),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),1)
            
            imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()

            window['image1.4'].update(data= imgbytes2)

        if myindex1 == 6:
            show2 = cv2.resize(show2, (int(image_width_display/2.2),int(image_height_display/1.4)), interpolation = cv2.INTER_AREA)
            cv2.putText(show2, '6',(30,70),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),1)
            
            imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()

            window['image1.6'].update(data= imgbytes2)  


        print('---------------------------------------------')




def program_camera3(model,size,conf):
    read_7102 = readdata('DM7102') 
    read_7100 = readdata('DM7100')
    global img1_orgin
    global myindex1

    #window['result_cam1'].update(value= '', text_color='red')    

    if read_7102 == 1 and read_7100  == 3: 

        writedata('DM7002.U',1)

        img1_orgin = my_callback1.image 
        img1_save = img1_orgin
        print('MODEL 3')

        writedata('DM7102.U',0) 
        writedata('DM7100.U',0)

        t1 = time.time()
        center_x = 0
        center_y = 0

        img1_orgin = cv2.cvtColor(img1_orgin, cv2.COLOR_BGR2RGB)

        result1 = model(img1_orgin,size= size,conf = conf) 
        table1 = result1.pandas().xyxy[0]
        area_remove1 = []
        myresult1 =0 
        writedata('DM7002.U',0)

        for item in range(len(table1.index)):

            width1 = table1['xmax'][item] - table1['xmin'][item]
            height1 = table1['ymax'][item] - table1['ymin'][item]
            conf1 = table1['confidence'][item] * 100
            if item == 0:
                center_x = int(float(table1['xmin'][item]) + float(table1['xmax'][item] - table1['xmin'][item])/2)
                center_y = int(float(table1['ymin'][item]) + float(table1['ymax'][item] - table1['ymin'][item])/2)
                print('x: ',center_x)
                print('y: ',center_y)

            #area1 = width1*height1
            label_name = table1['name'][item]
            for i3 in range(len(model3.names)):
                if values[f'{model3.names[i3]}_3'] == True:
                    #if values[f'{model1.names[i1]}_WH'] == True:
                    if label_name == model3.names[i3]:
                        if width1 < int(values[f'{model3.names[i3]}_Wn_3']): 
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)
                        elif width1 > int(values[f'{model3.names[i3]}_Wx_3']): 
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)
                        elif height1 < int(values[f'{model3.names[i3]}_Hn_3']): 
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)
                        elif height1 > int(values[f'{model3.names[i3]}_Hx_3']): 
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)
                        elif conf1 < int(values[f'{model3.names[i3]}_Conf_3']):
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)                                             
                if values[f'{model3.names[i3]}_3'] == False:
                    if label_name == model3.names[i3]:
                        table1.drop(item, axis=0, inplace=True)
                        area_remove1.append(item)

        names1 = list(table1['name'])

        show1 = np.squeeze(result1.render(area_remove1))
        show1 = cv2.resize(show1, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
        show1 = cv2.cvtColor(show1, cv2.COLOR_BGR2RGB)

        # #ta = time.time()
        # for i1 in range(len(model1.names)):
        #     register_ng = str(values[f'{model3.names[i3]}_PLC_3'])

        #     if values[f'{model3.names[i3]}_OK_3'] == True:
        #         len_name1 = 0
        #         for name1 in names1:
        #             if name1 == model3.names[i3]:
        #                 len_name1 +=1
        #         if len_name1 != int(values[f'{model3.names[i3]}_Num_3']):
        #             print('NG')
        #             writedata('DM'+ register_ng +'.U',1) 

        #             #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register_ng,b'\x00\x01',1)
        #             t2 = time.time() - t1
        #             print(t2) 
        #             cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
        #             window['result_cam1'].update(value= 'NG', text_color='red')
                    
        #             myresult1 = 1
           

        #     if values[f'{model3.names[i3]}_NG_3'] == True:
        #         if model3.names[i3] in names3:
        #             print('NG')
        #             writedata('DM'+ register_ng +'.U',1) 

        #             t2 = time.time() - t1
        #             print(t2) 
        #             cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
        #             window['result_cam1'].update(value= 'NG', text_color='red')    
                    
        #             myresult1 = 1         
        

        if myresult1 == 0:
            # print('OK')
            # writedata('DM'+ str(values['OK_PLC_3']) +'.U',1) 

            # #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,(3000).to_bytes(2, byteorder='big') + b'\x00',b'\x00\x01',1)
            # t2 = time.time() - t1
            # ##print(t2) 
            # cv2.putText(show1, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
            # window['result_cam1'].update(value= 'OK', text_color='green')

            if values['have_save_OK_1']:
                name_folder_ok = time_to_name()
                cv2.imwrite(values['save_OK_1']  + '/' + name_folder_ok + '.jpg',img1_save)
        else:
            if values['have_save_NG_1']:
                name_folder_ng = time_to_name()
                cv2.imwrite(values['save_NG_1']  + '/' + name_folder_ng + '.jpg',img1_save)

        writedata('DM7012.U',center_x) 

        writedata('DM7014.U',center_y) 



        #time_cam1 = str(int(t2*1000)) + 'ms'
        #window['time_cam1'].update(value= time_cam1, text_color='black') 
        cv2.putText(show1, ' M3',(100,100),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)

        imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
        window['image1'].update(data= imgbytes1)


        print('---------------------------------------------')




def program_camera1_FH(model,size,conf):
    read_4000 = fins_instance.memory_area_read(FinsPLCMemoryAreas().DATA_MEMORY_WORD,b'\x0F\xA0\x00') # doc thanh ghi 5000
    #print(read_4000)
    if read_4000 == b'\xc0\x00\x02\x00\x19\x00\x00\x01\x00`\x01\x01\x00\x00\x00\x05':  # gia tri 5
        directory1 = 'C:/FH/camera1/'
        if os.listdir(directory1) == []:
            print('folder 1 empty')
        else:
            print('received folder 1')

            for filename1 in glob.glob('C:/FH/camera1/*'):
                for path1 in glob.glob(filename1 + '/*'):
                    name = path1[-18:]
                    if name == 'Input0_Camera0.jpg':
                        img1_orgin = cv2.imread(path1)
                        while type(img1_orgin) == type(None):
                            print('loading img 1...')
                            for path1 in glob.glob(filename1 + '/*'):
                                img1_orgin = cv2.imread(path1)

                        img1_save = img1_orgin
                        #img1_orgin = cv2.resize(img1_orgin,(640,480))                   
                        print('CAM 1')
                        t1 = time.time()

                        # ghi vao D4000 gia tri 0
                        fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,b'\x0F\xA0\x00',b'\x00\x00',1)

                        img1_orgin = cv2.cvtColor(img1_orgin, cv2.COLOR_BGR2RGB)

                        result1 = model(img1_orgin,size= size,conf = conf) 
                        table1 = result1.pandas().xyxy[0]
                        area_remove1 = []

                        myresult1 =0 

                        for item in range(len(table1.index)):
                            width1 = table1['xmax'][item] - table1['xmin'][item]
                            height1 = table1['ymax'][item] - table1['ymin'][item]
                            conf1 = table1['confidence'][item] * 100

                            #area1 = width1*height1
                            label_name = table1['name'][item]
                            for i1 in range(len(model1.names)):
                                if values[f'{model1.names[i1]}_1'] == True:
                                    #if values[f'{model1.names[i1]}_WH'] == True:
                                    if label_name == model1.names[i1]:
                                        if width1 < int(values[f'{model1.names[i1]}_Wn_1']): 
                                            table1.drop(item, axis=0, inplace=True)
                                            area_remove1.append(item)
                                        elif width1 > int(values[f'{model1.names[i1]}_Wx_1']): 
                                            table1.drop(item, axis=0, inplace=True)
                                            area_remove1.append(item)
                                        elif height1 < int(values[f'{model1.names[i1]}_Hn_1']): 
                                            table1.drop(item, axis=0, inplace=True)
                                            area_remove1.append(item)
                                        elif height1 > int(values[f'{model1.names[i1]}_Hx_1']): 
                                            table1.drop(item, axis=0, inplace=True)
                                            area_remove1.append(item)
                                        elif conf1 < int(values[f'{model1.names[i1]}_Conf_1']):
                                            table1.drop(item, axis=0, inplace=True)
                                            area_remove1.append(item)                                             
                                if values[f'{model1.names[i1]}_1'] == False:
                                    if label_name == model1.names[i1]:
                                        table1.drop(item, axis=0, inplace=True)
                                        area_remove1.append(item)

                        names1 = list(table1['name'])

                        show1 = np.squeeze(result1.render(area_remove1))
                        show1 = cv2.resize(show1, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
                        show1 = cv2.cvtColor(show1, cv2.COLOR_BGR2RGB)

                        #ta = time.time()
                        for i1 in range(len(model1.names)):
                            register_ng = (int(values[f'{model1.names[i1]}_PLC_1'])).to_bytes(2, byteorder='big') + b'\x00'
                            if values[f'{model1.names[i1]}_OK_1'] == True:
                                len_name1 = 0
                                for name1 in names1:
                                    if name1 == model1.names[i1]:
                                        len_name1 +=1
                                if len_name1 != int(values[f'{model1.names[i1]}_Num_1']):
                                    print('NG')
                                    fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register_ng,b'\x00\x01',1)
                                    #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register_ng,b'\x00\x01',1)
                                    t2 = time.time() - t1
                                    print(t2) 
                                    cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                    window['result_cam1'].update(value= 'NG', text_color='red')
                                    if values['have_save_NG_1']:
                                        name_folder_ng = time_to_name()
                                        cv2.imwrite(values['save_NG_1']  + '/' + name_folder_ng + '.jpg',img1_save)
                                    myresult1 = 1
                                    break

                            if values[f'{model1.names[i1]}_NG_1'] == True:
                                if model1.names[i1] in names1:
                                    print('NG')
                                    fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register_ng,b'\x00\x01',1)
                                    t2 = time.time() - t1
                                    print(t2) 
                                    cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                    window['result_cam1'].update(value= 'NG', text_color='red')    
                                    if values['have_save_NG_1']:
                                        name_folder_ng = time_to_name()
                                        cv2.imwrite(values['save_NG_1']  + '/' + name_folder_ng + '.jpg',img1_save)
                                    myresult1 = 1         
                                    break    

                        if myresult1 == 0:
                            print('OK')
                            fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,(int(values['OK_PLC_1'])).to_bytes(2, byteorder='big') + b'\x00',b'\x00\x01',1)
                            #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,(3000).to_bytes(2, byteorder='big') + b'\x00',b'\x00\x01',1)
                            t2 = time.time() - t1
                            print(t2) 
                            cv2.putText(show1, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                            window['result_cam1'].update(value= 'OK', text_color='green')
                            if values['have_save_OK_1']:
                                name_folder_ng = time_to_name()
                                cv2.imwrite(values['save_OK_1']  + '/' + name_folder_ng + '.jpg',img1_save)


                        time_cam1 = str(int(t2*1000)) + 'ms'
                        window['time_cam1'].update(value= time_cam1, text_color='black') 
                    

                        imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
                        window['image1'].update(data= imgbytes1)
                        print('---------------------------------------------')

                    if os.path.isfile(path1):
                        os.remove(path1)
                while os.path.isdir(filename1):
                    try:
                        shutil.rmtree(filename1)
                    except:
                        print('Error delete folder 1')


def program_camera2_FH(model,size,conf):
    read_4002 = fins_instance.memory_area_read(FinsPLCMemoryAreas().DATA_MEMORY_WORD,b'\x0F\xA2\x00') # doc thanh ghi 4002
    if read_4002 ==b'\xc0\x00\x02\x00\x19\x00\x00\x01\x00`\x01\x01\x00\x00\x00\x05':  # gia tri 5
        directory2 = 'C:/FH/camera2/'
        if os.listdir(directory2) == []:
            print('folder 2 empty')
        else:
            print('received folder 2')

            for filename2 in glob.glob('C:/FH/camera2/*'):
                for path2 in glob.glob(filename2 + '/*'):
                    name = path2[-18:]
                    if name == 'Input0_Camera0.jpg':
                        img2_orgin = cv2.imread(path2)
                        while type(img2_orgin) == type(None):
                            print('loading img 2...')
                            for path2 in glob.glob(filename2 + '/*'):
                                img2_orgin = cv2.imread(path2)

                        #img2_orgin = cv2.resize(img2_orgin,(640,480))
                        img2_save = img2_orgin
                        #edit
                        print('CAM 2')
                        t1 = time.time()

                        # ghi vao D4002 gia tri 0
                        fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,b'\x0F\xA2\x00',b'\x00\x00',1)


                        img2_orgin = cv2.cvtColor(img2_orgin, cv2.COLOR_BGR2RGB)

                        result2 = model(img2_orgin,size= size,conf = conf) 
                        table2 = result2.pandas().xyxy[0]
                        area_remove2 = []

                        myresult2 =0 

                        for item in range(len(table2.index)):
                            width2 = table2['xmax'][item] - table2['xmin'][item]
                            height2 = table2['ymax'][item] - table2['ymin'][item]
                            conf2 = table2['confidence'][item] * 100

                            label_name = table2['name'][item]
                            for i2 in range(len(model2.names)):
                                if values[f'{model2.names[i2]}_2'] == True:
                                    if label_name == model2.names[i2]:
                                        if width2 < int(values[f'{model2.names[i2]}_Wn_2']): 
                                            table2.drop(item, axis=0, inplace=True)
                                            area_remove2.append(item)
                                        elif width2 > int(values[f'{model2.names[i2]}_Wx_2']): 
                                            table2.drop(item, axis=0, inplace=True)
                                            area_remove2.append(item)
                                        elif height2 < int(values[f'{model2.names[i2]}_Hn_2']): 
                                            table2.drop(item, axis=0, inplace=True)
                                            area_remove2.append(item)
                                        elif height2 > int(values[f'{model2.names[i2]}_Hx_2']): 
                                            table2.drop(item, axis=0, inplace=True)
                                            area_remove2.append(item)
                                        elif conf2 < int(values[f'{model2.names[i2]}_Conf_2']):
                                            table2.drop(item, axis=0, inplace=True)
                                            area_remove2.append(item)    

                                if values[f'{model2.names[i2]}_2'] == False:
                                    if label_name == model2.names[i2]:
                                        table2.drop(item, axis=0, inplace=True)
                                        area_remove2.append(item)

                        names2 = list(table2['name'])

                        show2 = np.squeeze(result2.render(area_remove2))
                        show2 = cv2.resize(show2, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
                        show2 = cv2.cvtColor(show2, cv2.COLOR_BGR2RGB)
                        #ta = time.time()
                        for i2 in range(len(model2.names)):
                            register_ng = (int(values[f'{model2.names[i2]}_PLC_2'])).to_bytes(2, byteorder='big') + b'\x00'
                            if values[f'{model2.names[i2]}_OK_2'] == True:
                                len_name2 = 0
                                for name2 in names2:
                                    if name2 == model2.names[i2]:
                                        len_name2 +=1
                                if len_name2 != int(values[f'{model2.names[i2]}_Num_2']):
                                    print('NG')
                                    fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register_ng,b'\x00\x01',1)
                                    t2 = time.time() - t1
                                    print(t2) 
                                    cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                    window['result_cam2'].update(value= 'NG', text_color='red')
                                    if values['have_save_NG_2']:
                                        name_folder_ng = time_to_name()
                                        cv2.imwrite(values['save_NG_2']  + '/' + name_folder_ng + '.jpg',img2_save)
                                    myresult2 = 1
                                    break

                            if values[f'{model2.names[i2]}_NG_2'] == True:
                                if model2.names[i2] in names2:
                                    print('NG')
                                    fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register_ng,b'\x00\x01',1)
                                    t2 = time.time() - t1
                                    print(t2) 
                                    cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                    window['result_cam2'].update(value= 'NG', text_color='red')    
                                    if values['have_save_NG_2']:
                                        name_folder_ng = time_to_name()
                                        cv2.imwrite(values['save_NG_2']  + '/' + name_folder_ng + '.jpg',img2_save)
                                    myresult2 = 1         
                                    break    

                        if myresult2 == 0:
                            print('OK')
                            fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,(int(values['OK_PLC_2'])).to_bytes(2, byteorder='big') + b'\x00',b'\x00\x01',1)
                            t2 = time.time() - t1
                            print(t2) 
                            cv2.putText(show2, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                            window['result_cam2'].update(value= 'OK', text_color='green')
                            if values['have_save_NG_2']:
                                name_folder_ng = time_to_name()
                                cv2.imwrite(values['save_NG_2']  + '/' + name_folder_ng + '.jpg',img2_save)


                        time_cam2 = str(int(t2*1000)) + 'ms'
                        window['time_cam2'].update(value= time_cam2, text_color='black') 
                    

                        imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()
                        window['image2'].update(data= imgbytes2)
                        print('---------------------------------------------')
                    if os.path.isfile(path2):
                        os.remove(path2)
                while os.path.isdir(filename2):
                    try:
                        shutil.rmtree(filename2)
                    except:
                        print('Error delete folder 2')



def make_window(theme):
    sg.theme(theme)

    #file_img = [("JPEG (*.jpg)",("*jpg","*.png"))]

    file_weights = [('Weights (*.pt)', ('*.pt'))]

    # menu = [['Application', ['Connect PLC','Interrupt Connect PLC','Exit']],
    #         ['Tool', ['Check Cam','Change Theme']],
    #         ['Help',['About']]]

    right_click_menu = [[], ['Exit','Administrator','Change Theme']]


    layout_main = [

        # [
        # sg.Text('CAM 1',justification='center' ,font= ('Helvetica',30),text_color='red',expand_x=True),

        # ],

        [

        #1
        sg.Frame('',[
            [sg.Image(filename='', size=(image_width_display,image_height_display),key='image1',background_color='black'),

            sg.Frame('',[
                
                [sg.Button('Webcam', size=(10,1),  font=('Helvetica',14),disabled=True ,key= 'Webcam1'),sg.Text(' '), sg.Button('Snap', size=(10,1), font=('Helvetica',14),disabled=True ,key= 'Snap1')],
                [sg.Text('')],
                [sg.Button('Stop', size=(10,1), font=('Helvetica',14),disabled=True ,key= 'Stop1'), sg.Text(' '),sg.Button('Pic', size=(10,1), font=('Helvetica',14),disabled=True,key= 'Pic1')],
                [sg.Text('')],
                [sg.Button('Detect 1', size=(10,1), font=('Helvetica',14),disabled=True ,key= 'Detect1'), sg.Text(' '),sg.Button('Detect 2', size=(10,1), font=('Helvetica',14),disabled=True,key= 'Detect2')],
                [sg.Text('')],
                [sg.Button('Back', size=(10,1), font=('Helvetica',14) ,key= 'back',enable_events=True,disabled=True ), sg.Text(' '),sg.Button('Next', size=(10,1), font=('Helvetica',14),key= 'next',disabled=True )], 
                [sg.Text('')],
                [sg.Checkbox('Check',size=(2,1),font=('Helvetica',14), key='check_model1',enable_events=True,expand_x=True, expand_y=True),sg.Input('0',size=(8,1),font=('Helvetica',15),key= 'Num_index',text_color='navy',enable_events=True)],
                [sg.Text('')],
                [sg.Text(' '), sg.Combo(values=['1','2','3','4','5','6','7','8','9'], default_value='1',font=('Helvetica',20),size=(5, 100),text_color='navy',enable_events= True, key='choose_model')],
                [sg.Text('',font=('Helvetica',90), justification='center', key='result_cam1',expand_x=True)],
                [sg.Text('',font=('Helvetica',35), justification='center', key='time_cam1', expand_x=True)],
                ],element_justification='center', vertical_alignment='top', relief= sg.RELIEF_FLAT),
            ],           
        ]),
        ],
        # [
        # sg.Text('',justification='center' ,font= ('Helvetica',15),text_color='pink',expand_x=True,key='name_file'),
        # ],
        # [sg.Image(filename='', size=(int(image_width_display/2.2),int(image_height_display/1.4)),key='image1.1',background_color='black'),

        #     sg.Image(filename='', size=(int(image_width_display/2.2),int(image_height_display/1.4)),key='image1.2',background_color='black'),

        #     sg.Image(filename='', size=(int(image_width_display/2.2),int(image_height_display/1.4)),key='image1.3',background_color='black'),

        #     sg.Image(filename='', size=(int(image_width_display/2.2),int(image_height_display/1.4)),key='image1.4',background_color='black'),

        #     sg.Image(filename='', size=(int(image_width_display/2.2),int(image_height_display/1.4)),key='image1.5',background_color='black'),

        #     sg.Image(filename='', size=(int(image_width_display/2.2),int(image_height_display/1.4)),key='image1.6',background_color='black'),],
       
       
       
        [sg.Text('Index ',justification='center' ,font= ('Helvetica',15),text_color='navy'),sg.Input('0',size=(8,1),font=('Helvetica',15),key= 'Index_image',text_color='navy',enable_events=True,readonly = True),]
 
    
    ] 

    layout_image = [
            [sg.Image(filename='', size=(int(image_width_display/2.2),int(image_height_display/1.4)),key='image1.1',background_color='black'),

            sg.Image(filename='', size=(int(image_width_display/2.2),int(image_height_display/1.4)),key='image1.2',background_color='black'),

            sg.Image(filename='', size=(int(image_width_display/2.2),int(image_height_display/1.4)),key='image1.3',background_color='black')],

            [sg.Image(filename='', size=(int(image_width_display/2.2),int(image_height_display/1.4)),key='image1.4',background_color='black'),

            sg.Image(filename='', size=(int(image_width_display/2.2),int(image_height_display/1.4)),key='image1.5',background_color='black'),

            sg.Image(filename='', size=(int(image_width_display/2.2),int(image_height_display/1.4)),key='image1.6',background_color='black'),],
            # [sg.Text('Index ',justification='center' ,font= ('Helvetica',15),text_color='navy'),sg.Input('0',size=(8,1),font=('Helvetica',15),key= 'Index_image',text_color='navy',enable_events=True,readonly = True),
            # sg.Text('Signal Model(7100) ',justification='center' ,font= ('Helvetica',15),text_color='navy'),sg.Input('0',size=(8,1),font=('Helvetica',15),key= 'signal_model',text_color='navy',enable_events=True,readonly = True),
            # sg.Text('Signal Sub(7102) ',justification='center' ,font= ('Helvetica',15),text_color='navy'),sg.Input('0',size=(8,1),font=('Helvetica',15),key= 'signal_sub',text_color='navy',enable_events=True,readonly = True),]


    ]
    # layout_main = [

    #     [
    #     sg.Text('CAM 1',justification='center' ,font= ('Helvetica',30),text_color='red', expand_y=True),
    #     sg.Text('CAM 2',justification='center' ,font= ('Helvetica',30),text_color='red',expand_x=True),
    #     ],
    #     # sg.Frame('',[
    #     #     [sg.Text('CAM 2',justification='center' ,font= ('Helvetica',30),text_color='red'),
    #     #     sg.Text('CAM 1',justification='center' ,font= ('Helvetica',30),text_color='red')],
    #     # ]),

    #     [
    #     #1
    #     sg.Frame('',[
    #         #[sg.Image(filename='', size=(640,480),key='image1',background_color='black')],
    #         [sg.Image(filename='', size=(image_width_display,image_height_display),key='image1',background_color='black')],
    #         [sg.Frame('',
    #         [
    #             [sg.Text('',font=('Helvetica',120), justification='center', key='result_cam1',expand_x=True)],
    #             [sg.Text('',font=('Helvetica',30), justification='center', key='time_cam1', expand_x=True)],
    #         ], vertical_alignment='top',size=(int(560*1.2),int(250*1.2))),
    #         sg.Frame('',[
    #             #[sg.Text('')],
    #             [sg.Button('Webcam', size=(8,1),  font=('Helvetica',14),disabled=True ,key= 'Webcam1')],
    #             [sg.Text('')],
    #             [sg.Button('Stop', size=(8,1), font=('Helvetica',14),disabled=True ,key= 'Stop1')],
    #             [sg.Text('')],
    #             #[sg.Button('Continue', size=(8,1),  font=('Helvetica',14), disabled=True, key= 'Continue1')],
    #             #[sg.Text('')],
    #             [sg.Button('Snap', size=(8,1), font=('Helvetica',14),disabled=True ,key= 'Snap1')],
    #             [sg.Text('')],
    #             [sg.Checkbox('Check',size=(6,1),font=('Helvetica',14), key='check_model1',enable_events=True,expand_x=True, expand_y=True)],
    #             #],element_justification='center', vertical_alignment='top', relief= sg.RELIEF_FLAT),
    #             ],element_justification='center', vertical_alignment='top', relief= sg.RELIEF_FLAT),
                
    #         sg.Frame('',[   
    #             [sg.Button('Change', size=(8,1), font=('Helvetica',14), disabled= True, key= 'Change1')],
    #             [sg.Text('')],
    #             [sg.Button('Pic', size=(8,1), font=('Helvetica',14),disabled=True,key= 'Pic1')],
    #             [sg.Text('')],
    #             [sg.Button('Detect', size=(8,1), font=('Helvetica',14),disabled=True,key= 'Detect1')],
    #             #[sg.Text('',size=(4,2))],
    #             [sg.Text('',size=(4,1))],
    #             [sg.Combo(values=['A22','A19','3','4','5','6','7','8','9'], default_value='A22',font=('Helvetica',20),size=(5, 100),text_color='navy',enable_events= True, key='choose_model'),],
    #             #[sg.Button('SaveData', size=(8,1), font=('Helvetica',14),disabled=True,key= 'SaveData1')],

    #             ],element_justification='center', vertical_alignment='top', relief= sg.RELIEF_FLAT),
    #         ],
                
    #     ]),
    
    #     # 2
    #     sg.Frame('',[
    #         #[sg.Image(filename='', size=(640,480),key='image1',background_color='black')],
    #         [sg.Image(filename='', size=(image_width_display,image_height_display),key='image2',background_color='black')],
    #         [sg.Frame('',
    #         [
    #             [sg.Text('',font=('Helvetica',120), justification='center', key='result_cam2',expand_x=True)],
    #             [sg.Text('',font=('Helvetica',30),justification='center', key='time_cam2',expand_x=True)],
    #         ], vertical_alignment='top',size=(int(560*1.2),int(250*1.2))),
    #         sg.Frame('',[
    #             #[sg.Text('')],
    #             [sg.Button('Webcam', size=(8,1),  font=('Helvetica',14),disabled=True,key= 'Webcam2',auto_size_button=True)],
    #             [sg.Text('')],
    #             [sg.Button('Stop', size=(8,1), font=('Helvetica',14),disabled=True  ,key='Stop2')],
    #             [sg.Text('')],
    #             #[sg.Button('Continue', size=(8,1),  font=('Helvetica',14),disabled=True ,key='Continue2')],
    #             #[sg.Text('')],
    #             [sg.Button('Snap', size=(8,1), font=('Helvetica',14),disabled=True  ,key='Snap2')],
    #             [sg.Text('')],
    #             [sg.Checkbox('Check',size=(6,1),font=('Helvetica',14), key='check_model2',enable_events=True,expand_x=True, expand_y=True)],
    #             ],element_justification='center', vertical_alignment='top', relief= sg.RELIEF_FLAT),

    #         sg.Frame('',[   
    #             [sg.Button('Change', size=(8,1), font=('Helvetica',14), disabled= True, key= 'Change2')],
    #             [sg.Text('')],
    #             [sg.Button('Pic', size=(8,1), font=('Helvetica',14),disabled=True,key='Pic2')],
    #             [sg.Text('')],
    #             [sg.Button('Detect', size=(8,1), font=('Helvetica',14),disabled=True ,key= 'Detect2')],
    #             [sg.Text('',size=(4,2))],
    #             [sg.Text('',size=(4,1))],
    #             #[sg.Button('SaveData', size=(8,1), font=('Helvetica',14),disabled=True ,key= 'SaveData2')],

    #             ],element_justification='center', vertical_alignment='top', relief= sg.RELIEF_FLAT),
    #         ]
    #     ], expand_y= True),

    # ]] 



    my_option1 = [
        [sg.Frame('',[
        [sg.Frame('',
        [   
            #[sg.Text('Location', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Input(size=(60,1), font=('Helvetica',12), key='location_weights1',readonly= True, text_color='navy',enable_events= True),
            #sg.FolderBrowse(size=(15,1), font=('Helvetica',10),key= 'folder_browse1',enable_events=True)],
            [sg.Text('Weights', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Input(size=(60,1), font=('Helvetica',12), key='file_weights1',readonly= True, text_color='navy',enable_events= True),
            #[sg.Text('Weights', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Combo(values='', font=('Helvetica',12),size=(59, 30),text_color='navy',enable_events= True, key='file_weights1'),],
            sg.Frame('',[
                [sg.FileBrowse(file_types= file_weights, size=(12,1), font=('Helvetica',10),key= 'file_browse1',enable_events=True, disabled=True)]
            ], relief= sg.RELIEF_FLAT),
            sg.Frame('',[
                [sg.Button('Change Model', size=(14,1), font=('Helvetica',10), disabled= True, key= 'Change_1')]
            ], relief= sg.RELIEF_FLAT),],
            [sg.Text('Confidence',size=(12,1),font=('Helvetica',15), text_color='red'), sg.Slider(range=(1,100),orientation='h',size=(60,20),font=('Helvetica',11),disabled=True, key= 'conf_thres1'),]
        ], relief=sg.RELIEF_FLAT),
        ],
        [sg.Frame('',[
            [sg.Text('Name',size=(15,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Join',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('OK',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Num',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('NG',size=(8,1),font=('Helvetica',15), text_color='red'),  
            sg.Text('Width Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Width Max',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Max',size=(12,1),font=('Helvetica',15), text_color='red'),
            sg.Text('PLC',size=(11,1),font=('Helvetica',15), text_color='red'),
            sg.Text('Confidence',size=(11,1),font=('Helvetica',15), text_color='red')],
        ], relief=sg.RELIEF_FLAT)],
        [sg.Frame('',[
            [
                sg.Text(f'{model1.names[i1]}_1',size=(15,1),font=('Helvetica',15), text_color='yellow'), 
                sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key=f'{model1.names[i1]}_1',enable_events=True, disabled=True), 
                #sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model1.names[i1]}_OK_1',enable_events=True, disabled=True), 
                sg.Radio('',group_id=f'group_1_{i1}',size=(5,5),font=('Helvetica',15),  key=f'{model1.names[i1]}_OK_1',enable_events=True, disabled=True), 
                
                sg.Input('1',size=(2,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Num_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(4,1),font=('Helvetica',15), text_color='red'), 
                #sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model1.names[i1]}_NG_1',enable_events=True, disabled=True), 
                sg.Radio('',group_id=f'group_1_{i1}',size=(5,5),font=('Helvetica',15),  key=f'{model1.names[i1]}_NG_1',enable_events=True, disabled=True), 
                
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Wn_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Wx_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Hn_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Hx_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model1.names[i1]}_PLC_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Slider(range=(1,100),default_value=25,orientation='h',size=(30,20),font=('Helvetica',11), key= f'{model1.names[i1]}_Conf_1'),                  
            ] for i1 in range(len(model1.names))
        ], relief=sg.RELIEF_FLAT)],
        [sg.Text('  OK',size=(15,1),font=('Helvetica',15), text_color='yellow'),
        sg.Text(' '*230), 
        sg.Input('0',size=(8,1),font=('Helvetica',15),key= 'OK_PLC_1',text_color='navy',enable_events=True)],
        [sg.Text(' ')],
        [sg.Text(' '*250), sg.Button('Save Data', size=(12,1),  font=('Helvetica',12),key='SaveData1',enable_events=True)] 
        ])]
    ]
    
    

    my_option2 = [
        [sg.Frame('',[
        [sg.Frame('',[
            [sg.Text('Weights', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Input(size=(60,1), font=('Helvetica',12), key='file_weights2',readonly= True, text_color='navy',enable_events= True),
            sg.Frame('',[
                [sg.FileBrowse(file_types= file_weights, size=(12,1), font=('Helvetica',10),key= 'file_browse2',enable_events=True, disabled=True)]
            ], relief= sg.RELIEF_FLAT),
            sg.Frame('',[
                [sg.Button('Change Model', size=(14,1), font=('Helvetica',10), disabled= True, key= 'Change_2')]
            ], relief= sg.RELIEF_FLAT),],
            [sg.Text('Confidence',size=(12,1),font=('Helvetica',15), text_color='red'), sg.Slider(range=(1,100),orientation='h',size=(60,20),font=('Helvetica',11),disabled=True, key= 'conf_thres2')],

        ], relief=sg.RELIEF_FLAT, expand_y= True),],
        [sg.Frame('',[
            [sg.Text('Name',size=(15,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Join',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('OK',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Num',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('NG',size=(8,1),font=('Helvetica',15), text_color='red'),  
            sg.Text('Width Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Width Max',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Max',size=(12,1),font=('Helvetica',15), text_color='red'),
            sg.Text('PLC',size=(11,1),font=('Helvetica',15), text_color='red'),
            sg.Text('Confidence',size=(11,1),font=('Helvetica',15), text_color='red')],
        ], relief=sg.RELIEF_FLAT)],
        [sg.Frame('',[
            [
                sg.Text(f'{model2.names[i2]}_2',size=(15,1),font=('Helvetica',15), text_color='yellow'), 
                sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key=f'{model2.names[i2]}_2',enable_events=True, disabled=True), 
                #sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model2.names[i2]}_OK_2',enable_events=True, disabled=True), 
                sg.Radio('', group_id=f'group_2_{i2}',size=(5,5),font=('Helvetica',15),  key=f'{model2.names[i2]}_OK_2',enable_events=True, disabled=True), 
                
                sg.Input('1',size=(2,1),font=('Helvetica',15),key= f'{model2.names[i2]}_Num_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(4,1),font=('Helvetica',15), text_color='red'), 
                #sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model2.names[i2]}_NG_2',enable_events=True, disabled=True), 
                sg.Radio('',group_id=f'group_2_{i2}',size=(5,5),font=('Helvetica',15),  key=f'{model2.names[i2]}_NG_2',enable_events=True, disabled=True), 
                
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model2.names[i2]}_Wn_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model2.names[i2]}_Wx_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model2.names[i2]}_Hn_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model2.names[i2]}_Hx_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model2.names[i2]}_PLC_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Slider(range=(1,100),default_value=25,orientation='h',size=(30,20),font=('Helvetica',11), key= f'{model2.names[i2]}_Conf_2'),                      
            ] for i2 in range(len(model2.names))
        ], relief=sg.RELIEF_FLAT)],
        [sg.Text('  OK',size=(15,1),font=('Helvetica',15), text_color='yellow'),
        sg.Text(' '*230), 
        sg.Input('0',size=(8,1),font=('Helvetica',15),key= 'OK_PLC_2',text_color='navy',enable_events=True)],
        [sg.Text(' ')],
        [sg.Text(' '*250), sg.Button('Save Data', size=(12,1),  font=('Helvetica',12),key='SaveData2',enable_events=True)] 
        ])]
    ]


    my_option3 = [
        [sg.Frame('',[
        [sg.Frame('',
        [   
            #[sg.Text('Location', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Input(size=(60,1), font=('Helvetica',12), key='location_weights4',readonly= True, text_color='navy',enable_events= True),
            #sg.FolderBrowse(size=(15,1), font=('Helvetica',10),key= 'folder_browse3',enable_events=True)],
            [sg.Text('Weights', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Input(size=(60,1), font=('Helvetica',12), key='file_weights3',readonly= True, text_color='navy',enable_events= True),
            #[sg.Text('Weights', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Combo(values='', font=('Helvetica',12),size=(59, 30),text_color='navy',enable_events= True, key='file_weights3'),],
            sg.Frame('',[
                [sg.FileBrowse(file_types= file_weights, size=(12,1), font=('Helvetica',10),key= 'file_browse3',enable_events=True, disabled=True)]
            ], relief= sg.RELIEF_FLAT),
            sg.Frame('',[
                [sg.Button('Change Model', size=(14,1), font=('Helvetica',10), disabled= True, key= 'Change_3')]
            ], relief= sg.RELIEF_FLAT),],
            [sg.Text('Confidence',size=(12,1),font=('Helvetica',15), text_color='red'), sg.Slider(range=(1,100),orientation='h',size=(60,20),font=('Helvetica',11),disabled=True, key= 'conf_thres3'),]
        ], relief=sg.RELIEF_FLAT),
        ],
        [sg.Frame('',[
            [sg.Text('Name',size=(15,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Join',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('OK',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Num',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('NG',size=(8,1),font=('Helvetica',15), text_color='red'),  
            sg.Text('Width Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Width Max',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Max',size=(12,1),font=('Helvetica',15), text_color='red'),
            sg.Text('PLC',size=(11,1),font=('Helvetica',15), text_color='red'),
            sg.Text('Confidence',size=(11,1),font=('Helvetica',15), text_color='red')],
        ], relief=sg.RELIEF_FLAT)],
        [sg.Frame('',[
            [
                sg.Text(f'{model3.names[i3]}_3',size=(15,1),font=('Helvetica',15), text_color='yellow'), 
                sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key=f'{model3.names[i3]}_3',enable_events=True, disabled=True), 
                #sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model3.names[i3]}_OK_3',enable_events=True, disabled=True), 
                sg.Radio('', group_id=f'group_3_{i3}',size=(5,5),font=('Helvetica',15),  key=f'{model3.names[i3]}_OK_3',enable_events=True, disabled=True), 
                
                sg.Input('1',size=(2,1),font=('Helvetica',15),key= f'{model3.names[i3]}_Num_3',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(4,1),font=('Helvetica',15), text_color='red'), 
                #sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model3.names[i3]}_NG_3',enable_events=True, disabled=True), 
                sg.Radio('',group_id=f'group_3_{i3}',size=(5,5),font=('Helvetica',15),  key=f'{model3.names[i3]}_NG_3',enable_events=True, disabled=True), 
                
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model3.names[i3]}_Wn_3',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model3.names[i3]}_Wx_3',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model3.names[i3]}_Hn_3',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model3.names[i3]}_Hx_3',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model3.names[i3]}_PLC_3',text_color='navy',enable_events=True, disabled=False), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Slider(range=(1,100),default_value=25,orientation='h',size=(30,20),font=('Helvetica',11), key= f'{model3.names[i3]}_Conf_3'),            
            ] for i3 in range(len(model3.names))
        ], relief=sg.RELIEF_FLAT)],
        [sg.Text('  OK',size=(15,1),font=('Helvetica',15), text_color='yellow'),
        sg.Text(' '*230), 
        sg.Input('0',size=(8,1),font=('Helvetica',15),key= 'OK_PLC_3',text_color='navy',enable_events=True)],        
        [sg.Text(' ')],
        [sg.Text(' '*250), sg.Button('Save Data', size=(12,1),  font=('Helvetica',12),key='SaveData3',enable_events=True)] 
        ])]
    ]

    
    layout_option1 = [[sg.Column(my_option1,size = (1800,900),scrollable = True)]]
    layout_option2 = [[sg.Column(my_option2,size = (1800,900),scrollable = True)]]
    layout_option3 = [[sg.Column(my_option3,size = (1800,900),scrollable = True)]]



    layout_saveimg = [
        [sg.Frame('',[
                [sg.Text('Have save folder image OK for camera 1',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=False,font=('Helvetica',15),  key='have_save_OK_1',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image OK for camera 1', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='C:/Cam1/OK' ,font=('Helvetica',12), key='save_OK_1',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_OK_1',enable_events=True) ],
                [sg.Text('')],
                [sg.Text('Have save folder image OK for camera 2',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=False,font=('Helvetica',15),  key='have_save_OK_2',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image OK for camera 2', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='C:/Cam2/OK' , font=('Helvetica',12), key='save_OK_2',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_OK_2',enable_events=True) ],
                [sg.Text('')],
                [sg.Text('Have save folder image NG for camera 1',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key='have_save_NG_1',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image NG for camera 1', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='C:/Cam1/NG' , font=('Helvetica',12), key='save_NG_1',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_NG_1',enable_events=True) ],
                [sg.Text('')],
                [sg.Text('Have save folder image NG for camera 2',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key='have_save_NG_2',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image NG for camera 2', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='C:/Cam2/NG' , font=('Helvetica',12), key='save_NG_2',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_NG_2',enable_events=True) ],
        ], relief=sg.RELIEF_FLAT),
            ],
        ]

    layout_terminal = [[sg.Text("Anything printed will display here!")],
                      [sg.Multiline( font=('Helvetica',14), write_only=True, autoscroll=True, auto_refresh=True,reroute_stdout=True, reroute_stderr=True, echo_stdout_stderr=True,expand_x=True,expand_y=True)]
                      ]
    
    layout = [[sg.TabGroup([[  sg.Tab('Main', layout_main),
                               sg.Tab('Image', layout_image),
                               sg.Tab('Option for model 1', layout_option1),
                               sg.Tab('Option for model 2', layout_option2),
                               sg.Tab('Option for model kiem', layout_option3),
                               
                               sg.Tab('Save Image', layout_saveimg),
                               sg.Tab('Output', layout_terminal)]])
               ]]

    #layout[-1].append(sg.Sizegrip())
    window = sg.Window('HuynhLeVu', layout, location=(0,0),right_click_menu=right_click_menu,resizable=True).Finalize()
    #window.bind('<Configure>',"Configure")
    window.Maximize()

    return window


# image_width_display = 650
# image_height_display = 410

# result_width_display = 470
# result_height_display = 100

image_width_display = int(760*1.2)
image_height_display = int(480*1.2)

result_width_display = 700
result_height_display = 100


file_name_img = [("Img(*.jpg,*.png)",("*jpg","*.png"))]


recording1 = False

error_cam1 = True
a = 0

#window['result_cam1'].update(value= 'Wait', text_color='yellow')
#window['result_cam2'].update(value= 'Wait', text_color='yellow')

list_path = []
for path1 in glob.glob('C:/Check1/*'):
    list_path.append(path1)


connected = False
while connected == False:
    connected = socket_connect('192.168.0.10',8501)
print("connected") 



mypath1 = load_model(1)
model1 = torch.hub.load('./levu','custom', path= mypath1, source='local',force_reload =False)

img1_test = os.path.join(os.getcwd(), 'img/imgtest.jpg')
result1 = model1(img1_test,416,0.25) 
print('model1 already')


mypath2 = load_model(2)
model2 = torch.hub.load('./levu','custom', path= mypath2, source='local',force_reload =False)

img2_test = os.path.join(os.getcwd(), 'img/imgtest.jpg')
result2 = model2(img2_test,416,0.25) 

print('model2 already')


mypath3 = load_model(3)
model3 = torch.hub.load('./levu','custom', path= mypath3, source='local',force_reload =False)

img1_test = os.path.join(os.getcwd(), 'img/imgtest.jpg')
result3 = model3(img1_test,416,0.25) 
print('model3 already')


choose_model = load_choosemodel()

themes = load_theme()
theme = themes[0]
window = make_window(theme)

window['choose_model'].update(value=choose_model)


try:
    load_all_sql(1,choose_model)
except:
    print(traceback.format_exc())
    window['time_cam1'].update(value= "Error data") 
False

try:
    load_all_sql(2,choose_model)
except:
    print(traceback.format_exc())
    window['time_cam2'].update(value= "Error data") 

try:
    load_all_sql(3,choose_model)
except:
    print(traceback.format_exc())

connect_camera1 = False
connect_camera2 = False

index_path = 0
index_show = 1

myindex1 = 0

connect_total = False

#removefile()

try:
    my_callback1 = CMyCallback()
    cb_func1 = my_callback1.datastream_callback1
    connect_camera1 = True

except Exception as exception:
    print('Error 1: ',exception)
    window['result_cam1'].update(value= 'Error', text_color='red')

# try:
#     my_callback2 = CMyCallback()
#     cb_func2 = my_callback2.datastream_callback2
#     connect_camera2 = True

# except Exception as exception:
#     print('Error 2: ',exception)
#     window['result_cam2'].update(value= 'Error', text_color='red')


# try:
#     my_callback3 = CMyCallback()
#     cb_func3 = my_callback3.datastream_callback3
#     connect_camera3 = True

# except Exception as exception:
#     print('Error 3: ',exception)
#     window['result_cam3'].update(value= 'Error', text_color='red')

# try:
#     my_callback4 = CMyCallback()
#     cb_func4 = my_callback4.datastream_callback4
#     connect_camera4 = True

# except Exception as exception:
#     print('Error 4: ',exception)
#     window['result_cam4'].update(value= 'Error', text_color='red')


try:
    st.initialize()
    st_system = st.create_system()
    connect_total = True
    writedata('DM7000.U',1)
except Exception as exception:
    print('Error total: ',exception)
    window['result_cam1'].update(value= 'Error', text_color='red')
#     window['result_cam2'].update(value= 'Error', text_color='red')
#     window['result_cam3'].update(value= 'Error', text_color='red')
#     window['result_cam4'].update(value= 'Error', text_color='red')

# st_datastream2, st_device2, remote_nodemap2= setup_camera2_stc()

st_datastream1, st_device1, remote_nodemap1= setup_camera1_stc()
# st_datastream2, st_device2, remote_nodemap2= setup_camera2_stc()
# st_datastream3, st_device3, remote_nodemap3= setup_camera3_stc()
# st_datastream4, st_device4, remote_nodemap4= setup_camera4_stc()


if connect_camera1 == True and connect_total == True:
    window['result_cam1'].update(value= 'Done', text_color='blue')
if connect_camera2 == True and connect_total == True:
    window['result_cam2'].update(value= 'Done', text_color='blue')

try:
    while True:

        event, values = window.read(timeout=20)

        for i1 in range(len(model1.names)):
            #if event == f'{model1.names[i1]}_1':
            if values[f'{model1.names[i1]}_1'] == False:
                window[f'{model1.names[i1]}_OK_1'].update(disabled=True)
                window[f'{model1.names[i1]}_Num_1'].update(disabled=True)
                window[f'{model1.names[i1]}_NG_1'].update(disabled=True)
                window[f'{model1.names[i1]}_Wn_1'].update(disabled=True)
                window[f'{model1.names[i1]}_Wx_1'].update(disabled=True)
                window[f'{model1.names[i1]}_Hn_1'].update(disabled=True)
                window[f'{model1.names[i1]}_Hx_1'].update(disabled=True)
                window[f'{model1.names[i1]}_PLC_1'].update(disabled=True)

            elif values[f'{model1.names[i1]}_1'] == True:
                window[f'{model1.names[i1]}_OK_1'].update(disabled=False)
                window[f'{model1.names[i1]}_Num_1'].update(disabled=False)
                window[f'{model1.names[i1]}_NG_1'].update(disabled=False)
                window[f'{model1.names[i1]}_Wn_1'].update(disabled=False)
                window[f'{model1.names[i1]}_Wx_1'].update(disabled=False)
                window[f'{model1.names[i1]}_Hn_1'].update(disabled=False)
                window[f'{model1.names[i1]}_Hx_1'].update(disabled=False)
                window[f'{model1.names[i1]}_PLC_1'].update(disabled=False)

        for i1 in range(len(model1.names)):
            if event == f'{model1.names[i1]}_OK_1':
                if values[f'{model1.names[i1]}_OK_1'] == True:
                    window[f'{model1.names[i1]}_NG_1'].update(disabled=True)
                else:
                    window[f'{model1.names[i1]}_NG_1'].update(disabled=False)
            if event == f'{model1.names[i1]}_NG_1':
                if values[f'{model1.names[i1]}_NG_1'] == True:
                    window[f'{model1.names[i1]}_OK_1'].update(disabled=True)
                else:
                    window[f'{model1.names[i1]}_OK_1'].update(disabled=False)


        for i2 in range(len(model2.names)):
            #if event == f'{model2.names[i2]}_2':
            if values[f'{model2.names[i2]}_2'] == False:
                window[f'{model2.names[i2]}_OK_2'].update(disabled=True)
                window[f'{model2.names[i2]}_Num_2'].update(disabled=True)
                window[f'{model2.names[i2]}_NG_2'].update(disabled=True)
                window[f'{model2.names[i2]}_Wn_2'].update(disabled=True)
                window[f'{model2.names[i2]}_Wx_2'].update(disabled=True)
                window[f'{model2.names[i2]}_Hn_2'].update(disabled=True)
                window[f'{model2.names[i2]}_Hx_2'].update(disabled=True)
                window[f'{model2.names[i2]}_PLC_2'].update(disabled=True)
                window[f'{model2.names[i2]}_Conf_2'].update(disabled=True)

            elif values[f'{model2.names[i2]}_2'] == True:
                window[f'{model2.names[i2]}_OK_2'].update(disabled=False)
                window[f'{model2.names[i2]}_Num_2'].update(disabled=False)
                window[f'{model2.names[i2]}_NG_2'].update(disabled=False)
                window[f'{model2.names[i2]}_Wn_2'].update(disabled=False)
                window[f'{model2.names[i2]}_Wx_2'].update(disabled=False)
                window[f'{model2.names[i2]}_Hn_2'].update(disabled=False)
                window[f'{model2.names[i2]}_Hx_2'].update(disabled=False)
                window[f'{model2.names[i2]}_PLC_2'].update(disabled=False)
                window[f'{model2.names[i2]}_Conf_2'].update(disabled=False)

        for i2 in range(len(model2.names)):
            if event == f'{model2.names[i2]}_OK_2':
                if values[f'{model2.names[i2]}_OK_2'] == True:
                    window[f'{model2.names[i2]}_NG_2'].update(disabled=True)
                else:
                    window[f'{model2.names[i2]}_NG_2'].update(disabled=False)
            if event == f'{model2.names[i2]}_NG_2':
                if values[f'{model2.names[i2]}_NG_2'] == True:
                    window[f'{model2.names[i2]}_OK_2'].update(disabled=True)
                else:
                    window[f'{model2.names[i2]}_OK_2'].update(disabled=False)


        for i3 in range(len(model3.names)):
            #if event == f'{model3.names[i3]}_3':
            if values[f'{model3.names[i3]}_3'] == False:
                window[f'{model3.names[i3]}_OK_3'].update(disabled=True)
                window[f'{model3.names[i3]}_Num_3'].update(disabled=True)
                window[f'{model3.names[i3]}_NG_3'].update(disabled=True)
                window[f'{model3.names[i3]}_Wn_3'].update(disabled=True)
                window[f'{model3.names[i3]}_Wx_3'].update(disabled=True)
                window[f'{model3.names[i3]}_Hn_3'].update(disabled=True)
                window[f'{model3.names[i3]}_Hx_3'].update(disabled=True)

            elif values[f'{model3.names[i3]}_3'] == True:
                window[f'{model3.names[i3]}_OK_3'].update(disabled=False)
                window[f'{model3.names[i3]}_Num_3'].update(disabled=False)
                window[f'{model3.names[i3]}_NG_3'].update(disabled=False)
                window[f'{model3.names[i3]}_Wn_3'].update(disabled=False)
                window[f'{model3.names[i3]}_Wx_3'].update(disabled=False)
                window[f'{model3.names[i3]}_Hn_3'].update(disabled=False)
                window[f'{model3.names[i3]}_Hx_3'].update(disabled=False)

        for i3 in range(len(model3.names)):
            if event == f'{model3.names[i3]}_OK_3':
                if values[f'{model3.names[i3]}_OK_3'] == True:
                    window[f'{model3.names[i3]}_NG_3'].update(disabled=True)
                else:
                    window[f'{model3.names[i3]}_NG_3'].update(disabled=False)
            if event == f'{model3.names[i3]}_NG_3':
                if values[f'{model3.names[i3]}_NG_3'] == True:
                    window[f'{model3.names[i3]}_OK_3'].update(disabled=True)
                else:
                    window[f'{model3.names[i3]}_OK_3'].update(disabled=False)




        if event =='Exit' or event == sg.WINDOW_CLOSED :
            break

        # if event == 'Configure':
        #     if window.TKroot.state() == 'zoomed':
        #         #print(window['image1'].get_size()[0])
        #         image_width_display = window['image1'].get_size()[0]
        #         image_height_display = window['image1'].get_size()[1]
        #         result_width_display = image_width_display - 190
        #         result_height_display = 100 


        if event =='Administrator':
            login_password = 'vu123'  # helloworld
            password = sg.popup_get_text(
                'Enter Password: ', password_char='*') 
            if password == login_password:
                sg.popup_ok('Login Successed!!! ',text_color='green', font=('Helvetica',14))  

                window['conf_thres1'].update(disabled= False)
                window['conf_thres2'].update(disabled= False)


                window['file_browse1'].update(disabled= False,button_color='turquoise')
                window['file_browse2'].update(disabled= False,button_color='turquoise')


                window['SaveData1'].update(disabled= False,button_color='turquoise')
                window['SaveData2'].update(disabled= False,button_color='turquoise')

                window['Webcam1'].update(disabled= False,button_color='turquoise')
                window['Stop1'].update(disabled= False,button_color='turquoise')
                window['Pic1'].update(disabled= False,button_color='turquoise')
                window['Snap1'].update(disabled= False,button_color='turquoise')
                window['next'].update(disabled= False,button_color='turquoise')
                window['back'].update(disabled= False,button_color='turquoise')
                #window['Change_1'].update(button_color='turquoise',disabled=False)
                #window['Change_2'].update(button_color='turquoise',disabled=False)
                window['Detect1'].update(button_color='turquoise',disabled=False)
                window['Detect2'].update(button_color='turquoise',disabled=False)


                window['have_save_OK_1'].update(disabled=False)
                window['have_save_NG_1'].update(disabled=False)
                window['have_save_OK_2'].update(disabled=False)
                window['have_save_NG_2'].update(disabled=False)

                window['save_OK_1'].update(disabled=False)
                window['save_NG_1'].update(disabled=False)
                window['save_OK_2'].update(disabled=False)
                window['save_NG_2'].update(disabled=False)

                window['save_folder_OK_1'].update(disabled= False,button_color='turquoise')
                window['save_folder_NG_1'].update(disabled= False,button_color='turquoise')
                window['save_folder_OK_2'].update(disabled= False,button_color='turquoise')
                window['save_folder_NG_2'].update(disabled= False,button_color='turquoise')

                window['conf_thres3'].update(disabled= False)

                window['file_browse3'].update(disabled= False,button_color='turquoise')

                window['SaveData3'].update(disabled= False,button_color='turquoise')



                for i1 in range(len(model1.names)):
                    window[f'{model1.names[i1]}_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_OK_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Num_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_NG_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Wn_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Wx_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Hn_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Hx_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_PLC_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Conf_1'].update(disabled=False)

                for i2 in range(len(model2.names)):
                    window[f'{model2.names[i2]}_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_OK_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Num_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_NG_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Wn_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Wx_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Hn_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Hx_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_PLC_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Conf_2'].update(disabled=False)

                for i3 in range(len(model3.names)):
                    window[f'{model3.names[i3]}_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_OK_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_Num_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_NG_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_Wn_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_Wx_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_Hn_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_Hx_3'].update(disabled=False)


            else:
                sg.popup_cancel('Wrong Password!!!',text_color='red', font=('Helvetica',14))


        if event == 'Change Theme':
            layout_theme = [
                [sg.Listbox(values= sg.theme_list(), size = (30,20),auto_size_text=18,default_values='Dark',key='theme', enable_events=True)],
                [
                    [sg.Button('Apply'),
                    sg.Button('Cancel')]
                ]
            ] 
            window_theme = sg.Window('Change Theme', layout_theme, location=(50,50),keep_on_top=True).Finalize()
            window_theme.set_min_size((300,400))   

            while True:
                event_theme, values_theme = window_theme.read(timeout=20)
                if event_theme == sg.WIN_CLOSED:
                    break

                if event_theme == 'Apply':
                    theme_choose = values_theme['theme'][0]
                    if theme_choose == 'Default':
                        continue
                    window.close()
                    window = make_window(theme_choose)
                    save_theme(theme_choose)
                    #print(theme_choose)
                if event_theme == 'Cancel':
                    answer = sg.popup_yes_no('Do you want to exit?')
                    if answer == 'Yes':
                        break
                    if answer == 'No':
                        continue
            window_theme.close()



        if event == 'file_browse1': 
            window['file_weights1'].update(value=values['file_browse1'])
            if values['file_browse1']:
                window['Change_1'].update(disabled=False,button_color='turquoise')

        if event == 'file_browse2': 
            window['file_weights2'].update(value=values['file_browse2'])
            if values['file_browse2']:
                window['Change_2'].update(disabled=False,button_color='turquoise')

        if event == 'file_browse3': 
            window['file_weights3'].update(value=values['file_browse3'])
            if values['file_browse3']:
                window['Change_3'].update(disabled=False,button_color='turquoise')




        if event == 'choose_model':
            mychoose = values['choose_model']
            weight1 = ''
            conf_thres1 = 1


            OK_Cam1 = False

            NG_Cam1 = True

            Folder_OK_Cam1 = 'C:/Cam1/OK'
           
            Folder_NG_Cam1 = 'C:/Cam1/NG'
       

 


            conn = sqlite3.connect('modeldb_2_PLC_Conf.db')
            cursor = conn.execute("SELECT ChooseModel,Camera,Weights,Confidence,OK_Cam1,NG_Cam1,Folder_OK_Cam1,Folder_NG_Cam1,Joined,Ok,Num,NG,WidthMin,WidthMax,HeightMin,HeightMax from MYMODEL")
            for row in cursor:
                if row[0] == values['choose_model']:
 
                    mychoose = values['choose_model']
                    row1_a, row1_b = row[1].strip().split('_')
                    if row1_a == '1' and row1_b == '0':
                        weight1 = row[2]
                        conf_thres1 = row[3]
                        OK_Cam1 = str2bool(row[4])
                     
               
                        NG_Cam1 = str2bool(row[5])
                
             
                        Folder_OK_Cam1 = row[6]
                       
                  
                        Folder_NG_Cam1 = row[7]
                        
                     
                        model1 = torch.hub.load('./levu','custom', path= row[2], source='local',force_reload =False)

                    if row1_a == '2' and row1_b == '0':
                        weight2 = row[2]
                        conf_thres2 = row[3]
                        OK_Cam1 = str2bool(row[4])
     
                      

                        NG_Cam1 = str2bool(row[5])
        
                      
            
                        Folder_OK_Cam1 = row[6]

             
                        Folder_NG_Cam1 = row[7]
        
         
            
                        model2 = torch.hub.load('./levu','custom', path= row[2], source='local',force_reload =False)

 
        
            window.close() 
            window = make_window(theme)

            window['file_weights1'].update(value=weight1)
            window['conf_thres1'].update(value=conf_thres1)
      
            window['choose_model'].update(value=mychoose)

            window['have_save_OK_1'].update(value=OK_Cam1)
        
            window['have_save_NG_1'].update(value=NG_Cam1)
         

            window['save_OK_1'].update(value=Folder_OK_Cam1)
       
            window['save_NG_1'].update(value=Folder_NG_Cam1)
   


            window['choose_model'].update(value=mychoose)


   


            cursor = conn.execute("SELECT ChooseModel,Camera,Weights,Confidence,OK_Cam1,NG_Cam1,Folder_OK_Cam1,Folder_NG_Cam1,Joined,Ok,Num,NG,WidthMin,WidthMax,HeightMin,HeightMax,PLC_NG,PLC_OK,Conf from MYMODEL")
            for row in cursor:
                if row[0] == values['choose_model']:
                    row1_a, row1_b = row[1].strip().split('_')
                    if row1_a == '1':
                        for item in range(len(model1.names)):
                            if int(row1_b) == item:
                                window[f'{model1.names[item]}_1'].update(value=str2bool(row[8]))
                                window[f'{model1.names[item]}_OK_1'].update(value=str2bool(row[9]))
                                window[f'{model1.names[item]}_Num_1'].update(value=str(row[10]))
                                window[f'{model1.names[item]}_NG_1'].update(value=str2bool(row[11]))
                                window[f'{model1.names[item]}_Wn_1'].update(value=str(row[12]))
                                window[f'{model1.names[item]}_Wx_1'].update(value=str(row[13]))
                                window[f'{model1.names[item]}_Hn_1'].update(value=str(row[14]))
                                window[f'{model1.names[item]}_Hx_1'].update(value=str(row[15]))
                                window[f'{model1.names[item]}_PLC_1'].update(value=str(row[16]))
                                window['OK_PLC_1'].update(value=str(row[17]))
                                window[f'{model1.names[item]}_Conf_1'].update(value=str(row[18]))


                    if row1_a == '2':
                        for item in range(len(model2.names)):
                            if int(row1_b) == item:
                                window[f'{model2.names[item]}_2'].update(value=str2bool(row[8]))
                                window[f'{model2.names[item]}_OK_2'].update(value=str2bool(row[9]))
                                window[f'{model2.names[item]}_Num_2'].update(value=str(row[10]))
                                window[f'{model2.names[item]}_NG_2'].update(value=str2bool(row[11]))
                                window[f'{model2.names[item]}_Wn_2'].update(value=str(row[12]))
                                window[f'{model2.names[item]}_Wx_2'].update(value=str(row[13]))
                                window[f'{model2.names[item]}_Hn_2'].update(value=str(row[14]))
                                window[f'{model2.names[item]}_Hx_2'].update(value=str(row[15]))
                                window[f'{model2.names[item]}_PLC_2'].update(value=str(row[16]))
                                window['OK_PLC_2'].update(value=str(row[17]))
                                window[f'{model2.names[item]}_Conf_2'].update(value=str(row[18]))

                    if row1_a == '3':
                        for item in range(len(model3.names)):
                            if int(row1_b) == item:
                                window[f'{model3.names[item]}_3'].update(value=str2bool(row[20]))
                                window[f'{model3.names[item]}_OK_3'].update(value=str2bool(row[21]))
                                window[f'{model3.names[item]}_Num_3'].update(value=str(row[22]))
                                window[f'{model3.names[item]}_NG_3'].update(value=str2bool(row[23]))
                                window[f'{model3.names[item]}_Wn_3'].update(value=str(row[24]))
                                window[f'{model3.names[item]}_Wx_3'].update(value=str(row[25]))
                                window[f'{model3.names[item]}_Hn_3'].update(value=str(row[26]))
                                window[f'{model3.names[item]}_Hx_3'].update(value=str(row[27]))
                                window[f'{model3.names[item]}_PLC_3'].update(value=str(row[28]))
                                window['OK_PLC_3'].update(value=str(row[29]))
                                window[f'{model3.names[item]}_Conf_3'].update(value=str(row[30]))


            conn.close()

        if event == 'SaveData1':

            save_all_sql(model1,1,str(values['choose_model']))
            save_choosemodel(values['choose_model'])
            save_model(1,values['file_weights1'])
            sg.popup('Saved param model 1 successed',font=('Helvetica',15), text_color='green',keep_on_top= True)

        if event == 'SaveData2':
            save_all_sql(model2,2,str(values['choose_model']))
            save_choosemodel(values['choose_model'])
            save_model(2,values['file_weights2'])
            sg.popup('Saved param model 2 successed',font=('Helvetica',15), text_color='green',keep_on_top= True)

        if event == 'SaveData3':

            save_all_sql(model3,3,str(values['choose_model']))
            save_choosemodel(values['choose_model'])
            save_model(3,values['file_weights3'])
            sg.popup('Saved param model 3 successed',font=('Helvetica',15), text_color='green',keep_on_top= True)





        try:
            #HomNay
            if readdata('DM6050') == 1:
                excel_sang()

            # #Dem
            if readdata('DM6050') == 2:
                excel_dem()


            if readdata('DM7020') == 1:
                today = date.today()
                d1 = today.strftime("%d/%m/%Y")
                now = datetime.datetime.now()
                t1 = now.strftime("%H:%M:%S")

                mydate = today.strftime("%Y_%m_%d")
            
                hour = int(now.strftime("%H"))
        
                if 7<= hour <=18:
                    if not os.path.isfile(f"excel/{mydate}_Ngay.xlsx"):
                        excel_sang()
                    wb = openpyxl.load_workbook(f"excel/{mydate}_Ngay.xlsx")

                if 19 <= hour <= 23:
                    if not os.path.isfile(f"excel/{mydate}_Dem.xlsx"):
                        excel_dem()
                    wb = openpyxl.load_workbook(f"excel/{mydate}_Dem.xlsx")

                if 0 <= hour <= 6:
                    Previous_Date = datetime.datetime.today() - datetime.timedelta(days=1)
                    Previous_Date = Previous_Date.strftime("%Y_%m_%d")
                    if not os.path.isfile(f"excel/{Previous_Date}_Dem.xlsx"):
                        excel_dem()

                    wb = openpyxl.load_workbook(f"excel/{Previous_Date}_Dem.xlsx")


                ws = wb.active
                col1 = d1
                col2 = t1
                col3 = int(readdata('DM7218'))
                col4 = int(readdata('DM7200'))
                col5 = int(readdata('DM7202'))
                col6 = int(readdata('DM7204'))
                col7 = int(readdata('DM7206'))
                col8 = int(readdata('DM7208'))
                col9 = int(readdata('DM7212'))
                col10 = int(readdata('DM7214'))
                col11 = int(readdata('DM7210'))
                col12 = int(readdata('DM7220'))
                #col13 = int(readdata('DM6042'))
                ws.append([col1, col2,col3, col4,col5, col6,col7, col8,col9, col10,col11, col12])

                for row in range(ws.max_row, ws.max_row+1):
                    for col in range(1, ws.max_column+1):
                        #print(col[row].value)
                        d = ws.cell(row = row, column = col)
                        #currentCell = ws.cell(col[row])
                        d.alignment = Alignment(horizontal='center')
                        #d.style.alignment.horizontal = 'center'
                        d.font = Font(name= 'Calibri', size=12)

                if 7<= hour <=18:
                    
                    wb.save(f"excel/{mydate}_Ngay.xlsx")
                    try:
                        shutil.copy(f"excel/{mydate}_Ngay.xlsx", f"C:/excel/{mydate}_Ngay.xlsx")
                    except:
                        pass
                    
                    try:
                        wb1 = openpyxl.load_workbook(f"C:/excel/{mydate}_Ngay.xlsx")

                        ws1 = wb1.active
                        all_cell = []


                        for row in range(1, ws1.max_row+1):
                            all_col = []    
                            for col in range(1, 13):
                                cell_obj = ws1.cell(row = row, column = col)
                                all_col.append(cell_obj.value)
                            all_cell.append(all_col)




                        wb2 = openpyxl.load_workbook("C:/excel/Now.xlsx")
                        ws2 = wb2.active

                        index_row = ws2.max_row


                        if ws1.max_row > ws2.max_row:
                            for row in range(4,ws1.max_row+1):
                                for col in range(1, 13):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                    
                        else:
                            for row in range(4,ws2.max_row+1):
                                for col in range(1, 13):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                                    else:
                                        ws2.cell(row = row, column = col).value = None
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)   

                        wb2.save("C:/excel/Now.xlsx")



                    except:
                        pass
                if 19 <= hour <= 23:
                    wb.save(f"excel/{mydate}_Dem.xlsx")
                    try:
                        shutil.copy(f"excel/{mydate}_Dem.xlsx", f"C:/excel/{mydate}_Dem.xlsx")
                    except:
                        pass

                    try:
                        wb1 = openpyxl.load_workbook(f"C:/excel/{mydate}_Dem.xlsx")

                        ws1 = wb1.active
                        all_cell = []


                        for row in range(1, ws1.max_row+1):
                            all_col = []    
                            for col in range(1, 13):
                                cell_obj = ws1.cell(row = row, column = col)
                                all_col.append(cell_obj.value)
                            all_cell.append(all_col)




                        wb2 = openpyxl.load_workbook("C:/excel/Now.xlsx")
                        ws2 = wb2.active

                        index_row = ws2.max_row


                        if ws1.max_row > ws2.max_row:
                            for row in range(4,ws1.max_row+1):
                                for col in range(1, 13):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                    
                        else:
                            for row in range(4,ws2.max_row+1):
                                for col in range(1, 13):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                                    else:
                                        ws2.cell(row = row, column = col).value = None
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)   

                        wb2.save("C:/excel/Now.xlsx")



                    except:
                        pass


                if 0 <= hour <= 6:
                    Previous_Date = datetime.datetime.today() - datetime.timedelta(days=1)
                    Previous_Date = Previous_Date.strftime("%Y_%m_%d")
                    wb.save(f"excel/{Previous_Date}_Dem.xlsx")
                    try:
                        shutil.copy(f"excel/{Previous_Date}_Dem.xlsx", f"C:/excel/{Previous_Date}_Dem.xlsx")
                    except:
                        pass

                    try:
                        wb1 = openpyxl.load_workbook(f"C:/excel/{Previous_Date}_Dem.xlsx")

                        ws1 = wb1.active
                        all_cell = []


                        for row in range(1, ws1.max_row+1):
                            all_col = []    
                            for col in range(1, 13):
                                cell_obj = ws1.cell(row = row, column = col)
                                all_col.append(cell_obj.value)
                            all_cell.append(all_col)




                        wb2 = openpyxl.load_workbook("C:/excel/Now.xlsx")
                        ws2 = wb2.active

                        index_row = ws2.max_row


                        if ws1.max_row > ws2.max_row:
                            for row in range(4,ws1.max_row+1):
                                for col in range(1, 13):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                    
                        else:
                            for row in range(4,ws2.max_row+1):
                                for col in range(1, 13):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                                    else:
                                        ws2.cell(row = row, column = col).value = None
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)   

                        wb2.save("C:/excel/Now.xlsx")



                    except:
                        pass


                writedata('DM7020.U',0) 
            

            if readdata('DM6054') == 1:
                today = date.today()
                d1 = today.strftime("%d/%m/%Y")
                now = datetime.datetime.now()
                t1 = now.strftime("%H:%M:%S")

                wb = openpyxl.load_workbook("excel/All.xlsx")
                ws = wb.active
                col1 = d1
                col2 = t1
                col3 = int(readdata('DM7218'))
                col4 = int(readdata('DM7200'))
                col5 = int(readdata('DM7202'))
                col6 = int(readdata('DM7204'))
                col7 = int(readdata('DM7206'))
                col8 = int(readdata('DM7208'))
                col9 = int(readdata('DM7212'))
                col10 = int(readdata('DM7214'))
                col11 = int(readdata('DM7210'))
                col12 = int(readdata('DM7220'))
                #col13 = int(readdata('DM6042'))
                ws.append([col1, col2,col3, col4,col5, col6,col7, col8,col9, col10,col11, col12])

                for row in range(ws.max_row, ws.max_row+1):
                    for col in range(1, ws.max_column+1):
                        #print(col[row].value)
                        d = ws.cell(row = row, column = col)
                        #currentCell = ws.cell(col[row])
                        d.alignment = Alignment(horizontal='center')
                        #d.style.alignment.horizontal = 'center'
                        d.font = Font(name= 'Calibri', size=12)


                wb.save("excel/All.xlsx")
                try:
                    shutil.copy("excel/All.xlsx", "C:/excel/A.xlsx")



                    wb1 = openpyxl.load_workbook("C:/excel/A.xlsx")

                    ws1 = wb1.active
                    all_cell = []
                    all_cell_date = []
                    all_cell_time = []

                    for row in range(1, ws1.max_row+1):
                        all_col = []    
                        for col in range(1, 13):
                            cell_obj = ws1.cell(row = row, column = col)
                            all_col.append(cell_obj.value)
                        all_cell.append(all_col)
                        all_cell_date.append(all_col[0])
                        all_cell_time.append(all_col[1])



                    wb2 = openpyxl.load_workbook("C:/excel/All.xlsx")
                    ws2 = wb2.active

                    index_row = ws2.max_row
                    for row in range(4, ws2.max_row+1):
                        if ws2.cell(row=row,column=1).value == None:
                            index_row = row
                            break

                    for row in range(index_row,ws1.max_row+1):
                        for col in range(1, 13):
                            ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                            d = ws2.cell(row = row, column = col)
                            d.alignment = Alignment(horizontal='center')
                            d.font = Font(name= 'Calibri', size=12)
                    wb2.save("C:/excel/All.xlsx")


                except:
                    pass
                
                writedata('DM6054.U',0) 
            
        except:
            pass

     









        try:
            #program_camera1_test(model=model1,size= 416,conf= values['conf_thres1']/100)
            #program_camera2_test(model=model2,size= 416,conf= values['conf_thres1']/100)


            program_camera1(model=model1,size= 416,conf= values['conf_thres1']/100)
            program_camera2(model=model2,size= 416,conf= values['conf_thres1']/100)
            program_camera3(model=model3,size= 416,conf= values['conf_thres1']/100)

        except:
            pass

        window['Index_image'].update(value = myindex1)

        #window['signal_model'].update(value = str(readdata('DM7100')))

        if readdata('DM7604') == 1:
            window['result_cam1'].update(value= 'OK', text_color='green')    
            #writedata('DM7604.U',0)
        
        if readdata('DM7606') == 1:
            window['result_cam1'].update(value= 'NG', text_color='red')    
            #writedata('DM7606.U',0)

        if readdata('DM7004') == 1:
            myindex1 = 0
            writedata('DM7004.U',0)
            imgbytes = np.zeros([100,100,3],dtype = np.uint8)
            imgbytes = cv2.resize(imgbytes, (int(image_width_display/2.2),int(image_height_display/1.4)), interpolation = cv2.INTER_AREA)
            imgbytes = cv2.imencode('.png',imgbytes)[1].tobytes()
            window['image1.1'].update(data= imgbytes)
            window['image1.2'].update(data= imgbytes)
            window['image1.3'].update(data= imgbytes)
            window['image1.4'].update(data= imgbytes)
            window['image1.5'].update(data= imgbytes)
            window['image1.6'].update(data= imgbytes)
        # #HomNay
        # if readdata('DM7050') == 1:
        #     today = date.today()
        #     mydate = today.strftime("%Y_%m_%d")
        #     wb = openpyxl.Workbook()

        #     HomNay = wb.create_sheet("Data")


        #     HomNay.merge_cells('A1:M1')
        #     HomNay.merge_cells('A2:A3')
        #     HomNay.merge_cells('B2:B3')
        #     HomNay.merge_cells('C2:M2')
        #     #HomNay.unmerge_cells('A2:D2')
        #     HomNay['A1'] = 'DỮ LIỆU MÁY NQVNHT RS656 A17'
        #     HomNay['A1'].alignment = Alignment(horizontal='center')
        #     HomNay['A1'].font = Font(name= 'Calibri', size=20)

        #     HomNay['A2'] = 'Ngày sản xuất'
        #     HomNay['A2'].alignment = Alignment(horizontal='center')
        #     HomNay['A2'].font = Font(name= 'Calibri', size=12)
        #     HomNay['B2'] = 'Giờ lưu dữ liệu'
        #     HomNay['B2'].alignment = Alignment(horizontal='center')
        #     HomNay['B2'].font = Font(name= 'Calibri', size=12)
        #     HomNay['C2'] = 'HẠNG MỤC PHẾ PHẨM'
        #     HomNay['C2'].alignment = Alignment(horizontal='center')
        #     HomNay['C2'].font = Font(name= 'Calibri', size=12)
        #     HomNay['C3'] = 'Tổng số lượng sản xuất'
        #     HomNay['D3'] = 'Cuộn cảm'
        #     HomNay['E3'] = 'Cacbon tay chổi'
        #     HomNay['F3'] = 'Hàn chổi'
        #     HomNay['G3'] = 'Hàn chấu '
        #     HomNay['H3'] = 'Đế vỏ nhỏ'
        #     HomNay['I3'] = 'Tụ điện'
        #     HomNay['J3'] = 'Cong chấu điện'
        #     HomNay['K3'] = 'Bụi chì'
        #     HomNay['L3'] = 'PP khác (nhiều hạng mục)'
        #     HomNay['M3'] = 'Tổng số lượng PP'

        #     for i in range(67,77):
        #         HomNay[f'{str(chr(i))}3'].alignment = Alignment(horizontal='center')
        #         HomNay[f'{str(chr(i))}3'].font = Font(name= 'Calibri', size=12) 

        #     HomNay.column_dimensions['A'].width = 20
        #     HomNay.column_dimensions['B'].width = 20
        #     HomNay.column_dimensions['C'].width = 25
        #     HomNay.column_dimensions['D'].width = 18
        #     HomNay.column_dimensions['E'].width = 18
        #     HomNay.column_dimensions['F'].width = 18
        #     HomNay.column_dimensions['G'].width = 18
        #     HomNay.column_dimensions['H'].width = 18
        #     HomNay.column_dimensions['I'].width = 18
        #     HomNay.column_dimensions['J'].width = 18
        #     HomNay.column_dimensions['K'].width = 18
        #     HomNay.column_dimensions['L'].width = 25
        #     HomNay.column_dimensions['M'].width = 25

        #     wb.remove(wb['Sheet'])

        #     wb.save(f"excel/{mydate}_Ngay.xlsx")
        #     try:
        #         shutil.copy(f"excel/{mydate}_Ngay.xlsx", f"C:/excel/{mydate}_Ngay.xlsx")
        #     except:
        #         pass
        #     writedata('DM6050.U',0) 

        # # #Dem
        # if readdata('DM7050') == 2:
        #     today = date.today()
        #     mydate = today.strftime("%Y_%m_%d")
        #     wb = openpyxl.Workbook()

        #     HomNay = wb.create_sheet("Data")


        #     HomNay.merge_cells('A1:M1')
        #     HomNay.merge_cells('A2:A3')
        #     HomNay.merge_cells('B2:B3')
        #     HomNay.merge_cells('C2:M2')
        #     #HomNay.unmerge_cells('A2:D2')
        #     HomNay['A1'] = 'DỮ LIỆU MÁY NQVNHT RS656 A17'
        #     HomNay['A1'].alignment = Alignment(horizontal='center')
        #     HomNay['A1'].font = Font(name= 'Calibri', size=20)

        #     HomNay['A2'] = 'Ngày sản xuất'
        #     HomNay['A2'].alignment = Alignment(horizontal='center')
        #     HomNay['A2'].font = Font(name= 'Calibri', size=12)
        #     HomNay['B2'] = 'Giờ lưu dữ liệu'
        #     HomNay['B2'].alignment = Alignment(horizontal='center')
        #     HomNay['B2'].font = Font(name= 'Calibri', size=12)
        #     HomNay['C2'] = 'HẠNG MỤC PHẾ PHẨM'
        #     HomNay['C2'].alignment = Alignment(horizontal='center')
        #     HomNay['C2'].font = Font(name= 'Calibri', size=12)
        #     HomNay['C3'] = 'Tổng số lượng sản xuất'
        #     HomNay['D3'] = 'Cuộn cảm'
        #     HomNay['E3'] = 'Cacbon tay chổi'
        #     HomNay['F3'] = 'Hàn chổi'
        #     HomNay['G3'] = 'Hàn chấu '
        #     HomNay['H3'] = 'Đế vỏ nhỏ'
        #     HomNay['I3'] = 'Tụ điện'
        #     HomNay['J3'] = 'Cong chấu điện'
        #     HomNay['K3'] = 'Bụi chì'
        #     HomNay['L3'] = 'PP khác (nhiều hạng mục)'
        #     HomNay['M3'] = 'Tổng số lượng PP'

        #     for i in range(67,77):
        #         HomNay[f'{str(chr(i))}3'].alignment = Alignment(horizontal='center')
        #         HomNay[f'{str(chr(i))}3'].font = Font(name= 'Calibri', size=12) 

        #     HomNay.column_dimensions['A'].width = 20
        #     HomNay.column_dimensions['B'].width = 20
        #     HomNay.column_dimensions['C'].width = 25
        #     HomNay.column_dimensions['D'].width = 18
        #     HomNay.column_dimensions['E'].width = 18
        #     HomNay.column_dimensions['F'].width = 18
        #     HomNay.column_dimensions['G'].width = 18
        #     HomNay.column_dimensions['H'].width = 18
        #     HomNay.column_dimensions['I'].width = 18
        #     HomNay.column_dimensions['J'].width = 18
        #     HomNay.column_dimensions['K'].width = 18
        #     HomNay.column_dimensions['L'].width = 25
        #     HomNay.column_dimensions['M'].width = 25

        #     wb.remove(wb['Sheet'])

        #     wb.save(f"excel/{mydate}_Dem.xlsx")
        #     try:
        #         shutil.copy(f"excel/{mydate}_Dem.xlsx", f"C:/excel/{mydate}_Dem.xlsx")
        #     except:
        #         pass
        #     writedata('DM7050.U',0) 


        # if readdata('DM1100') == 1:
        #     today = date.today()
        #     d1 = today.strftime("%d/%m/%Y")
        #     now = datetime.datetime.now()
        #     t1 = now.strftime("%H:%M:%S")

        #     mydate = today.strftime("%Y_%m_%d")
          
        #     hour = int(now.strftime("%H"))
    
        #     if 7<= hour <=18:
        #         wb = openpyxl.load_workbook(f"excel/{mydate}_Ngay.xlsx")

        #     if 19 <= hour <= 23:
        #         wb = openpyxl.load_workbook(f"excel/{mydate}_Dem.xlsx")

        #     if 0 <= hour <= 6:
        #         Previous_Date = datetime.datetime.today() - datetime.timedelta(days=1)
        #         Previous_Date = Previous_Date.strftime("%Y_%m_%d")
        #         wb = openpyxl.load_workbook(f"excel/{Previous_Date}_Dem.xlsx")


        #     ws = wb.active
        #     col1 = d1
        #     col2 = t1
        #     col3 = str(readdata('DM110'))
        #     col4 = str(readdata('DM6000'))
        #     col5 = str(readdata('DM6004'))
        #     col6 = str(readdata('DM6008'))
        #     col7 = str(readdata('DM6012'))
        #     col8 = str(readdata('DM6016'))
        #     col9 = str(readdata('DM6020'))
        #     col10 = str(readdata('DM6024'))
        #     col11 = str(readdata('DM6032'))
        #     col12 = str(readdata('DM6028'))
        #     col13 = str(readdata('DM6042'))
        #     ws.append([col1, col2,col3, col4,col5, col6,col7, col8,col9, col10,col11, col12,col13,])

        #     for row in range(ws.max_row, ws.max_row+1):
        #         for col in range(1, ws.max_column+1):
        #             #print(col[row].value)
        #             d = ws.cell(row = row, column = col)
        #             #currentCell = ws.cell(col[row])
        #             d.alignment = Alignment(horizontal='center')
        #             #d.style.alignment.horizontal = 'center'
        #             d.font = Font(name= 'Calibri', size=12)

        #     if 7<= hour <=18:
        #         wb.save(f"excel/{mydate}_Ngay.xlsx")
        #         try:
        #             shutil.copy(f"excel/{mydate}_Ngay.xlsx", f"C:/excel/{mydate}_Ngay.xlsx")
        #         except:
        #             pass
        #     if 19 <= hour <= 23:
        #         wb.save(f"excel/{mydate}_Dem.xlsx")
        #         try:
        #             shutil.copy(f"excel/{mydate}_Dem.xlsx", f"C:/excel/{mydate}_Dem.xlsx")
        #         except:
        #             pass
        #     if 0 <= hour <= 6:
        #         Previous_Date = datetime.datetime.today() - datetime.timedelta(days=1)
        #         Previous_Date = Previous_Date.strftime("%Y_%m_%d")
        #         wb.save(f"excel/{Previous_Date}_Dem.xlsx")
        #         try:
        #             shutil.copy(f"excel/{Previous_Date}_Dem.xlsx", f"C:/excel/{mydate}_Dem.xlsx")
        #         except:
        #             pass


        #     writedata('DM1100.U',0) 
        

        # if readdata('DM6054') == 1:
        #     today = date.today()
        #     d1 = today.strftime("%d/%m/%Y")
        #     now = datetime.datetime.now()
        #     t1 = now.strftime("%H:%M:%S")

        #     wb = openpyxl.load_workbook("excel/All.xlsx")
        #     ws = wb.active
        #     col1 = d1
        #     col2 = t1
        #     col3 = str(readdata('DM110'))
        #     col4 = str(readdata('DM6000'))
        #     col5 = str(readdata('DM6004'))
        #     col6 = str(readdata('DM6008'))
        #     col7 = str(readdata('DM6012'))
        #     col8 = str(readdata('DM6016'))
        #     col9 = str(readdata('DM6020'))
        #     col10 = str(readdata('DM6024'))
        #     col11 = str(readdata('DM6032'))
        #     col12 = str(readdata('DM6028'))
        #     col13 = str(readdata('DM6042'))
        #     ws.append([col1, col2,col3, col4,col5, col6,col7, col8,col9, col10,col11, col12,col13,])

        #     for row in range(ws.max_row, ws.max_row+1):
        #         for col in range(1, ws.max_column+1):
        #             #print(col[row].value)
        #             d = ws.cell(row = row, column = col)
        #             #currentCell = ws.cell(col[row])
        #             d.alignment = Alignment(horizontal='center')
        #             #d.style.alignment.horizontal = 'center'
        #             d.font = Font(name= 'Calibri', size=12)


        #     wb.save("excel/All.xlsx")
        #     try:
        #         shutil.copy("excel/All.xlsx", "C:/excel/All.xlsx")
        #     except:
        #         pass
            
        #     writedata('DM6054.U',0) 
        




        if values['check_model1'] == True:
            if type(index_path) == int and index_path < len(list_path):
                path1 = list_path[index_path]
                name = path1[10:]

                img1_orgin = cv2.imread(path1)


                #img1_orgin = cv2.resize(img1_orgin,(640,480))  

                img1_orgin = cv2.cvtColor(img1_orgin, cv2.COLOR_BGR2RGB)     


                result1 = model1(img1_orgin,size= 416,conf = values['conf_thres1']/100)

                table1 = result1.pandas().xyxy[0]

                area_remove1 = []

                myresult1 =0 

                for item in range(len(table1.index)):
                    width1 = table1['xmax'][item] - table1['xmin'][item]
                    height1 = table1['ymax'][item] - table1['ymin'][item]
                    conf1 = table1['confidence'][item] *100

                    #area1 = width1*height1
                    label_name = table1['name'][item]
                    for i1 in range(len(model1.names)):
                        if values[f'{model1.names[i1]}_1'] == True:
                            #if values[f'{model1.names[i1]}_WH'] == True:
                            if label_name == model1.names[i1]:
                                if width1 < int(values[f'{model1.names[i1]}_Wn_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif width1 > int(values[f'{model1.names[i1]}_Wx_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif height1 < int(values[f'{model1.names[i1]}_Hn_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif height1 > int(values[f'{model1.names[i1]}_Hx_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif conf1  < int(values[f'{model1.names[i1]}_Conf_1']):
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                    if values[f'{model1.names[i1]}_1'] == False:
                        if label_name == model1.names[i1]:
                            table1.drop(item, axis=0, inplace=True)
                            area_remove1.append(item)


                names1 = list(table1['name'])

                show1 = np.squeeze(result1.render(area_remove1))
                show1 = cv2.resize(show1, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)

                show1 = cv2.cvtColor(show1, cv2.COLOR_BGR2RGB)

                #ta = time.time()
                for i1 in range(len(model1.names)):
                    #register_ng = (3002 + i1*2).to_bytes(2, byteorder='big') + b'\x00'

                    if values[f'{model1.names[i1]}_OK_1'] == True:
                        len_name1 = 0
                        for name1 in names1:
                            if name1 == model1.names[i1]:
                                len_name1 +=1
                        if len_name1 != int(values[f'{model1.names[i1]}_Num_1']):
                            print('NG')
                            #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register_ng,b'\x00\x01',1)
                            cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 4,(0,0,255),5)
                            window['result_cam1'].update(value= 'NG', text_color='red')
                            myresult1 = 1
                            

                    elif values[f'{model1.names[i1]}_NG_1'] == True:
                        if model1.names[i1] in names1:
                            print('NG')
                            #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register_ng,b'\x00\x01',1)
                            cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 4,(0,0,255),5)
                            window['result_cam1'].update(value= 'NG', text_color='red')    
                            myresult1 = 1         
                                

                if myresult1 == 0:
                    print('OK')
                    check_ok = 1
                    #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,(3000).to_bytes(2, byteorder='big') + b'\x00',b'\x00\x01',1)
                    cv2.putText(show1, 'OK',(result_width_display+100,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 4,(0,255,0),5)
                    window['result_cam1'].update(value= 'OK', text_color='green')

                cv2.putText(show1, str(index_show),(50,60),cv2.FONT_HERSHEY_COMPLEX, 2,(255,0,0),2)
                imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
                window['image1'].update(data= imgbytes1)
                #window['name_file'].update(value= str(name))


        if event == 'Num_index':
            if values['Num_index'] != '' and values['Num_index'].isdigit() :
                index_path = int(values['Num_index']) - 1
                index_show = int(values['Num_index']) 

        if event == 'next':
            if index_path < len(list_path):
                index_path += 1
                index_show += 1
            else:
                sg.popup('Da het anh')
        if event == 'back':
            index_path -= 1
            index_show -= 1
       #time.sleep(1)
                    # if check_ok == 1:
                    #     break
        #print(index_path)

        if event == 'Webcam1':
            #cap1 = cv2.VideoCapture(0)
            recording1 = True


        elif event == 'Stop1':
            recording1 = False 
            imgbytes1 = np.zeros([100,100,3],dtype=np.uint8)
            imgbytes1 = cv2.resize(imgbytes1, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
            imgbytes1 = cv2.imencode('.png',imgbytes1)[1].tobytes()
            window['image1'].update(data=imgbytes1)
            window['result_cam1'].update(value='')




        if recording1:
            # if have_webcam == True:
            #     img1_orgin = my_callback1.image 
            #     img1_orgin = img1_orgin[50:530,70:710]
            #     img1_orgin = img1_orgin.copy()
            #     img1_orgin = cv2.cvtColor(img1_orgin, cv2.COLOR_BGR2RGB)                              
                # result1 = model1(img1_orgin,size= 416,conf= values['conf_thres1']/100)             
                # table1 = result1.pandas().xyxy[0]
                # area_remove1 = []

                # myresult1 =0 

                # for item in range(len(table1.index)):
                #     width1 = table1['xmax'][item] - table1['xmin'][item]
                #     height1 = table1['ymax'][item] - table1['ymin'][item]
                #     #area1 = width1*height1
                #     label_name = table1['name'][item]
                #     for i1 in range(len(model1.names)):
                #         if values[f'{model1.names[i1]}_1'] == True:
                #             #if values[f'{model1.names[i1]}_WH'] == True:
                #             if label_name == model1.names[i1]:
                #                 if width1 < int(values[f'{model1.names[i1]}_Wn_1']): 
                #                     table1.drop(item, axis=0, inplace=True)
                #                     area_remove1.append(item)
                #                 elif width1 > int(values[f'{model1.names[i1]}_Wx_1']): 
                #                     table1.drop(item, axis=0, inplace=True)
                #                     area_remove1.append(item)
                #                 elif height1 < int(values[f'{model1.names[i1]}_Hn_1']): 
                #                     table1.drop(item, axis=0, inplace=True)
                #                     area_remove1.append(item)
                #                 elif height1 > int(values[f'{model1.names[i1]}_Hx_1']): 
                #                     table1.drop(item, axis=0, inplace=True)
                #                     area_remove1.append(item)

                # names1 = list(table1['name'])

                # show1 = np.squeeze(result1.render(area_remove1))
                # show1 = cv2.resize(show1, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
        
                # #ta = time.time()
                # for i1 in range(len(model1.names)):
                #     if values[f'{model1.names[i1]}_OK_1'] == True:
                #         len_name1 = 0
                #         for name1 in names1:
                #             if name1 == model1.names[i1]:
                #                 len_name1 +=1
                #         if len_name1 != int(values[f'{model1.names[i1]}_Num_1']):
                #             print('NG')
                #             cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                #             window['result_cam1'].update(value= 'NG', text_color='red')
                #             myresult1 = 1
                #             break

                #     if values[f'{model1.names[i1]}_NG_1'] == True:
                #         if model1.names[i1] in names1:
                #             print('NG')
                #             cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                #             window['result_cam1'].update(value= 'NG', text_color='red')    
                #             myresult1 = 1         
                #             break    

                # if myresult1 == 0:
                #     print('OK')
                #     cv2.putText(show1, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                #     window['result_cam1'].update(value= 'OK', text_color='green')
                
                # imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
                # window['image1'].update(data= imgbytes1)

            #else:
            img1_orgin = my_callback1.image 
            #img1_orgin = img1_orgin[50:530,70:710]
            img1_orgin = img1_orgin.copy()
            #img1_orgin = cv2.cvtColor(img1_orgin, cv2.COLOR_BGR2RGB) 
            img1_resize = cv2.resize(img1_orgin,(image_width_display,image_height_display))
            if img1_orgin is not None:
                show1 = img1_resize
                imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
                window['image1'].update(data=imgbytes1)
                window['result_cam1'].update(value='')



        if event == 'Pic1':
            dir_img1 = sg.popup_get_file('Choose your image 1',file_types=file_name_img,keep_on_top= True)
            if dir_img1 not in ('',None):
                pic1 = Image.open(dir_img1)
                img1_resize = pic1.resize((image_width_display,image_height_display))
                imgbytes1 = ImageTk.PhotoImage(img1_resize)
                window['image1'].update(data= imgbytes1)
                window['Detect1'].update(disabled= False)      
                window['Detect2'].update(disabled= False)         


        if event == 'Change_1':


            #list_variable = [[0]*11]*len(model1.names)
            list_variable = [[0]*12 for i in range(len(model1.names))]

            for i,item in enumerate(range(len(model1.names))):
                list_variable[i][0] = model1.names[i]

                list_variable[i][1] = values[f'{model1.names[item]}_1']
                list_variable[i][2] = values[f'{model1.names[item]}_OK_1'] 
                list_variable[i][3] = values[f'{model1.names[item]}_Num_1'] 
                list_variable[i][4] = values[f'{model1.names[item]}_NG_1'] 
                list_variable[i][5] = values[f'{model1.names[item]}_Wn_1'] 
                list_variable[i][6] = values[f'{model1.names[item]}_Wx_1'] 
                list_variable[i][7] = values[f'{model1.names[item]}_Hn_1'] 
                list_variable[i][8] = values[f'{model1.names[item]}_Hx_1'] 
                list_variable[i][9] = values[f'{model1.names[item]}_PLC_1'] 
                list_variable[i][10] = values['OK_PLC_1']
                list_variable[i][11] = values[f'{model1.names[item]}_Conf_1'] 

    


            mypath1 = values['file_weights1']
            model1= torch.hub.load('./levu','custom',path=mypath1,source='local',force_reload=False)
            if mypath1[-7:] == 'edit.pt': 
                change_label(model1)
            mychoose = values['choose_model']
            weight1 = values['file_weights1']
            conf_thres1 = values['conf_thres1'] 

            OK_Cam1 = values['have_save_OK_1']

            NG_Cam1 = values['have_save_NG_1']
            
            Folder_OK_Cam1 = values['save_OK_1']
    
            Folder_NG_Cam1 = values['save_NG_1']




            window.close() 
            window = make_window(theme)

            window['choose_model'].update(value=mychoose)
            window['file_weights1'].update(value=weight1)
            window['conf_thres1'].update(value=conf_thres1)

            window['have_save_OK_1'].update(value=OK_Cam1)
            window['have_save_NG_1'].update(value=NG_Cam1)


            window['save_OK_1'].update(value=Folder_OK_Cam1)
    
            window['save_NG_1'].update(value=Folder_NG_Cam1)


            window['choose_model'].update(value=mychoose)



            for i, item in enumerate(range(len(model1.names))):
                for name_label in model1.names:
                    if len(model1.names) <= len(list_variable):
                        if name_label == list_variable[i][0]:

                            window[f'{model1.names[item]}_1'].update(value= list_variable[i][1])
                            window[f'{model1.names[item]}_OK_1'].update(value= list_variable[i][2])
                            window[f'{model1.names[item]}_Num_1'].update(value= list_variable[i][3])
                            window[f'{model1.names[item]}_NG_1'].update(value= list_variable[i][4])
                            window[f'{model1.names[item]}_Wn_1'].update(value= list_variable[i][5])
                            window[f'{model1.names[item]}_Wx_1'].update(value= list_variable[i][6])
                            window[f'{model1.names[item]}_Hn_1'].update(value= list_variable[i][7])
                            window[f'{model1.names[item]}_Hx_1'].update(value= list_variable[i][8])
                            window[f'{model1.names[item]}_PLC_1'].update(value= list_variable[i][9])
                            window['OK_PLC_1'].update(value= list_variable[i][10])
                            window[f'{model1.names[item]}_Conf_1'].update(value= list_variable[i][11])




        if event == 'Change_2':
            list_variable = [[0]*12 for i in range(len(model2.names))]

            for i,item in enumerate(range(len(model2.names))):
                list_variable[i][0] = model2.names[i]

                list_variable[i][1] = values[f'{model2.names[item]}_2']
                list_variable[i][2] = values[f'{model2.names[item]}_OK_2'] 
                list_variable[i][3] = values[f'{model2.names[item]}_Num_2'] 
                list_variable[i][4] = values[f'{model2.names[item]}_NG_2'] 
                list_variable[i][5] = values[f'{model2.names[item]}_Wn_2'] 
                list_variable[i][6] = values[f'{model2.names[item]}_Wx_2'] 
                list_variable[i][7] = values[f'{model2.names[item]}_Hn_2'] 
                list_variable[i][8] = values[f'{model2.names[item]}_Hx_2'] 
                list_variable[i][9] = values[f'{model2.names[item]}_PLC_2'] 
                list_variable[i][10] = values['OK_PLC_2']
                list_variable[i][11] = values[f'{model2.names[item]}_Conf_2'] 

            mypath2 = values['file_weights2']
            model2= torch.hub.load('./levu','custom',path=mypath2,source='local',force_reload=False)
            mychoose = values['choose_model']
            weight1 = values['file_weights1']
            conf_thres1 = values['conf_thres1'] 
            weight2 = values['file_weights2']
            conf_thres2 = values['conf_thres2'] 

            OK_Cam1 = values['have_save_OK_1']
            OK_Cam2 = values['have_save_OK_2']
            NG_Cam1 = values['have_save_NG_1']
            NG_Cam2 = values['have_save_NG_2']
            Folder_OK_Cam1 = values['save_OK_1']
            Folder_OK_Cam2 = values['save_OK_2']
            Folder_NG_Cam1 = values['save_NG_1']
            Folder_NG_Cam2 = values['save_NG_2']


       


            window.close() 
            window = make_window(theme)

            window['choose_model'].update(value=mychoose)
            window['file_weights1'].update(value=weight1)
            window['conf_thres1'].update(value=conf_thres1)
            window['file_weights2'].update(value=weight2)
            window['conf_thres2'].update(value=conf_thres2)

            window['have_save_OK_1'].update(value=OK_Cam1)
            window['have_save_OK_2'].update(value=OK_Cam2)
            window['have_save_NG_1'].update(value=NG_Cam1)
            window['have_save_NG_2'].update(value=NG_Cam2)

            window['save_OK_1'].update(value=Folder_OK_Cam1)
            window['save_OK_2'].update(value=Folder_OK_Cam2)
            window['save_NG_1'].update(value=Folder_NG_Cam1)
            window['save_NG_2'].update(value=Folder_NG_Cam2)


            window['choose_model'].update(value=mychoose)

            for i, item in enumerate(range(len(model2.names))):
                for name_label in model2.names:
                    if len(model2.names) <= len(list_variable):
                        if name_label == list_variable[i][0]:

                            window[f'{model2.names[item]}_2'].update(value= list_variable[i][1])
                            window[f'{model2.names[item]}_OK_2'].update(value= list_variable[i][2])
                            window[f'{model2.names[item]}_Num_2'].update(value= list_variable[i][3])
                            window[f'{model2.names[item]}_NG_2'].update(value= list_variable[i][4])
                            window[f'{model2.names[item]}_Wn_2'].update(value= list_variable[i][5])
                            window[f'{model2.names[item]}_Wx_2'].update(value= list_variable[i][6])
                            window[f'{model2.names[item]}_Hn_2'].update(value= list_variable[i][7])
                            window[f'{model2.names[item]}_Hx_2'].update(value= list_variable[i][8])
                            window[f'{model2.names[item]}_PLC_2'].update(value= list_variable[i][9])
                            window['OK_PLC_2'].update(value= list_variable[i][10])
                            window[f'{model2.names[item]}_Conf_2'].update(value= list_variable[i][11])


        if event == 'Change3' or event == 'Change_3':
            list_variable = [[0]*12 for i in range(len(model3.names))]

            for i,item in enumerate(range(len(model3.names))):
                list_variable[i][0] = model3.names[i]

                list_variable[i][1] = values[f'{model3.names[item]}_3']
                list_variable[i][2] = values[f'{model3.names[item]}_OK_3'] 
                list_variable[i][3] = values[f'{model3.names[item]}_Num_3'] 
                list_variable[i][4] = values[f'{model3.names[item]}_NG_3'] 
                list_variable[i][5] = values[f'{model3.names[item]}_Wn_3'] 
                list_variable[i][6] = values[f'{model3.names[item]}_Wx_3'] 
                list_variable[i][7] = values[f'{model3.names[item]}_Hn_3'] 
                list_variable[i][8] = values[f'{model3.names[item]}_Hx_3'] 
                list_variable[i][9] = values[f'{model3.names[item]}_PLC_3'] 
                list_variable[i][10] = values['OK_PLC_3']
                list_variable[i][11] = values[f'{model3.names[item]}_Conf_3'] 
            mypath3 = values['file_weights3']
            model3= torch.hub.load('./levu','custom',path=mypath3,source='local',force_reload=False)
            mychoose = values['choose_model']
            weight1 = values['file_weights1']
            conf_thres1 = values['conf_thres1'] 
            weight2 = values['file_weights2']
            conf_thres2 = values['conf_thres2'] 


            weight3 = values['file_weights3']
            conf_thres3 = values['conf_thres3'] 




            window.close() 
            window = make_window(theme)

            window['choose_model'].update(value=mychoose)
            window['file_weights1'].update(value=weight1)
            window['conf_thres1'].update(value=conf_thres1)
            window['file_weights2'].update(value=weight2)
            window['conf_thres2'].update(value=conf_thres2)

            window['file_weights3'].update(value=weight3)
            window['conf_thres3'].update(value=conf_thres3)
   



            for i, item in enumerate(range(len(model3.names))):
                for name_label in model3.names:
                    if len(model3.names) <= len(list_variable):
                        if name_label == list_variable[i][0]:

                            window[f'{model3.names[item]}_3'].update(value= list_variable[i][1])
                            window[f'{model3.names[item]}_OK_3'].update(value= list_variable[i][2])
                            window[f'{model3.names[item]}_Num_3'].update(value= list_variable[i][3])
                            window[f'{model3.names[item]}_NG_3'].update(value= list_variable[i][4])
                            window[f'{model3.names[item]}_Wn_3'].update(value= list_variable[i][5])
                            window[f'{model3.names[item]}_Wx_3'].update(value= list_variable[i][6])
                            window[f'{model3.names[item]}_Hn_3'].update(value= list_variable[i][7])
                            window[f'{model3.names[item]}_Hx_3'].update(value= list_variable[i][8])
                            window[f'{model3.names[item]}_PLC_3'].update(value= list_variable[i][9])
                            window['OK_PLC_3'].update(value= list_variable[i][10])
                            window[f'{model3.names[item]}_Conf_3'].update(value= list_variable[i][11])




            # window.close() 
            # window = make_window(theme)

            # window['choose_model'].update(value=mychoose)
            # window['file_weights1'].update(value=weight1)
            # window['conf_thres1'].update(value=conf_thres1)


            # window['have_save_OK_1'].update(value=OK_Cam1)
      
            # window['have_save_NG_1'].update(value=NG_Cam1)
      

            # window['save_OK_1'].update(value=Folder_OK_Cam1)

            # window['save_NG_1'].update(value=Folder_NG_Cam1)


            # window['choose_model'].update(value=mychoose)






        if event == 'Detect1':
            print('CAM 1 DETECT')
            t1 = time.time()
            try:
            
                result1 = model1(pic1,size= 416,conf = values['conf_thres1']/100)

                table1 = result1.pandas().xyxy[0]
                print(table1)
                area_remove1 = []

                myresult1 =0 

                for item in range(len(table1.index)):
                    width1 = table1['xmax'][item] - table1['xmin'][item]
                    height1 = table1['ymax'][item] - table1['ymin'][item]
                    conf1 = table1['confidence'][item] * 100

                    #area1 = width1*height1
                    label_name = table1['name'][item]
                    for i1 in range(len(model1.names)):
                        if values[f'{model1.names[i1]}_1'] == True:
                            #if values[f'{model1.names[i1]}_WH'] == True:
                            if label_name == model1.names[i1]:
                                if width1 < int(values[f'{model1.names[i1]}_Wn_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif width1 > int(values[f'{model1.names[i1]}_Wx_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif height1 < int(values[f'{model1.names[i1]}_Hn_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif height1 > int(values[f'{model1.names[i1]}_Hx_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif conf1 < int(values[f'{model1.names[i1]}_Conf_1']):
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)   
                        if values[f'{model1.names[i1]}_1'] == False:
                            if label_name == model1.names[i1]:
                                table1.drop(item, axis=0, inplace=True)
                                area_remove1.append(item)

                names1 = list(table1['name'])

                show1 = np.squeeze(result1.render(area_remove1))
                show1 = cv2.resize(show1, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
                show1 = cv2.cvtColor(show1, cv2.COLOR_BGR2RGB)
                #ta = time.time()
                for i1 in range(len(model1.names)):
                    if values[f'{model1.names[i1]}_1'] == True:
                        if values[f'{model1.names[i1]}_OK_1'] == True:
                            len_name1 = 0
                            for name1 in names1:
                                if name1 == model1.names[i1]:
                                    len_name1 +=1
                            if len_name1 != int(values[f'{model1.names[i1]}_Num_1']):
                                print('NG')
                                cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                window['result_cam1'].update(value= 'NG', text_color='red')
                                myresult1 = 1
                                break

                        if values[f'{model1.names[i1]}_NG_1'] == True:
                            if model1.names[i1] in names1:
                                print('NG')
                                cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                window['result_cam1'].update(value= 'NG', text_color='red')    
                                myresult1 = 1         
                                break    

                if myresult1 == 0:
                    print('OK')
                    cv2.putText(show1, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                    window['result_cam1'].update(value= 'OK', text_color='green')

                imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
                window['image1'].update(data= imgbytes1)

            
            except:
                print(traceback.format_exc())
                sg.popup_annoying("Don't have image or parameter wrong", font=('Helvetica',14),text_color='red')
            
            t2 = time.time() - t1
            print(t2)
            time_cam1 = str(int(t2*1000)) + 'ms'
            window['time_cam1'].update(value= time_cam1, text_color='black') 
            print('---------------------------------------------') 


            
        if event == 'Detect2':
            print('CAM 2 DETECT')
            t1 = time.time()
            try:
                result2 = model2(pic1,size= 416,conf = values['conf_thres2']/100)

                table2 = result2.pandas().xyxy[0]

                area_remove2 = []

                myresult2 =0 

                for item in range(len(table2.index)):
                    width2 = table2['xmax'][item] - table2['xmin'][item]
                    height2 = table2['ymax'][item] - table2['ymin'][item]
                    conf2 = table2['confidence'][item] * 100

                    #area2 = width2*height2
                    label_name = table2['name'][item]
                    for i2 in range(len(model2.names)):
                        if values[f'{model2.names[i2]}_2'] == True:
                            #if values[f'{model2.names[i2]}_WH'] == True:
                            if label_name == model2.names[i2]:
                                if width2 < int(values[f'{model2.names[i2]}_Wn_2']): 
                                    table2.drop(item, axis=0, inplace=True)
                                    area_remove2.append(item)
                                elif width2 > int(values[f'{model2.names[i2]}_Wx_2']): 
                                    table2.drop(item, axis=0, inplace=True)
                                    area_remove2.append(item)
                                elif height2 < int(values[f'{model2.names[i2]}_Hn_2']): 
                                    table2.drop(item, axis=0, inplace=True)
                                    area_remove2.append(item)
                                elif height2 > int(values[f'{model2.names[i2]}_Hx_2']): 
                                    table2.drop(item, axis=0, inplace=True)
                                    area_remove2.append(item)
                                elif conf2 < int(values[f'{model2.names[i2]}_Conf_2']):
                                    table2.drop(item, axis=0, inplace=True)
                                    area_remove2.append(item)  

                        if values[f'{model2.names[i2]}_2'] == False:
                            if label_name == model2.names[i2]:
                                table2.drop(item, axis=0, inplace=True)
                                area_remove2.append(item)

                names2 = list(table2['name'])

                show2 = np.squeeze(result2.render(area_remove2))
                show2 = cv2.resize(show2, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
                show2 = cv2.cvtColor(show2, cv2.COLOR_BGR2RGB)
                #ta = time.time()
                for i2 in range(len(model2.names)):
                    if values[f'{model2.names[i2]}_OK_2'] == True:
                        len_name2 = 0
                        for name2 in names2:
                            if name2 == model2.names[i2]:
                                len_name2 +=1
                        if len_name2 != int(values[f'{model2.names[i2]}_Num_2']):
                            print('NG')
                            cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                            window['result_cam1'].update(value= 'NG', text_color='red')
                            myresult2 = 1
                            break

                    if values[f'{model2.names[i2]}_NG_2'] == True:
                        if model2.names[i2] in names2:
                            print('NG')
                            cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                            window['result_cam2'].update(value= 'NG', text_color='red')    
                            myresult2 = 1      
                            break    

                if myresult2 == 0:
                    print('OK')
                    cv2.putText(show2, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                    window['result_cam1'].update(value= 'OK', text_color='green')

                imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()
                window['image1'].update(data= imgbytes2)

            
            except:
                print(traceback.format_exc())
                sg.popup_annoying("Don't have image or parameter wrong", font=('Helvetica',24),text_color='red')
            
            t2 = time.time() - t1
            print(t2)
            time_cam2 = str(int(t2*1000)) + 'ms'
            window['time_cam1'].update(value= time_cam2, text_color='black') 
            print('---------------------------------------------') 




    window.close() 

except Exception as e:
    print(traceback.print_exc())
    str_error = str(e)
    sg.popup(str_error,font=('Helvetica',15), text_color='red',keep_on_top= True)
              