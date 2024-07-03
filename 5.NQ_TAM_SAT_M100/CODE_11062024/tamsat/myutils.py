import os
import glob

import cv2
import torch
import numpy as np 
import time

import PySimpleGUI as sg

from PIL import Image,ImageTk
import datetime 
import shutil
import sqlite3
from PIL import Image
from yaml import load
import keyboard


import traceback
import socket
import math
import threading

import openpyxl
from openpyxl.styles import Alignment
from openpyxl import Workbook
from openpyxl.styles import Font

from datetime import date
import datetime

from udp import UDPFinsConnection
from initialization import FinsPLCMemoryAreas

import ctypes
import csv

import stapipy as st

# user32 = ctypes.windll.user32
# screen_width = user32.GetSystemMetrics(0)
# screen_height = user32.GetSystemMetrics(1)

image_width_display = int(650*1)
image_height_display = int(410*1)


result_width_display = int(450)
result_height_display = int(100)

SCALE_X_CAM1 = 1#1280/2048
SCALE_Y_CAM1 = 1#960/1536


SCALE_X_CAM2 = 1#640/1440
SCALE_Y_CAM2 = 1#480/1080
soc = socket.socket(socket.AF_INET, socket.SOCK_STREAM)


import math


class CMyCallback:
    """
    Class that contains a callback function.
    """

    def __init__(self):
        self._image = None
        self._lock = threading.Lock()

    @property
    def image(self):
        """Property: return PyIStImage of the grabbed image."""
        duplicate = None
        self._lock.acquire()
        if self._image is not None:
            duplicate = self._image.copy()
        self._lock.release()
        return duplicate

    def datastream_callback1(self, handle=None, context=None):
        """
        Callback to handle events from DataStream.

        :param handle: handle that `trigger` the callback.
        :param context: user data passed on during callback registration.
        """
        st_datastream = handle.module
        if st_datastream:
            with st_datastream.retrieve_buffer() as st_buffer:
                # Check if the acquired data contains image data.
                if st_buffer.info.is_image_present:
                    # Create an image object.
                    st_image = st_buffer.get_image()

                    # Check the pixelformat of the input image.
                    pixel_format = st_image.pixel_format
                    pixel_format_info = st.get_pixel_format_info(pixel_format)

                    # Only mono or bayer is processed.
                    if not(pixel_format_info.is_mono or pixel_format_info.is_bayer):
                        return

                    # Get raw image data.
                    data = st_image.get_image_data()

                    # Perform pixel value scaling if each pixel component is
                    # larger than 8bit. Example: 10bit Bayer/Mono, 12bit, etc.
                    if pixel_format_info.each_component_total_bit_count > 8:
                        nparr = np.frombuffer(data, np.uint16)
                        division = pow(2, pixel_format_info
                                       .each_component_valid_bit_count - 8)
                        nparr = (nparr / division).astype('uint8')
                    else:
                        nparr = np.frombuffer(data, np.uint8)

                    # Process image for display.
                    nparr = nparr.reshape(st_image.height, st_image.width, 1)

                    # Perform color conversion for Bayer.
                    if pixel_format_info.is_bayer:
                        bayer_type = pixel_format_info.get_pixel_color_filter()
                        if bayer_type == st.EStPixelColorFilter.BayerRG:
                            nparr = cv2.cvtColor(nparr, cv2.COLOR_BAYER_RG2RGB)
                        elif bayer_type == st.EStPixelColorFilter.BayerGR:
                            nparr = cv2.cvtColor(nparr, cv2.COLOR_BAYER_GR2RGB)
                        elif bayer_type == st.EStPixelColorFilter.BayerGB:
                            nparr = cv2.cvtColor(nparr, cv2.COLOR_BAYER_GB2RGB)
                        elif bayer_type == st.EStPixelColorFilter.BayerBG:
                            nparr = cv2.cvtColor(nparr, cv2.COLOR_BAYER_BG2RGB)

                    # Resize image and store to self._image.
                    nparr = cv2.resize(nparr, None,
                                       fx=SCALE_X_CAM1,
                                       fy=SCALE_Y_CAM1)
                    self._lock.acquire()
                    self._image = nparr
                    self._lock.release()


    def datastream_callback2(self, handle=None, context=None):
        """
        Callback to handle events from DataStream.

        :param handle: handle that trigger the callback.
        :param context: user data passed on during callback registration.
        """
        st_datastream = handle.module
        if st_datastream:
            with st_datastream.retrieve_buffer() as st_buffer:
                # Check if the acquired data contains image data.
                if st_buffer.info.is_image_present:
                    # Create an image object.
                    st_image = st_buffer.get_image()

                    # Check the pixelformat of the input image.
                    pixel_format = st_image.pixel_format
                    pixel_format_info = st.get_pixel_format_info(pixel_format)

                    # Only mono or bayer is processed.
                    if not(pixel_format_info.is_mono or pixel_format_info.is_bayer):
                        return

                    # Get raw image data.
                    data = st_image.get_image_data()

                    # Perform pixel value scaling if each pixel component is
                    # larger than 8bit. Example: 10bit Bayer/Mono, 12bit, etc.
                    if pixel_format_info.each_component_total_bit_count > 8:
                        nparr = np.frombuffer(data, np.uint16)
                        division = pow(2, pixel_format_info
                                       .each_component_valid_bit_count - 8)
                        nparr = (nparr / division).astype('uint8')
                    else:
                        nparr = np.frombuffer(data, np.uint8)

                    # Process image for display.
                    nparr = nparr.reshape(st_image.height, st_image.width, 1)

                    # Perform color conversion for Bayer.
                    if pixel_format_info.is_bayer:
                        bayer_type = pixel_format_info.get_pixel_color_filter()
                        if bayer_type == st.EStPixelColorFilter.BayerRG:
                            nparr = cv2.cvtColor(nparr, cv2.COLOR_BAYER_RG2RGB)
                        elif bayer_type == st.EStPixelColorFilter.BayerGR:
                            nparr = cv2.cvtColor(nparr, cv2.COLOR_BAYER_GR2RGB)
                        elif bayer_type == st.EStPixelColorFilter.BayerGB:
                            nparr = cv2.cvtColor(nparr, cv2.COLOR_BAYER_GB2RGB)
                        elif bayer_type == st.EStPixelColorFilter.BayerBG:
                            nparr = cv2.cvtColor(nparr, cv2.COLOR_BAYER_BG2RGB)

                    # Resize image and store to self._image.
                    nparr = cv2.resize(nparr, None,
                                       fx=SCALE_X_CAM2,
                                       fy=SCALE_Y_CAM2)
                    self._lock.acquire()
                    self._image = nparr
                    self._lock.release()


def set_enumeration(nodemap, enum_name, entry_name):
    enum_node = st.PyIEnumeration(nodemap.get_node(enum_name))
    entry_node = st.PyIEnumEntry(enum_node[entry_name])
    enum_node.set_entry_value(entry_node)



def setup_camera_stc(num_camera,st_system,cb_func,window):

    # try:
    st_device = st_system.create_first_device()
    print(f'Device{num_camera}=', st_device.info.display_name)
    st_datastream = st_device.create_datastream()
    callback = st_datastream.register_callback(cb_func)
    st_datastream.start_acquisition()
    st_device.acquisition_start()
    remote_nodemap = st_device.remote_port.nodemap
    set_enumeration(remote_nodemap,"TriggerMode", "Off")
    # set_enumeration(remote_nodemap,"ExposureTime", "10000")

    # error_cam = False
    return  st_datastream, st_device,remote_nodemap

    # except Exception as exception:
    #     print(f' Error Cam {num_camera}:', exception)
    #     str_error = "Error"
    #     window[f'result_cam{num_camera}'].update(value= str_error, text_color='red',)



# def setup_camera1_stc(st_system,window):
#     #lobal error_cam1
#     #while error_cam1 == True:
#     try:
#         st_device1 = st_system.create_first_device()
#         print('Device1=', st_device1.info.display_name)
#         st_datastream1 = st_device1.create_datastream()
#         callback1 = st_datastream1.register_callback(cb_func1)
#         st_datastream1.start_acquisition()
#         st_device1.acquisition_start()
#         remote_nodemap = st_device1.remote_port.nodemap
#         set_enumeration(remote_nodemap,"TriggerMode", "Off")
#         error_cam1 = False
#         return  st_datastream1, st_device1,remote_nodemap

#     except Exception as exception:
#         print(' Error Cam 1:', exception)
#         str_error = "Error"
#         window['result_cam1'].update(value= str_error, text_color='red',)



# def setup_camera2_stc(st_system,window):
#     #global error_cam2
#     #while error_cam2 == True:
#     try:
#         st_device2 = st_system.create_first_device()
#         print('Device2=', st_device2.info.display_name)
#         st_datastream2 = st_device2.create_datastream()
#         callback2 = st_datastream2.register_callback(cb_func2)
#         st_datastream2.start_acquisition()
#         st_device2.acquisition_start()
#         remote_nodemap2 = st_device2.remote_port.nodemap
#         set_enumeration(remote_nodemap2,"TriggerMode", "Off")
#         error_cam2 = False
#         return  st_datastream2, st_device2,remote_nodemap2
#     except Exception as exception:     
#         print('Error Cam 2:', exception)
#         str_error = "Error"
#         #sg.popup(str_error,font=('Helvetica',15), text_color='red',keep_on_top= True)
#         window['result_cam2'].update(value= str_error, text_color='red')




def config_off_auto(remote_nodemap):
    # Configure the ExposureMode
    node_name_eps = "ExposureMode"

    node_eps = remote_nodemap.get_node(node_name_eps)

    if not node_eps.is_writable:
        print('not node ExposureMode')
    enum_node_eps = st.PyIEnumeration(node_eps)
    enum_entries_eps = enum_node_eps.entries
    selection_eps = 1

    if selection_eps < len(enum_entries_eps):
        enum_entry_eps = enum_entries_eps[selection_eps]
        enum_node_eps.set_int_value(enum_entry_eps.value)


    #Configure the BalanceWhiteAuto
    node_name_bwa = "BalanceWhiteAuto"

    node_bwa = remote_nodemap.get_node(node_name_bwa)

    if not node_bwa.is_writable:
        print('not node BalanceWhiteAuto')
    enum_node_bwa = st.PyIEnumeration(node_bwa)
    enum_entries_bwa = enum_node_bwa.entries
    selection_bwa = 0

    if selection_bwa < len(enum_entries_bwa):
        enum_entry_bwa = enum_entries_bwa[selection_bwa]
        enum_node_bwa.set_int_value(enum_entry_bwa.value)
        



def BalanceWhiteAuto(remote_nodemap,choose_value_br,index):

    enum_name_brs = "BalanceRatioSelector"
    numeric_name_br = "BalanceRatio"
    node_brs = remote_nodemap.get_node(enum_name_brs)
    if not node_brs.is_writable:
        print('not node '+ enum_name_brs)

    enum_node_brs = st.PyIEnumeration(node_brs)
    enum_entries_brs = enum_node_brs.entries

    enum_entry_brs = enum_entries_brs[index]
    if enum_entry_brs.is_available:
        enum_node_brs.value = enum_entry_brs.value
        #print(st.PyIEnumEntry(enum_entry).symbolic_value)
        node_name_br = numeric_name_br
        node_br = remote_nodemap.get_node(node_name_br)

        if not node_br.is_writable:
            # print('not node '+ name)
            pass
        else:
            if node_br.principal_interface_type == st.EGCInterfaceType.IFloat:
                node_value_br = st.PyIFloat(node_br)
            elif node_br.principal_interface_type == st.EGCInterfaceType.IInteger:
                node_value_br = st.PyIInteger(node_br)
            value_br = choose_value_br

            if node_br.principal_interface_type == st.EGCInterfaceType.IFloat:
                value_br = float(value_br)
                pass
            else:
                value_br = int(value_br)
                pass
            if node_value_br.min <= value_br <= node_value_br.max:
                node_value_br.value = value_br


def set_exposure_or_gain(remote_nodemap,node_name,value):
    if remote_nodemap.get_node(node_name):
        node = remote_nodemap.get_node(node_name)
        if not node.is_writable:
            print('not node' + node_name)
        else:
            if node.principal_interface_type == st.EGCInterfaceType.IFloat:
                node_value = st.PyIFloat(node)
            elif node.principal_interface_type == st.EGCInterfaceType.IInteger:
                node_value = st.PyIInteger(node)
            value = float(value)
            if node.principal_interface_type == st.EGCInterfaceType.IFloat:
                value = float(value)
            else:
                value = int(value)
            if node_value.min <= value <= node_value.max:
                node_value.value = value
            


def set_balance_white_auto(remote_nodemap,index,value):
    enum_name_brs = "BalanceRatioSelector"
    numeric_name_br = "BalanceRatio"
    node_brs = remote_nodemap.get_node(enum_name_brs)
    if not node_brs.is_writable:
        print('not node  '+ enum_name_brs)

    enum_node_brs = st.PyIEnumeration(node_brs)
    enum_entries_brs = enum_node_brs.entries

    enum_entry_brs = enum_entries_brs[index]
    if enum_entry_brs.is_available:
        enum_node_brs.value = enum_entry_brs.value

        node_name_br = numeric_name_br
        node_br = remote_nodemap.get_node(node_name_br)

        if not node_br.is_writable:
            print('not node '+ numeric_name_br + index)
        else:
            if node_br.principal_interface_type == st.EGCInterfaceType.IFloat:
                node_value_br = st.PyIFloat(node_br)
            elif node_br.principal_interface_type == st.EGCInterfaceType.IInteger:
                node_value_br = st.PyIInteger(node_br)
            value_br = value

            if node_br.principal_interface_type == st.EGCInterfaceType.IFloat:
                value_br = float(value_br)
            else:
                value_br = int(value_br)
            if node_value_br.min <= value_br <= node_value_br.max:
                node_value_br.value = value_br
            print(st.PyIEnumEntry(enum_entry_brs).symbolic_value + ' : ' + str(node_value_br.value))

def set_config_init_camera(remote_nodemap,value_exposure_time, value_gain,select_balance, value_balance_red,value_balance_green, value_balance_blue):

    set_exposure_or_gain(remote_nodemap,"ExposureTime",value_exposure_time)
    set_exposure_or_gain(remote_nodemap,"ExposureTimeRaw",value_exposure_time)

    set_exposure_or_gain(remote_nodemap,"Gain",value_gain)
    set_exposure_or_gain(remote_nodemap,"GainRaw",value_gain)
    if select_balance == 'off':
        pass
    else:
        set_balance_white_auto(remote_nodemap,0,value_balance_red)
        set_balance_white_auto(remote_nodemap,3,value_balance_green)
        set_balance_white_auto(remote_nodemap,4,value_balance_blue)





def phuong_trinh_duong_thang(diem_A, diem_B):
    delta_x = diem_B[0] - diem_A[0]
    delta_y = diem_B[1] - diem_A[1]

    slope = delta_y / delta_x
    c = diem_A[1] - slope * diem_A[0]

    A = -slope
    B = 1
    C = -c

    return A, B, C

def khoang_cach_toi_duong_thang(A, B, C, diem_C):
    x0, y0 = diem_C[0], diem_C[1]
    khoang_cach = abs(A * x0 + B * y0 + C) / math.sqrt(A ** 2 + B ** 2)
    return khoang_cach


def phuong_trinh_duong_thang_vuong_goc(A, B, diem_C):
    # Kiểm tra điều kiện A khác 0 để tính hệ số góc slope
    if A != 0:
        C_vuong_goc =  - diem_C[0] + A * diem_C[1]
    else:

        C_vuong_goc = - diem_C[0]

    return B, -A, C_vuong_goc



def giao_diem_hai_duong_thang(A1, B1, C1, A2, B2, C2):
    # Tìm điểm giao nhau D của hai đường thẳng
    x = (C2 * B1 - C1 * B2) / (A1 * B2 - A2 * B1)
    y = (C1 * A2 - C2 * A1) / (A1 * B2 - A2 * B1)
    giao_diem = [x,y]
    return giao_diem

def connect_plc_keyence(host, port):
    try:
        soc.connect((host, port))
        return True
    except OSError:
        print("Can't connect to PLC")
        time.sleep(3)
        print("Reconnecting....")
        return False

def run_plc_keyence(host, port):
    connected = False
    while connected == False:
        connected = connect_plc_keyence(host, port)
    print("connected") 

def read_plc_keyence(data):
    a = 'RD '
    c = '\x0D'
    d = a+ data +c
    datasend = d.encode("UTF-8")
    soc.sendall(datasend)
    data = soc.recv(1024)
    datadeco = data.decode("UTF-8")
    data1 = int(datadeco)

    return data1

#Write data
def write_plc_keyence(register, data):
    a = 'WR '
    b = ' '
    c = '\x0D'
    d = a+ register + b + str(data) + c
    datasend  = d.encode("UTF-8")
    soc.sendall(datasend)
    datares = soc.recv(1024)
    # print(datares)

def read_plc_omron(register):
    register = (register).to_bytes(2, byteorder='big') + b'\x00'
    read_var = fins_instance.memory_area_read(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register)
    read_var = int.from_bytes(read_var[-2:], byteorder='big')  
    return read_var


def write_plc_omron(register, data):
    register = (register).to_bytes(2, byteorder='big') + b'\x00'
    fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register,b'\x00\x00',data)


def remove_file(directory):

    try:
        if os.listdir(directory) != []:
            for i in glob.glob(directory+'*'):
                for j in glob.glob(i+'/*'):
                    os.remove(j)
                os.rmdir(i)
        print('already delete folder')
        
    except:
        pass



def connect_plc_omron(host):
    global fins_instance
    try:
        fins_instance = UDPFinsConnection()
        fins_instance.connect(host)
        fins_instance.dest_node_add=1
        fins_instance.srce_node_add=25

        return True
    except:
        print("Can't connect to PLC")
        for i in range(100000000):
            pass
        #sleep(3)
        print("Reconnecting....")
        return False



def run_plc_omron(host):
    connected = False
    while connected == False:
        connected = connect_plc_omron(host)
        print('connecting ....')
        #event, values = window.read(timeout=20)

    print("connected plc") 


time_to_name = lambda: str(datetime.datetime.now()).replace(':', '-').replace(' ', '_').replace('.', '-')

load_theme = lambda: [line.strip().split(':')[1] for line in open('static/theme.txt')]

load_choosemodel = lambda: next(line.strip().split('=')[1] for line in open('static/choose_model.txt'))

save_theme = lambda name_theme: open('static/theme.txt', 'w').write(f'theme:{name_theme}')

save_choosemodel = lambda name_model: open('static/choose_model.txt','w').write(f"choose_model={name_model}")

load_model = lambda i: [line.strip().split('=')[1] for line in open(f'static/model{i}.txt')][0]

save_model = lambda i, name_model: open(f'static/model{i}.txt', 'w').write(f'model{i}={name_model}')

str2bool = lambda v: v.lower() in ("yes", "true", "t", "1")

def load_kc(window):
    values_all = []
    with open('static/kc.txt','r') as lines:
        for line in lines:
            _, name_all = line.strip().split('=')
            values_all.append(name_all)
    window['kc_tren_t'].update(value=values_all[0])
    window['kc_tren_c'].update(value=values_all[1])
    window['kc_duoi_t'].update(value=values_all[2])
    window['kc_duoi_c'].update(value=values_all[3])


def save_kc(values):
    with open('static/kc.txt','w') as f:
        f.write('kc_tren_t' + '=' + str(values['kc_tren_t']))
        f.write('\n')
        f.write('kc_tren_c' + '=' + str(values['kc_tren_c']))
        f.write('\n')
        f.write('kc_duoi_t' + '=' + str(values['kc_duoi_t']))
        f.write('\n')
        f.write('kc_duoi_c' + '=' + str(values['kc_duoi_c']))
        f.write('\n')



def run_model(num_camera):
    mypath = load_model(num_camera)
    # device2 = torch.device("cuda:1")
    # model = torch.hub.load('./levu','custom', path= mypath, source='local',force_reload =False).to(device2)
    model = torch.hub.load('./levu','custom', path= mypath, source='local',force_reload =False)
    img_test = os.path.join(os.getcwd(), 'img/imgtest.jpg')
    result = model(img_test,416,0.25) 
    print(f'model {num_camera} already')
    return model


def load_all_sql(num_camera, choose_model,window, nums_camera):
    conn = sqlite3.connect(f'modeldb_{nums_camera}_PLC_conf_date.db')
    cursor = conn.execute("SELECT * FROM MYMODEL")

    for row in cursor:
        if row[0] != choose_model:
            continue
        row1_a, row1_b = row[1].strip().split('_')
        if row1_a != str(num_camera):
            continue
        if row1_b == '0':
            window[f'file_weights' + str(num_camera)].update(value=row[2])
            window[f'conf_thres' + str(num_camera)].update(value=row[3])
            window[f'choose_size' + str(num_camera)].update(value=row[-1])

            for i in range(nums_camera):
                window[f'have_save_OK_{i+1}'].update(value=str2bool(row[4+i]))
                window[f'have_save_NG_{i+1}'].update(value=str2bool(row[4+i+nums_camera]))
                window[f'save_OK_{i+1}'].update(value=row[i+4+nums_camera*2])
                window[f'save_NG_{i+1}'].update(value=row[i+4+nums_camera*3])

            model = torch.hub.load('./levu', 'custom', path=row[2], source='local', force_reload=False)

        item = int(row1_b)

        keys = ['_', '_OK_', '_Num_', '_NG_', '_Wn_', '_Wx_', '_Hn_', '_Hx_', '_PLC_', 'PLC_OK_', '_Conf_']

        for j, key in enumerate(keys):
            value = j + 8 + 4 * (nums_camera -1)

            if j in [0,1,3]:
                window[f'{model.names[item]}{key}{num_camera}'].update(value=str2bool(row[value]))
            elif j == 9:
                window[f'PLC_OK_' + str(num_camera)].update(value=str(row[value]))
            else:
                window[f'{model.names[item]}{key}{num_camera}'].update(value=str(row[value]))
    
    conn.close()


def save_all_sql(model,num_camera,choose_model, nums_camera, values):
    conn = sqlite3.connect(f'modeldb_{nums_camera}_PLC_conf_date.db')
    cursor = conn.execute("SELECT * from MYMODEL")
    update = 0 
    d_list = ['OK_Cam','NG_Cam','Folder_OK_Cam','Folder_NG_Cam']
    saves_list = ["have_save_OK_","have_save_NG_","save_OK_","save_NG_"]
    mystring = ''
    update_string = ''
    question = ''
    for d_l in d_list:
        for num in range(1, nums_camera+1):
            mystring += f"{d_l}{num}, "
            update_string += f"{d_l}{num} = ?, "
            question += "?,"
    update_string = update_string[:-2]
    for row in cursor:
        if row[0] == choose_model:            
            row1_a, _ = row[1].strip().split('_')
            if row1_a == str(num_camera):
                conn.execute("DELETE FROM MYMODEL WHERE (ChooseModel = ? AND Camera LIKE ?)", (choose_model,str(num_camera) + '%'))


                for item in range(len(model.names)):
                    values_list = []
                    values_list += [str(values['choose_model']),str(num_camera)+ '_' +str(item) ,str(values['file_weights' + str(num_camera)]), int(values['conf_thres' + str(num_camera)])]
                    for s_l in saves_list:
                        for num in range(1, nums_camera+1):
                            values_list.append(str(values[f'{s_l}{num}']))  
                    values_list += [str(values[f'{model.names[item]}_' + str(num_camera)]), str(values[f'{model.names[item]}_OK_' + str(num_camera)]), int(values[f'{model.names[item]}_Num_' + str(num_camera)]), str(values[f'{model.names[item]}_NG_' + str(num_camera)]), int(values[f'{model.names[item]}_Wn_' + str(num_camera)]), int(values[f'{model.names[item]}_Wx_' + str(num_camera)]), int(values[f'{model.names[item]}_Hn_' + str(num_camera)]), int(values[f'{model.names[item]}_Hx_' + str(num_camera)]), int(values[f'{model.names[item]}_PLC_' + str(num_camera)]), int(values['PLC_OK_' + str(num_camera)]), int(values[f'{model.names[item]}_Conf_' + str(num_camera)])]
                    values_list += [str(values['choose_size' + str(num_camera)])]
                    conn.execute("INSERT INTO MYMODEL (ChooseModel,Camera, Weights,Confidence," + mystring + "Joined,Ok,Num,NG,WidthMin, WidthMax,HeightMin,HeightMax,PLC_NG,PLC_OK,Conf, Size) \
                        VALUES (?,?,?,?," + question + "?,?,?,?,?,?,?,?,?,?,?,?)", tuple(values_list))           
                    update = 1
                break

    if update == 0:
        for item in range(len(model.names)):
            values_list = []
            values_list += [str(values['choose_model']),str(num_camera)+ '_' +str(item) ,str(values['file_weights' + str(num_camera)]), int(values['conf_thres' + str(num_camera)])]
            for s_l in saves_list:
                for num in range(1, nums_camera+1):
                    values_list.append(str(values[f'{s_l}{num}']))   
            values_list += [str(values[f'{model.names[item]}_' + str(num_camera)]), str(values[f'{model.names[item]}_OK_' + str(num_camera)]), int(values[f'{model.names[item]}_Num_' + str(num_camera)]), str(values[f'{model.names[item]}_NG_' + str(num_camera)]), int(values[f'{model.names[item]}_Wn_' + str(num_camera)]), int(values[f'{model.names[item]}_Wx_' + str(num_camera)]), int(values[f'{model.names[item]}_Hn_' + str(num_camera)]), int(values[f'{model.names[item]}_Hx_' + str(num_camera)]), int(values[f'{model.names[item]}_PLC_' + str(num_camera)]), int(values['PLC_OK_' + str(num_camera)]), int(values[f'{model.names[item]}_Conf_' + str(num_camera)])]
            values_list += [str(values['choose_size' + str(num_camera)])]
            
            conn.execute("INSERT INTO MYMODEL (ChooseModel,Camera, Weights,Confidence," + mystring + "Joined,Ok,Num,NG,WidthMin, WidthMax,HeightMin,HeightMax, PLC_NG,PLC_OK,Conf, Size) \
                VALUES (?,?,?,?," + question + "?,?,?,?,?,?,?,?,?,?,?,?)", tuple(values_list))
        
    for row in cursor:
        values_list = []
        for s_l in saves_list:
            for num in range(1, nums_camera+1):
                values_list.append(str(values[f'{s_l}{num}']))  
        values_list.append(choose_model)
        if row[0] == choose_model:
            conn.execute("UPDATE MYMODEL SET " + update_string + " WHERE ChooseModel = ? ",tuple(values_list))


    conn.commit()
    conn.close()



def collect_dict_date(num_camera,nums_camera):
    list_choose_model = []
    dict_date = {}
    conn = sqlite3.connect(f'modeldb_{nums_camera}_PLC_conf_date.db')
    # cursor = conn.execute("SELECT ChooseModel,Camera,Weights,Confidence,OK_Cam1,NG_Cam1,Folder_OK_Cam1,Folder_NG_Cam1,Joined,Ok,Num,NG,WidthMin,WidthMax,HeightMin,HeightMax,PLC_NG,PLC_OK,Conf,Date_save from DATESAVE")
    cursor = conn.execute("SELECT * from DATESAVE")


    for row in cursor:
        list_choose_model.append(row[0])
    list_choose_model = list(set(list_choose_model))
    
    cursor = conn.execute("SELECT * from DATESAVE")
    for choose in list_choose_model:
        dict_date[choose] = []
    for row in cursor:
        row1_a, row1_b = row[1].strip().split('_')

        for choose in list_choose_model:
            # print(row1_a ,str(num_camera))
            if choose == row[0] and row1_a == str(num_camera) and row1_b == '0':
                dict_date[choose].append(row[-1])
    return dict_date


def save_all_sql_date(values,model,num_camera, nums_camera):
    conn = sqlite3.connect(f'modeldb_{nums_camera}_PLC_conf_date.db')

    d_list = ['OK_Cam','NG_Cam','Folder_OK_Cam','Folder_NG_Cam']
    saves_list = ["have_save_OK_","have_save_NG_","save_OK_","save_NG_"]
    mystring = ''
    update_string = ''
    question = ''
    for d_l in d_list:
        for num in range(1, nums_camera+1):
            mystring += f"{d_l}{num}, "
            update_string += f"{d_l}{num} = ?, "
            question += "?,"


    for item in range(len(model.names)):
        values_list = []
        values_list += [str(values['choose_model']),str(num_camera)+ '_' +str(item) ,str(values['file_weights' + str(num_camera)]), int(values['conf_thres' + str(num_camera)])]
        for s_l in saves_list:
            for num in range(1, nums_camera+1):
                values_list.append(str(values[f'{s_l}{num}']))   
        values_list += [str(values[f'{model.names[item]}_' + str(num_camera)]), str(values[f'{model.names[item]}_OK_' + str(num_camera)]), int(values[f'{model.names[item]}_Num_' + str(num_camera)]), str(values[f'{model.names[item]}_NG_' + str(num_camera)]), int(values[f'{model.names[item]}_Wn_' + str(num_camera)]), int(values[f'{model.names[item]}_Wx_' + str(num_camera)]), int(values[f'{model.names[item]}_Hn_' + str(num_camera)]), int(values[f'{model.names[item]}_Hx_' + str(num_camera)]), int(values[f'{model.names[item]}_PLC_' + str(num_camera)]), int(values['PLC_OK_' + str(num_camera)]), int(values[f'{model.names[item]}_Conf_' + str(num_camera)])]
        values_list += [str(values['choose_size' + str(num_camera)]), str(time_to_name()[:-7])]

        conn.execute("INSERT INTO DATESAVE (ChooseModel,Camera, Weights,Confidence," + mystring + "Joined,Ok,Num,NG,WidthMin, WidthMax,HeightMin,HeightMax, PLC_NG,PLC_OK,Conf,Size,Date_save) \
            VALUES (?,?,?,?," + question + "?,?,?,?,?,?,?,?,?,?,?,?,?)", tuple(values_list))

    conn.commit()
    conn.close()



def read_plc(plc_name,register):
    if plc_name == 'k':
        read_var = read_plc_keyence(f'DM{register}') 
    elif plc_name == 'o':
        read_var = read_plc_omron(register)
    return read_var


def write_plc(plc_name,register,data):
    if plc_name == 'k':
        write_plc_keyence(f'DM{register}', data)
    elif plc_name == 'o':
        write_plc_omron(register, data)
    else:
        pass


def load_image(path, filename):
    img_orgin = cv2.imread(path)
    while type(img_orgin) == type(None):
        print(f'loading img {path}...')
        for path in glob.glob(filename + '/*'):
            img_orgin = cv2.imread(path)

    # print('CAM received image')
    return img_orgin              


def handle_image(img_orgin, model,size,conf,num_camera,values, kiem_guong=False,center_x_min = 0, center_x_max =0,center_y_min =0, center_y_max=0):
    if kiem_guong:
        center_x = 0
        center_y = 0

    img_orgin = cv2.cvtColor(img_orgin, cv2.COLOR_BGR2RGB)

    result = model(img_orgin,size= size,conf = conf) 
    table = result.pandas().xyxy[0]
    area_remove = []
    for item in range(len(table.index)):
        width = table['xmax'][item] - table['xmin'][item]
        height = table['ymax'][item] - table['ymin'][item]
        conf = table['confidence'][item] * 100
        if kiem_guong:

            if item == 0:
                center_x = int(float(table['xmin'][item]) + float(table['xmax'][item] - table['xmin'][item])/2)
                center_y = int(float(table['ymin'][item]) + float(table['ymax'][item] - table['ymin'][item])/2)

                

        label_name = table['name'][item]
        for i in range(len(model.names)):
            if values[f'{model.names[i]}_{num_camera}'] == True:

                if label_name == model.names[i]:

                    if (width < int(values[f'{label_name}_Wn_{num_camera}'])) \
                        or (width > int(values[f'{label_name}_Wx_{num_camera}'])) \
                        or (height < int(values[f'{label_name}_Hn_{num_camera}'])) \
                        or (height > int(values[f'{label_name}_Hx_{num_camera}'])) \
                        or (conf < int(values[f'{label_name}_Conf_{num_camera}'])):

                        table.drop(item, axis=0, inplace=True)
                        area_remove.append(item)


            if values[f'{model.names[i]}_{num_camera}'] == False:
                if label_name == model.names[i]:
                    table.drop(item, axis=0, inplace=True)
                    area_remove.append(item)

    names = list(table['name'])

    show = np.squeeze(result.render(area_remove))
    if kiem_guong == 1:
        if center_x_min < center_x < center_x_max:
            cv2.putText(show, f'Tam x: {center_x}',(50,150),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),2)
        else:
            cv2.putText(show, f'Tam x: {center_x}',(50,150),cv2.FONT_HERSHEY_COMPLEX, 1,(255,0,0),2)
        if center_y_min < center_y < center_y_max:
            cv2.putText(show, f'Tam y: {center_y}',(50,200),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),2)
        else:
            cv2.putText(show, f'Tam y: {center_y}',(50,200),cv2.FONT_HERSHEY_COMPLEX, 1,(255,0,0),2)

    show = cv2.resize(show, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
    show = cv2.cvtColor(show, cv2.COLOR_BGR2RGB)

    if kiem_guong == 1:

        print('x: ',center_x)
        print('y: ',center_y)
        return names, show, center_x, center_y
    return names, show



def show_and_save(window, result,show,num_camera,img_orgin,values,okng=True):
    global hang_muc_ng_1

    if num_camera == 31 or num_camera == 32:
        num_camera = 3
        
    if values[f'have_save_{result}_{num_camera}']:
        # name_folder_ng = time_to_name()
        # cv2.imwrite(values[f'save_{result}_{num_camera}']  + '/' + name_folder_ng + '.jpg',img_orgin)

        name_folder= time_to_name()
        today = datetime.date.today()
        if not os.path.isdir(values[f'save_{result}_{num_camera}']  + "/" + str(today)):
            os.mkdir(values[f'save_{result}_{num_camera}'] + "/" + str(today))
        cv2.imwrite(values[f'save_{result}_{num_camera}']  + "/" + str(today) + '/' + name_folder + '.jpg',img_orgin)


    # if num_camera ==3:
    #     num_camera =2

    if result == 'OK':
        if okng:
            cv2.putText(show, f'{result}',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
            window[f'result_cam{num_camera}'].update(value= f'{result}', text_color='green')
    else:
        if okng:
            cv2.putText(show, f'{result}',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
            k=1
            for i in hang_muc_ng_1:
                cv2.putText(show, i ,(5,30*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                k+=1
                hang_muc_ng_1 = []
            window[f'result_cam{num_camera}'].update(value= f'{result}', text_color='red')


    

bien_ket_qua_1 = []
bien_ket_qua_2 = []
hang_muc_ng_1 = []



def handle_result(window, model,names, show, num_camera, plc_name,img_orgin, values,okng=True):
    global bien_ket_qua_1
    global bien_ket_qua_2
    global hang_muc_ng_1


    myresult =0 
    kg = 0
    if num_camera == 31:
        num_camera = 3
        kg = 1
    if num_camera == 32:
        num_camera = 3
        kg = 2
    print(num_camera)
    for i in range(len(model.names)):
        register_ng = int(values[f'{model.names[i]}_PLC_{num_camera}'])
        if values[f'{model.names[i]}_OK_{num_camera}'] == True:
            len_name = names.count(model.names[i])

            if len_name != int(values[f'{model.names[i]}_Num_{num_camera}']):
                print(f'NG cam {num_camera}: {model.names[i]}')
                hang_muc_ng_1.append(model.names[i])
                if num_camera == 1:
                    bien_ket_qua_1.append(register_ng)
                if num_camera == 2:
                    bien_ket_qua_2.append(register_ng)
                  
                # if num_camera ==1:
                #     write_plc(plc_name,register_ng,1)
                # if plc_name:
                #     write_plc(plc_name,register_ng,1)
                # show_and_save(window, 'NG',show,num_camera,img_orgin,values)                                                                                                                                                                
                myresult = 1
            
        if values[f'{model.names[i]}_NG_{num_camera}'] == True:
            if model.names[i] in names:
                print(f'NG cam {num_camera}: {model.names[i]}')
                hang_muc_ng_1.append(model.names[i])
                # if num_camera ==1:
                #     write_plc(plc_name,register_ng,1)
                # bien_ket_qua.append(register_ng)
                if num_camera == 1:
                    bien_ket_qua_1.append(register_ng)
                if num_camera == 2:
                    bien_ket_qua_2.append(register_ng)
                # if plc_name:
                #     write_plc(plc_name,register_ng,1)
                # show_and_save(window, 'NG',show,num_camera,img_orgin,values)                                                                                                                                                                
                myresult = 1         
        
    if kg ==1:
        num_camera =31
    if kg ==2:
        num_camera =32
    if myresult == 0:
        print(f'OK cam {num_camera}')
        # if num_camera ==1:
        #     register_ok = int(values[f'PLC_OK_{num_camera}'])
        #     write_plc(plc_name,register_ok,1)
        # register_ok = str(values[f'PLC_OK_{num_camera}'])
    
        # if plc_name:
        #     write_plc(plc_name,register_ok,1)
        show_and_save(window, 'OK',show,num_camera,img_orgin,values,okng)           
    else:
        show_and_save(window, 'NG',show,num_camera,img_orgin,values,okng)  
    
    return myresult

loi_kc = 0
def program_camera_FH_c4(window,values,model,directory,plc_name,value_plc_trigger, value_plc_done): 
    global loi_kc
    num_camera = 4
    size = values[f'choose_size{num_camera}']
    conf = values[f'conf_thres{num_camera}']/100
    if read_plc(plc_name,value_plc_trigger) == 1:   
        if os.listdir(directory) == []:
            print(f'folder {directory} empty')
        else:

            print(f'received folder {directory}')

            for filename in glob.glob(directory + '*'):
                for path in glob.glob(filename + '/*'):
                    name = path[-18:]
                    if name == 'Input0_Camera0.jpg':
                        img_orgin = load_image(path,filename)   
                        t1 = time.time()

                        write_plc(plc_name,value_plc_trigger,0)

                        img_orgin1 = cv2.cvtColor(img_orgin, cv2.COLOR_BGR2RGB)     


                        result1 = model(img_orgin1,size= size,conf = conf)

                        table1 = result1.pandas().xyxy[0]
                        area_remove1 = []

                        myresult1 =0 

                        a1 =0
                        a2 =0
                        a3 =0
                        # print(table1)
                        for item in range(len(table1.index)):

                            width1 = table1['xmax'][item] - table1['xmin'][item]
                            height1 = table1['ymax'][item] - table1['ymin'][item]
                            #area1 = width1*height1
                            label_name = table1['name'][item]
                            conf1 = table1['confidence'][item] *100

                            for i1 in range(len(model.names)):
                                if values[f'{model.names[i1]}_4'] == True:
                                #if values[f'{model.names[i1]}_WH'] == True:
                                    if label_name == model.names[i1]:
                                        if width1 < int(values[f'{model.names[i1]}_Wn_4']): 
                                            table1.drop(item, axis=0, inplace=True)
                                            area_remove1.append(item)
                                        elif width1 > int(values[f'{model.names[i1]}_Wx_4']): 
                                            table1.drop(item, axis=0, inplace=True)
                                            area_remove1.append(item)
                                        elif height1 < int(values[f'{model.names[i1]}_Hn_4']): 
                                            table1.drop(item, axis=0, inplace=True)
                                            area_remove1.append(item)
                                        elif height1 > int(values[f'{model.names[i1]}_Hx_4']): 
                                            table1.drop(item, axis=0, inplace=True)
                                            area_remove1.append(item)
                                        elif conf1  < int(values[f'{model.names[i1]}_Conf_4']):
                                            table1.drop(item, axis=0, inplace=True)
                                            area_remove1.append(item)
                                        else:
                                        
                                            if str(table1['name'][item]) == '1':
                                                diem1_tt = [table1['xmin'][item] + width1/2  ,table1['ymin'][item] + height1/2]

                                                h1 = table1['ymin'][item] + (table1['ymax'][item] - table1['ymin'][item])/2  
                                                a1=1  

                                            if str(table1['name'][item]) == '2':
                                                diem2_min = [table1['xmin'][item],table1['ymin'][item]]
                                                diem2_max = [table1['xmax'][item],table1['ymax'][item]]
                                                h2 = table1['ymin'][item] + (table1['ymax'][item] - table1['ymin'][item])/2   
                                                a2=1

                                            if str(table1['name'][item]) == '3':
                                                diem3_min = [table1['xmin'][item],table1['ymin'][item]]
                                                diem3_max = [table1['xmax'][item],table1['ymax'][item]]
                                                h3 = table1['ymin'][item] + (table1['ymax'][item] - table1['ymin'][item])/2
                                                a3=1     


                                if values[f'{model.names[i1]}_4'] == False:
                                    if label_name == model.names[i1]:
                                        table1.drop(item, axis=0, inplace=True)
                                        area_remove1.append(item)


                        names1 = list(table1['name'])

                        show1 = np.squeeze(result1.render(area_remove1))

                        show1 = cv2.cvtColor(show1, cv2.COLOR_BGR2RGB)

                    #ta = time.time()
                        k = 1 
                        my_kc_t = 0
                        my_kc_d =0
                        loi_kc = 0

                        kc_tren_mm = 0
                        kc_duoi_mm = 0

                        if a1 == 1 and a2 ==1:
                            # my_kc_t = h1 - h2
                            # print('t: ',my_kc_t)
                            A, B, C = phuong_trinh_duong_thang(diem2_max, diem2_min)

                            my_kc_t = khoang_cach_toi_duong_thang(A, B, C, diem1_tt)

                            A1,B1,C1 = phuong_trinh_duong_thang_vuong_goc(A, B, diem1_tt)

                            giao_diem = giao_diem_hai_duong_thang(A, B, C ,A1, B1, C1)

                            mypixel = round(my_kc_t,3)
                            mymm = round(my_kc_t * 0.0048,3)
                            kc_tren_mm = str(mymm)
                            resultkc = f't: {mypixel} pixel = {mymm} mm'

                            if  my_kc_t < int(values['kc_tren_t']) or my_kc_t > int(values['kc_tren_c']):
                                if my_kc_t < int(values['kc_tren_t']):
                                    write_plc(plc_name,'2416',1)
                                else:
                                    write_plc(plc_name,'2417',1)


                                print('NG')
                                write_plc(plc_name,'2414',1)
                                loi_kc = 1
                                        
                                show1 = cv2.circle(show1,(int(diem1_tt[0]),int(diem1_tt[1])), 5,(0,0,255),-1)
                                show1 = cv2.line(show1, (int(diem2_max[0]),int(diem2_max[1])), (int(diem2_min[0]),int(diem2_min[1])), (0,255,255), 5)
                                show1 = cv2.line(show1, (int(giao_diem[0]),int(giao_diem[1])), (int(diem1_tt[0]),int(diem1_tt[1])), (0,255,255), 5)


                                cv2.putText(show1,resultkc,(30,1100),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),1)
                                # cv2.putText(show1,f"kc_tren: {int(my_kc_t)}",(30,50*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                                window[f'result_cam{num_camera}'].update(value= 'NG', text_color='red')
                                k +=1 
                                myresult1 = 1    

                            else:
                                show1 = cv2.circle(show1,(int(diem1_tt[0]),int(diem1_tt[1])), 5,(0,0,255),-1)
                                show1 = cv2.line(show1, (int(diem2_max[0]),int(diem2_max[1])), (int(diem2_min[0]),int(diem2_min[1])), (0,255,255), 5)
                                show1 = cv2.line(show1, (int(giao_diem[0]),int(giao_diem[1])), (int(diem1_tt[0]),int(diem1_tt[1])), (0,255,255), 5)

                                cv2.putText(show1,resultkc,(30,1100),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),1)
                                # cv2.putText(show1,f"t: {int(my_kc_t)}",(30,50*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                                k +=1 

                        else:
                            cv2.putText(show1,f"no t",(30,1100),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),1)
                            # cv2.putText(show1,f"no t",(30,50*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                            k +=1 


                        # print(a1, a2, a3)
                        if a1 == 1 and a3 ==1:
                            # my_kc_d = h3 -h1
                            #print('d: ',my_kc_d)
                            A, B, C = phuong_trinh_duong_thang(diem3_max, diem3_min)

                            my_kc_d = khoang_cach_toi_duong_thang(A, B, C, diem1_tt)

                            A1,B1,C1 = phuong_trinh_duong_thang_vuong_goc(A, B, diem1_tt)

                            giao_diem = giao_diem_hai_duong_thang(A, B, C ,A1, B1, C1)
                            # my_kc_d = h3 -h1
                            print(f'd: {round(my_kc_d,3)} pixel = {round(my_kc_d * 0.0048,3)} mm')
                            kc_duoi_mm = str(round(my_kc_d * 0.0048,3))

                            if  my_kc_d < int(values['kc_duoi_t']) or my_kc_d > int(values['kc_duoi_c']):
                                if my_kc_d < int(values['kc_duoi_t']):
                                    write_plc(plc_name,'2416',1)
                                else:
                                    write_plc(plc_name,'2417',1)
                                print('NG')
                                write_plc(plc_name,'2414',1)
                                loi_kc = 1
                                show1 = cv2.circle(show1,(int(diem1_tt[0]),int(diem1_tt[1])), 5,(0,0,255),-1)
                                show1 = cv2.line(show1, (int(diem3_max[0]),int(diem3_max[1])), (int(diem3_min[0]),int(diem3_min[1])), (255,0,0), 5)
                                show1 = cv2.line(show1, (int(giao_diem[0]),int(giao_diem[1])), (int(diem1_tt[0]),int(diem1_tt[1])), (255,0,0), 5)


                                cv2.putText(show1,f'd: {round(my_kc_d,3)} pixel = {round(my_kc_d * 0.0048,3)} mm',(30,1150),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),1)

                                #cv2.putText(show1,f"kc_duoi: {int(my_kc_d)}",(30,50*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                                window[f'result_cam{num_camera}'].update(value= 'NG', text_color='red')
                                k +=1 
                                myresult1 = 1  
                            else:
                                show1 = cv2.circle(show1,(int(diem1_tt[0]),int(diem1_tt[1])), 5,(0,0,255),-1)
                                show1 = cv2.line(show1, (int(diem3_max[0]),int(diem3_max[1])), (int(diem3_min[0]),int(diem3_min[1])), (255,0,0), 5)
                                show1 = cv2.line(show1, (int(giao_diem[0]),int(giao_diem[1])), (int(diem1_tt[0]),int(diem1_tt[1])), (255,0,0), 5)


                                cv2.putText(show1,f'd: {round(my_kc_d,3)} pixel = {round(my_kc_d * 0.0048,3)} mm',(30,1150),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),1)
                                # cv2.putText(show1,f"d: {int(my_kc_d)}",(30,50*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                                k +=1 

                        else:
                            cv2.putText(show1,f"no d",(30,1150),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),1)

                            # cv2.putText(show1,f"no d",(30,50*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                            k +=1 



                        for i1 in range(len(model.names)):
                            register_ng = int(values[f'{model.names[i1]}_PLC_{num_camera}'])

                            if values[f'{model.names[i1]}_OK_4'] == True:
                                len_name1 = 0
                                for name1 in names1:
                                    if name1 == model.names[i1]:
                                        len_name1 +=1
                                if len_name1 != int(values[f'{model.names[i1]}_Num_4']):
                                    print('NG')
                                    write_plc(plc_name,register_ng,1)
                                    if values[f'have_save_NG_{num_camera}']:
                                        # name_folder_ng = time_to_name()
                                        # cv2.imwrite(values[f'save_{result}_{num_camera}']  + '/' + name_folder_ng + '.jpg',img_orgin)

                                        name_folder= time_to_name()
                                        today = datetime.date.today()
                                        if not os.path.isdir(values[f'save_NG_{num_camera}']  + "/" + str(today)):
                                            os.mkdir(values[f'save_NG_{num_camera}'] + "/" + str(today))
                                        cv2.imwrite(values[f'save_NG_{num_camera}']  + "/" + str(today) + '/' + name_folder + '.jpg',img_orgin)

 
                                    # cv2.putText(show1,model.names[i1],(30,50*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                                    window[f'result_cam{num_camera}'].update(value= 'NG', text_color='red')
                                    k +=1 
                                    myresult1 = 1
                                

                            elif values[f'{model.names[i1]}_NG_4'] == True:
                                if model.names[i1] in names1:
                                    print('NG')
                                    if register_ng == 2411:
                                        write_plc(plc_name,2413,1)
                                        write_plc(plc_name,2414,1)
                                    if register_ng == 2413:
                                        write_plc(plc_name,2411,1)
                                        write_plc(plc_name,2414,1)
                                    write_plc(plc_name,register_ng,1)

                                    if values[f'have_save_NG_{num_camera}']:
                                        # name_folder_ng = time_to_name()
                                        # cv2.imwrite(values[f'save_{result}_{num_camera}']  + '/' + name_folder_ng + '.jpg',img_orgin)

                                        name_folder= time_to_name()
                                        today = datetime.date.today()
                                        if not os.path.isdir(values[f'save_NG_{num_camera}']  + "/" + str(today)):
                                            os.mkdir(values[f'save_NG_{num_camera}'] + "/" + str(today))
                                        cv2.imwrite(values[f'save_NG_{num_camera}']  + "/" + str(today) + '/' + name_folder + '.jpg',img_orgin)

 
                                    # cv2.putText(show1,model.names[i1],(30,50*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                                    window[f'result_cam{num_camera}'].update(value= 'NG', text_color='red')
                                    k +=1     
                                    myresult1 = 1         
                                    

                        if myresult1 == 0 and loi_kc == 0:
                            print('OK')
                            register_ok = int(values[f'PLC_OK_{num_camera}'])
                            write_plc(plc_name,register_ok,1)
                            if values[f'have_save_OK_{num_camera}']:
                                # name_folder_ng = time_to_name()
                                # cv2.imwrite(values[f'save_{result}_{num_camera}']  + '/' + name_folder_ng + '.jpg',img_orgin)

                                name_folder= time_to_name()
                                today = datetime.date.today()
                                if not os.path.isdir(values[f'save_OK_{num_camera}']  + "/" + str(today)):
                                    os.mkdir(values[f'save_OK_{num_camera}'] + "/" + str(today))
                                cv2.imwrite(values[f'save_OK_{num_camera}']  + "/" + str(today) + '/' + name_folder + '.jpg',img_orgin)

 
                            cv2.putText(show1, 'OK',(result_width_display+100,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 4,(0,255,0),5)
                            window[f'result_cam{num_camera}'].update(value= 'OK', text_color='green')

                        write_plc(plc_name, value_plc_done,1)
                        t2 = time.time() - t1
                        time_cam = str(int(t2*1000)) + 'ms'
                        window[f'time_cam{num_camera}'].update(value= time_cam, text_color='black') 
                        show1 = cv2.resize(show1, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)

                        imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
                        window[f'image{num_camera}'].update(data= imgbytes1)

                        todaycsv = date.today()
                        name_csv = todaycsv.strftime("%d_%m_%Y")
                        time_csv = datetime.datetime.now()
                        time_csv = time_csv.strftime("%H_%M_%S")
                        csv_file_path = f"kc/kc_{name_csv}.csv"
                        data_kc = [kc_tren_mm,kc_duoi_mm,time_csv]
                        with open(csv_file_path, mode='a', newline='', encoding='utf-8') as file:
                            writer = csv.writer(file)
                            writer.writerow(data_kc)

                    if os.path.isfile(path):
                        os.remove(path)
                while os.path.isdir(filename):
                    try:
                        shutil.rmtree(filename)
                    except:
                        print('Error delete folder {num_camera}')


# list_directory = []
# chau1 = 0
# chau2 = 0
# choi1 = 0
# choi2 = 0
def program_camera_FH(window,values,model,directory,plc_name,num_camera,value_plc_trigger, value_plc_done):
    # global chau1
    # global chau2
    # global choi1
    # global choi2
    # global bien_ket_qua
    # global list_directory
    # global image_width_display
    # global image_height_display 
 
    size = values[f'choose_size{num_camera}']
    conf = values[f'conf_thres{num_camera}']/100
    if read_plc(plc_name,value_plc_trigger) == 1:     
        # if chau1 == 0 and chau2 ==0 and choi1 ==0 and choi2 ==0: 
        #     imgbytes = np.zeros([100,100,3],dtype=np.uint8)
        #     imgbytes = cv2.resize(imgbytes, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
        #     imgbytes = cv2.imencode('.png',imgbytes)[1].tobytes()
        #     window["image_chau1"].update(data=imgbytes)
        #     window["image_chau2"].update(data=imgbytes)
        #     window["image_choi1"].update(data=imgbytes)
        #     window["image_choi2"].update(data=imgbytes)

        if os.listdir(directory) == []:
            time.sleep(0.001)
            print(f'folder {directory} empty')
        else:
            # print(f'len: {len(os.listdir(directory))}')
            # print(f'received folder {directory}')

            for filename in glob.glob(directory + '*'):
                # if directory == 'C:/FH/camera2/chau1/' and chau1 == 1:
                #     while os.path.isdir(filename):
                #         try:
                #             shutil.rmtree(filename)
                #         except:
                #             print('Error delete folder {num_camera}')
                #     continue
                # if directory == 'C:/FH/camera2/chau2/'and chau2 == 1:
                #     while os.path.isdir(filename):
                #         try:
                #             shutil.rmtree(filename)
                #         except:
                #             print('Error delete folder {num_camera}')
                #     continue
                    
                # if directory == 'C:/FH/camera2/choi1/' and choi1 == 1:
                #     while os.path.isdir(filename):
                #         try:
                #             shutil.rmtree(filename)
                #         except:
                #             print('Error delete folder {num_camera}')
                #     continue
                    
                # if directory == 'C:/FH/camera2/choi2/' and choi2 == 1:
                #     while os.path.isdir(filename):
                #         try:
                #             shutil.rmtree(filename)
                #         except:
                #             print('Error delete folder {num_camera}')
                #     continue
                for path in glob.glob(filename + '/*'):
                    name = path[-18:]
                    if name == 'Input0_Camera0.jpg':
                        img_orgin = load_image(path,filename)              
                        t1 = time.time()
                        # if num_camera ==1:
                        #     write_plc(plc_name,value_plc_trigger,0)
                        write_plc(plc_name,value_plc_trigger,0)

                        names, show = handle_image(img_orgin, model,size,conf,num_camera,values)

                        handle_result(window, model,names,show, num_camera, plc_name,img_orgin, values)
                        # if num_camera ==1:
                        #     write_plc(plc_name, value_plc_done,1)  
                        #                       
                        write_plc(plc_name, value_plc_done,1)
                        t2 = time.time() - t1
                        time_cam = str(int(t2*1000)) + 'ms'

                        # if num_camera ==3:
                        #         num_camera =2

                        window[f'time_cam{num_camera}'].update(value= time_cam, text_color='black') 
                        imgbytes = cv2.imencode('.png',show)[1].tobytes()
                        window[f'image{num_camera}'].update(data= imgbytes)

                        # if directory == 'C:/FH/camera2/chau1/':
                        #     window['image_chau1'].update(data= imgbytes)
                        #     chau1 = 1

                        # if directory == 'C:/FH/camera2/chau2/':
                        #     window['image_chau2'].update(data= imgbytes)
                        #     chau2 = 1

                        # if directory == 'C:/FH/camera2/choi1/':
                        #     window['image_choi1'].update(data= imgbytes)
                        #     choi1 = 1

                        # if directory == 'C:/FH/camera2/choi2/':
                        #     window['image_choi2'].update(data= imgbytes)
                        #     choi2 = 1

                        print('---------------------------------------------')

                    if os.path.isfile(path):
                        os.remove(path)
                while os.path.isdir(filename):
                    try:
                        shutil.rmtree(filename)
                    except:
                        print('Error delete folder {num_camera}')

        # # print("chau1: ",chau1, " chau2: ", chau2, " choi1: ", choi1, " choi2: ", choi2)
        # if chau1 == 1 and chau2 == 1 and choi1 == 1 and choi2 ==1:
        #     print('Done 2------------------------')
        #     write_plc(plc_name,value_plc_trigger,0)

        #     chau1 = 0
        #     chau2 = 0
        #     choi1 = 0
        #     choi2 = 0

        #     if bien_ket_qua == []:
        #         register_ok = int(values[f'PLC_OK_{num_camera}'])
        #         # print('bien_ok_2: ',register_ok)

        #         write_plc(plc_name,register_ok,1)
        #     else:
        #         for register_ng in bien_ket_qua:
        #             write_plc(plc_name,register_ng,1)
        #         bien_ket_qua = []

        #     write_plc(plc_name, value_plc_done,1)




# def layout_main(num_camera):
#     layout_main = [

#         [
#         sg.Text(f'CAM {num_camera}',justification='center' ,font= ('Helvetica',font_size*2),text_color='red',expand_x=True),
#         ],

#         [
#         sg.Frame('',[
#             [sg.Image(filename='', size=(image_width_display,image_height_display),key=f'image{num_camera}',background_color='black'),

#             sg.Frame('',[
#                 [sg.Button('Webcam', size=(size_x,1),  font=('Helvetica',font_size),disabled=False ,key=f'Webcam{num_camera}'),sg.Text(' '), sg.Button('Change', size=(size_x,1), font=('Helvetica',font_size), disabled= True, key= 'Change{num_camera}')],
#                 [sg.Text('')],
#                 [sg.Button('Stop', size=(size_x,1), font=('Helvetica',font_size),disabled=False ,key=f'Stop{num_camera}'), sg.Text(' '),sg.Button('Pic', size=(size_x,1), font=('Helvetica',font_size),disabled=False,key= 'Pic{num_camera}')],
#                 [sg.Text('')],
#                 [sg.Button('Snap', size=(size_x,1), font=('Helvetica',font_size),disabled=False ,key=f'Snap{num_camera}'), sg.Text(' '),sg.Button('Detect', size=(size_x,1), font=('Helvetica',font_size),disabled=False,key= 'Detect{num_camera}')],
#                 [sg.Text('')],
#                 [sg.Checkbox('Check',size=(size_x-2,1),font=('Helvetica',font_size), key=f'check_model{num_camera}',enable_events=True,expand_x=True, expand_y=True),sg.Text(' '), sg.Combo(values=['1','2','3','4','5','6','7','8','9'], default_value='1',font=('Helvetica',font_size+6),size=(size_x-3, 100),text_color='navy',enable_events= True, key=f'choose_model')],
#                 [sg.Text('',font=('Helvetica',font_size * 4), justification='center', key=f'result_cam{num_camera}',expand_x=True)],
#                 [sg.Text('',font=('Helvetica',font_size*2 -5), justification='center', key=f'time_cam{num_camera}', expand_x=True)],
#                 ],element_justification='center', vertical_alignment='top', relief= sg.RELIEF_FLAT),
#             ],           
#         ]),

#         ],
    
#     ] 
#     layout_main = sg.Frame('', layout_main)
#     return layout_main

dem_cam_1 = 0
dem_cam_2 = 0
def program_camera(window,values,model,plc_name,num_camera,my_callback,value_plc_trigger, value_plc_done):
    global bien_ket_qua_1
    global bien_ket_qua_2
    global dem_cam_1
    global dem_cam_2
 
    if read_plc('k',"10050") == 1: #Clear bien dem
        dem_cam_1 = 0
        dem_cam_2 = 0
        write_plc('k',"10050",0) 

    if read_plc(plc_name,value_plc_trigger) == 1: 
        print('10000')
        size = values[f'choose_size{num_camera}']
        conf = values[f'conf_thres{num_camera}']/100    
        if num_camera == 1:
            dem_cam_1 +=1
        if num_camera == 2:
            dem_cam_2 +=1
        img_orgin = my_callback.image 
        img_orgin = img_orgin[182:182+846,504:504+1416]
        img_orgin = img_orgin.copy()           
        t1 = time.time()
        # if num_camera ==1:
        #     write_plc(plc_name,value_plc_trigger,0)
        # write_plc(plc_name,value_plc_trigger,0)
        write_plc(plc_name,value_plc_trigger,0)
        names, show = handle_image(img_orgin, model,size,conf,num_camera,values)
       
        handle_result(window, model,names,show, num_camera, plc_name,img_orgin, values)
        # if num_camera ==1:
        #     write_plc(plc_name, value_plc_done,1)  
        #        
        print('dem_cam_1',dem_cam_1,bien_ket_qua_1)
        print('dem_cam_2',dem_cam_2,bien_ket_qua_2)
        if dem_cam_1 == 3:    
            if bien_ket_qua_1 == []:
                register_ok = int(values[f'PLC_OK_{num_camera}'])
                write_plc(plc_name,register_ok,1)
                print(f'da gui bien OK giá tri {register_ok}')
            else:
                for register_ng in bien_ket_qua_1:
                    write_plc(plc_name,register_ng,1)
                bien_ket_qua_1 = []

            write_plc(plc_name, value_plc_done,1) 
            dem_cam_1 = 0 
            
        if dem_cam_2 == 3:    
            if bien_ket_qua_2 == []:
                register_ok = int(values[f'PLC_OK_{num_camera}'])
                write_plc(plc_name,register_ok,1)
                print(f'da gui bien OK giá tri {register_ok}')
            else:
                for register_ng in bien_ket_qua_2:
                    write_plc(plc_name,register_ng,1)
                bien_ket_qua_2 = []

            write_plc(plc_name, value_plc_done,1) 
            dem_cam_2 = 0 

        

        t2 = time.time() - t1
        time_cam = str(int(t2*1000)) + 'ms'

        # if num_camera ==3:
        #         num_camera =2

        window[f'time_cam{num_camera}'].update(value= time_cam, text_color='black') 
        imgbytes = cv2.imencode('.png',show)[1].tobytes()
        window[f'image{num_camera}'].update(data= imgbytes)


        print('---------------------------------------------')





def program_camera_test(window,values,model,plc_name,num_camera,my_callback,mykeyboard):
 
    if keyboard.is_pressed(mykeyboard):
        size = values[f'choose_size{num_camera}']
        conf = values[f'conf_thres{num_camera}']/100    

        img_orgin = my_callback.image 
        img_orgin = img_orgin[182:182+846,504:504+1416]

        img_orgin = img_orgin.copy() 
        t1 = time.time()


        names, show = handle_image(img_orgin, model,size,conf,num_camera,values)

        handle_result(window, model,names,show, num_camera, plc_name,img_orgin, values)

        t2 = time.time() - t1
        time_cam = str(int(t2*1000)) + 'ms'

        window[f'time_cam{num_camera}'].update(value= time_cam, text_color='black') 
        imgbytes = cv2.imencode('.png',show)[1].tobytes()
        window[f'image{num_camera}'].update(data= imgbytes)


        print('---------------------------------------------')







def layout_show(len_camera):
    global image_width_display
    global image_height_display
    if len_camera <= 2:
        image_width_display = 600
        image_height_display = 510


    elif len_camera <= 4:
        image_width_display = int(720*1.4)
        image_height_display = int(480*1.4)


    layout_show_img = [

        [
        sg.Frame('',[
            [
                sg.Image(filename='', size=(image_width_display,image_height_display),key='image_chau1',background_color='black'),
                sg.Image(filename='', size=(image_width_display,image_height_display),key='image_chau2',background_color='black'),

            ],    
            [
                sg.Image(filename='', size=(image_width_display,image_height_display),key='image_choi1',background_color='black'),
                sg.Image(filename='', size=(image_width_display,image_height_display),key='image_choi2',background_color='black'),

            ],       
        ]),

        ],
    
    ] 

    # layout_show_img = sg.Frame('', layout_show_img)
    return layout_show_img

def layout_main(num_camera, len_camera, choose_model):
    global image_width_display
    global image_height_display
    if len_camera <= 2:
        image_width_display = 590
        image_height_display = 410
        size_x = 8
        font_size = 14

    elif len_camera <= 4:
        image_width_display = 650
        image_height_display = 410
        size_x = 8
        font_size = 14
    elif len_camera <= 6:
        image_width_display = int(650*0.73)
        image_height_display = 410
        size_x = 7
        font_size = 13
    # elif len_camera <= 8:
    #     image_width_display = int(650*0.5)
    #     image_height_display = 410
    #     size_x = 6
    #     font_size = 12
    elif len_camera <= 8:
        image_width_display = int(650*0.66)
        image_height_display = int(410*0.66)
        size_x = 6
        font_size = 12
    if num_camera == 1:
        layout_main = [

            [
            sg.Text(f'CAM {num_camera}',justification='center' ,font= ('Helvetica',font_size*2),text_color='red',expand_x=True),
            ],

            [
            sg.Frame('',[
                [sg.Image(filename='', size=(image_width_display,image_height_display),key=f'image{num_camera}',background_color='black')],

                [sg.Frame('',[
                    [sg.Text(' '),sg.Button('Pic', size=(size_x,1), font=('Helvetica',font_size),disabled=False,key= f'Pic{num_camera}'),
                    # [sg.Text('')],
                    sg.Text(' '),sg.Button('Detect', size=(size_x,1), font=('Helvetica',font_size),disabled=False,key= f'Detect{num_camera}'),
                    # sg.Text(' '),sg.Button('Reconnect', size=(size_x,1), font=('Helvetica',font_size),disabled=True,key= f'Reconnect{num_camera}'),

                    # [sg.Text('')],
                    sg.Text(' '), sg.Combo(values=['1','2','3','4','5','6','7','8','9'], default_value=choose_model,font=('Helvetica',font_size),size=(size_x-3, 100),text_color='navy',enable_events= True, key='choose_model', disabled = True)],
                    [sg.Text('',font=('Helvetica',font_size * 4), justification='center', key=f'result_cam{num_camera}',expand_x=True)],
                    [sg.Text('',font=('Helvetica',font_size*2 -5), justification='center', key=f'time_cam{num_camera}', expand_x=True)],
                    ],element_justification='center', vertical_alignment='top', relief= sg.RELIEF_FLAT),
                ],           
            ]),

            ],
        
        ] 

    # elif num_camera == 3:
    #     pass

    else:
        layout_main = [

            [
            sg.Text(f'CAM {num_camera}',justification='center' ,font= ('Helvetica',font_size*2),text_color='red',expand_x=True),
            ],

            [
            sg.Frame('',[
                [sg.Image(filename='', size=(image_width_display,image_height_display),key=f'image{num_camera}',background_color='black')],

                [sg.Frame('',[
                    [sg.Text(' '),sg.Button('Pic', size=(size_x,1), font=('Helvetica',font_size),disabled=False,key= f'Pic{num_camera}'),
                    #[sg.Text('')],
                    sg.Text(' '),sg.Button('Detect', size=(size_x,1), font=('Helvetica',font_size),disabled=False,key= f'Detect{num_camera}')],
                    # sg.Text(' '),sg.Button('Reconnect', size=(size_x,1), font=('Helvetica',font_size),disabled=True,key= f'Reconnect{num_camera}')],

                    # [sg.Text('')],
                    [sg.Text('',font=('Helvetica',font_size * 4), justification='center', key=f'result_cam{num_camera}',expand_x=True)],
                    [sg.Text('',font=('Helvetica',font_size*2 -5), justification='center', key=f'time_cam{num_camera}', expand_x=True)],
                    ],element_justification='center', vertical_alignment='top', relief= sg.RELIEF_FLAT),
                ],           
            ]),

            ],
        
        ]  
    layout_main = sg.Frame('', layout_main)
    return layout_main


def layout_main_cam(nums_camera, choose_model):

    # layout_main_cam = [[layout_main(num_camera,nums_camera,choose_model) for num_camera in range(1,nums_camera-1)]]

    if nums_camera <= 2:
        layout_main_cam = [[layout_main(num_camera,nums_camera,choose_model) for num_camera in range(1,nums_camera+1)]]


    else:
        layout_main_cam = [[layout_main(num_camera,nums_camera,choose_model) for num_camera in range(1, math.ceil(nums_camera/2) +1)],
                        [layout_main(num_camera,nums_camera,choose_model) for num_camera in range(math.ceil(nums_camera/2) +1, nums_camera +1)]]

    return layout_main_cam



def layout_option(model, num_camera):
    file_weights = [('Weights (*.pt)', ('*.pt'))]

    # if num_camera == 4:
    #     layout_option = [
    #         [sg.Frame('',[
    #         [sg.Frame('',
    #         [  
    #             [
    #             sg.Text('Date Save', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Combo(values= [], font=('Helvetica',12),size=(59, 30),text_color='navy',enable_events= True, key=f'date_save4'),
    #             ],
    #             [sg.Text('Weights', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Input(size=(60,1), font=('Helvetica',12), key=f'file_weights{num_camera}',readonly= True, text_color='navy',enable_events= True),
    #             sg.Frame('',[
    #                 [sg.FileBrowse(file_types= file_weights, size=(12,1), font=('Helvetica',10),key= f'file_browse{num_camera}',enable_events=True, disabled=False)]
    #             ], relief= sg.RELIEF_FLAT),
    #             sg.Frame('',[
    #                 [sg.Button('Change Model', size=(14,1), font=('Helvetica',10), disabled= True, key= f'Change_{num_camera}')]
    #             ], relief= sg.RELIEF_FLAT),],
    #             [sg.Text('Confidence',size=(12,1),font=('Helvetica',15), text_color='red'), sg.Slider(range=(1,100),orientation='h',size=(60,20),font=('Helvetica',11),disabled=False, key= f'conf_thres{num_camera}'),
    #             sg.Text('Size',size=(12,1),font=('Helvetica',15), text_color='red'),
    #             sg.InputCombo((416,512,608,768,896,1024,1280,1408,1536),size=(10,200),font=('Helvetica',11),disabled=False,text_color='navy',default_value=416,key=f'choose_size{num_camera}',)],
    #             [sg.Text('KC tren',size=(12,1),font=('Helvetica',15), text_color='red'),sg.Input('',size=(8,1),font=('Helvetica',15),key= 'kc_tren_t',text_color='navy',enable_events=True, disabled=False),sg.Input('',size=(8,1),font=('Helvetica',15),key= 'kc_tren_c',text_color='navy',enable_events=True, disabled=False)],
    #             [sg.Text('KC duoi',size=(12,1),font=('Helvetica',15), text_color='red'),sg.Input('',size=(8,1),font=('Helvetica',15),key= 'kc_duoi_t',text_color='navy',enable_events=True, disabled=False),sg.Input('',size=(8,1),font=('Helvetica',15),key= 'kc_duoi_c',text_color='navy',enable_events=True, disabled=False)]

    #         ], relief=sg.RELIEF_FLAT),
    #         ],
    #         [sg.Frame('',[
    #             [sg.Text('Name',size=(15,1),font=('Helvetica',15), text_color='red'), 
    #             sg.Text('Join',size=(7,1),font=('Helvetica',15), text_color='red'), 
    #             sg.Text('OK',size=(7,1),font=('Helvetica',15), text_color='red'), 
    #             sg.Text('Num',size=(7,1),font=('Helvetica',15), text_color='red'), 
    #             sg.Text('NG',size=(8,1),font=('Helvetica',15), text_color='red'),  
    #             sg.Text('Width Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
    #             sg.Text('Width Max',size=(11,1),font=('Helvetica',15), text_color='red'), 
    #             sg.Text('Height Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
    #             sg.Text('Height Max',size=(12,1),font=('Helvetica',15), text_color='red'),
    #             sg.Text('PLC',size=(11,1),font=('Helvetica',15), text_color='red'),
    #             sg.Text('Confidence',size=(11,1),font=('Helvetica',15), text_color='red')],
    #         ], relief=sg.RELIEF_FLAT)],
    #         [sg.Frame('',[
    #             [
    #                 sg.Text(f'{model.names[i]}_{num_camera}',size=(15,1),font=('Helvetica',15), text_color='yellow'), 
    #                 sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key=f'{model.names[i]}_{num_camera}',enable_events=True, disabled=False), 
    #                 sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model.names[i]}_OK_{num_camera}',enable_events=True, disabled=False), 
    #                 sg.Input('1',size=(2,1),font=('Helvetica',15),key= f'{model.names[i]}_Num_{num_camera}',text_color='navy',enable_events=True, disabled=False), 
    #                 sg.Text('',size=(4,1),font=('Helvetica',15), text_color='red'), 
    #                 sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model.names[i]}_NG_{num_camera}',enable_events=True, disabled=False), 
    #                 sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model.names[i]}_Wn_{num_camera}',text_color='navy',enable_events=True, disabled=False), 
    #                 sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
    #                 sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model.names[i]}_Wx_{num_camera}',text_color='navy',enable_events=True, disabled=False), 
    #                 sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
    #                 sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model.names[i]}_Hn_{num_camera}',text_color='navy',enable_events=True, disabled=False), 
    #                 sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
    #                 sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model.names[i]}_Hx_{num_camera}',text_color='navy',enable_events=True, disabled=False), 
    #                 sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
    #                 sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model.names[i]}_PLC_{num_camera}',text_color='navy',enable_events=True, disabled=False), 
    #                 sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
    #                 sg.Slider(range=(1,100),default_value=25,orientation='h',size=(30,20),font=('Helvetica',11), key= f'{model.names[i]}_Conf_{num_camera}', disabled=False),            
    #             ] for i in range(len(model.names))
    #         ], relief=sg.RELIEF_FLAT)],
    #         [sg.Text('  OK',size=(15,1),font=('Helvetica',15), text_color='yellow'),
    #         sg.Text(' '*230), 
    #         sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'PLC_OK_{num_camera}',text_color='navy',enable_events=True, disabled=False)],
    #         [sg.Text(' ')],
    #         [sg.Text(' '*250), sg.Button('Save Data', size=(12,1),  font=('Helvetica',12),key=f'SaveData{num_camera}',enable_events=True, disabled=False)] 
    #         ], relief= sg.RELIEF_FLAT)]
    #     ]

    # else:
    layout_option = [
        [sg.Frame('',[
        [sg.Frame('',
        [   
            [
            sg.Text('Date Save', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Combo(values= [], font=('Helvetica',12),size=(59, 30),text_color='navy',enable_events= True, key=f'date_save{num_camera}'),
            ],
            [sg.Text('Weights', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Input(size=(60,1), font=('Helvetica',12), key=f'file_weights{num_camera}',readonly= True, text_color='navy',enable_events= True),
            sg.Frame('',[
                [sg.FileBrowse(file_types= file_weights, size=(12,1), font=('Helvetica',10),key= f'file_browse{num_camera}',enable_events=True, disabled=False)]
            ], relief= sg.RELIEF_FLAT),
            sg.Frame('',[
                [sg.Button('Change Model', size=(14,1), font=('Helvetica',10), disabled= True, key= f'Change_{num_camera}')]
            ], relief= sg.RELIEF_FLAT),],
            [sg.Text('Confidence',size=(12,1),font=('Helvetica',15), text_color='red'), sg.Slider(range=(1,100),orientation='h',size=(60,20),font=('Helvetica',11),disabled=False, key= f'conf_thres{num_camera}'),
            sg.Text('Size',size=(12,1),font=('Helvetica',15), text_color='red'),
            sg.InputCombo((416,512,608,768,896,1024,1280,1408,1536),size=(10,200),font=('Helvetica',11),disabled=False,text_color='navy',default_value=416,key=f'choose_size{num_camera}',)]
        ], relief=sg.RELIEF_FLAT),
        ],
        [sg.Frame('',[
            [sg.Text('Name',size=(15,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Join',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('OK',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Num',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('NG',size=(8,1),font=('Helvetica',15), text_color='red'),  
            sg.Text('Width Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Width Max',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Max',size=(12,1),font=('Helvetica',15), text_color='red'),
            sg.Text('PLC',size=(11,1),font=('Helvetica',15), text_color='red'),
            sg.Text('Confidence',size=(11,1),font=('Helvetica',15), text_color='red')],
        ], relief=sg.RELIEF_FLAT)],
        [sg.Frame('',[
            [
                sg.Text(f'{model.names[i]}_{num_camera}',size=(15,1),font=('Helvetica',15), text_color='yellow'), 
                sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key=f'{model.names[i]}_{num_camera}',enable_events=True, disabled=False), 
                sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model.names[i]}_OK_{num_camera}',enable_events=True, disabled=False), 
                sg.Input('1',size=(2,1),font=('Helvetica',15),key= f'{model.names[i]}_Num_{num_camera}',text_color='navy',enable_events=True, disabled=False), 
                sg.Text('',size=(4,1),font=('Helvetica',15), text_color='red'), 
                sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model.names[i]}_NG_{num_camera}',enable_events=True, disabled=False), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model.names[i]}_Wn_{num_camera}',text_color='navy',enable_events=True, disabled=False), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model.names[i]}_Wx_{num_camera}',text_color='navy',enable_events=True, disabled=False), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model.names[i]}_Hn_{num_camera}',text_color='navy',enable_events=True, disabled=False), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model.names[i]}_Hx_{num_camera}',text_color='navy',enable_events=True, disabled=False), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('4000',size=(8,1),font=('Helvetica',15),key= f'{model.names[i]}_PLC_{num_camera}',text_color='navy',enable_events=True, disabled=False), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Slider(range=(1,100),default_value=25,orientation='h',size=(30,20),font=('Helvetica',11), key= f'{model.names[i]}_Conf_{num_camera}', disabled=False),            
            ] for i in range(len(model.names))
        ], relief=sg.RELIEF_FLAT)],
        [sg.Text('  OK',size=(15,1),font=('Helvetica',15), text_color='yellow'),
        sg.Text(' '*230), 
        sg.Input('4000',size=(8,1),font=('Helvetica',15),key= f'PLC_OK_{num_camera}',text_color='navy',enable_events=True, disabled=False)],
        [sg.Text(' ')],
        [sg.Text(' '*250), sg.Button('Save Data', size=(12,1),  font=('Helvetica',12),key=f'SaveData{num_camera}',enable_events=True, disabled=False)] 
        ], relief= sg.RELIEF_FLAT)]
    ]

    layout_option = [[sg.Column(layout_option,scrollable=True, expand_x=True, expand_y=True)]]
    return layout_option
    


def layout_part_saveimg(status,num_camera):
    layout_part_saveimg = [
        [sg.Text(f'Have save folder image {status} for camera {num_camera}',size=(35,1),font=('Helvetica',12), text_color='yellow'),
         sg.Checkbox('',size=(5,5),default=False,font=('Helvetica',12),  key=f'have_save_{status}_{num_camera}',enable_events=True, disabled=False)], 
        [sg.T(f'Choose folder save image {status} for camera {num_camera}', font=('Helvetica',12), text_color = 'green')],
        [sg.Input(size=(50,1),default_text=f'C:/Cam{num_camera}/{status}' ,font=('Helvetica',12), key=f'save_{status}_{num_camera}',readonly= True, text_color='navy',enable_events= True),
         sg.FolderBrowse(size=(10,1), font=('Helvetica',12),key=f'save_folder_{status}_{num_camera}',enable_events=True) ],

    ]
    layout_part_saveimg = sg.Frame('', layout_part_saveimg)

    return layout_part_saveimg

# def layout_saveimg_4cam():
#     return [
#         [sg.Frame('',[
#             [layout_saveimg("OK",1)],
#             [layout_saveimg("OK",2)],
#             [layout_saveimg("NG",1)],
#             [layout_saveimg("NG",2)],
#         ], relief=sg.RELIEF_FLAT),
#         sg.Frame('',[
#             [layout_saveimg("OK",3)],
#             [layout_saveimg("OK",4)],
#             [layout_saveimg("NG",3)],
#             [layout_saveimg("NG",4)],
#         ], relief=sg.RELIEF_FLAT)],
#     ]


def layout_save_img(num_camera):
    layout_save_img =  [[layout_part_saveimg(status, i) for status in ["OK", "NG"]] for i in range(1, num_camera+1)]
    return layout_save_img




def layout_termi_nal():
    return [
        [sg.Text("Anything printed will display here!")]#,
        # [sg.Multiline(font=('Helvetica',14), write_only=True, autoscroll=True, auto_refresh=True,reroute_stdout=True, reroute_stderr=True, echo_stdout_stderr=True,expand_x=True,expand_y=True)]
    ]


def layout_theme(window, models, nums_camera, choose_model):
    layout_theme = [
        [sg.Listbox(values= sg.theme_list(), size = (30,20),auto_size_text=18,default_values='Dark',key='theme', enable_events=True)],
        [
            [sg.Button('Apply'),
            sg.Button('Cancel')]
        ]
    ] 
    window_theme = sg.Window('Change Theme', layout_theme, location=(50,50),keep_on_top=True).Finalize()
    window_theme.set_min_size((300,400))   

    while True:
        event_theme, values_theme = window_theme.read(timeout=20)
        if event_theme == sg.WIN_CLOSED:
            break

        if event_theme == 'Apply':
            theme_choose = values_theme['theme'][0]
            if theme_choose == 'Default':
                continue
            window.close()
            window = make_window(theme_choose, models, nums_camera, choose_model)
            save_theme(theme_choose)
            #print(theme_choose)
        if event_theme == 'Cancel':
            break
            # answer = sg.popup_yes_no('Do you want to exit?')
            # if answer == 'Yes':
            #     break
            # if answer == 'No':
            #     continue
    window_theme.close()

    return window



def make_window(theme, models, nums_camera,choose_model):
    sg.theme(theme)

    right_click_menu = [[], ['Exit','Administrator','Change Theme']]

    layout_main = layout_main_cam(nums_camera,choose_model)


    layouts_option = []

    for i in range(1,nums_camera+2):            
        
        layouts_option.append(layout_option(models[i-1],i))

    layout_saveimg = layout_save_img(nums_camera+1)
    layout_terminal = layout_termi_nal()
    # layout_show_img = layout_show(4)

    tab_list = [
        sg.Tab('Main', layout_main), 
        # sg.Tab('Image show', layout_show_img), 

        sg.Tab('Save Image', layout_saveimg), 
        sg.Tab('Output', layout_terminal)
        ]

    # for i, layout in enumerate(layouts_option):
    #     tab_list.insert(i+2, sg.Tab(f'Option for model {i+1}', layout))

    for i, layout in enumerate(layouts_option):
        tab_list.insert(i+1, sg.Tab(f'Option for model {i+1}', layout))

    layout = [[sg.TabGroup([tab_list])]]

    # window = sg.Window('HuynhLeVu', layout, location=(0,screen_height//2),right_click_menu=right_click_menu,resizable=True,size = (screen_width, screen_height//2)).Finalize()
    # # window.Maximize()
    window = sg.Window('HuynhLeVu', layout, location=(0,0),right_click_menu=right_click_menu,resizable=True).Finalize()

    # window = sg.Window('HuynhLeVu', layout, location=(0,0),right_click_menu=right_click_menu,resizable=True, finalize= True)
    # window.Maximize()

    return window



def update_disabled_control(window, model, event, values, num_camera):
    for i_model in range(len(model.names)):
        is_enabled = values[f'{model.names[num_camera]}_{num_camera}']
        window[f'{model.names[i_model]}_OK_{num_camera}'].update(disabled=not is_enabled)
        window[f'{model.names[i_model]}_Num_{num_camera}'].update(disabled=not is_enabled)
        window[f'{model.names[i_model]}_NG_{num_camera}'].update(disabled=not is_enabled)
        window[f'{model.names[i_model]}_Wn_{num_camera}'].update(disabled=not is_enabled)
        window[f'{model.names[i_model]}_Wx_{num_camera}'].update(disabled=not is_enabled)
        window[f'{model.names[i_model]}_Hn_{num_camera}'].update(disabled=not is_enabled)
        window[f'{model.names[i_model]}_Hx_{num_camera}'].update(disabled=not is_enabled)
        window[f'{model.names[i_model]}_PLC_{num_camera}'].update(disabled=not is_enabled)
        window[f'{model.names[i_model]}_Conf_{num_camera}'].update(disabled=not is_enabled)



    for i in range(len(model.names)):
        if event == f'{model.names[i_model]}_OK_{num_camera}':
            window[f'{model.names[i_model]}_NG_{num_camera}'].update(disabled=not values[f'{model.names[i_model]}_OK_{num_camera}'])
        elif event == f'{model.names[i_model]}_NG_{num_camera}':
            window[f'{model.names[i_model]}_OK_{num_camera}'].update(disabled=not values[f'{model.names[i_model]}_NG_{num_camera}'])




def change_disabled_administrator(window,event, values, num_camera, model):
    if num_camera <=2:
        window[f'choose_model'].update(disabled= False)

        window[f'conf_thres{num_camera}'].update(disabled= False)
        window[f'choose_size{num_camera}'].update(disabled= False)


        window[f'file_browse{num_camera}'].update(disabled= False,button_color='turquoise')

        window[f'SaveData{num_camera}'].update(disabled= False,button_color='turquoise')

        # window[f'Webcam{num_camera}'].update(disabled= False,button_color='turquoise')
        # window[f'Stop{num_camera}'].update(disabled= False,button_color='turquoise')
        # window[f'Change{num_camera}'].update(button_color='turquoise')
        # window[f'Snap{num_camera}'].update(disabled= False,button_color='turquoise')

        window[f'Change_{num_camera}'].update(button_color='turquoise')


    window[f'have_save_OK_{num_camera}'].update(disabled=False)
    window[f'have_save_NG_{num_camera}'].update(disabled=False)

    window[f'save_OK_{num_camera}'].update(disabled=False)
    window[f'save_NG_{num_camera}'].update(disabled=False)

    window[f'save_folder_OK_{num_camera}'].update(disabled= False,button_color='turquoise')
    window[f'save_folder_NG_{num_camera}'].update(disabled= False,button_color='turquoise')


    for i in range(len(model.names)):
        window[f'{model.names[i]}_{num_camera}'].update(disabled=False)
        window[f'{model.names[i]}_OK_{num_camera}'].update(disabled=False)
        window[f'{model.names[i]}_Num_{num_camera}'].update(disabled=False)
        window[f'{model.names[i]}_NG_{num_camera}'].update(disabled=False)
        window[f'{model.names[i]}_Wn_{num_camera}'].update(disabled=False)
        window[f'{model.names[i]}_Wx_{num_camera}'].update(disabled=False)
        window[f'{model.names[i]}_Hn_{num_camera}'].update(disabled=False)
        window[f'{model.names[i]}_Hx_{num_camera}'].update(disabled=False)
        window[f'{model.names[i]}_PLC_{num_camera}'].update(disabled=False)
        window[f'{model.names[i]}_Conf_{num_camera}'].update(disabled=False)
        window[f'PLC_OK_{num_camera}'].update(disabled=False)



    for i_model in range(len(model.names)):
        is_enabled = values[f'{model.names[num_camera]}_{num_camera}']
        window[f'{model.names[i_model]}_OK_{num_camera}'].update(disabled=not is_enabled)
        window[f'{model.names[i_model]}_Num_{num_camera}'].update(disabled=not is_enabled)
        window[f'{model.names[i_model]}_NG_{num_camera}'].update(disabled=not is_enabled)
        window[f'{model.names[i_model]}_Wn_{num_camera}'].update(disabled=not is_enabled)
        window[f'{model.names[i_model]}_Wx_{num_camera}'].update(disabled=not is_enabled)
        window[f'{model.names[i_model]}_Hn_{num_camera}'].update(disabled=not is_enabled)
        window[f'{model.names[i_model]}_Hx_{num_camera}'].update(disabled=not is_enabled)
        window[f'{model.names[i_model]}_PLC_{num_camera}'].update(disabled=not is_enabled)
        window[f'{model.names[i_model]}_Conf_{num_camera}'].update(disabled=not is_enabled)
        window[f'PLC_OK_{num_camera}'].update(disabled=not is_enabled)



    for i in range(len(model.names)):
        if event == f'{model.names[i_model]}_OK_{num_camera}':
            window[f'{model.names[i_model]}_NG_{num_camera}'].update(disabled=not values[f'{model.names[i_model]}_OK_{num_camera}'])
        elif event == f'{model.names[i_model]}_NG_{num_camera}':
            window[f'{model.names[i_model]}_OK_{num_camera}'].update(disabled=not values[f'{model.names[i_model]}_NG_{num_camera}'])

    # if num_camera ==3:
    #     num_camera =2
    window[f'Pic{num_camera}'].update(disabled= False,button_color='turquoise')
    window[f'Detect{num_camera}'].update(button_color='turquoise')

def login_administrator(window, event, values, models):
    if event =='Administrator':
        for i, model in enumerate(models):
            change_disabled_administrator(window,event, values, i+1, model)

        # login_password = '1'
        # password = sg.popup_get_text(
        #     'Enter Password: ', password_char='*') 
        # if password == login_password:
        #     sg.popup_ok('Login Successed!!! ',text_color='green', font=('Helvetica',14))      
        #     for i, model in enumerate(models):
        #         change_disabled_administrator(window,event, values, i+1, model)

        # else:
        #     sg.popup_cancel('Wrong Password!!!',text_color='red', font=('Helvetica',14))


def add_file_browser(event, window, values, num_camera):
    if event == f'file_browse{num_camera}':
        window[f'file_weights{num_camera}'].update(value=values[f'file_browse{num_camera}'])
        if values[f'file_browse{num_camera}']:
            # window[f'Change{num_camera}'].update(disabled=False)
            window[f'Change_{num_camera}'].update(disabled=False)


def choose_model_event(window, values, nums_camera,theme,mychoose,havedate):

    weights, conf_thress, choose_sizes, OK_Cams, NG_Cams, Folder_OK_Cams, Folder_NG_Cams = set_param_before_choose_model_cam(nums_camera+1)
    models,weights, conf_thress, choose_sizes, OK_Cams, NG_Cams, Folder_OK_Cams, Folder_NG_Cams  = row_for_var(values,nums_camera+1, weights,conf_thress,choose_sizes, OK_Cams, NG_Cams, Folder_OK_Cams, Folder_NG_Cams, mychoose, havedate)  
    window.close() 
    window = make_window(theme, models, nums_camera, mychoose)
    var_for_interface(window,nums_camera+1,weights, conf_thress, choose_sizes, OK_Cams, NG_Cams, Folder_OK_Cams, Folder_NG_Cams)

    update_value_model(window, values,models, nums_camera+1,mychoose,havedate)
    return window, models


def set_param_before_choose_model(num_camera):
    weight = ''
    conf_thres = 1
    choose_size = 416
    OK_Cam = False
    NG_Cam = True
    Folder_OK_Cam = f'C:/Cam{num_camera}/OK'
    Folder_NG_Cam = f'C:/Cam{num_camera}/NG'
    return weight, conf_thres,choose_size, OK_Cam, NG_Cam, Folder_OK_Cam, Folder_NG_Cam




def set_param_before_choose_model_cam(nums_camera):
    weights = []
    conf_thress = []
    choose_sizes = []
    OK_Cams = []
    NG_Cams = []
    Folder_OK_Cams = []
    Folder_NG_Cams = []

    for i in range(1,nums_camera+1):
        weight, conf_thres,choose_size , OK_Cam, NG_Cam, Folder_OK_Cam, Folder_NG_Cam = set_param_before_choose_model(i)
        weights.append(weight)
        conf_thress.append(conf_thres)
        choose_sizes.append(choose_size)
        OK_Cams.append(OK_Cam)
        NG_Cams.append(NG_Cam)
        Folder_OK_Cams.append(Folder_OK_Cam)
        Folder_NG_Cams.append(Folder_NG_Cam)

    return weights, conf_thress, choose_sizes, OK_Cams, NG_Cams, Folder_OK_Cams, Folder_NG_Cams


def row_for_var(values,nums_camera, weights,conf_thress,choose_sizes, OK_Cams, NG_Cams, Folder_OK_Cams, Folder_NG_Cams,mychoose, havedate):
    conn = sqlite3.connect(f'modeldb_{nums_camera}_PLC_conf_date.db')
    if date:
        cursor = conn.execute("SELECT * from DATESAVE")
    else:
        cursor = conn.execute("SELECT * from MYMODEL")

    # models = []
    models = [''] * nums_camera
    empty_models = [''] * nums_camera
    names_models = [''] * nums_camera
    list_models = []
    for row in cursor:
        if row[0] == mychoose:
            row1_a, row1_b = row[1].strip().split('_')
            for num_camera in range(nums_camera):
                if row1_a == f'{num_camera+1}' and row1_b == '0' and (not havedate or values[f'date_save{num_camera+1}'] == row[-1]):
                    weights[num_camera] = row[2]
                    conf_thress[num_camera] = row[3]
                    choose_sizes[num_camera] = row[-2]

                    for i in range(1,5):
                        OK_Cams[num_camera] = str2bool(row[4+num_camera])
                        NG_Cams[num_camera] = str2bool(row[8+num_camera])
                        Folder_OK_Cams[num_camera] = row[12+num_camera]
                        Folder_NG_Cams[num_camera] = row[16+num_camera]

                    # models.append(torch.hub.load('./levu','custom', path= row[2], source='local',force_reload =False))
                    models[num_camera] = torch.hub.load('./levu','custom', path= row[2], source='local',force_reload =False)
                    names_models[num_camera] = row[2]
                    list_models.append(num_camera+1)
                    # print(row[2])
    if '' not in models:
        pass
    else:
        missing_numbers = []
        for num in range(1, nums_camera+1):
            if num not in list_models:
                missing_numbers.append(num)
        # print("missing_numbers: ",missing_numbers)
        for num in missing_numbers:
            index = num - 1  # Vị trí chèn là num - 1
            # weights.insert(index, '')
            # conf_thress.insert(index, '1')
            # choose_sizes.insert(index, '416')
            models[index] = torch.hub.load('./levu','custom', path= 'best.pt', source='local',force_reload =False)
            names_models[index] = 'best.pt'
        


    
    if models == empty_models:
        for num_camera in range(nums_camera):
            models[num_camera] = torch.hub.load('./levu','custom', path= 'best.pt', source='local',force_reload =False)   
            names_models[num_camera] = 'best.pt'


    # print("weights: ",weights)
    # print("names_models: ",names_models)

    return models,weights, conf_thress, choose_sizes, OK_Cams, NG_Cams, Folder_OK_Cams, Folder_NG_Cams
    

def var_for_interface(window,nums_camera,weights, conf_thress, choose_sizes, OK_Cams, NG_Cams, Folder_OK_Cams, Folder_NG_Cams):
    # window['choose_model'].update(value=mychoose)

    for num_camera in range(nums_camera):
        window[f'file_weights{num_camera+1}'].update(value=weights[num_camera])
        window[f'conf_thres{num_camera+1}'].update(value=conf_thress[num_camera])
        window[f'choose_size{num_camera+1}'].update(value=choose_sizes[num_camera])



        window[f'have_save_OK_{num_camera+1}'].update(value=OK_Cams[num_camera])
        window[f'have_save_NG_{num_camera+1}'].update(value=NG_Cams[num_camera])

        window[f'save_OK_{num_camera+1}'].update(value=Folder_OK_Cams[num_camera])
        window[f'save_NG_{num_camera+1}'].update(value=Folder_NG_Cams[num_camera])


def update_value_model(window, values, models, nums_camera,mychoose, havedate):
    conn = sqlite3.connect(f'modeldb_{nums_camera}_PLC_conf_date.db')
    if havedate:
        cursor = conn.execute("SELECT * from DATESAVE")
    else:
        cursor = conn.execute("SELECT * from MYMODEL")

    for row in cursor:
        if row[0] == mychoose:
            

            for model, num_camera in zip(models,range(1,nums_camera+1)):
                row1_a, row1_b = row[1].strip().split('_')
                if row1_a == str(num_camera) and (not havedate or values[f'date_save{num_camera}'] == row[-1]):
                    for item in range(len(model.names)):

                        if int(row1_b) == item:
                            stt = 8 + 4 * (nums_camera -1)
                            # print(f'{model.names[item]}_Wn_{num_camera}', ' ',  str(row[1]), '',str(row[stt+4]))

                            window[f'{model.names[item]}_{num_camera}'].update(value=str2bool(row[stt]))
                            window[f'{model.names[item]}_OK_{num_camera}'].update(value=str2bool(row[stt+1]))
                            window[f'{model.names[item]}_Num_{num_camera}'].update(value=str(row[stt+2]))
                            window[f'{model.names[item]}_NG_{num_camera}'].update(value=str2bool(row[stt+3]))
                            window[f'{model.names[item]}_Wn_{num_camera}'].update(value=str(row[stt+4]))
                            window[f'{model.names[item]}_Wx_{num_camera}'].update(value=str(row[stt+5]))
                            window[f'{model.names[item]}_Hn_{num_camera}'].update(value=str(row[stt+6]))
                            window[f'{model.names[item]}_Hx_{num_camera}'].update(value=str(row[stt+7]))
                            window[f'{model.names[item]}_PLC_{num_camera}'].update(value=str(row[stt+8]))
       
                            window[f'PLC_OK_{num_camera}'].update(value=str(row[stt+9]))

                            window[f'{model.names[item]}_Conf_{num_camera}'].update(value=str(row[stt+10]))
                            # stt = 8 + 4 * (nums_camera -1)
                            # window[f'{model.names[item]}_{num_camera}'].update(value=str2bool(row[12]))
                            # window[f'{model.names[item]}_OK_{num_camera}'].update(value=str2bool(row[13]))
                            # window[f'{model.names[item]}_Num_{num_camera}'].update(value=str(row[14]))
                            # window[f'{model.names[item]}_NG_{num_camera}'].update(value=str2bool(row[15]))
                            # window[f'{model.names[item]}_Wn_{num_camera}'].update(value=str(row[16]))
                            # window[f'{model.names[item]}_Wx_{num_camera}'].update(value=str(row[17]))
                            # window[f'{model.names[item]}_Hn_{num_camera}'].update(value=str(row[18]))
                            # window[f'{model.names[item]}_Hx_{num_camera}'].update(value=str(row[19]))
                            # window[f'{model.names[item]}_PLC_{num_camera}'].update(value=str(row[20]))
                            # window[f'PLC_OK_{num_camera}'].update(value=str(row[21]))
                            # window[f'{model.names[item]}_Conf_{num_camera}'].update(value=str(row[22]))
    conn.close()
    for num_camera in range(1, nums_camera+1):
        try:
            dict_date = collect_dict_date(num_camera, nums_camera)
            window[f'date_save{num_camera}'].update(values=dict_date[mychoose])

        except:
            pass


def save_data(window,event ,model, num_camera,choose_model, nums_camera, values):
    if event == f'SaveData{num_camera}':
        save_all_sql(model,num_camera,choose_model, nums_camera, values)
        save_choosemodel(choose_model)
        # save_model(num_camera,values[f'file_weights{num_camera}'], choose_model)
        save_model(num_camera,values[f'file_weights{num_camera}'])
        save_all_sql_date(values,model,num_camera, nums_camera)
        try:
            dict_date = collect_dict_date(num_camera, nums_camera)
            window[f'date_save{num_camera}'].update(values=dict_date[choose_model])
        except:
            pass
        sg.popup(f'Saved param model {num_camera} successed',font=('Helvetica',15), text_color='green',keep_on_top= True)

    if event == 'SaveData4':
        save_kc(values) 


# def program_camera_FH(model,size,conf,directory,plc_name,num_camera,window,values):
#     if read_plc(plc_name,4000) == 1:     

#         if os.listdir(directory) == []:
#             print(f'folder {directory} empty')
#         else:
#             print(f'received folder {directory}')

#             for filename in glob.glob(directory + '*'):
#                 for path in glob.glob(filename + '/*'):
#                     name = path[-18:]
#                     if name == 'Input0_Camera0.jpg':
#                         img_orgin = load_image(path,filename)              
#                         t1 = time.time()

#                         write_plc(plc_name,4000,0)

#                         names, show = handle_image(img_orgin, model,size,conf,1,values)

#                         handle_result(window, model,names,show, num_camera, plc_name,img_orgin)
                        
#                         write_plc(plc_name, 5000,1)
#                         t2 = time.time() - t1
#                         time_cam = str(int(t2*1000)) + 'ms'
#                         window[f'time_cam{num_camera}'].update(value= time_cam, text_color='black') 
                    

#                         imgbytes = cv2.imencode('.png',show)[1].tobytes()
#                         window[f'image{num_camera}'].update(data= imgbytes)
#                         print('---------------------------------------------')

#                     if os.path.isfile(path):
#                         os.remove(path)
#                 while os.path.isdir(filename):
#                     try:
#                         shutil.rmtree(filename)
#                     except:
#                         print('Error delete folder {num_camera}')




def check_model(window, event, values, directory, model,size,conf,num_camera):
    if event == f'check_model{num_camera}' and values[f'check_model{num_camera}'] == True:
        if os.listdir(directory) == []:
            print(f'folder {directory} empty')
        else:
            print(f'received folder {directory}')

            for path in glob.glob(directory + '/*'):
                name = os.path.basename(path)

                img_orgin = load_image(path,directory)              
                t1 = time.time()

                names, show = handle_image(img_orgin, model,size,conf,num_camera,values)

                handle_result(window, model,names,show, num_camera, "",img_orgin, values)
                t2 = time.time() - t1

                time_cam = str(int(t2*1000)) + 'ms'
                window[f'time_cam{num_camera}'].update(value= time_cam, text_color='black') 

                imgbytes = cv2.imencode('.png',show)[1].tobytes()
                window[f'image{num_camera}'].update(data= imgbytes)
                print('---------------------------------------------')



def stop_webcam(window, num_camera):
    imgbytes = np.zeros([100,100,3],dtype=np.uint8)
    imgbytes = cv2.resize(imgbytes, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
    imgbytes = cv2.imencode('.png',imgbytes)[1].tobytes()
    window[f'image{num_camera}'].update(data=imgbytes)
    window[f'result_cam{num_camera}'].update(value='')



def webcam(window,num_camera):

    # img_orgin = img_orgin[50:530,dir_img70:710]
    img_orgin = img_orgin.copy()
    img_orgin = cv2.cvtColor(img_orgin, cv2.COLOR_BGR2RGB) 
    img_resize = cv2.resize(img_orgin,(image_width_display,image_height_display))
    if img_orgin is not None:
        show = img_resize
        imgbytes = cv2.imencode('.png',show)[1].tobytes()
        window[f'image{num_camera}'].update(data=imgbytes)
        window[f'result_cam{num_camera}'].update(value='')


def camera(window, event,nums_camera):
    for num_camera in range(1,nums_camera+1):
        if event == f'Webcam{num_camera}':
            webcam(window,num_camera)
        if event == f'Stop{num_camera}':
            stop_webcam(window, num_camera)


def picture(window, num_camera):
    
    file_name_img = [("Img(*.jpg,*.png)",("*jpg","*.png"))]
    dir_img = sg.popup_get_file(f'Choose your image {num_camera}',file_types=file_name_img,keep_on_top= True)
    if dir_img not in ('',None):
        pic = Image.open(dir_img)
        img_resize = pic.resize((image_width_display,image_height_display))
        imgbytes = ImageTk.PhotoImage(img_resize)
        window[f'image{num_camera}'].update(data= imgbytes)
        window[f'Detect{num_camera}'].update(disabled= False)  
        return dir_img



def temp_param_model(model, values, num_camera):
    list_variable = [[0]*12 for i in range(len(model.names))]

    for i,item in enumerate(range(len(model.names))):
        list_variable[i][0] = model.names[i]

        list_variable[i][1] = values[f'{model.names[item]}_{num_camera}']
        list_variable[i][2] = values[f'{model.names[item]}_OK_{num_camera}'] 
        list_variable[i][3] = values[f'{model.names[item]}_Num_{num_camera}'] 
        list_variable[i][4] = values[f'{model.names[item]}_NG_{num_camera}'] 
        list_variable[i][5] = values[f'{model.names[item]}_Wn_{num_camera}'] 
        list_variable[i][6] = values[f'{model.names[item]}_Wx_{num_camera}'] 
        list_variable[i][7] = values[f'{model.names[item]}_Hn_{num_camera}'] 
        list_variable[i][8] = values[f'{model.names[item]}_Hx_{num_camera}'] 
        list_variable[i][9] = values[f'{model.names[item]}_PLC_{num_camera}'] 
        list_variable[i][10] = values[f'PLC_OK_{num_camera}']
        list_variable[i][11] = values[f'{model.names[item]}_Conf_{num_camera}'] 
    return list_variable





def interface_for_var(values, num_camera, nums_camera, models):
    mypath = values[f'file_weights{num_camera}']
    models[num_camera-1] = torch.hub.load('./levu','custom',path=mypath,source='local',force_reload=False)
    mychoose = values['choose_model']

    weights = []
    conf_thress = []
    choose_sizes = []
    OK_Cams = []
    NG_Cams = []
    Folder_OK_Cams = []
    Folder_NG_Cams = []

    for i_cam in range(1,nums_camera+1):
        weights.append(values[f'file_weights{num_camera}'])
        conf_thress.append(values[f'conf_thres{num_camera}']) 
        choose_sizes.append(values[f'choose_size{num_camera}']) 
        OK_Cams.append(values[f'have_save_OK_{i_cam}'])
        NG_Cams.append(values[f'have_save_NG_{i_cam}'])
        Folder_OK_Cams.append(values[f'save_OK_{i_cam}'])
        Folder_NG_Cams.append(values[f'save_NG_{i_cam}'])
        
    return models, mychoose, weights, conf_thress,choose_sizes, OK_Cams, NG_Cams, Folder_OK_Cams, Folder_NG_Cams

 
def change_model(window,values, num_camera, nums_camera, theme, models,mychoose):
    list_variable = temp_param_model(models[num_camera-1], values, num_camera)
    models, mychoose, weights, conf_thress,choose_sizes, OK_Cams, NG_Cams, Folder_OK_Cams, Folder_NG_Cams =interface_for_var(values, num_camera, nums_camera+1, models)
    window.close() 
    window = make_window(theme,models, nums_camera, mychoose)
    var_for_interface(window,nums_camera+1,weights, conf_thress,choose_sizes, OK_Cams, NG_Cams, Folder_OK_Cams, Folder_NG_Cams)
    list_variable_for_interface(models[num_camera-1], num_camera, list_variable, window, nums_camera+1,mychoose)
    return window

def list_variable_for_interface(model, num_camera, list_variable, window, nums_camera,mychoose):
    for i, item in enumerate(range(len(model.names))):
        for name_label in model.names:
            if len(model.names) <= len(list_variable):
                if name_label == list_variable[i][0]:
                    window[f'{model.names[item]}_{num_camera}'].update(value= list_variable[i][1])
                    window[f'{model.names[item]}_OK_{num_camera}'].update(value= list_variable[i][2])
                    window[f'{model.names[item]}_Num_{num_camera}'].update(value= list_variable[i][3])
                    window[f'{model.names[item]}_NG_{num_camera}'].update(value= list_variable[i][4])
                    window[f'{model.names[item]}_Wn_{num_camera}'].update(value= list_variable[i][5])
                    window[f'{model.names[item]}_Wx_{num_camera}'].update(value= list_variable[i][6])
                    window[f'{model.names[item]}_Hn_{num_camera}'].update(value= list_variable[i][7])
                    window[f'{model.names[item]}_Hx_{num_camera}'].update(value= list_variable[i][8])
                    window[f'{model.names[item]}_PLC_{num_camera}'].update(value= list_variable[i][9])
                    window[f'PLC_OK_{num_camera}'].update(value= list_variable[i][10])
                    window[f'{model.names[item]}_Conf_{num_camera}'].update(value= list_variable[i][11])

    for num_camera in range(1, nums_camera+1):
        try:
            dict_date = collect_dict_date(num_camera, nums_camera)
            window[f'date_save{num_camera}'].update(values=dict_date[mychoose])

        except:
            pass
    for num_camera in range(1, nums_camera+1):
        try:
            dict_date = collect_dict_date(num_camera, nums_camera)
            window[f'date_save{num_camera}'].update(values=dict_date[mychoose])

        except:
            pass
def detect_model(window,dir_img, model,size,conf,num_camera,values): 
    
    print(f'CAM {num_camera} DETECT')
    try:
        t1 = time.time()
        
        img_orgin = cv2.imread(dir_img)
        names, show = handle_image(img_orgin, model,size,conf,num_camera,values)

        handle_result(window, model,names,show, num_camera, "",img_orgin, values)
        t2 = time.time() - t1

        time_cam = str(int(t2*1000)) + 'ms'
        window[f'time_cam{num_camera}'].update(value= time_cam, text_color='black') 

        imgbytes = cv2.imencode('.png',show)[1].tobytes()
        window[f'image{num_camera}'].update(data= imgbytes)
        print('---------------------------------------------')

    except:
        print(traceback.format_exc())
        sg.popup_annoying("Don't have image or parameter wrong", font=('Helvetica',14),text_color='red')
            


def detect_model_4(window,dir_img, model,size,conf,num_camera,values): 
    num_camera = 4
    size = values[f'choose_size{num_camera}']
    conf = values[f'conf_thres{num_camera}']/100
    t1 = time.time()

    img_orgin = cv2.imread(dir_img)


    img_orgin = cv2.cvtColor(img_orgin, cv2.COLOR_BGR2RGB)     


    result1 = model(img_orgin,size= size,conf = conf)

    table1 = result1.pandas().xyxy[0]
    area_remove1 = []

    myresult1 =0 

    a1 =0
    a2 =0
    a3 =0
    # print(table1)
    for item in range(len(table1.index)):

        width1 = table1['xmax'][item] - table1['xmin'][item]
        height1 = table1['ymax'][item] - table1['ymin'][item]
        #area1 = width1*height1
        label_name = table1['name'][item]
        conf1 = table1['confidence'][item] *100

        for i1 in range(len(model.names)):
            if values[f'{model.names[i1]}_4'] == True:
            #if values[f'{model.names[i1]}_WH'] == True:
                if label_name == model.names[i1]:
                    if width1 < int(values[f'{model.names[i1]}_Wn_4']): 
                        table1.drop(item, axis=0, inplace=True)
                        area_remove1.append(item)
                    elif width1 > int(values[f'{model.names[i1]}_Wx_4']): 
                        table1.drop(item, axis=0, inplace=True)
                        area_remove1.append(item)
                    elif height1 < int(values[f'{model.names[i1]}_Hn_4']): 
                        table1.drop(item, axis=0, inplace=True)
                        area_remove1.append(item)
                    elif height1 > int(values[f'{model.names[i1]}_Hx_4']): 
                        table1.drop(item, axis=0, inplace=True)
                        area_remove1.append(item)
                    elif conf1  < int(values[f'{model.names[i1]}_Conf_4']):
                        table1.drop(item, axis=0, inplace=True)
                        area_remove1.append(item)
                    else:
                    
                        if str(table1['name'][item]) == '1':
                            diem1_tt = [table1['xmin'][item] + width1/2  ,table1['ymin'][item] + height1/2]

                            h1 = table1['ymin'][item] + (table1['ymax'][item] - table1['ymin'][item])/2  
                            a1=1  

                        if str(table1['name'][item]) == '2':
                            diem2_min = [table1['xmin'][item],table1['ymin'][item]]
                            diem2_max = [table1['xmax'][item],table1['ymax'][item]]
                            h2 = table1['ymin'][item] + (table1['ymax'][item] - table1['ymin'][item])/2   
                            a2=1

                        if str(table1['name'][item]) == '3':
                            diem3_min = [table1['xmin'][item],table1['ymin'][item]]
                            diem3_max = [table1['xmax'][item],table1['ymax'][item]]
                            h3 = table1['ymin'][item] + (table1['ymax'][item] - table1['ymin'][item])/2
                            a3=1     


            if values[f'{model.names[i1]}_4'] == False:
                if label_name == model.names[i1]:
                    table1.drop(item, axis=0, inplace=True)
                    area_remove1.append(item)


    names1 = list(table1['name'])

    show1 = np.squeeze(result1.render(area_remove1))

    show1 = cv2.cvtColor(show1, cv2.COLOR_BGR2RGB)

#ta = time.time()
    k = 1 
    my_kc_t = 0
    my_kc_d =0
    if a1 == 1 and a2 ==1:
        # my_kc_t = h1 - h2
        # print('t: ',my_kc_t)
        A, B, C = phuong_trinh_duong_thang(diem2_max, diem2_min)

        my_kc_t = khoang_cach_toi_duong_thang(A, B, C, diem1_tt)

        A1,B1,C1 = phuong_trinh_duong_thang_vuong_goc(A, B, diem1_tt)

        giao_diem = giao_diem_hai_duong_thang(A, B, C ,A1, B1, C1)


        mypixel = round(my_kc_t,3)
        mymm = round(my_kc_t * 0.0048,3)
        resultkc = f't: {mypixel} pixel = {mymm} mm'
        if  my_kc_t < int(values['kc_tren_t']) or my_kc_t > int(values['kc_tren_c']):
            print('NG')
            show1 = cv2.circle(show1,(int(diem1_tt[0]),int(diem1_tt[1])), 5,(0,0,255),-1)
            show1 = cv2.line(show1, (int(diem2_max[0]),int(diem2_max[1])), (int(diem2_min[0]),int(diem2_min[1])), (0,255,255), 5)
            show1 = cv2.line(show1, (int(giao_diem[0]),int(giao_diem[1])), (int(diem1_tt[0]),int(diem1_tt[1])), (0,255,255), 5)


            cv2.putText(show1,resultkc,(30,1100),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),1)
            # cv2.putText(show1,f"kc_tren: {int(my_kc_t)}",(30,50*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
            window[f'result_cam{num_camera}'].update(value= 'NG', text_color='red')
            k +=1 
            myresult1 = 1    

        else:
            show1 = cv2.circle(show1,(int(diem1_tt[0]),int(diem1_tt[1])), 5,(0,0,255),-1)
            show1 = cv2.line(show1, (int(diem2_max[0]),int(diem2_max[1])), (int(diem2_min[0]),int(diem2_min[1])), (0,255,255), 5)
            show1 = cv2.line(show1, (int(giao_diem[0]),int(giao_diem[1])), (int(diem1_tt[0]),int(diem1_tt[1])), (0,255,255), 5)

            cv2.putText(show1,resultkc,(30,1100),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),1)
            # cv2.putText(show1,f"t: {int(my_kc_t)}",(30,50*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
            k +=1 

    else:
        cv2.putText(show1,f"no t",(30,1100),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),1)

        # cv2.putText(show1,f"no t",(30,50*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
        k +=1 


    print(a1, a2, a3)
    if a1 == 1 and a3 ==1:
        # my_kc_d = h3 -h1
        # print('d: ',my_kc_d)
        A, B, C = phuong_trinh_duong_thang(diem3_max, diem3_min)

        my_kc_d = khoang_cach_toi_duong_thang(A, B, C, diem1_tt)

        A1,B1,C1 = phuong_trinh_duong_thang_vuong_goc(A, B, diem1_tt)

        giao_diem = giao_diem_hai_duong_thang(A, B, C ,A1, B1, C1)
        # my_kc_d = h3 -h1
        print(f'd: {round(my_kc_d,3)} pixel = {round(my_kc_d * 0.0048,3)} mm')
        if  my_kc_d < int(values['kc_duoi_t']) or my_kc_d > int(values['kc_duoi_c']):
            print('NG')
            # cv2.putText(show1,f"kc_duoi: {int(my_kc_d)}",(30,50*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
            show1 = cv2.circle(show1,(int(diem1_tt[0]),int(diem1_tt[1])), 5,(0,0,255),-1)
            show1 = cv2.line(show1, (int(diem3_max[0]),int(diem3_max[1])), (int(diem3_min[0]),int(diem3_min[1])), (255,0,0), 5)
            show1 = cv2.line(show1, (int(giao_diem[0]),int(giao_diem[1])), (int(diem1_tt[0]),int(diem1_tt[1])), (255,0,0), 5)


            cv2.putText(show1,f'd: {round(my_kc_d,3)} pixel = {round(my_kc_d * 0.0048,3)} mm',(30,1150),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),1)
            
            window[f'result_cam{num_camera}'].update(value= 'NG', text_color='red')
            k +=1 
            myresult1 = 1  
        else:
            show1 = cv2.circle(show1,(int(diem1_tt[0]),int(diem1_tt[1])), 5,(0,0,255),-1)
            show1 = cv2.line(show1, (int(diem3_max[0]),int(diem3_max[1])), (int(diem3_min[0]),int(diem3_min[1])), (255,0,0), 5)
            show1 = cv2.line(show1, (int(giao_diem[0]),int(giao_diem[1])), (int(diem1_tt[0]),int(diem1_tt[1])), (255,0,0), 5)
            cv2.putText(show1,f'd: {round(my_kc_d,3)} pixel = {round(my_kc_d * 0.0048,3)} mm',(30,1150),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),1)
            
            # cv2.putText(show1,f"d: {int(my_kc_d)}",(30,50*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
            k +=1 

    else:
        cv2.putText(show1,f"no d",(30,1150),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),1)

        # cv2.putText(show1,f"no d",(30,50*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
        k +=1 



    for i1 in range(len(model.names)):

        if values[f'{model.names[i1]}_OK_4'] == True:
            len_name1 = 0
            for name1 in names1:
                if name1 == model.names[i1]:
                    len_name1 +=1
            if len_name1 != int(values[f'{model.names[i1]}_Num_4']):
                print('NG')

                cv2.putText(show1,model.names[i1],(30,50*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                window[f'result_cam{num_camera}'].update(value= 'NG', text_color='red')
                k +=1 
                myresult1 = 1
            

        elif values[f'{model.names[i1]}_NG_4'] == True:
            if model.names[i1] in names1:
                print('NG')

                cv2.putText(show1,model.names[i1],(30,50*k),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),0)
                window[f'result_cam{num_camera}'].update(value= 'NG', text_color='red')
                k +=1     
                myresult1 = 1         
                

    if myresult1 == 0:
        print('OK')
        cv2.putText(show1, 'OK',(result_width_display+100,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 4,(0,255,0),5)
        window[f'result_cam{num_camera}'].update(value= 'OK', text_color='green')

    t2 = time.time() - t1
    time_cam = str(int(t2*1000)) + 'ms'
    window[f'time_cam{num_camera}'].update(value= time_cam, text_color='black') 
    show1 = cv2.resize(show1, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)

    imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
    window[f'image{num_camera}'].update(data= imgbytes1)



def excel_hangmuc(ngaydem):
    # if read_plc(plc_name,register) == 1:
    today = date.today()
    mydate = today.strftime("%Y_%m_%d")
    wb = openpyxl.Workbook()

    HomNay = wb.create_sheet("Data")


    HomNay.merge_cells('A1:O1')
    HomNay.merge_cells('A2:A3')
    HomNay.merge_cells('B2:B3')
    HomNay.merge_cells('C2:O2')
    #HomNay.unmerge_cells('A2:D2')
    HomNay['A1'] = 'DỮ LIỆU PHẾ PHẨM HẠNG MỤC TÂM SẮT M100 A75'
    HomNay['A1'].alignment = Alignment(horizontal='center')
    HomNay['A1'].font = Font(name= 'Calibri', size=20)

    HomNay['A2'] = 'Ngày sản xuất'
    HomNay['A2'].alignment = Alignment(horizontal='center')
    HomNay['A2'].font = Font(name= 'Calibri', size=12)
    HomNay['B2'] = 'Giờ lưu dữ liệu'
    HomNay['B2'].alignment = Alignment(horizontal='center')
    HomNay['B2'].font = Font(name= 'Calibri', size=12)
    HomNay['C2'] = 'HẠNG MỤC PHẾ PHẨM'
    HomNay['C2'].alignment = Alignment(horizontal='center')
    HomNay['C2'].font = Font(name= 'Calibri', size=12)
    HomNay['C3'] = 'Tổng số lượng sản xuất'
    HomNay['D3'] = 'Keo dính trục c1'
    HomNay['E3'] = 'Keo dính comi c1'
    HomNay['F3'] = 'Keo ít và không đều c1'
    HomNay['G3'] = 'Keo thủng lỗ c1'
    HomNay['H3'] = 'Pp khác c1'

    HomNay['I3'] = 'Keo dính trục c2'
    HomNay['J3'] = 'Keo dính comi c2'
    HomNay['K3'] = 'Keo ít và không đều c2'
    HomNay['L3'] = 'Keo thủng lỗ c2'
    HomNay['M3'] = 'Pp khác c2'

    HomNay['N3'] = 'Tổng số lượng PP c1'
    HomNay['O3'] = 'Tổng số lượng PP c2'



    #HomNay['M3'] = 'Tổng số lượng PP'

    for i in range(67,67+20):
        # HomNay[f'{str(chr(i))}3'].alignment = Alignment(horizontal='center')
        # HomNay[f'{str(chr(i))}3'].font = Font(name= 'Calibri', size=12)

        HomNay.column_dimensions[f'{str(chr(i))}'].width = 25




    wb.remove(wb['Sheet'])

    if ngaydem =='Ngay':

        wb.save(f"excel/HANG_MUC/{mydate}_Ngay_hangmuc.xlsx")
        try:
            shutil.copy(f"excel/HANG_MUC/{mydate}_Ngay_hangmuc.xlsx", f"C:/excel/HANG_MUC/{mydate}_Ngay_hangmuc.xlsx")
        except:
            pass

    elif ngaydem == 'Dem':
        wb.save(f"excel/HANG_MUC/{mydate}_Dem_hangmuc.xlsx")
        try:
            shutil.copy(f"excel/HANG_MUC/{mydate}_Dem_hangmuc.xlsx", f"C:/excel/HANG_MUC/{mydate}_Dem_hangmuc.xlsx")
        except:
            pass

    # write_plc(plc_name,register,0)


def excel_ray(ngaydem):
    

    today = date.today()
    mydate = today.strftime("%Y_%m_%d")
    wb = openpyxl.Workbook()

    HomNay = wb.create_sheet("Data")


    HomNay.merge_cells('A1:K1')
    HomNay.merge_cells('A2:A3')
    HomNay.merge_cells('B2:B3')
    HomNay.merge_cells('C2:K2')
    #HomNay.unmerge_cells('A2:D2')
    HomNay['A1'] = 'DỮ LIỆU PHẾ PHẨM RAY '
    HomNay['A1'].alignment = Alignment(horizontal='center')
    HomNay['A1'].font = Font(name= 'Calibri', size=20)

    HomNay['A2'] = 'Ngày sản xuất'
    HomNay['A2'].alignment = Alignment(horizontal='center')
    HomNay['A2'].font = Font(name= 'Calibri', size=12)
    HomNay['B2'] = 'Giờ lưu dữ liệu'
    HomNay['B2'].alignment = Alignment(horizontal='center')
    HomNay['B2'].font = Font(name= 'Calibri', size=12)
    HomNay['C2'] = 'RAY PHẾ PHẨM'
    HomNay['C2'].alignment = Alignment(horizontal='center')
    HomNay['C2'].font = Font(name= 'Calibri', size=12)
    HomNay['C3'] = 'Ray keo trục và comi c1'
    HomNay['D3'] = 'Ray keo không đạt c1'
    HomNay['E3'] = 'Ray khác c1'

    HomNay['F3'] = 'Ray keo trục và comi c2'
    HomNay['G3'] = 'Ray keo không đạt c2'
    HomNay['H3'] = 'Ray khác c2'

    HomNay['I3'] = 'Ray keo trục và comi total'
    HomNay['J3'] = 'Ray keo không đạt total'
    HomNay['K3'] = 'Ray khác total'
    # HomNay['L3'] = 'Ray 5 Cam 2'

    # HomNay['L3'] = 'Tổng số lượng PP'
    #HomNay['M3'] = 'Tổng số lượng PP'

    for i in range(67,77):
        # HomNay[f'{str(chr(i))}3'].alignment = Alignment(horizontal='center')
        # HomNay[f'{str(chr(i))}3'].font = Font(name= 'Calibri', size=12)

        HomNay.column_dimensions[f'{str(chr(i))}'].width = 25



    wb.remove(wb['Sheet'])

    if ngaydem =='Ngay':

        wb.save(f"excel/RAY/{mydate}_Ngay_ray.xlsx")
        try:
            shutil.copy(f"excel/RAY/{mydate}_Ngay_ray.xlsx", f"C:/excel/RAY/{mydate}_Ngay_ray.xlsx")
        except:
            pass

    elif ngaydem == 'Dem':
        wb.save(f"excel/RAY/{mydate}_Dem_ray.xlsx")
        try:
            shutil.copy(f"excel/RAY/{mydate}_Dem_ray.xlsx", f"C:/excel/RAY/{mydate}_Dem_ray.xlsx")
        except:
            pass

    # write_plc(plc_name,register,0)



def excel_handle_pphangmuc(plc_name, register):
    if read_plc(plc_name,register) == 1 or keyboard.is_pressed('shift+g'):
        today = date.today()
        d1 = today.strftime("%d/%m/%Y")
        now = datetime.datetime.now()
        t1 = now.strftime("%H:%M:%S")

        mydate = today.strftime("%Y_%m_%d")
    
        hour = int(now.strftime("%H"))

        if 7<= hour <=18:
            if not os.path.isfile(f"excel/HANG_MUC/{mydate}_Ngay_hangmuc.xlsx"):
                excel_hangmuc('Ngay')
            wb = openpyxl.load_workbook(f"excel/HANG_MUC/{mydate}_Ngay_hangmuc.xlsx")

        if 19 <= hour <= 23:
            if not os.path.isfile(f"excel/HANG_MUC/{mydate}_Dem_hangmuc.xlsx"):
                excel_hangmuc('Dem')
            wb = openpyxl.load_workbook(f"excel/HANG_MUC/{mydate}_Dem_hangmuc.xlsx")

        if 0 <= hour <= 6:
            Previous_Date = datetime.datetime.today() - datetime.timedelta(days=1)
            Previous_Date = Previous_Date.strftime("%Y_%m_%d")
            if not os.path.isfile(f"excel/HANG_MUC/{Previous_Date}_Dem_hangmuc.xlsx"):
                excel_hangmuc('Dem')

            wb = openpyxl.load_workbook(f"excel/HANG_MUC/{Previous_Date}_Dem_hangmuc.xlsx")


        ws = wb.active
        col1 = d1
        col2 = t1
        col3 = int(read_plc(plc_name,20260))
        col4 = int(read_plc(plc_name,20210))
        col5 = int(read_plc(plc_name,20212))
        col6 = int(read_plc(plc_name,20214))
        col7 = int(read_plc(plc_name,20216))
        col8 = int(read_plc(plc_name,20218))

        col9 = int(read_plc(plc_name,20310))
        col10 = int(read_plc(plc_name,20312))
        col11 = int(read_plc(plc_name,20314))
        col12 = int(read_plc(plc_name,20316))
        col13 = int(read_plc(plc_name,20318))

        col14 = int(read_plc(plc_name,20256))
        col15 = int(read_plc(plc_name,20258))
        # col16 = int(read_plc(plc_name,7030))
        # col17 = int(read_plc(plc_name,7032))

        # col18 = int(read_plc(plc_name,7040))
        # col19 = int(read_plc(plc_name,7042))
        # col20 = int(read_plc(plc_name,7044))
        # col21 = int(read_plc(plc_name,7046))
        # col22 = int(read_plc(plc_name,7048))
        # col23 = int(read_plc(plc_name,7050))
        # col24 = int(read_plc(plc_name,7052))

        # col25 = int(read_plc(plc_name,7060))
        # col26 = int(read_plc(plc_name,7062))
        # col27 = int(read_plc(plc_name,7064))
        # col28 = int(read_plc(plc_name,7066))
        # col29 = int(read_plc(plc_name,7068))
        # col30 = int(read_plc(plc_name,7070))
        # col31 = int(read_plc(plc_name,7072))

        # col32 = int(read_plc(plc_name,7080))
        # col33 = int(read_plc(plc_name,7082))
        # col34 = int(read_plc(plc_name,7084))
        # col35 = int(read_plc(plc_name,7086))
        # col36 = int(read_plc(plc_name,7088))
        # col37 = int(read_plc(plc_name,7090))
        # col38 = int(read_plc(plc_name,7092))

        # col39 = int(read_plc(plc_name,7100))
        # col40 = int(read_plc(plc_name,7102))
        # col41 = int(read_plc(plc_name,7104))
        # col42 = int(read_plc(plc_name,7106))
        # col43 = int(read_plc(plc_name,7108))
        # col44 = int(read_plc(plc_name,7110))
        # col45 = int(read_plc(plc_name,7112))

        # col46 = int(read_plc(plc_name,7120))
        # col47 = int(read_plc(plc_name,7122))
        # col48 = int(read_plc(plc_name,7124))
        # col49 = int(read_plc(plc_name,7126))
        # col50 = int(read_plc(plc_name,7128))
        # col51 = int(read_plc(plc_name,7130))
        # col52 = int(read_plc(plc_name,7132))

        # col53 = int(read_plc(plc_name,5400))
        # col54 = int(read_plc(plc_name,5402))
        # col55 = int(read_plc(plc_name,5404))
        # col56 = int(read_plc(plc_name,5406))
        # col57 = int(read_plc(plc_name,5408))
        # col58 = int(read_plc(plc_name,5410))
        # col59 = int(read_plc(plc_name,5412))

        len_col = 15

        ws.append([col1, col2,col3, col4,col5, col6,col7, col8,col9, col10, col11,col12,col13,col14,col15])#col16,col17,col18,col19,col20,col21,col22,col23,col24,col25,col26,col27,col28,col29,col30,col31,col32,col33,col34,col35,col36,col37,col38,col39,col40,col41,col42,col43,col44,col45,col46,col47,col48,col49,col50,col51,col52,col53,col54,col55,col56,col57,col58,col59])

        for row in range(ws.max_row, ws.max_row+1):
            for col in range(1, ws.max_column+1):
                #print(col[row].value)
                d = ws.cell(row = row, column = col)
                #currentCell = ws.cell(col[row])
                d.alignment = Alignment(horizontal='center')
                #d.style.alignment.horizontal = 'center'
                d.font = Font(name= 'Calibri', size=12)

        if 7<= hour <=18:
            
            wb.save(f"excel/HANG_MUC/{mydate}_Ngay_hangmuc.xlsx")
            try:
                shutil.copy(f"excel/HANG_MUC/{mydate}_Ngay_hangmuc.xlsx", f"C:/excel/HANG_MUC/{mydate}_Ngay_hangmuc.xlsx")
            except Exception as e:
                print(e)
            
            try:
                wb1 = openpyxl.load_workbook(f"C:/excel/HANG_MUC/{mydate}_Ngay_hangmuc.xlsx")

                ws1 = wb1.active
                all_cell = []


                for row in range(1, ws1.max_row+1):
                    all_col = []    
                    for col in range(1, len_col + 1):
                        cell_obj = ws1.cell(row = row, column = col)
                        all_col.append(cell_obj.value)
                    all_cell.append(all_col)




                wb2 = openpyxl.load_workbook("C:/excel/HANG_MUC/Now_hangmuc.xlsx")
                ws2 = wb2.active

                index_row = ws2.max_row


                if ws1.max_row > ws2.max_row:
                    for row in range(4,ws1.max_row+1):
                        for col in range(1, len_col + 1):

                            if len(all_cell) >= row:
                                ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)
            
                else:
                    for row in range(4,ws2.max_row+1):
                        for col in range(1, len_col + 1):

                            if len(all_cell) >= row:
                                ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)
                            else:
                                ws2.cell(row = row, column = col).value = None
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)   

                wb2.save("C:/excel/HANG_MUC/Now_hangmuc.xlsx")



            except Exception as e:
                print(e)
        if 19 <= hour <= 23:
            wb.save(f"excel/HANG_MUC/{mydate}_Dem_hangmuc.xlsx")
            try:
                shutil.copy(f"excel/HANG_MUC/{mydate}_Dem_hangmuc.xlsx", f"C:/excel/HANG_MUC/{mydate}_Dem_hangmuc.xlsx")
            except Exception as e:
                print(e)

            try:
                wb1 = openpyxl.load_workbook(f"C:/excel/HANG_MUC/{mydate}_Dem_hangmuc.xlsx")

                ws1 = wb1.active
                all_cell = []


                for row in range(1, ws1.max_row+1):
                    all_col = []    
                    for col in range(1, len_col + 1):
                        cell_obj = ws1.cell(row = row, column = col)
                        all_col.append(cell_obj.value)
                    all_cell.append(all_col)




                wb2 = openpyxl.load_workbook("C:/excel/HANG_MUC/Now_hangmuc.xlsx")
                ws2 = wb2.active

                index_row = ws2.max_row


                if ws1.max_row > ws2.max_row:
                    for row in range(4,ws1.max_row+1):
                        for col in range(1, len_col + 1):

                            if len(all_cell) >= row:
                                ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)
            
                else:
                    for row in range(4,ws2.max_row+1):
                        for col in range(1, len_col + 1):

                            if len(all_cell) >= row:
                                ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)
                            else:
                                ws2.cell(row = row, column = col).value = None
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)   

                wb2.save("C:/excel/HANG_MUC/Now_hangmuc.xlsx")



            except Exception as e:
                print(e)


        if 0 <= hour <= 6:
            Previous_Date = datetime.datetime.today() - datetime.timedelta(days=1)
            Previous_Date = Previous_Date.strftime("%Y_%m_%d")
            wb.save(f"excel/HANG_MUC/{Previous_Date}_Dem_hangmuc.xlsx")
            try:
                shutil.copy(f"excel/HANG_MUC/{Previous_Date}_Dem_hangmuc.xlsx", f"C:/excel/HANG_MUC/{Previous_Date}_Dem_hangmuc.xlsx")
            except Exception as e:
                print(e)

            try:
                wb1 = openpyxl.load_workbook(f"C:/excel/HANG_MUC/{Previous_Date}_Dem_hangmuc.xlsx")

                ws1 = wb1.active
                all_cell = []


                for row in range(1, ws1.max_row+1):
                    all_col = []    
                    for col in range(1, len_col + 1):
                        cell_obj = ws1.cell(row = row, column = col)
                        all_col.append(cell_obj.value)
                    all_cell.append(all_col)




                wb2 = openpyxl.load_workbook("C:/excel/HANG_MUC/Now_hangmuc.xlsx")
                ws2 = wb2.active

                index_row = ws2.max_row


                if ws1.max_row > ws2.max_row:
                    for row in range(4,ws1.max_row+1):
                        for col in range(1, len_col + 1):

                            if len(all_cell) >= row:
                                ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)
            
                else:
                    for row in range(4,ws2.max_row+1):
                        for col in range(1, len_col + 1):

                            if len(all_cell) >= row:
                                ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)
                            else:
                                ws2.cell(row = row, column = col).value = None
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)   

                wb2.save("C:/excel/HANG_MUC/Now_hangmuc.xlsx")



            except Exception as e:
                print(e)
        print('g2')

        write_plc(plc_name,register,0)

    

def excel_handle_ppray(plc_name, register):
    if read_plc(plc_name,register) == 1  or keyboard.is_pressed('shift+h'):
        today = date.today()
        d1 = today.strftime("%d/%m/%Y")
        now = datetime.datetime.now()
        t1 = now.strftime("%H:%M:%S")

        mydate = today.strftime("%Y_%m_%d")

        hour = int(now.strftime("%H"))

        if 7<= hour <=18:
            if not os.path.isfile(f"excel/RAY/{mydate}_Ngay_ray.xlsx"):
                excel_ray('Ngay')
            wb = openpyxl.load_workbook(f"excel/RAY/{mydate}_Ngay_ray.xlsx")

        if 19 <= hour <= 23:
            if not os.path.isfile(f"excel/RAY/{mydate}_Dem_ray.xlsx"):
                excel_ray('Dem')
            wb = openpyxl.load_workbook(f"excel/RAY/{mydate}_Dem_ray.xlsx")

        if 0 <= hour <= 6:
            Previous_Date = datetime.datetime.today() - datetime.timedelta(days=1)
            Previous_Date = Previous_Date.strftime("%Y_%m_%d")
            if not os.path.isfile(f"excel/{Previous_Date}_Dem_ray.xlsx"):
                excel_ray('Dem')

            wb = openpyxl.load_workbook(f"excel/RAY/{Previous_Date}_Dem_ray.xlsx")


        ws = wb.active
        col1 = d1
        col2 = t1
        col3 = int(read_plc(plc_name,20220))
        col4 = int(read_plc(plc_name,20222))
        col5 = int(read_plc(plc_name,20224))
        col7 = int(read_plc(plc_name,20320))
        col6 = int(read_plc(plc_name,20322))
        col8 = int(read_plc(plc_name,20324))
        col9 = int(read_plc(plc_name,20250))
        col10 = int(read_plc(plc_name,20252))
        col11 = int(read_plc(plc_name,20254))
        # col12 = int(read_plc(plc_name,7008))

        len_col = 11

        # col11 = int(read_plc(plc_name,'DM7210'))
        # col12 = int(read_plc(plc_name,'DM7220'))
        #col13 = int(read_plc(plc_name,'DM6042'))
        ws.append([col1, col2,col3, col4,col5, col6,col7, col8,col9, col10,col11])

        for row in range(ws.max_row, ws.max_row+1):
            for col in range(1, ws.max_column+1):
                #print(col[row].value)
                d = ws.cell(row = row, column = col)
                #currentCell = ws.cell(col[row])
                d.alignment = Alignment(horizontal='center')
                #d.style.alignment.horizontal = 'center'
                d.font = Font(name= 'Calibri', size=12)

        if 7<= hour <=18:
            
            wb.save(f"excel/RAY/{mydate}_Ngay_ray.xlsx")
            try:
                shutil.copy(f"excel/RAY/{mydate}_Ngay_ray.xlsx", f"C:/excel/RAY/{mydate}_Ngay_ray.xlsx")
            except:
                pass
            
            try:
                wb1 = openpyxl.load_workbook(f"C:/excel/RAY/{mydate}_Ngay_ray.xlsx")

                ws1 = wb1.active
                all_cell = []


                for row in range(1, ws1.max_row+1):
                    all_col = []    
                    for col in range(1, len_col + 1):
                        cell_obj = ws1.cell(row = row, column = col)
                        all_col.append(cell_obj.value)
                    all_cell.append(all_col)




                wb2 = openpyxl.load_workbook("C:/excel/RAY/Now_ray.xlsx")
                ws2 = wb2.active

                index_row = ws2.max_row


                if ws1.max_row > ws2.max_row:
                    for row in range(4,ws1.max_row+1):
                        for col in range(1, len_col + 1):

                            if len(all_cell) >= row:
                                ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)
            
                else:
                    for row in range(4,ws2.max_row+1):
                        for col in range(1, len_col + 1):

                            if len(all_cell) >= row:
                                ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)
                            else:
                                ws2.cell(row = row, column = col).value = None
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)   

                wb2.save("C:/excel/RAY/Now_ray.xlsx")



            except:
                print(traceback.print_exc())
        if 19 <= hour <= 23:
            wb.save(f"excel/RAY/{mydate}_Dem_ray.xlsx")
            try:
                shutil.copy(f"excel/RAY/{mydate}_Dem_ray.xlsx", f"C:/excel/RAY/{mydate}_Dem_ray.xlsx")
            except:
                pass

            try:
                wb1 = openpyxl.load_workbook(f"C:/excel/RAY/{mydate}_Dem_ray.xlsx")

                ws1 = wb1.active
                all_cell = []


                for row in range(1, ws1.max_row+1):
                    all_col = []    
                    for col in range(1, len_col + 1):
                        cell_obj = ws1.cell(row = row, column = col)
                        all_col.append(cell_obj.value)
                    all_cell.append(all_col)




                wb2 = openpyxl.load_workbook("C:/excel/RAY/Now_ray.xlsx")
                ws2 = wb2.active

                index_row = ws2.max_row


                if ws1.max_row > ws2.max_row:
                    for row in range(4,ws1.max_row+1):
                        for col in range(1, len_col + 1):

                            if len(all_cell) >= row:
                                ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)
            
                else:
                    for row in range(4,ws2.max_row+1):
                        for col in range(1, len_col + 1):

                            if len(all_cell) >= row:
                                ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)
                            else:
                                ws2.cell(row = row, column = col).value = None
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)   

                wb2.save("C:/excel/RAY/Now_ray.xlsx")



            except:
                print(traceback.print_exc())


        if 0 <= hour <= 6:
            Previous_Date = datetime.datetime.today() - datetime.timedelta(days=1)
            Previous_Date = Previous_Date.strftime("%Y_%m_%d")
            wb.save(f"excel/RAY/{Previous_Date}_Dem_ray.xlsx")
            try:
                shutil.copy(f"excel/RAY/{Previous_Date}_Dem_ray.xlsx", f"C:/excel/RAY/{Previous_Date}_Dem_ray.xlsx")
            except:
                pass

            try:
                wb1 = openpyxl.load_workbook(f"C:/excel/RAY/{Previous_Date}_Dem_ray.xlsx")

                ws1 = wb1.active
                all_cell = []


                for row in range(1, ws1.max_row+1):
                    all_col = []    
                    for col in range(1, len_col + 1):
                        cell_obj = ws1.cell(row = row, column = col)
                        all_col.append(cell_obj.value)
                    all_cell.append(all_col)




                wb2 = openpyxl.load_workbook("C:/excel/RAY/Now_ray.xlsx")
                ws2 = wb2.active

                index_row = ws2.max_row


                if ws1.max_row > ws2.max_row:
                    for row in range(4,ws1.max_row+1):
                        for col in range(1, len_col + 1):

                            if len(all_cell) >= row:
                                ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)
            
                else:
                    for row in range(4,ws2.max_row+1):
                        for col in range(1, len_col + 1):

                            if len(all_cell) >= row:
                                ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)
                            else:
                                ws2.cell(row = row, column = col).value = None
                                d = ws2.cell(row = row, column = col)
                                d.alignment = Alignment(horizontal='center')
                                d.font = Font(name= 'Calibri', size=12)   

                wb2.save("C:/excel/RAY/Now_ray.xlsx")

            except:
                print(traceback.print_exc())


        write_plc(plc_name,register,0)


def excel_handle_all(plc_name, register):
    if read_plc(plc_name,register) == 1:
        today = date.today()
        d1 = today.strftime("%d/%m/%Y")
        now = datetime.datetime.now()
        t1 = now.strftime("%H:%M:%S")

        wb = openpyxl.load_workbook("excel/All.xlsx")

        ws = wb.active
        col1 = d1
        col2 = t1
        col3 = int(read_plc(plc_name,110))
        col4 = int(read_plc(plc_name,7000))
        col5 = int(read_plc(plc_name,7002))
        col6 = int(read_plc(plc_name,7004))
        col7 = int(read_plc(plc_name,7006))
        col8 = int(read_plc(plc_name,7008))
        col9 = int(read_plc(plc_name,7010))
        col10 = int(read_plc(plc_name,7012))

        col11 = int(read_plc(plc_name,7020))
        col12 = int(read_plc(plc_name,7022))
        col13 = int(read_plc(plc_name,7024))
        col14 = int(read_plc(plc_name,7026))
        col15 = int(read_plc(plc_name,7028))
        col16 = int(read_plc(plc_name,7030))
        col17 = int(read_plc(plc_name,7032))

        col18 = int(read_plc(plc_name,7040))
        col19 = int(read_plc(plc_name,7042))
        col20 = int(read_plc(plc_name,7044))
        col21 = int(read_plc(plc_name,7046))
        col22 = int(read_plc(plc_name,7048))
        col23 = int(read_plc(plc_name,7050))
        col24 = int(read_plc(plc_name,7052))

        col25 = int(read_plc(plc_name,7060))
        col26 = int(read_plc(plc_name,7062))
        col27 = int(read_plc(plc_name,7064))
        col28 = int(read_plc(plc_name,7066))
        col29 = int(read_plc(plc_name,7068))
        col30 = int(read_plc(plc_name,7070))
        col31 = int(read_plc(plc_name,7072))

        col32 = int(read_plc(plc_name,7080))
        col33 = int(read_plc(plc_name,7082))
        col34 = int(read_plc(plc_name,7084))
        col35 = int(read_plc(plc_name,7086))
        col36 = int(read_plc(plc_name,7088))
        col37 = int(read_plc(plc_name,7090))
        col38 = int(read_plc(plc_name,7092))

        col39 = int(read_plc(plc_name,7100))
        col40 = int(read_plc(plc_name,7102))
        col41 = int(read_plc(plc_name,7104))
        col42 = int(read_plc(plc_name,7106))
        col43 = int(read_plc(plc_name,7108))
        col44 = int(read_plc(plc_name,7110))
        col45 = int(read_plc(plc_name,7112))

        col46 = int(read_plc(plc_name,7120))
        col47 = int(read_plc(plc_name,7122))
        col48 = int(read_plc(plc_name,7124))
        col49 = int(read_plc(plc_name,7126))
        col50 = int(read_plc(plc_name,7128))
        col51 = int(read_plc(plc_name,7130))
        col52 = int(read_plc(plc_name,7132))

        col53 = int(read_plc(plc_name,5400))
        col54 = int(read_plc(plc_name,5402))
        col55 = int(read_plc(plc_name,5404))
        col56 = int(read_plc(plc_name,5406))
        col57 = int(read_plc(plc_name,5408))
        col58 = int(read_plc(plc_name,5410))
        col59 = int(read_plc(plc_name,5412))

        len_col = 59

        ws.append([col1, col2,col3, col4,col5, col6,col7, col8,col9, col10, col11,col12,col13,col14,col15,col16,col17,col18,col19,col20,col21,col22,col23,col24,col25,col26,col27,col28,col29,col30,col31,col32,col33,col34,col35,col36,col37,col38,col39,col40,col41,col42,col43,col44,col45,col46,col47,col48,col49,col50,col51,col52,col53,col54,col55,col56,col57,col58,col59])





        for row in range(ws.max_row, ws.max_row+1):
            for col in range(1, ws.max_column+1):
                #print(col[row].value)
                d = ws.cell(row = row, column = col)
                #currentCell = ws.cell(col[row])
                d.alignment = Alignment(horizontal='center')
                #d.style.alignment.horizontal = 'center'
                d.font = Font(name= 'Calibri', size=12)


        wb.save("excel/All.xlsx")
        try:
            shutil.copy("excel/All.xlsx", "C:/excel/A.xlsx")



            wb1 = openpyxl.load_workbook("C:/excel/A.xlsx")

            ws1 = wb1.active
            all_cell = []
            all_cell_date = []
            all_cell_time = []

            for row in range(1, ws1.max_row+1):
                all_col = []    
                for col in range(1, len_col + 1):
                    cell_obj = ws1.cell(row = row, column = col)
                    all_col.append(cell_obj.value)
                all_cell.append(all_col)
                all_cell_date.append(all_col[0])
                all_cell_time.append(all_col[1])



            wb2 = openpyxl.load_workbook("C:/excel/All.xlsx")
            ws2 = wb2.active

            index_row = ws2.max_row
            for row in range(4, ws2.max_row+1):
                if ws2.cell(row=row,column=1).value == None:
                    index_row = row
                    break

            for row in range(index_row,ws1.max_row+1):
                for col in range(1, len_col + 1):
                    ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                    d = ws2.cell(row = row, column = col)
                    d.alignment = Alignment(horizontal='center')
                    d.font = Font(name= 'Calibri', size=12)
            wb2.save("C:/excel/All.xlsx")


        except:
            pass
        
        write_plc(plc_name,register,0)
            


def excel_ppbydate(var_plc,name):
    today = date.today()
    mydate = today.strftime("%Y_%m_%d")
    if not os.path.isfile(f"excel/pp/{mydate}_ppbydate.xlsx"):

        wb = openpyxl.Workbook()

        HomNay = wb.create_sheet("Phepham")

        HomNay['A1'] = 'Date'
        HomNay['A1'].alignment = Alignment(horizontal='center')
        HomNay['A1'].font = Font(name= 'Calibri', size=12)

        HomNay['B1'] = 'Message'
        HomNay['B1'].alignment = Alignment(horizontal='center')
        HomNay['B1'].font = Font(name= 'Calibri', size=12)

        HomNay.column_dimensions['A'].width = 35
        HomNay.column_dimensions['B'].width = 35

        wb.remove(wb['Sheet'])

        wb.save(f"excel/pp/{mydate}_ppbydate.xlsx")
        try:
            shutil.copy(f"excel/pp/{mydate}_ppbydate.xlsx", f"C:/excel/pp/{mydate}_ppbydate.xlsx")
        except:
            pass


    wb = openpyxl.load_workbook(f"excel/pp/{mydate}_ppbydate.xlsx")
    ws = wb.active

    now = datetime.datetime.now()
    col1 = now.strftime("%d/%m/%Y %H:%M:%S")

    phe_pham = ["BẠC TRỤC", "ĐẾ VỎ NHỎ","ĐỆM + TAY CHỔI","KC CHỔI", "CHẤU ĐIỆN"]
    all_cam = [name]
    # all_cam = []

    if name == 'CAM 1':
        tt = 0
    if name == 'CAM 2':
        tt = 10
    if name == 'CAM 3':
        tt = 20
    if name == 'CAM 4':
        tt = 30
    if name == 'CAM 5':
        tt = 40
    if name == 'CAM 6':
        tt = 50
    if name == 'CAM 7':
        tt = 60
 
    for pp in range(len(phe_pham)):
        for al in range(len(all_cam)):
            if tt == var_plc:
                col2 = all_cam[al] + " " + phe_pham[pp]
            tt+=1
    # tt = 0
    # for pp in range(len(phe_pham)):
    #     # for al in range(len(all_cam)):
    #     if tt == var_plc:
    #         col2 = phe_pham[pp]
    #     tt+=2

    ws.append([col1, col2])

    for row in range(ws.max_row, ws.max_row+1):
        for col in range(1, ws.max_column+1):
            #print(col[row].value)
            d = ws.cell(row = row, column = col)
            #currentCell = ws.cell(col[row])
            d.alignment = Alignment(horizontal='center')
            #d.style.alignment.horizontal = 'center'
            d.font = Font(name= 'Calibri', size=12)
    
    wb.save(f"excel/pp/{mydate}_ppbydate.xlsx")
    try:
        shutil.copy(f"excel/pp/{mydate}_ppbydate.xlsx", f"C:/excel/pp/{mydate}_ppbydate.xlsx")
    except:
        pass


def logging(window,values,model,directory,plc_name,num_camera,value_plc):

    size = values[f'choose_size{num_camera}']
    conf = values[f'conf_thres{num_camera}']/100
    logging = 1

    if os.listdir(directory) == []:
        print('không có hình trong folder')
    else:
        for path in glob.glob(directory + '/*.jpg'):

            img_orgin = load_image(path,directory)              
            t1 = time.time()

            # write_plc(plc_name,value_plc_trigger,0)

            names, show = handle_image(img_orgin, model,size,conf,num_camera,values)

            myresult = handle_result(window, model,names,show, num_camera, plc_name,img_orgin, values)
            

            t2 = time.time() - t1
            time_cam = str(int(t2*1000)) + 'ms'

            window[f'time_cam{num_camera}'].update(value= time_cam, text_color='black') 
            imgbytes = cv2.imencode('.png',show)[1].tobytes()

            window[f'image{num_camera}'].update(data=imgbytes)
            _,values = window.read(timeout=20)


            if myresult == 0:
                logging = 0
                write_plc(plc_name,value_plc,2)
                print(f'DONE LOGGING CAM {num_camera}, ONE OF THEM IS OK')
                break

        if logging == 1:
            write_plc(plc_name,value_plc,1)
            print(f'DONE LOGGING CAM {num_camera}, ALL OF THEM IS NG')


# https://stackoverflow.com/questions/14243472/estimate-brightness-of-an-image-opencv/22020098#22020098
from numpy.linalg import norm
                
def brightness(img):
    if len(img.shape) == 3:
        return np.average(norm(img, axis=2)) / np.sqrt(3)
    else:
        return np.average(img)

# https://stackoverflow.com/questions/7765810/is-there-a-way-to-detect-if-an-image-is-blurry
import numpy
def laplace(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    return numpy.max(cv2.convertScaleAbs(cv2.Laplacian(gray, 3)))


def brightness_and_laplace(img,do_sang_min,do_sang_max, do_net_min,do_net_max):
    do_sang = int(brightness(img))
    do_sac_net = int(laplace(img))
    if do_sang_min < do_sang < do_sang_max:
        cv2.putText(img, f'Do sang: {do_sang}',(50,50),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),2)
    else:
        cv2.putText(img, f'Do sang: {do_sang}',(50,50),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),2)
    if do_net_min < do_sac_net < do_net_max:
        cv2.putText(img, f'Do sac net: {do_sac_net}',(50,100),cv2.FONT_HERSHEY_COMPLEX, 1,(0,255,0),2)
    else:
        cv2.putText(img, f'Do sac net: {do_sac_net}',(50,100),cv2.FONT_HERSHEY_COMPLEX, 1,(0,0,255),2)

    print('Độ sáng: ', do_sang)
    print('Độ sắc nét: ', do_sac_net)
    if do_sang_min < do_sang < do_sang_max and do_net_min < do_sac_net < do_net_max:
        return img, True
    return img, False


def kiem_guong(window,values,model,plc_name,kg_camera,my_callback,value_plc_trigger, value_plc_result,do_sang_min,do_sang_max, do_net_min,do_net_max,center_x_min,center_x_max,center_y_min,center_y_max):
    num_camera = 3

    
    if (read_plc(plc_name,value_plc_trigger) == 1) or (keyboard.is_pressed('shift+e') and kg_camera ==31) or (keyboard.is_pressed('shift+f') and kg_camera ==32):     
        if kg_camera ==31:
            num_camera = 1
        elif kg_camera ==32:
            num_camera = 2

        check_center = 0
        size = values[f'choose_size{num_camera}']
        conf = values[f'conf_thres{num_camera}']/100
        img_orgin = my_callback.image 
        if num_camera ==1:
            img_orgin = img_orgin[288:288+860,736:736+928]
        elif num_camera ==2:
            # img_orgin = img_orgin[412:412+788,832:832+880]
            img_orgin = img_orgin[412:412+788,730:730+880]
        write_plc(plc_name,value_plc_trigger,0)

        img_orgin = img_orgin.copy()           
        t1 = time.time()
        # if num_camera ==1:
        #     write_plc(plc_name,value_plc_trigger,0)
        img_orgin,sang_va_net = brightness_and_laplace(img_orgin,do_sang_min,do_sang_max, do_net_min,do_net_max)

        names, show,center_x, center_y = handle_image(img_orgin, model,size,conf,3,values,True,center_x_min,center_x_max,center_y_min,center_y_max)

        handle_result(window, model,names,show, kg_camera, "",img_orgin, values,False)
        if center_x_min < center_x < center_x_max and center_y_min < center_y < center_y_max:
            check_center = 1

        t2 = time.time() - t1
        time_cam = str(int(t2*1000)) + 'ms'

        # if num_camera ==3:
        #         num_camera =2

        window[f'time_cam{num_camera}'].update(value= time_cam, text_color='black') 

        print('a: ', sang_va_net, check_center)
        if sang_va_net and check_center:

            cv2.putText(show, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
            window[f'result_cam{num_camera}'].update(value= 'OK', text_color='green')

            write_plc(plc_name,value_plc_result,1)
        else:
            cv2.putText(show, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
            window[f'result_cam{num_camera}'].update(value= 'NG', text_color='red')

            write_plc(plc_name,value_plc_result,2)
        imgbytes = cv2.imencode('.png',show)[1].tobytes()
        window[f'image{num_camera}'].update(data= imgbytes)

        print('---------------------------------------------')
